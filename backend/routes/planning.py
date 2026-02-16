"""
Routes API pour le module Planning
==================================

STATUT: ACTIF
Ce module gère le planning des gardes : assignations, exports, rapports d'heures.

Routes Assignations:
- GET    /{tenant_slug}/planning/{semaine_debut}                - Obtenir planning d'une semaine
- GET    /{tenant_slug}/planning/assignations/{semaine_debut}   - Liste assignations semaine
- POST   /{tenant_slug}/planning/assignation                    - Créer assignation
- DELETE /{tenant_slug}/planning/assignation/{assignation_id}   - Supprimer assignation

Routes Rapports:
- GET    /{tenant_slug}/planning/mes-heures                     - Mes heures (employé)
- GET    /{tenant_slug}/planning/rapport-heures                 - Rapport heures global

Routes Outils:
- POST   /{tenant_slug}/planning/recalculer-durees-gardes       - Recalculer durées
- GET    /{tenant_slug}/planning/rapport-assignations-invalides - Rapport assignations invalides
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
from io import BytesIO
import uuid
import logging

# Import de la fonction de calcul d'équipe de garde
from routes.equipes_garde import get_equipe_garde_du_jour_sync
import json
import asyncio
import asyncio
import time

from routes.dependencies import (
    db,
    get_current_user,
    get_tenant_from_slug,
    clean_mongo_doc,
    User,
    creer_activite,
    creer_notification
)

# Import des helpers PDF partagés
from utils.pdf_helpers import (
    create_branded_pdf,
    get_modern_pdf_styles,
    create_pdf_footer_text
)

# Import de la classe ParametresRemplacements
from routes.remplacements import ParametresRemplacements

router = APIRouter(tags=["Planning"])
logger = logging.getLogger(__name__)

import os
import resend


def send_planning_notification_email(user_email: str, user_name: str, gardes_list: list, tenant_slug: str, periode: str, tenant_nom: str = None, stats: dict = None):
    """
    Envoie un email détaillé avec les gardes assignées pour le mois
    """
    resend_api_key = os.environ.get('RESEND_API_KEY')
    
    if not resend_api_key:
        logger.warning(f"RESEND_API_KEY non configurée - Email NON envoyé à {user_email}")
        return False
    
    try:
        sender_email = os.environ.get('SENDER_EMAIL', 'noreply@profiremanager.ca')
        caserne_nom = tenant_nom or tenant_slug.title()
        
        # URL de l'application (utiliser FRONTEND_URL ou construire depuis REACT_APP_BACKEND_URL)
        frontend_url = os.environ.get('FRONTEND_URL', os.environ.get('REACT_APP_BACKEND_URL', 'https://www.profiremanager.ca'))
        # Construire l'URL vers le planning avec le paramètre de page
        planning_url = f"{frontend_url}/{tenant_slug}?page=planning"
        
        # Extraire le mois de la période
        mois_noms = ["janvier", "février", "mars", "avril", "mai", "juin", 
                     "juillet", "août", "septembre", "octobre", "novembre", "décembre"]
        try:
            date_debut = datetime.strptime(periode.split(" au ")[0], "%Y-%m-%d")
            mois_texte = f"{mois_noms[date_debut.month - 1]} {date_debut.year}"
        except:
            mois_texte = periode
        
        subject = f"📅 Votre planning validé - {mois_texte}"
        
        # Utiliser les stats fournies ou calculer
        if stats is None:
            stats = {
                "par_type": {},
                "heures_internes": 0,
                "heures_externes": 0,
                "total_gardes": len(gardes_list)
            }
            for g in gardes_list:
                t = g.get('type_garde', 'Garde')
                stats["par_type"][t] = stats["par_type"].get(t, 0) + 1
                duree = g.get('duree_heures', 0) or 0
                if g.get('est_externe', False):
                    stats["heures_externes"] += duree
                else:
                    stats["heures_internes"] += duree
        
        nb_gardes = stats.get("total_gardes", len(gardes_list))
        heures_internes = stats.get("heures_internes", 0)
        heures_externes = stats.get("heures_externes", 0)
        total_heures = heures_internes + heures_externes
        par_type = stats.get("par_type", {})
        
        # Résumé par type de garde
        resume_types_html = ""
        for type_nom, count in par_type.items():
            resume_types_html += f"""
                <div style="display: inline-block; background: #f1f5f9; border-radius: 20px; padding: 8px 16px; margin: 4px; font-size: 14px;">
                    <strong style="color: #dc2626;">{count}</strong> <span style="color: #475569;">{type_nom}</span>
                </div>
            """
        
        # Bloc heures
        heures_html = ""
        if heures_internes > 0 or heures_externes > 0:
            heures_html = '<div style="display: flex; justify-content: center; gap: 30px; margin-top: 20px; flex-wrap: wrap;">'
            if heures_internes > 0:
                heures_html += f'''
                    <div style="text-align: center; background: #f0fdf4; border-radius: 12px; padding: 15px 25px;">
                        <span style="font-size: 28px; font-weight: bold; color: #16a34a;">{heures_internes:.1f}h</span>
                        <br>
                        <span style="color: #166534; font-size: 13px;">Heures internes</span>
                    </div>
                '''
            if heures_externes > 0:
                heures_html += f'''
                    <div style="text-align: center; background: #fef3c7; border-radius: 12px; padding: 15px 25px;">
                        <span style="font-size: 28px; font-weight: bold; color: #d97706;">{heures_externes:.1f}h</span>
                        <br>
                        <span style="color: #92400e; font-size: 13px;">Heures externes</span>
                    </div>
                '''
            if heures_internes > 0 and heures_externes > 0:
                heures_html += f'''
                    <div style="text-align: center; background: #eff6ff; border-radius: 12px; padding: 15px 25px;">
                        <span style="font-size: 28px; font-weight: bold; color: #2563eb;">{total_heures:.1f}h</span>
                        <br>
                        <span style="color: #1e40af; font-size: 13px;">Total</span>
                    </div>
                '''
            heures_html += "</div>"
        
        # Liste des gardes en HTML
        gardes_html = ''
        for garde in gardes_list:
            collegues_str = ', '.join(garde.get('collegues', [])) if garde.get('collegues') else 'Seul(e)'
            jour = garde.get('jour', '')
            horaire = garde.get('horaire', 'Horaire non défini')
            
            try:
                date_obj = datetime.strptime(garde['date'], "%Y-%m-%d")
                date_formatee = date_obj.strftime("%d/%m/%Y")
            except:
                date_formatee = garde['date']
            
            gardes_html += f"""
                <tr style="border-bottom: 1px solid #e5e7eb;">
                    <td style="padding: 12px; font-weight: 600; color: #1e293b;">
                        {jour}<br>
                        <span style="font-weight: normal; color: #64748b; font-size: 0.9rem;">{date_formatee}</span>
                    </td>
                    <td style="padding: 12px;">
                        <strong style="color: #dc2626;">{garde['type_garde']}</strong><br>
                        <span style="color: #64748b; font-size: 0.9rem;">{horaire}</span>
                    </td>
                    <td style="padding: 12px; color: #64748b; font-size: 0.9rem;">
                        {collegues_str}
                    </td>
                </tr>
            """
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
        </head>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333; max-width: 650px; margin: 0 auto; padding: 20px; background-color: #f3f4f6;">
            
            <div style="background: linear-gradient(135deg, #dc2626 0%, #991b1b 100%); color: white; padding: 30px; text-align: center; border-radius: 12px 12px 0 0;">
                <h1 style="margin: 0; font-size: 24px;">📅 Planning Validé</h1>
                <p style="margin: 10px 0 0 0; opacity: 0.9; font-size: 16px;">{mois_texte}</p>
            </div>
            
            <div style="background: white; padding: 30px; border-radius: 0 0 12px 12px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                
                <p style="font-size: 16px;">Bonjour <strong>{user_name}</strong>,</p>
                
                <p>Votre planning pour le mois de <strong>{mois_texte}</strong> a été validé par votre administrateur.</p>
                
                <div style="background: linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%); border-radius: 12px; padding: 25px; margin: 25px 0; text-align: center;">
                    <h3 style="color: #1e40af; margin: 0 0 20px 0;">📊 Récapitulatif</h3>
                    
                    <div style="margin-bottom: 20px;">
                        <span style="font-size: 42px; font-weight: bold; color: #dc2626;">{nb_gardes}</span>
                        <br>
                        <span style="color: #64748b; font-size: 14px;">garde(s) assignée(s)</span>
                    </div>
                    
                    <div style="margin-bottom: 15px;">
                        {resume_types_html}
                    </div>
                    
                    {heures_html}
                </div>
                
                <div style="background: #f0fdf4; border-left: 4px solid #22c55e; padding: 15px; margin: 20px 0; border-radius: 0 8px 8px 0;">
                    <strong style="color: #166534;">✅ Vos gardes pour {mois_texte} :</strong>
                </div>
                
                <table style="width: 100%; border-collapse: collapse; margin: 20px 0; background: #fafafa; border-radius: 8px; overflow: hidden;">
                    <thead>
                        <tr style="background: #f1f5f9;">
                            <th style="padding: 12px; text-align: left; color: #475569; font-weight: 600;">Jour</th>
                            <th style="padding: 12px; text-align: left; color: #475569; font-weight: 600;">Type de garde</th>
                            <th style="padding: 12px; text-align: left; color: #475569; font-weight: 600;">Collègues</th>
                        </tr>
                    </thead>
                    <tbody>
                        {gardes_html}
                    </tbody>
                </table>
                
                <div style="background: #fef3c7; border: 1px solid #fcd34d; border-radius: 8px; padding: 15px; margin: 25px 0;">
                    <h4 style="color: #92400e; margin: 0 0 10px 0;">📢 Rappels importants :</h4>
                    <ul style="color: #78350f; margin: 0; padding-left: 20px;">
                        <li>Ce planning a été validé par votre administrateur</li>
                        <li>Des ajustements peuvent survenir en cas de remplacements</li>
                        <li>Consultez régulièrement l'application pour les mises à jour</li>
                        <li>En cas d'absence imprévue, signalez-le immédiatement</li>
                    </ul>
                </div>
                
                <div style="text-align: center; margin: 30px 0;">
                    <a href="{planning_url}" 
                       style="background: #dc2626; color: white; padding: 14px 30px; text-decoration: none; border-radius: 8px; font-weight: bold; display: inline-block; font-size: 16px;">
                        Consulter mon planning
                    </a>
                </div>
                
                <p style="color: #64748b; margin-top: 30px;">
                    Cordialement,<br>
                    <strong>L'équipe {caserne_nom}</strong>
                </p>
            </div>
            
            <div style="text-align: center; padding: 20px; color: #9ca3af; font-size: 12px;">
                <p style="margin: 0;">Ceci est un message automatique de ProFireManager.</p>
                <p style="margin: 5px 0 0 0;">© {datetime.now().year} {caserne_nom}</p>
            </div>
        </body>
        </html>
        """
        
        resend.api_key = resend_api_key
        
        params = {
            "from": f"{caserne_nom} <{sender_email}>",
            "to": [user_email],
            "subject": subject,
            "html": html_content
        }
        
        response = resend.Emails.send(params)
        logger.info(f"✅ Email de planning envoyé à {user_email} (ID: {response.get('id', 'N/A')})")
        
        # Log email (synchrone via create_task)
        try:
            from routes.emails_history import log_email_sent
            import asyncio
            asyncio.create_task(log_email_sent(
                type_email="planning_gardes",
                destinataire_email=user_email,
                destinataire_nom=user_name,
                sujet=subject,
                tenant_slug=tenant_slug,
                metadata={"nb_gardes": len(gardes_list), "periode": periode}
            ))
        except Exception as log_err:
            logger.warning(f"Erreur log email: {log_err}")
        
        return True
            
    except Exception as e:
        logger.error(f"❌ Erreur envoi email planning à {user_email}: {str(e)}")
        return False


# ==================== SYSTÈME DE PROGRESSION TEMPS RÉEL ====================
# Dictionnaire global pour stocker les progressions des attributions auto
attribution_progress_store: Dict[str, Dict[str, Any]] = {}

class AttributionProgress:
    """Classe pour gérer la progression d'une attribution automatique"""
    
    def __init__(self, task_id: str):
        self.task_id = task_id
        self.start_time = time.time()
        self.current_step = ""
        self.progress_percentage = 0
        self.total_gardes = 0
        self.gardes_traitees = 0
        self.assignations_creees = 0
        self.status = "en_cours"  # en_cours, termine, erreur
        self.error_message = None
        self.expires_at = time.time() + 3600  # Expire après 1 heure
        
    def update(self, step: str, progress: int, gardes_traitees: int = 0, assignations: int = 0):
        """Met à jour la progression"""
        self.current_step = step
        self.progress_percentage = min(progress, 100)
        self.gardes_traitees = gardes_traitees
        if assignations > 0:
            self.assignations_creees = assignations
        attribution_progress_store[self.task_id] = self.to_dict()
    
    def complete(self, assignations_totales: int):
        """Marque la tâche comme terminée"""
        self.status = "termine"
        self.progress_percentage = 100
        self.assignations_creees = assignations_totales
        elapsed_time = time.time() - self.start_time
        self.current_step = f"✅ Terminé en {elapsed_time:.1f}s - {assignations_totales} assignations créées"
        attribution_progress_store[self.task_id] = self.to_dict()
    
    def error(self, message: str):
        """Marque la tâche en erreur"""
        self.status = "erreur"
        self.error_message = message
        self.current_step = f"❌ Erreur: {message}"
        attribution_progress_store[self.task_id] = self.to_dict()
    
    def to_dict(self):
        """Convertit en dictionnaire pour JSON"""
        elapsed = time.time() - self.start_time
        return {
            "task_id": self.task_id,
            "status": self.status,
            "current_step": self.current_step,
            "progress_percentage": self.progress_percentage,
            "total_gardes": self.total_gardes,
            "gardes_traitees": self.gardes_traitees,
            "assignations_creees": self.assignations_creees,
            "elapsed_time": f"{elapsed:.1f}s",
            "error_message": self.error_message
        }


async def progress_event_generator(task_id: str):
    """Générateur SSE pour streamer les mises à jour de progression"""
    try:
        # Attendre que la tâche soit créée
        for _ in range(50):  # Attendre max 5 secondes
            if task_id in attribution_progress_store:
                break
            await asyncio.sleep(0.1)
        
        # Streamer les mises à jour
        last_data = None
        while True:
            if task_id in attribution_progress_store:
                current_data = attribution_progress_store[task_id]
                
                # Envoyer seulement si les données ont changé
                if current_data != last_data:
                    yield f"data: {json.dumps(current_data)}\n\n"
                    last_data = current_data.copy()
                
                # Si terminé ou en erreur, arrêter le stream
                if current_data.get("status") in ["termine", "erreur"]:
                    break
            
            await asyncio.sleep(0.5)  # Mise à jour toutes les 500ms
            
    except asyncio.CancelledError:
        pass


# ==================== MODÈLES ====================

class Planning(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    semaine_debut: str  # Format: YYYY-MM-DD
    semaine_fin: str
    assignations: Dict[str, Any] = {}  # jour -> type_garde -> assignation
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class Assignation(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    user_id: str
    type_garde_id: str
    date: str
    statut: str = "planifie"  # planifie, confirme, remplacement_demande
    assignation_type: str = "auto"  # auto, manuel, manuel_avance
    justification: Optional[Dict[str, Any]] = None
    notes_admin: Optional[str] = None
    justification_historique: Optional[List[Dict[str, Any]]] = None


class AssignationCreate(BaseModel):
    user_id: str
    type_garde_id: str
    date: str
    assignation_type: str = "manuel"
    notes_admin: Optional[str] = None


# ==================== FONCTIONS UTILITAIRES ====================

def get_semaine_range(semaine_debut: str) -> tuple:
    """Retourne les dates de début et fin de semaine"""
    date_debut = datetime.strptime(semaine_debut, "%Y-%m-%d")
    date_fin = date_debut + timedelta(days=6)
    return date_debut, date_fin


async def get_types_garde(tenant_id: str) -> List[Dict]:
    """Récupère les types de garde d'un tenant"""
    types = await db.types_garde.find(
        {"tenant_id": tenant_id},
        {"_id": 0}
    ).sort("ordre", 1).to_list(100)
    return types


async def get_user_info(user_id: str) -> Optional[Dict]:
    """Récupère les informations d'un utilisateur"""
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "mot_de_passe_hash": 0})
    return user


# ==================== ROUTES PLANNING ====================

# NOTE: La route générique /{tenant_slug}/planning/{semaine_debut} 
# est définie à la FIN du fichier pour éviter les conflits avec les routes spécifiques


@router.get("/{tenant_slug}/planning/assignations/{date_debut}")
async def get_assignations_periode(
    tenant_slug: str,
    date_debut: str,
    mode: str = Query(default="semaine", description="Mode d'affichage: 'semaine' ou 'mois'"),
    current_user: User = Depends(get_current_user)
):
    """
    Récupérer les assignations pour une période donnée.
    - mode='semaine': Récupère 7 jours à partir de date_debut
    - mode='mois': Récupère tout le mois de la date fournie
    """
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if mode == "mois":
        # Parser la date et calculer le premier et dernier jour du mois
        date_obj = datetime.strptime(date_debut, "%Y-%m-%d")
        year = date_obj.year
        month = date_obj.month
        
        # Premier jour du mois
        debut_mois = datetime(year, month, 1)
        # Dernier jour du mois
        if month == 12:
            fin_mois = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            fin_mois = datetime(year, month + 1, 1) - timedelta(days=1)
        
        date_debut_str = debut_mois.strftime("%Y-%m-%d")
        date_fin_str = fin_mois.strftime("%Y-%m-%d")
    else:
        # Mode semaine (par défaut): 7 jours à partir de la date
        date_obj, date_fin_obj = get_semaine_range(date_debut)
        date_debut_str = date_debut
        date_fin_str = date_fin_obj.strftime("%Y-%m-%d")
    
    assignations = await db.assignations.find({
        "tenant_id": tenant.id,
        "date": {
            "$gte": date_debut_str,
            "$lte": date_fin_str
        }
    }, {"_id": 0}).sort("date", 1).to_list(1000)
    
    return assignations


@router.post("/{tenant_slug}/planning/assignation")
async def create_assignation(
    tenant_slug: str,
    assignation_data: AssignationCreate,
    current_user: User = Depends(get_current_user)
):
    """Créer une nouvelle assignation"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Vérifier que l'utilisateur existe
    user = await db.users.find_one({"id": assignation_data.user_id, "tenant_id": tenant.id})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    # Vérifier que le type de garde existe
    type_garde = await db.types_garde.find_one({"id": assignation_data.type_garde_id, "tenant_id": tenant.id})
    if not type_garde:
        raise HTTPException(status_code=404, detail="Type de garde non trouvé")
    
    # Vérifier s'il n'y a pas déjà une assignation pour ce user/jour/type (évite les doublons)
    existing = await db.assignations.find_one({
        "tenant_id": tenant.id,
        "user_id": assignation_data.user_id,
        "date": assignation_data.date,
        "type_garde_id": assignation_data.type_garde_id
    })
    
    if existing:
        # Mettre à jour l'assignation existante pour ce même utilisateur
        await db.assignations.update_one(
            {"id": existing["id"]},
            {"$set": {
                "assignation_type": assignation_data.assignation_type,
                "notes_admin": assignation_data.notes_admin,
                "updated_at": datetime.now(timezone.utc)
            }}
        )
        updated = await db.assignations.find_one({"id": existing["id"]}, {"_id": 0})
        return updated
    
    # Compter le nombre d'assignations actuelles pour cette garde/date
    current_count = await db.assignations.count_documents({
        "tenant_id": tenant.id,
        "date": assignation_data.date,
        "type_garde_id": assignation_data.type_garde_id
    })
    
    personnel_requis = type_garde.get("personnel_requis", 1)
    warning_message = None
    
    # Vérifier si on va dépasser le personnel requis
    if current_count >= personnel_requis:
        warning_message = f"Cette garde nécessite {personnel_requis} personne(s), vous en avez maintenant {current_count + 1}."
    
    # Créer une nouvelle assignation (permet plusieurs personnes sur le même créneau)
    assignation = Assignation(
        tenant_id=tenant.id,
        user_id=assignation_data.user_id,
        type_garde_id=assignation_data.type_garde_id,
        date=assignation_data.date,
        assignation_type=assignation_data.assignation_type,
        notes_admin=assignation_data.notes_admin
    )
    
    await db.assignations.insert_one(assignation.dict())
    
    # Créer une notification pour l'employé assigné
    await creer_notification(
        tenant_id=tenant.id,
        user_id=assignation_data.user_id,
        type_notification="planning_assignation",
        titre="Nouvelle assignation",
        message=f"Vous avez été assigné(e) le {assignation_data.date} - {type_garde.get('nom', 'Garde')}",
        lien=f"/planning?date={assignation_data.date}"
    )
    
    # Créer un log d'activité
    await creer_activite(
        tenant_id=tenant.id,
        type_activite="planning",
        description=f"Nouvelle assignation créée pour {user.get('prenom', '')} {user.get('nom', '')}",
        user_id=current_user.id,
        user_nom=f"{current_user.prenom} {current_user.nom}",
        data={"assignation_id": assignation.id, "date": assignation_data.date}
    )
    
    # Retourner l'assignation avec un éventuel avertissement
    result = clean_mongo_doc(assignation.dict())
    if warning_message:
        result["warning"] = warning_message
    
    return result


@router.delete("/{tenant_slug}/planning/assignation/{assignation_id}")
async def delete_assignation(
    tenant_slug: str,
    assignation_id: str,
    current_user: User = Depends(get_current_user)
):
    """Supprimer une assignation"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    assignation = await db.assignations.find_one({
        "id": assignation_id,
        "tenant_id": tenant.id
    })
    
    if not assignation:
        raise HTTPException(status_code=404, detail="Assignation non trouvée")
    
    # Notifier l'employé de la suppression
    await creer_notification(
        tenant_id=tenant.id,
        user_id=assignation.get("user_id"),
        type_notification="planning_suppression",
        titre="Assignation annulée",
        message=f"Votre assignation du {assignation.get('date', '')} a été annulée",
        lien="/planning"
    )
    
    await db.assignations.delete_one({"id": assignation_id})
    
    return {"message": "Assignation supprimée avec succès"}


# ==================== ROUTES RAPPORTS D'HEURES ====================

@router.get("/{tenant_slug}/planning/mes-heures")
async def get_mes_heures(
    tenant_slug: str,
    mois: Optional[str] = None,  # Format: YYYY-MM
    current_user: User = Depends(get_current_user)
):
    """
    Récupérer mes heures pour un mois donné (pour l'employé connecté)
    """
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Si pas de mois spécifié, prendre le mois courant
    if not mois:
        mois = datetime.now(timezone.utc).strftime("%Y-%m")
    
    # Calculer les dates du mois
    year, month = map(int, mois.split('-'))
    date_debut = datetime(year, month, 1, tzinfo=timezone.utc)
    if month == 12:
        date_fin = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
    else:
        date_fin = datetime(year, month + 1, 1, tzinfo=timezone.utc)
    
    # Récupérer les assignations de l'utilisateur pour ce mois
    assignations = await db.assignations.find({
        "tenant_id": tenant.id,
        "user_id": current_user.id,
        "date": {
            "$gte": date_debut.strftime("%Y-%m-%d"),
            "$lt": date_fin.strftime("%Y-%m-%d")
        }
    }, {"_id": 0}).to_list(100)
    
    # Récupérer les types de garde pour calculer les heures
    types_garde = await get_types_garde(tenant.id)
    types_garde_dict = {t["id"]: t for t in types_garde}
    
    total_heures = 0
    detail_par_type = {}
    
    for assignation in assignations:
        type_garde_id = assignation.get("type_garde_id")
        if type_garde_id in types_garde_dict:
            tg = types_garde_dict[type_garde_id]
            duree = tg.get("duree_heures", 0)
            total_heures += duree
            
            if type_garde_id not in detail_par_type:
                detail_par_type[type_garde_id] = {
                    "nom": tg.get("nom", ""),
                    "couleur": tg.get("couleur", "#3B82F6"),
                    "heures": 0,
                    "count": 0
                }
            
            detail_par_type[type_garde_id]["heures"] += duree
            detail_par_type[type_garde_id]["count"] += 1
    
    return {
        "mois": mois,
        "total_heures": total_heures,
        "nombre_gardes": len(assignations),
        "detail_par_type": list(detail_par_type.values()),
        "assignations": assignations
    }


@router.get("/{tenant_slug}/planning/rapport-heures")
async def get_rapport_heures(
    tenant_slug: str,
    mois: Optional[str] = None,  # Format: YYYY-MM
    date_debut: Optional[str] = None,
    date_fin: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    Rapport d'heures global pour tous les employés (admin/superviseur)
    
    Les heures sont catégorisées ainsi:
    - Heures INTERNES = gardes de type "interne" (est_garde_externe = False)
    - Heures EXTERNES = gardes de type "externe" (est_garde_externe = True)
    """
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Calculer les dates de la période
    if date_debut and date_fin:
        # Utiliser les dates fournies
        debut_str = date_debut
        fin_str = date_fin
    elif mois:
        # Calculer les dates du mois
        year, month = map(int, mois.split('-'))
        debut_str = f"{year}-{month:02d}-01"
        if month == 12:
            fin_str = f"{year + 1}-01-01"
        else:
            fin_str = f"{year}-{month + 1:02d}-01"
    else:
        # Mois courant par défaut
        now = datetime.now(timezone.utc)
        debut_str = f"{now.year}-{now.month:02d}-01"
        if now.month == 12:
            fin_str = f"{now.year + 1}-01-01"
        else:
            fin_str = f"{now.year}-{now.month + 1:02d}-01"
    
    # Récupérer toutes les assignations de la période
    assignations = await db.assignations.find({
        "tenant_id": tenant.id,
        "date": {
            "$gte": debut_str,
            "$lt": fin_str
        }
    }, {"_id": 0}).to_list(10000)
    
    # Récupérer les types de garde avec leur catégorisation interne/externe
    types_garde = await get_types_garde(tenant.id)
    types_garde_dict = {t["id"]: t for t in types_garde}
    
    # Récupérer tous les employés actifs (temps plein ET temps partiel)
    users = await db.users.find(
        {"tenant_id": tenant.id, "statut": "Actif"},
        {"_id": 0, "mot_de_passe_hash": 0}
    ).to_list(1000)
    users_dict = {u["id"]: u for u in users}
    
    # Calculer les heures par employé
    rapport = {}
    for assignation in assignations:
        user_id = assignation.get("user_id")
        type_garde_id = assignation.get("type_garde_id")
        
        if not user_id or user_id not in users_dict:
            continue
        
        if user_id not in rapport:
            user = users_dict.get(user_id, {})
            nom = user.get("nom", "")
            prenom = user.get("prenom", "")
            type_emploi = user.get("type_emploi", "")
            
            rapport[user_id] = {
                "user_id": user_id,
                "nom": nom,
                "prenom": prenom,
                "nom_complet": f"{prenom} {nom}".strip(),
                "grade": user.get("grade", ""),
                "type_emploi": type_emploi,
                "total_heures": 0,
                "heures_internes": 0,  # Heures de gardes INTERNES (pas externe)
                "heures_externes": 0,  # Heures de gardes EXTERNES
                "nombre_gardes": 0,
                "detail_par_type": {}
            }
        
        if type_garde_id in types_garde_dict:
            tg = types_garde_dict[type_garde_id]
            duree = tg.get("duree_heures", 0) or 0
            est_garde_externe = tg.get("est_garde_externe", False)
            
            rapport[user_id]["total_heures"] += duree
            rapport[user_id]["nombre_gardes"] += 1
            
            # Catégoriser selon le TYPE DE GARDE (pas le type d'emploi!)
            if est_garde_externe:
                rapport[user_id]["heures_externes"] += duree
            else:
                rapport[user_id]["heures_internes"] += duree
            
            if type_garde_id not in rapport[user_id]["detail_par_type"]:
                rapport[user_id]["detail_par_type"][type_garde_id] = {
                    "nom": tg.get("nom", ""),
                    "heures": 0,
                    "count": 0,
                    "est_externe": est_garde_externe
                }
            
            rapport[user_id]["detail_par_type"][type_garde_id]["heures"] += duree
            rapport[user_id]["detail_par_type"][type_garde_id]["count"] += 1
    
    # Ajouter les employés sans assignation (pour voir qui n'a pas d'heures)
    for user_id, user in users_dict.items():
        if user_id not in rapport:
            rapport[user_id] = {
                "user_id": user_id,
                "nom": user.get("nom", ""),
                "prenom": user.get("prenom", ""),
                "nom_complet": f"{user.get('prenom', '')} {user.get('nom', '')}".strip(),
                "grade": user.get("grade", ""),
                "type_emploi": user.get("type_emploi", ""),
                "total_heures": 0,
                "heures_internes": 0,
                "heures_externes": 0,
                "nombre_gardes": 0,
                "detail_par_type": {}
            }
    
    # Convertir en liste triée
    rapport_list = list(rapport.values())
    rapport_list.sort(key=lambda x: x["total_heures"], reverse=True)
    
    # Calculer les statistiques globales
    total_heures_planifiees = sum([r["total_heures"] for r in rapport_list])
    total_heures_internes = sum([r["heures_internes"] for r in rapport_list])
    total_heures_externes = sum([r["heures_externes"] for r in rapport_list])
    nb_employes = len(rapport_list)
    
    # Compter employés avec heures internes/externes
    nb_avec_heures_internes = len([r for r in rapport_list if r["heures_internes"] > 0])
    nb_avec_heures_externes = len([r for r in rapport_list if r["heures_externes"] > 0])
    
    return {
        "mois": mois if mois else f"{debut_str[:7]}",
        "date_debut": debut_str,
        "date_fin": fin_str,
        "total_assignations": len(assignations),
        "employes": rapport_list,
        "statistiques": {
            "total_heures_planifiees": round(total_heures_planifiees, 2),
            "total_heures_internes": round(total_heures_internes, 2),
            "total_heures_externes": round(total_heures_externes, 2),
            "moyenne_heures_internes": round(total_heures_internes / nb_avec_heures_internes, 2) if nb_avec_heures_internes > 0 else 0,
            "moyenne_heures_externes": round(total_heures_externes / nb_avec_heures_externes, 2) if nb_avec_heures_externes > 0 else 0,
            "nb_employes": nb_employes,
            "nb_avec_heures_internes": nb_avec_heures_internes,
            "nb_avec_heures_externes": nb_avec_heures_externes
        }
    }


# ==================== ROUTES OUTILS ====================

@router.get("/{tenant_slug}/planning/rapport-assignations-invalides")
async def get_assignations_invalides(
    tenant_slug: str,
    current_user: User = Depends(get_current_user)
):
    """
    Rapport des assignations invalides (employés inactifs, types de garde supprimés, etc.)
    """
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Récupérer toutes les assignations futures
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    assignations = await db.assignations.find({
        "tenant_id": tenant.id,
        "date": {"$gte": today}
    }, {"_id": 0}).to_list(10000)
    
    # Récupérer les employés actifs
    users_actifs = await db.users.find(
        {"tenant_id": tenant.id, "statut": "Actif"},
        {"_id": 0, "id": 1}
    ).to_list(1000)
    users_actifs_ids = set([u["id"] for u in users_actifs])
    
    # Récupérer les types de garde actifs
    types_garde = await db.types_garde.find(
        {"tenant_id": tenant.id},
        {"_id": 0, "id": 1}
    ).to_list(100)
    types_garde_ids = set([t["id"] for t in types_garde])
    
    invalides = []
    for assignation in assignations:
        problemes = []
        
        if assignation.get("user_id") not in users_actifs_ids:
            problemes.append("Employé inactif ou supprimé")
        
        if assignation.get("type_garde_id") not in types_garde_ids:
            problemes.append("Type de garde supprimé")
        
        if problemes:
            invalides.append({
                "assignation_id": assignation.get("id"),
                "date": assignation.get("date"),
                "user_id": assignation.get("user_id"),
                "type_garde_id": assignation.get("type_garde_id"),
                "problemes": problemes
            })
    
    return {
        "total_invalides": len(invalides),
        "assignations_invalides": invalides
    }


@router.post("/{tenant_slug}/planning/recalculer-durees-gardes")
async def recalculer_durees_gardes(
    tenant_slug: str,
    current_user: User = Depends(get_current_user)
):
    """
    Recalculer les durées de toutes les gardes selon les types de garde actuels
    """
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Récupérer les types de garde
    types_garde = await get_types_garde(tenant.id)
    types_garde_dict = {t["id"]: t for t in types_garde}
    
    # Récupérer toutes les assignations
    assignations = await db.assignations.find(
        {"tenant_id": tenant.id},
        {"_id": 0}
    ).to_list(100000)
    
    updated_count = 0
    for assignation in assignations:
        type_garde_id = assignation.get("type_garde_id")
        if type_garde_id in types_garde_dict:
            tg = types_garde_dict[type_garde_id]
            duree = tg.get("duree_heures", 0)
            
            # Mettre à jour si différent
            if assignation.get("duree_heures") != duree:
                await db.assignations.update_one(
                    {"id": assignation["id"]},
                    {"$set": {"duree_heures": duree}}
                )
                updated_count += 1
    
    return {
        "message": "Recalcul des durées terminé",
        "total_assignations": len(assignations),
        "mises_a_jour": updated_count
    }


# ==================== ROUTE GÉNÉRIQUE (DOIT ÊTRE À LA FIN) ====================
# Cette route capture tout ce qui n'a pas été capturé par les routes spécifiques ci-dessus

@router.get("/{tenant_slug}/planning/{semaine_debut}")
async def get_planning_semaine(
    tenant_slug: str,
    semaine_debut: str,
    current_user: User = Depends(get_current_user)
):
    """
    Récupérer le planning d'une semaine complète
    Retourne les assignations avec les informations des employés et types de garde
    
    IMPORTANT: Cette route DOIT être définie en DERNIER car {semaine_debut} 
    est un paramètre générique qui capturerait sinon les autres routes.
    """
    tenant = await get_tenant_from_slug(tenant_slug)
    
    date_debut, date_fin = get_semaine_range(semaine_debut)
    semaine_fin = date_fin.strftime("%Y-%m-%d")
    
    # Récupérer les assignations de la semaine
    assignations = await db.assignations.find({
        "tenant_id": tenant.id,
        "date": {
            "$gte": semaine_debut,
            "$lte": semaine_fin
        }
    }, {"_id": 0}).to_list(1000)
    
    # Récupérer les types de garde
    types_garde = await get_types_garde(tenant.id)
    types_garde_dict = {t["id"]: t for t in types_garde}
    
    # Récupérer les infos des employés assignés
    user_ids = list(set([a["user_id"] for a in assignations if a.get("user_id")]))
    users = await db.users.find(
        {"id": {"$in": user_ids}},
        {"_id": 0, "mot_de_passe_hash": 0}
    ).to_list(1000)
    users_dict = {u["id"]: u for u in users}
    
    # Enrichir les assignations
    for assignation in assignations:
        user_id = assignation.get("user_id")
        type_garde_id = assignation.get("type_garde_id")
        
        if user_id and user_id in users_dict:
            user = users_dict[user_id]
            assignation["user_nom"] = f"{user.get('prenom', '')} {user.get('nom', '')}"
            assignation["user_grade"] = user.get("grade", "")
        
        if type_garde_id and type_garde_id in types_garde_dict:
            tg = types_garde_dict[type_garde_id]
            assignation["type_garde_nom"] = tg.get("nom", "")
            assignation["type_garde_couleur"] = tg.get("couleur", "#3B82F6")
    
    # Construire l'objet Planning
    planning_obj = Planning(semaine_debut=semaine_debut, semaine_fin=semaine_fin)
    
    return {
        "id": planning_obj.id,
        "semaine_debut": semaine_debut,
        "semaine_fin": semaine_fin,
        "assignations": assignations,
        "types_garde": types_garde
    }


# ==================== ROUTES PLANNING AVANCÉES (MIGRÉES DE SERVER.PY) ====================


# DELETE formater-mois
@router.delete("/{tenant_slug}/planning/formater-mois")
async def formater_planning_mois(
    tenant_slug: str,
    mois: str,  # Format: YYYY-MM
    current_user: User = Depends(get_current_user)
):
    """
    Formate (vide) le planning d'un mois spécifique
    UNIQUEMENT pour le tenant demo
    Supprime: assignations, demandes de remplacement
    """
    # 1. Vérifier que c'est le tenant demo
    if tenant_slug != "demo":
        raise HTTPException(status_code=403, detail="Cette fonctionnalité est réservée au tenant demo")
    
    # 2. Vérifier que l'utilisateur est admin
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # 3. Valider le format du mois
    try:
        year, month = map(int, mois.split('-'))
        if month < 1 or month > 12:
            raise ValueError()
    except:
        raise HTTPException(status_code=400, detail="Format de mois invalide. Utilisez YYYY-MM")
    
    # 4. Calculer les dates de début et fin du mois
    date_debut = datetime(year, month, 1, tzinfo=timezone.utc)
    if month == 12:
        date_fin = datetime(year + 1, 1, 1, tzinfo=timezone.utc) - timedelta(seconds=1)
    else:
        date_fin = datetime(year, month + 1, 1, tzinfo=timezone.utc) - timedelta(seconds=1)
    
    # 5. Supprimer les assignations du mois
    result_assignations = await db.assignations.delete_many({
        "tenant_id": tenant.id,
        "date": {
            "$gte": date_debut.isoformat(),
            "$lte": date_fin.isoformat()
        }
    })
    
    # 6. Supprimer les demandes de remplacement du mois
    result_remplacements = await db.demandes_remplacement.delete_many({
        "tenant_id": tenant.id,
        "date_garde": {
            "$gte": date_debut.isoformat(),
            "$lte": date_fin.isoformat()
        }
    })
    
    return {
        "message": f"Planning formaté avec succès pour {mois}",
        "mois": mois,
        "assignations_supprimees": result_assignations.deleted_count,
        "demandes_supprimees": result_remplacements.deleted_count
    }


# GET export-pdf
@router.get("/{tenant_slug}/planning/export-pdf")
async def export_planning_pdf(
    tenant_slug: str, 
    periode: str,
    type: str,
    current_user: User = Depends(get_current_user)
):
    """Export du planning en PDF"""
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        from io import BytesIO
        
        tenant = await get_tenant_from_slug(tenant_slug)
        
        # Calculer la période
        if type == 'semaine':
            date_debut = datetime.strptime(periode, '%Y-%m-%d')
            date_fin = date_debut + timedelta(days=6)
        else:  # mois
            year, month = map(int, periode.split('-'))
            date_debut = datetime(year, month, 1)
            if month == 12:
                date_fin = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                date_fin = datetime(year, month + 1, 1) - timedelta(days=1)
        
        # Récupérer les données
        assignations_list = await db.assignations.find({
            "tenant_id": tenant.id,
            "date": {
                "$gte": date_debut.strftime('%Y-%m-%d'),
                "$lte": date_fin.strftime('%Y-%m-%d')
            }
        }).to_list(length=None)
        
        types_garde_list = await db.types_garde.find({"tenant_id": tenant.id}).to_list(length=None)
        users_list = await db.users.find({"tenant_id": tenant.id}).to_list(length=None)
        
        types_map = {t['id']: t for t in types_garde_list}
        users_map = {u['id']: u for u in users_list}
        
        # Créer le PDF avec branding
        buffer, doc, elements = create_branded_pdf(
            tenant,
            pagesize=landscape(letter),
            leftMargin=0.5*inch,
            rightMargin=0.5*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch
        )
        styles = getSampleStyleSheet()
        modern_styles = get_modern_pdf_styles(styles)
        
        # Titre principal
        titre = f"PLANNING DES GARDES - V4"
        elements.append(Paragraph(titre, modern_styles['title']))
        
        # Sous-titre avec période
        type_label = "Semaine" if type == "semaine" else "Mois"
        periode_str = f"{type_label} du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        elements.append(Paragraph(periode_str, modern_styles['subheading']))
        elements.append(Spacer(1, 0.1*inch))
        
        # Ligne de séparation
        from reportlab.platypus import HRFlowable
        elements.append(HRFlowable(width="100%", thickness=1, color=modern_styles['grid'], spaceAfter=0.3*inch))
        
        # Jours français
        jours_fr = ['Lun', 'Mar', 'Mer', 'Jeu', 'Ven', 'Sam', 'Dim']
        
        # Couleurs pour la grille
        HEADER_BG = colors.HexColor('#1F2937')
        PRIMARY_RED = colors.HexColor('#DC2626')
        COMPLETE_GREEN = colors.HexColor('#D1FAE5')
        PARTIAL_YELLOW = colors.HexColor('#FEF3C7')
        VACANT_RED = colors.HexColor('#FEE2E2')
        LIGHT_GRAY = colors.HexColor('#F3F4F6')
        BORDER_COLOR = colors.HexColor('#E5E7EB')
        
        # Trier les types de garde par heure
        types_garde_sorted = sorted(types_garde_list, key=lambda x: x.get('heure_debut', '00:00'))
        
        if type == 'semaine':
            # ===== FORMAT GRILLE SEMAINE =====
            from reportlab.lib.enums import TA_LEFT
            
            # Styles pour les cellules - taille augmentée et leading adapté
            header_style_white = ParagraphStyle('HeaderWhite', fontSize=9, alignment=TA_CENTER, textColor=colors.white, leading=12, wordWrap='CJK')
            garde_cell_style = ParagraphStyle('GardeCell', fontSize=8, alignment=TA_LEFT, textColor=colors.white, leading=11, wordWrap='CJK')
            day_cell_style = ParagraphStyle('DayCell', fontSize=8, alignment=TA_CENTER, leading=11, textColor=colors.HexColor('#1F2937'), wordWrap='CJK')
            
            # Debug: logger les assignations
            logging.warning(f"DEBUG PDF: {len(assignations_list)} assignations trouvées pour {date_debut.strftime('%Y-%m-%d')} à {date_fin.strftime('%Y-%m-%d')}")
            logging.warning(f"DEBUG PDF: {len(types_garde_sorted)} types de garde")
            
            # En-tête : Type de garde + 7 jours
            header_row = [Paragraph("<b>Type de garde</b>", header_style_white)]
            for i in range(7):
                d = date_debut + timedelta(days=i)
                header_row.append(Paragraph(f"<b>{jours_fr[d.weekday()]}</b><br/>{d.strftime('%d/%m')}", header_style_white))
            
            table_data = [header_row]
            cell_colors = []
            
            # Une ligne par type de garde
            for type_garde in types_garde_sorted:
                garde_nom = type_garde.get('nom', 'N/A')
                heure_debut_garde = type_garde.get('heure_debut', '??:??')
                heure_fin_garde = type_garde.get('heure_fin', '??:??')
                personnel_requis = type_garde.get('personnel_requis', 1)
                jours_app = type_garde.get('jours_application', [])
                type_garde_id = type_garde.get('id')
                
                # Première colonne - nom complet sans troncature
                row = [Paragraph(f"<b>{garde_nom}</b><br/><font size='7'>{heure_debut_garde}-{heure_fin_garde}</font>", garde_cell_style)]
                row_colors = [PRIMARY_RED]
                
                for i in range(7):
                    d = date_debut + timedelta(days=i)
                    date_str = d.strftime('%Y-%m-%d')
                    day_name = d.strftime('%A').lower()
                    
                    # Vérifier si applicable ce jour
                    if jours_app and day_name not in jours_app:
                        row.append(Paragraph("-", day_cell_style))
                        row_colors.append(LIGHT_GRAY)
                        continue
                    
                    # Trouver les assignations pour ce jour et ce type de garde
                    assignations_jour = [a for a in assignations_list 
                                        if a.get('date') == date_str and a.get('type_garde_id') == type_garde_id]
                    
                    # Construire les noms des pompiers
                    noms = []
                    for a in assignations_jour:
                        user_id = a.get('user_id')
                        if user_id and user_id in users_map:
                            u = users_map[user_id]
                            prenom = u.get('prenom', '')
                            nom = u.get('nom', '')
                            if prenom or nom:
                                noms.append(f"{prenom[:1]}. {nom}")
                    
                    if noms:
                        cell_text = "<br/>".join(noms[:4])
                        if len(noms) > 4:
                            cell_text += f"<br/><font size='6'>+{len(noms)-4}</font>"
                        row.append(Paragraph(cell_text, day_cell_style))
                        
                        if len(noms) >= personnel_requis:
                            row_colors.append(COMPLETE_GREEN)
                        else:
                            row_colors.append(PARTIAL_YELLOW)
                    else:
                        row.append(Paragraph("<font color='#B91C1C'>Vacant</font>", day_cell_style))
                        row_colors.append(VACANT_RED)
                
                table_data.append(row)
                cell_colors.append(row_colors)
            
            # Largeurs de colonnes - première colonne plus large pour les noms de garde
            page_width = landscape(letter)[0]
            available_width = page_width - 1*inch
            first_col = 1.8*inch
            day_col = (available_width - first_col) / 7
            col_widths = [first_col] + [day_col] * 7
            
            # Hauteur minimale des lignes pour permettre l'affichage des noms
            row_heights = [0.5*inch] + [0.7*inch] * (len(table_data) - 1)
            
            # Créer la table avec hauteurs de lignes
            table = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
            
            # Style de base - WORDWRAP activé pour permettre le retour à la ligne
            style_commands = [
                ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ]
            
            # Appliquer les couleurs par cellule
            for row_idx, colors_row in enumerate(cell_colors, start=1):
                for col_idx, bg_color in enumerate(colors_row):
                    style_commands.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), bg_color))
                    if col_idx == 0:
                        style_commands.append(('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), colors.white))
                        style_commands.append(('ALIGN', (col_idx, row_idx), (col_idx, row_idx), 'LEFT'))
                        style_commands.append(('LEFTPADDING', (col_idx, row_idx), (col_idx, row_idx), 8))
            
            table.setStyle(TableStyle(style_commands))
            elements.append(table)
            
            # Légende
            elements.append(Spacer(1, 0.3*inch))
            legend_style = ParagraphStyle('Legend', fontSize=9, alignment=TA_CENTER, textColor=colors.HexColor('#6B7280'))
            elements.append(Paragraph("Legende: Vert = Complet | Jaune = Partiel | Rouge = Vacant | Gris = Non applicable", legend_style))
            
        elif type == 'mois':
            # ===== FORMAT GRILLE MOIS - Une grille par semaine =====
            from reportlab.platypus import PageBreak
            
            current = date_debut
            semaine_num = 1
            page_width = landscape(letter)[0]
            
            # Style pour titre de semaine
            semaine_style = ParagraphStyle(
                'SemaineStyle',
                parent=styles['Heading2'],
                fontSize=12,
                textColor=PRIMARY_RED,
                spaceBefore=10,
                spaceAfter=10,
                fontName='Helvetica-Bold'
            )
            
            while current <= date_fin:
                fin_semaine = min(current + timedelta(days=6), date_fin)
                nb_jours = (fin_semaine - current).days + 1
                
                # Titre de la semaine
                elements.append(Paragraph(
                    f"Semaine {semaine_num} - Du {current.strftime('%d/%m')} au {fin_semaine.strftime('%d/%m/%Y')}",
                    semaine_style
                ))
                
                # Styles pour le format mois - avec retour à la ligne
                from reportlab.lib.enums import TA_LEFT
                header_style_mois = ParagraphStyle('HeaderMois', fontSize=8, alignment=TA_CENTER, textColor=colors.white, leading=10, wordWrap='CJK')
                garde_cell_style_mois = ParagraphStyle('GardeCellMois', fontSize=7, alignment=TA_LEFT, textColor=colors.white, leading=9, wordWrap='CJK')
                day_cell_style_mois = ParagraphStyle('DayCellMois', fontSize=7, alignment=TA_CENTER, leading=9, textColor=colors.HexColor('#1F2937'), wordWrap='CJK')
                
                # En-tête avec des Paragraph pour le retour à la ligne
                header_row = [Paragraph("<b>Type de garde</b>", header_style_mois)]
                for i in range(nb_jours):
                    d = current + timedelta(days=i)
                    header_row.append(Paragraph(f"<b>{jours_fr[d.weekday()]}</b><br/>{d.strftime('%d')}", header_style_mois))
                
                table_data = [header_row]
                cell_colors_mois = []
                
                for type_garde in types_garde_sorted:
                    garde_nom = type_garde.get('nom', 'N/A')  # NE PLUS TRONQUER
                    heure_debut_garde = type_garde.get('heure_debut', '')
                    heure_fin_garde = type_garde.get('heure_fin', '')
                    personnel_requis = type_garde.get('personnel_requis', 1)
                    jours_app = type_garde.get('jours_application', [])
                    type_garde_id = type_garde.get('id')
                    
                    # Première colonne avec nom complet et horaires
                    row = [Paragraph(f"<b>{garde_nom}</b><br/><font size='6'>{heure_debut_garde}-{heure_fin_garde}</font>", garde_cell_style_mois)]
                    row_colors = [PRIMARY_RED]
                    
                    for i in range(nb_jours):
                        d = current + timedelta(days=i)
                        date_str = d.strftime('%Y-%m-%d')
                        day_name = d.strftime('%A').lower()
                        
                        if jours_app and day_name not in jours_app:
                            row.append(Paragraph("-", day_cell_style_mois))
                            row_colors.append(LIGHT_GRAY)
                            continue
                        
                        assignations_jour = [a for a in assignations_list 
                                            if a['date'] == date_str and a['type_garde_id'] == type_garde_id]
                        
                        # Construire les noms des pompiers
                        noms = []
                        for a in assignations_jour:
                            user_id = a.get('user_id')
                            if user_id and user_id in users_map:
                                u = users_map[user_id]
                                prenom = u.get('prenom', '')
                                nom = u.get('nom', '')
                                if prenom or nom:
                                    noms.append(f"{prenom[:1]}. {nom}")
                        
                        if noms:
                            cell_text = "<br/>".join(noms[:3])  # Max 3 noms pour le mois
                            if len(noms) > 3:
                                cell_text += f"<br/><font size='5'>+{len(noms)-3}</font>"
                            row.append(Paragraph(cell_text, day_cell_style_mois))
                            
                            if len(noms) >= personnel_requis:
                                row_colors.append(COMPLETE_GREEN)
                            else:
                                row_colors.append(PARTIAL_YELLOW)
                        else:
                            row.append(Paragraph("<font color='#B91C1C'>Vacant</font>", day_cell_style_mois))
                            row_colors.append(VACANT_RED)
                    
                    table_data.append(row)
                    cell_colors_mois.append(row_colors)
                
                # Largeurs - première colonne plus large pour afficher le nom complet
                available_width = page_width - 1*inch
                first_col = 1.6*inch  # Augmenté pour les noms de garde complets
                day_col = (available_width - first_col) / nb_jours
                col_widths = [first_col] + [day_col] * nb_jours
                
                # Hauteurs de lignes - plus hautes pour afficher les noms des pompiers
                row_heights = [0.4*inch] + [0.6*inch] * (len(table_data) - 1)
                
                table = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
                
                style_commands = [
                    ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('GRID', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ('LEFTPADDING', (0, 0), (-1, -1), 3),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ]
                
                for row_idx, colors_row in enumerate(cell_colors_mois, start=1):
                    for col_idx, bg_color in enumerate(colors_row):
                        style_commands.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), bg_color))
                        if col_idx == 0:
                            style_commands.append(('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), colors.white))
                            style_commands.append(('ALIGN', (col_idx, row_idx), (col_idx, row_idx), 'LEFT'))
                            style_commands.append(('LEFTPADDING', (col_idx, row_idx), (col_idx, row_idx), 6))
                
                table.setStyle(TableStyle(style_commands))
                elements.append(table)
                elements.append(Spacer(1, 0.2*inch))
                
                current = fin_semaine + timedelta(days=1)
                semaine_num += 1
                
                if current <= date_fin:
                    elements.append(PageBreak())
        
        # Footer
        def add_footer(canvas, doc_obj):
            canvas.saveState()
            canvas.setStrokeColor(colors.HexColor('#e2e8f0'))
            canvas.setLineWidth(1)
            canvas.line(0.5*inch, 0.5*inch, landscape(letter)[0] - 0.5*inch, 0.5*inch)
            canvas.setFont('Helvetica', 9)
            canvas.setFillColor(colors.HexColor('#64748b'))
            footer_text = f"ProFireManager - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            canvas.drawCentredString(landscape(letter)[0] / 2, 0.35*inch, footer_text)
            canvas.setFont('Helvetica', 8)
            canvas.drawRightString(landscape(letter)[0] - 0.5*inch, 0.35*inch, f"Page {doc_obj.page}")
            canvas.restoreState()
        
        doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=planning_{type}_{periode}.pdf"}
        )
        
    except Exception as e:
        import traceback
        logging.error(f"Erreur export PDF: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Erreur export PDF: {str(e)}")



# GET export-excel
@router.get("/{tenant_slug}/planning/export-excel")
async def export_planning_excel(
    tenant_slug: str, 
    periode: str,
    type: str,
    current_user: User = Depends(get_current_user)
):
    """Export du planning en Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO
        
        tenant = await get_tenant_from_slug(tenant_slug)
        
        # Calculer la période
        if type == 'semaine':
            date_debut = datetime.strptime(periode, '%Y-%m-%d')
            date_fin = date_debut + timedelta(days=6)
        else:
            year, month = map(int, periode.split('-'))
            date_debut = datetime(year, month, 1)
            if month == 12:
                date_fin = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                date_fin = datetime(year, month + 1, 1) - timedelta(days=1)
        
        assignations_list = await db.assignations.find({
            "tenant_id": tenant.id,
            "date": {
                "$gte": date_debut.strftime('%Y-%m-%d'),
                "$lte": date_fin.strftime('%Y-%m-%d')
            }
        }).to_list(length=None)
        
        types_garde_list = await db.types_garde.find({"tenant_id": tenant.id}).to_list(length=None)
        users_list = await db.users.find({"tenant_id": tenant.id}).to_list(length=None)
        
        types_map = {t['id']: t for t in types_garde_list}
        users_map = {u['id']: u for u in users_list}
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Planning {type}"
        
        header_fill = PatternFill(start_color="FCA5A5", end_color="FCA5A5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws.merge_cells('A1:H1')
        ws['A1'] = f"Planning des Gardes - {type.capitalize()}"
        ws['A1'].font = Font(bold=True, size=16, color="EF4444")
        ws['A1'].alignment = center_alignment
        
        ws.merge_cells('A2:H2')
        ws['A2'] = f"Du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        ws['A2'].alignment = center_alignment
        
        row = 4
        if type == 'semaine':
            headers = ['Type de Garde', 'Horaires'] + [(date_debut + timedelta(days=i)).strftime('%a %d/%m') for i in range(7)]
        else:
            headers = ['Date', 'Jour', 'Type de Garde', 'Horaires', 'Personnel', 'Requis', 'Assignés', 'Statut']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        row += 1
        
        if type == 'semaine':
            for type_garde in sorted(types_garde_list, key=lambda x: x.get('heure_debut', '')):
                ws.cell(row=row, column=1, value=type_garde['nom'])
                ws.cell(row=row, column=2, value=f"{type_garde.get('heure_debut', '')} - {type_garde.get('heure_fin', '')}")
                
                for i in range(7):
                    current_date = (date_debut + timedelta(days=i)).strftime('%Y-%m-%d')
                    assignations_jour = [a for a in assignations_list if a['date'] == current_date and a['type_garde_id'] == type_garde['id']]
                    
                    noms = [f"{users_map[a['user_id']]['prenom']} {users_map[a['user_id']]['nom']}" 
                           for a in assignations_jour if a['user_id'] in users_map]
                    
                    cell_text = '\n'.join(noms) if noms else 'Vacant'
                    cell = ws.cell(row=row, column=3+i, value=cell_text)
                    cell.alignment = center_alignment
                    cell.border = border
                    
                    if len(noms) >= type_garde.get('personnel_requis', 1):
                        cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    elif noms:
                        cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                
                row += 1
        else:
            current = date_debut
            while current <= date_fin:
                date_str = current.strftime('%Y-%m-%d')
                jour_fr = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche'][current.weekday()]
                
                for type_garde in types_garde_list:
                    assignations_jour = [a for a in assignations_list if a['date'] == date_str and a['type_garde_id'] == type_garde['id']]
                    
                    noms = [f"{users_map[a['user_id']]['prenom']} {users_map[a['user_id']]['nom']}" 
                           for a in assignations_jour if a['user_id'] in users_map]
                    
                    personnel_str = ', '.join(noms) if noms else 'Aucun'
                    requis = type_garde.get('personnel_requis', 1)
                    assignes = len(noms)
                    statut = 'Complet' if assignes >= requis else 'Partiel' if noms else 'Vacant'
                    
                    ws.cell(row=row, column=1, value=current.strftime('%d/%m/%Y'))
                    ws.cell(row=row, column=2, value=jour_fr)
                    ws.cell(row=row, column=3, value=type_garde['nom'])
                    ws.cell(row=row, column=4, value=f"{type_garde.get('heure_debut', '')} - {type_garde.get('heure_fin', '')}")
                    ws.cell(row=row, column=5, value=personnel_str)
                    ws.cell(row=row, column=6, value=requis)
                    ws.cell(row=row, column=7, value=assignes)
                    status_cell = ws.cell(row=row, column=8, value=statut)
                    
                    if statut == 'Complet':
                        status_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    elif statut == 'Partiel':
                        status_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                    else:
                        status_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    
                    for col in range(1, 9):
                        ws.cell(row=row, column=col).border = border
                        ws.cell(row=row, column=col).alignment = center_alignment
                    
                    row += 1
                
                current += timedelta(days=1)
        
        # Définir les largeurs de colonnes fixes pour éviter les erreurs avec MergedCell
        column_widths = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        default_widths = [12, 12, 18, 15, 25, 10, 10, 12, 12]
        for i, col_letter in enumerate(column_widths):
            if i < len(default_widths):
                ws.column_dimensions[col_letter].width = default_widths[i]
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=planning_{type}_{periode}.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur export Excel: {str(e)}")

# ===== RAPPORT D'HEURES =====



# GET rapport-heures/debug
@router.get("/{tenant_slug}/planning/rapport-heures/debug/{user_id}")
async def debug_rapport_heures_user(
    tenant_slug: str,
    user_id: str,
    date_debut: str,
    date_fin: str,
    current_user: User = Depends(get_current_user)
):
    """Endpoint de diagnostic pour comprendre le calcul des heures d'un utilisateur"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Récupérer l'utilisateur
    user = await db.users.find_one({"id": user_id, "tenant_id": tenant.id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    # Récupérer TOUTES les assignations (avec doublons éventuels)
    assignations_brutes = await db.assignations.find({
        "user_id": user_id,
        "tenant_id": tenant.id,
        "date": {"$gte": date_debut, "$lte": date_fin}
    }, {"_id": 0}).to_list(10000)
    
    # Déduplication
    assignations_uniques = {}
    doublons = []
    for a in assignations_brutes:
        key = f"{a['user_id']}_{a['type_garde_id']}_{a['date']}"
        if key not in assignations_uniques:
            assignations_uniques[key] = a
        else:
            doublons.append(a)
    
    assignations = list(assignations_uniques.values())
    
    # Récupérer les types de garde
    types_garde = await db.types_garde.find({"tenant_id": tenant.id}, {"_id": 0}).to_list(1000)
    types_garde_map = {t["id"]: t for t in types_garde}
    
    # Calculer les détails
    details = []
    total_heures = 0
    total_heures_calculees = 0
    
    for a in assignations:
        type_garde = types_garde_map.get(a["type_garde_id"])
        if type_garde:
            duree_stored = type_garde.get("duree_heures", None)
            
            # Calculer la durée réelle à partir des horaires
            duree_calculee = None
            if type_garde.get("heure_debut") and type_garde.get("heure_fin"):
                try:
                    from datetime import datetime
                    debut = datetime.strptime(type_garde["heure_debut"], "%H:%M")
                    fin = datetime.strptime(type_garde["heure_fin"], "%H:%M")
                    if fin < debut:
                        fin = fin.replace(day=debut.day + 1)
                    delta = (fin - debut).total_seconds() / 3600
                    duree_calculee = round(delta, 2)
                except:
                    duree_calculee = None
            
            # Durée utilisée par le code (celle dans le rapport)
            duree_utilisee = duree_stored if duree_stored is not None else 8
            
            total_heures += duree_utilisee
            if duree_calculee:
                total_heures_calculees += duree_calculee
            
            details.append({
                "date": a["date"],
                "type_garde_nom": type_garde.get("nom"),
                "type_garde_id": a["type_garde_id"],
                "heure_debut": type_garde.get("heure_debut"),
                "heure_fin": type_garde.get("heure_fin"),
                "duree_stored_bd": duree_stored,
                "duree_calculee_horaires": duree_calculee,
                "duree_utilisee_rapport": duree_utilisee,
                "est_garde_externe": type_garde.get("est_garde_externe", False)
            })
    
    return {
        "user": {
            "id": user["id"],
            "nom_complet": f"{user.get('prenom')} {user.get('nom')}",
            "email": user.get("email"),
            "heures_max_semaine": user.get("heures_max_semaine", 40)
        },
        "periode": f"{date_debut} au {date_fin}",
        "compteurs": {
            "assignations_brutes": len(assignations_brutes),
            "assignations_uniques": len(assignations),
            "doublons_detectes": len(doublons),
            "total_heures_rapport": total_heures,
            "total_heures_reelles_calculees": total_heures_calculees
        },
        "assignations_details": sorted(details, key=lambda x: x["date"]),
        "doublons": doublons[:10] if doublons else []
    }


# POST supprimer-assignations-invalides
@router.post("/{tenant_slug}/planning/supprimer-assignations-invalides")
async def supprimer_assignations_invalides(
    tenant_slug: str,
    current_user: User = Depends(get_current_user)
):
    """Supprime toutes les assignations qui ne respectent pas les jours_application"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé - Admin uniquement")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # D'abord, récupérer le rapport des invalides
    assignations = await db.assignations.find({"tenant_id": tenant.id}, {"_id": 0}).to_list(10000)
    types_garde = await db.types_garde.find({"tenant_id": tenant.id}, {"_id": 0}).to_list(1000)
    types_garde_map = {t["id"]: t for t in types_garde}
    
    jours_fr_to_en = {
        0: "monday", 1: "tuesday", 2: "wednesday", 3: "thursday",
        4: "friday", 5: "saturday", 6: "sunday"
    }
    
    ids_to_delete = []
    
    for a in assignations:
        type_garde = types_garde_map.get(a["type_garde_id"])
        if not type_garde:
            continue
        
        jours_application = type_garde.get("jours_application", [])
        if not jours_application:
            continue
        
        try:
            from datetime import datetime
            date_obj = datetime.strptime(a["date"], "%Y-%m-%d")
            jour_semaine_en = jours_fr_to_en[date_obj.weekday()]
            
            if jour_semaine_en not in jours_application:
                ids_to_delete.append(a["id"])
        except:
            pass
    
    # Supprimer
    if ids_to_delete:
        result = await db.assignations.delete_many({
            "id": {"$in": ids_to_delete},
            "tenant_id": tenant.id
        })
        deleted_count = result.deleted_count
    else:
        deleted_count = 0
    
    return {
        "message": f"Supprimé {deleted_count} assignations invalides",
        "deleted_count": deleted_count,
        "ids_supprimes": ids_to_delete[:50]  # Max 50 pour la réponse
    }


# GET rapport-heures/export-pdf
@router.get("/{tenant_slug}/planning/rapport-heures/export-pdf")
async def export_rapport_heures_pdf(
    tenant_slug: str,
    date_debut: str,
    date_fin: str,
    current_user: User = Depends(get_current_user)
):
    """Génère le PDF du rapport d'heures pour impression"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Convertir date_debut en format mois pour appeler get_rapport_heures
    mois = date_debut[:7]  # YYYY-MM
    
    # Récupérer les données du rapport en créant un faux current_user pour l'appel interne
    # On va récupérer les données directement ici
    from datetime import datetime as dt
    debut_dt = dt.strptime(date_debut, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    fin_dt = dt.strptime(date_fin, "%Y-%m-%d").replace(tzinfo=timezone.utc) + timedelta(days=1)
    
    # Récupérer toutes les assignations de la période
    assignations = await db.assignations.find({
        "tenant_id": tenant.id,
        "date": {"$gte": debut_dt.strftime("%Y-%m-%d"), "$lt": fin_dt.strftime("%Y-%m-%d")}
    }).to_list(10000)
    
    # Récupérer tous les utilisateurs (actifs ou pas - pour afficher tous ceux qui ont des assignations)
    users = await db.users.find({"tenant_id": tenant.id}).to_list(1000)
    users_dict = {u["id"]: u for u in users}
    
    # Récupérer les types de garde pour déterminer est_externe et duree
    types_garde = await db.types_garde.find({"tenant_id": tenant.id}).to_list(1000)
    types_garde_dict = {tg["id"]: tg for tg in types_garde}
    
    # Calculer les heures par employé
    heures_par_employe = {}
    for assign in assignations:
        uid = assign.get("user_id")
        if uid not in heures_par_employe:
            heures_par_employe[uid] = {"internes": 0, "externes": 0}
        
        # Récupérer la durée depuis l'assignation OU depuis le type de garde
        type_garde_id = assign.get("type_garde_id")
        type_garde = types_garde_dict.get(type_garde_id, {})
        duree = assign.get("duree_heures") or type_garde.get("duree_heures", 0)
        
        # Déterminer si externe depuis l'assignation OU depuis le type de garde
        est_externe = assign.get("est_externe") or type_garde.get("est_garde_externe", False)
        
        if est_externe:
            heures_par_employe[uid]["externes"] += duree
        else:
            heures_par_employe[uid]["internes"] += duree
    
    # Construire la liste des employés
    employes = []
    for uid, heures in heures_par_employe.items():
        user = users_dict.get(uid, {})
        employes.append({
            "nom_complet": f"{user.get('prenom', '')} {user.get('nom', '')}".strip() or "Inconnu",
            "type_emploi": user.get("type_emploi", "temps_plein"),
            "grade": user.get("grade", "N/A"),
            "heures_internes": heures["internes"],
            "heures_externes": heures["externes"],
            "total_heures": heures["internes"] + heures["externes"]
        })
    
    # Trier par nom
    employes.sort(key=lambda x: x["nom_complet"])
    
    # Statistiques
    total_internes = sum(e["heures_internes"] for e in employes)
    total_externes = sum(e["heures_externes"] for e in employes)
    nb_employes = len(employes)
    
    statistiques = {
        "nombre_employes": nb_employes,
        "total_heures_planifiees": total_internes + total_externes,
        "moyenne_heures_internes": round(total_internes / nb_employes, 1) if nb_employes > 0 else 0,
        "moyenne_heures_externes": round(total_externes / nb_employes, 1) if nb_employes > 0 else 0
    }
    
    rapport_response = {"employes": employes, "statistiques": statistiques}
    
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    
    # Utiliser la fonction helper pour créer un PDF brandé
    buffer, doc, elements = create_branded_pdf(tenant, pagesize=A4)
    styles = getSampleStyleSheet()
    modern_styles = get_modern_pdf_styles(styles)
    
    # Titre
    elements.append(Paragraph("Rapport d'Heures", modern_styles['title']))
    
    # Période
    debut_dt = datetime.strptime(date_debut, "%Y-%m-%d")
    fin_dt = datetime.strptime(date_fin, "%Y-%m-%d")
    periode_text = f"Période: {debut_dt.strftime('%d/%m/%Y')} - {fin_dt.strftime('%d/%m/%Y')}"
    elements.append(Paragraph(periode_text, modern_styles['subheading']))
    
    # Tableau des employés
    table_data = [
        ['Employé', 'Type', 'Grade', 'H. Internes', 'H. Externes', 'Total']
    ]
    
    for emp in rapport_response["employes"]:
        type_emploi = emp.get("type_emploi", "temps_plein")
        if type_emploi == "temps_plein":
            type_emploi_abbr = "TP"
        elif type_emploi == "temporaire":
            type_emploi_abbr = "Tempo"
        else:
            type_emploi_abbr = "TPart"
        table_data.append([
            emp["nom_complet"],
            type_emploi_abbr,
            emp.get("grade", "N/A"),
            f"{emp['heures_internes']}h",
            f"{emp['heures_externes']}h",
            f"{emp['total_heures']}h"
        ])
    
    table = Table(table_data, colWidths=[2.5*inch, 0.6*inch, 1.2*inch, 1*inch, 1*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), modern_styles['primary_color']),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, modern_styles['grid']),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, modern_styles['bg_light']])
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Statistiques
    stats = rapport_response["statistiques"]
    stats_text = f"""
    <b>Statistiques Globales</b><br/>
    Nombre d'employés: {stats['nombre_employes']}<br/>
    Total heures planifiées: {stats['total_heures_planifiees']}h<br/>
    Moyenne heures internes: {stats['moyenne_heures_internes']}h<br/>
    Moyenne heures externes: {stats['moyenne_heures_externes']}h
    """
    elements.append(Paragraph(stats_text, styles['Normal']))
    
    # Ajouter footer ProFireManager
    elements.append(Spacer(1, 0.5*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    footer_text = create_pdf_footer_text(tenant)
    if footer_text:
        elements.append(Paragraph(footer_text, footer_style))
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=rapport_heures_{date_debut}_{date_fin}.pdf"}
    )


# GET rapport-heures/export-excel
@router.get("/{tenant_slug}/planning/rapport-heures/export-excel")
async def export_rapport_heures_excel(
    tenant_slug: str,
    date_debut: str,
    date_fin: str,
    current_user: User = Depends(get_current_user)
):
    """Génère l'Excel du rapport d'heures"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Récupérer les données directement (même logique que pour le PDF)
    from datetime import datetime as dt
    debut_dt = dt.strptime(date_debut, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    fin_dt = dt.strptime(date_fin, "%Y-%m-%d").replace(tzinfo=timezone.utc) + timedelta(days=1)
    
    assignations = await db.assignations.find({
        "tenant_id": tenant.id,
        "date": {"$gte": debut_dt.strftime("%Y-%m-%d"), "$lt": fin_dt.strftime("%Y-%m-%d")}
    }).to_list(10000)
    
    users = await db.users.find({"tenant_id": tenant.id, "actif": True}).to_list(1000)
    users_dict = {u["id"]: u for u in users}
    
    # Récupérer les types de garde pour déterminer est_externe et duree
    types_garde = await db.types_garde.find({"tenant_id": tenant.id}).to_list(1000)
    types_garde_dict = {tg["id"]: tg for tg in types_garde}
    
    heures_par_employe = {}
    for assign in assignations:
        uid = assign.get("user_id")
        if uid not in heures_par_employe:
            heures_par_employe[uid] = {"internes": 0, "externes": 0}
        
        # Récupérer la durée depuis l'assignation OU depuis le type de garde
        type_garde_id = assign.get("type_garde_id")
        type_garde = types_garde_dict.get(type_garde_id, {})
        duree = assign.get("duree_heures") or type_garde.get("duree_heures", 0)
        
        # Déterminer si externe depuis l'assignation OU depuis le type de garde
        est_externe = assign.get("est_externe") or type_garde.get("est_garde_externe", False)
        
        if est_externe:
            heures_par_employe[uid]["externes"] += duree
        else:
            heures_par_employe[uid]["internes"] += duree
    
    employes = []
    for uid, heures in heures_par_employe.items():
        user = users_dict.get(uid, {})
        employes.append({
            "nom_complet": f"{user.get('prenom', '')} {user.get('nom', '')}".strip() or "Inconnu",
            "type_emploi": user.get("type_emploi", "temps_plein"),
            "grade": user.get("grade", "N/A"),
            "heures_internes": heures["internes"],
            "heures_externes": heures["externes"],
            "total_heures": heures["internes"] + heures["externes"]
        })
    
    employes.sort(key=lambda x: x["nom_complet"])
    
    total_internes = sum(e["heures_internes"] for e in employes)
    total_externes = sum(e["heures_externes"] for e in employes)
    nb_employes = len(employes)
    
    statistiques = {
        "nombre_employes": nb_employes,
        "total_heures_planifiees": total_internes + total_externes,
        "moyenne_heures_internes": round(total_internes / nb_employes, 1) if nb_employes > 0 else 0,
        "moyenne_heures_externes": round(total_externes / nb_employes, 1) if nb_employes > 0 else 0
    }
    
    rapport_response = {"employes": employes, "statistiques": statistiques}
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport Heures"
    
    # Styles
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Titre
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "Rapport d'Heures"
    title_cell.font = Font(size=16, bold=True, color="DC2626")
    title_cell.alignment = center_alignment
    
    # Période
    ws.merge_cells('A2:F2')
    periode_cell = ws['A2']
    debut_dt = datetime.strptime(date_debut, "%Y-%m-%d")
    fin_dt = datetime.strptime(date_fin, "%Y-%m-%d")
    periode_cell.value = f"Période: {debut_dt.strftime('%d/%m/%Y')} - {fin_dt.strftime('%d/%m/%Y')}"
    periode_cell.alignment = center_alignment
    
    # En-têtes du tableau
    headers = ['Employé', 'Type', 'Grade', 'H. Internes', 'H. Externes', 'Total']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_alignment
    
    # Données
    row = 5
    for emp in rapport_response["employes"]:
        type_emploi = emp.get("type_emploi", "temps_plein")
        if type_emploi == "temps_plein":
            type_emploi_abbr = "TP"
        elif type_emploi == "temporaire":
            type_emploi_abbr = "Tempo"
        else:
            type_emploi_abbr = "TPart"
        ws.cell(row=row, column=1, value=emp["nom_complet"])
        ws.cell(row=row, column=2, value=type_emploi_abbr)
        ws.cell(row=row, column=3, value=emp.get("grade", "N/A"))
        ws.cell(row=row, column=4, value=emp["heures_internes"])
        ws.cell(row=row, column=5, value=emp["heures_externes"])
        ws.cell(row=row, column=6, value=emp["total_heures"])
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = center_alignment
        
        row += 1
    
    # Statistiques
    stats = rapport_response["statistiques"]
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    stats_cell = ws.cell(row=row, column=1)
    stats_cell.value = "Statistiques Globales"
    stats_cell.font = Font(bold=True, size=12)
    
    row += 1
    ws.cell(row=row, column=1, value="Nombre d'employés:")
    ws.cell(row=row, column=2, value=stats['nombre_employes'])
    
    row += 1
    ws.cell(row=row, column=1, value="Total heures planifiées:")
    ws.cell(row=row, column=2, value=stats['total_heures_planifiees'])
    
    row += 1
    ws.cell(row=row, column=1, value="Moyenne heures internes:")
    ws.cell(row=row, column=2, value=stats['moyenne_heures_internes'])
    
    row += 1
    ws.cell(row=row, column=1, value="Moyenne heures externes:")
    ws.cell(row=row, column=2, value=stats['moyenne_heures_externes'])
    
    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=rapport_heures_{date_debut}_{date_fin}.xlsx"}
    )


# POST assignation-avancee
@router.post("/{tenant_slug}/planning/assignation-avancee")
async def assignation_manuelle_avancee(
    tenant_slug: str,
    assignation_data: dict,
    current_user: User = Depends(get_current_user)
):
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    # Vérifier le tenant
    tenant = await get_tenant_from_slug(tenant_slug)
    
    try:
        user_id = assignation_data.get("user_id")
        type_garde_id = assignation_data.get("type_garde_id")
        recurrence_type = assignation_data.get("recurrence_type", "unique")
        date_debut = datetime.strptime(assignation_data.get("date_debut"), "%Y-%m-%d").date()
        date_fin = datetime.strptime(assignation_data.get("date_fin", assignation_data.get("date_debut")), "%Y-%m-%d").date()
        jours_semaine = assignation_data.get("jours_semaine", [])
        bi_hebdomadaire = assignation_data.get("bi_hebdomadaire", False)
        recurrence_intervalle = assignation_data.get("recurrence_intervalle", 1)
        recurrence_frequence = assignation_data.get("recurrence_frequence", "jours")
        
        assignations_creees = []
        
        if recurrence_type == "unique":
            # Assignation unique
            assignation_obj = Assignation(
                user_id=user_id,
                type_garde_id=type_garde_id,
                date=date_debut.strftime("%Y-%m-%d"),
                assignation_type="manuel_avance",
                tenant_id=tenant.id
            )
            await db.assignations.insert_one(assignation_obj.dict())
            assignations_creees.append(assignation_obj.dict())
            
        elif recurrence_type == "hebdomadaire":
            # Récurrence hebdomadaire (avec option bi-hebdomadaire)
            current_date = date_debut
            jours_semaine_index = {
                'monday': 0, 'tuesday': 1, 'wednesday': 2, 'thursday': 3,
                'friday': 4, 'saturday': 5, 'sunday': 6
            }
            
            # Pour bi-hebdomadaire : calculer le numéro de semaine ISO de la date de début
            def get_iso_week_number(date):
                # Python's isocalendar() retourne (année, semaine, jour_semaine)
                return date.isocalendar()[1]
            
            reference_week = get_iso_week_number(date_debut)
            
            while current_date <= date_fin:
                day_name = current_date.strftime("%A").lower()
                
                # Vérifier si c'est un jour sélectionné
                if day_name in jours_semaine:
                    # Si bi-hebdomadaire, vérifier la différence de semaines
                    current_week = get_iso_week_number(current_date)
                    weeks_difference = current_week - reference_week
                    
                    if not bi_hebdomadaire or weeks_difference % 2 == 0:
                        # Vérifier qu'il n'y a pas déjà une assignation
                        existing = await db.assignations.find_one({
                            "user_id": user_id,
                            "type_garde_id": type_garde_id,
                            "date": current_date.strftime("%Y-%m-%d"),
                            "tenant_id": tenant.id
                        })
                        
                        if not existing:
                            assignation_obj = Assignation(
                                user_id=user_id,
                                type_garde_id=type_garde_id,
                                date=current_date.strftime("%Y-%m-%d"),
                                assignation_type="manuel_avance",
                                tenant_id=tenant.id
                            )
                            await db.assignations.insert_one(assignation_obj.dict())
                            assignations_creees.append(assignation_obj.dict())
                
                current_date += timedelta(days=1)
        
        elif recurrence_type == "bihebdomadaire":
            # Récurrence bi-hebdomadaire (toutes les 2 semaines)
            current_date = date_debut
            
            # Calculer le numéro de semaine ISO de référence
            def get_iso_week_number(date):
                return date.isocalendar()[1]
            
            reference_week = get_iso_week_number(date_debut)
            
            while current_date <= date_fin:
                day_name = current_date.strftime("%A").lower()
                
                # Calculer la différence de semaines
                current_week = get_iso_week_number(current_date)
                weeks_difference = current_week - reference_week
                
                # Vérifier si c'est un jour sélectionné et une semaine paire
                if day_name in jours_semaine and weeks_difference % 2 == 0:
                    existing = await db.assignations.find_one({
                        "user_id": user_id,
                        "type_garde_id": type_garde_id,
                        "date": current_date.strftime("%Y-%m-%d"),
                        "tenant_id": tenant.id
                    })
                    
                    if not existing:
                        assignation_obj = Assignation(
                            user_id=user_id,
                            type_garde_id=type_garde_id,
                            date=current_date.strftime("%Y-%m-%d"),
                            assignation_type="manuel_avance",
                            tenant_id=tenant.id
                        )
                        await db.assignations.insert_one(assignation_obj.dict())
                        assignations_creees.append(assignation_obj.dict())
                
                current_date += timedelta(days=1)
                
        elif recurrence_type == "mensuel" or recurrence_type == "mensuelle":
            # Récurrence mensuelle (même jour du mois)
            jour_mois = date_debut.day
            current_month = date_debut.replace(day=1)
            
            while current_month <= date_fin:
                try:
                    # Essayer de créer la date pour ce mois
                    target_date = current_month.replace(day=jour_mois)
                    
                    if date_debut <= target_date <= date_fin:
                        existing = await db.assignations.find_one({
                            "user_id": user_id,
                            "type_garde_id": type_garde_id,
                            "date": target_date.strftime("%Y-%m-%d"),
                            "tenant_id": tenant.id
                        })
                        
                        if not existing:
                            assignation_obj = Assignation(
                                user_id=user_id,
                                type_garde_id=type_garde_id,
                                date=target_date.strftime("%Y-%m-%d"),
                                assignation_type="manuel_avance",
                                tenant_id=tenant.id
                            )
                            await db.assignations.insert_one(assignation_obj.dict())
                            assignations_creees.append(assignation_obj.dict())
                            
                except ValueError:
                    # Jour n'existe pas dans ce mois (ex: 31 février)
                    pass
                
                # Passer au mois suivant
                if current_month.month == 12:
                    current_month = current_month.replace(year=current_month.year + 1, month=1)
                else:
                    current_month = current_month.replace(month=current_month.month + 1)
        
        elif recurrence_type == "annuelle":
            # Récurrence annuelle (même jour et mois chaque année)
            jour_mois = date_debut.day
            mois = date_debut.month
            current_year = date_debut.year
            
            while True:
                try:
                    target_date = date(current_year, mois, jour_mois)
                    
                    if target_date > date_fin:
                        break
                    
                    if target_date >= date_debut:
                        existing = await db.assignations.find_one({
                            "user_id": user_id,
                            "type_garde_id": type_garde_id,
                            "date": target_date.strftime("%Y-%m-%d"),
                            "tenant_id": tenant.id
                        })
                        
                        if not existing:
                            assignation_obj = Assignation(
                                user_id=user_id,
                                type_garde_id=type_garde_id,
                                date=target_date.strftime("%Y-%m-%d"),
                                assignation_type="manuel_avance",
                                tenant_id=tenant.id
                            )
                            await db.assignations.insert_one(assignation_obj.dict())
                            assignations_creees.append(assignation_obj.dict())
                    
                    current_year += 1
                except ValueError:
                    # Jour n'existe pas (ex: 29 février dans une année non bissextile)
                    current_year += 1
        
        elif recurrence_type == "personnalisee":
            # Récurrence personnalisée
            current_date = date_debut
            
            if recurrence_frequence == "jours":
                delta = timedelta(days=recurrence_intervalle)
            elif recurrence_frequence == "semaines":
                delta = timedelta(weeks=recurrence_intervalle)
            else:
                # Pour mois et ans, on gérera différemment
                delta = None
            
            if delta:
                while current_date <= date_fin:
                    existing = await db.assignations.find_one({
                        "user_id": user_id,
                        "type_garde_id": type_garde_id,
                        "date": current_date.strftime("%Y-%m-%d"),
                        "tenant_id": tenant.id
                    })
                    
                    if not existing:
                        assignation_obj = Assignation(
                            user_id=user_id,
                            type_garde_id=type_garde_id,
                            date=current_date.strftime("%Y-%m-%d"),
                            assignation_type="manuel_avance",
                            tenant_id=tenant.id
                        )
                        await db.assignations.insert_one(assignation_obj.dict())
                        assignations_creees.append(assignation_obj.dict())
                    
                    current_date += delta
            else:
                # Pour mois et ans
                current_date = date_debut
                while current_date <= date_fin:
                    existing = await db.assignations.find_one({
                        "user_id": user_id,
                        "type_garde_id": type_garde_id,
                        "date": current_date.strftime("%Y-%m-%d"),
                        "tenant_id": tenant.id
                    })
                    
                    if not existing:
                        assignation_obj = Assignation(
                            user_id=user_id,
                            type_garde_id=type_garde_id,
                            date=current_date.strftime("%Y-%m-%d"),
                            assignation_type="manuel_avance",
                            tenant_id=tenant.id
                        )
                        await db.assignations.insert_one(assignation_obj.dict())
                        assignations_creees.append(assignation_obj.dict())
                    
                    if recurrence_frequence == "mois":
                        # Ajouter X mois
                        month = current_date.month + recurrence_intervalle
                        year = current_date.year
                        while month > 12:
                            month -= 12
                            year += 1
                        try:
                            current_date = current_date.replace(year=year, month=month)
                        except ValueError:
                            # Jour invalide pour ce mois
                            break
                    elif recurrence_frequence == "ans":
                        # Ajouter X ans
                        try:
                            current_date = current_date.replace(year=current_date.year + recurrence_intervalle)
                        except ValueError:
                            # Jour invalide (29 février)
                            break
        
        return {
            "message": "Assignation avancée créée avec succès",
            "assignations_creees": len(assignations_creees),
            "recurrence": recurrence_type,
            "periode": f"{date_debut.strftime('%Y-%m-%d')} à {date_fin.strftime('%Y-%m-%d')}"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur assignation avancée: {str(e)}")

# Mode démo spécial - Attribution automatique agressive pour impression client

# POST attribution-auto-demo
@router.post("/{tenant_slug}/planning/attribution-auto-demo")
async def attribution_automatique_demo(tenant_slug: str, semaine_debut: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    # Vérifier le tenant
    tenant = await get_tenant_from_slug(tenant_slug)
    
    try:
        # Get all available users and types de garde pour ce tenant
        users = await db.users.find({"statut": "Actif", "tenant_id": tenant.id}).to_list(1000)
        types_garde = await db.types_garde.find({"tenant_id": tenant.id}).to_list(1000)
        
        # Get existing assignations for the week
        semaine_fin = (datetime.strptime(semaine_debut, "%Y-%m-%d") + timedelta(days=6)).strftime("%Y-%m-%d")
        existing_assignations = await db.assignations.find({
            "date": {
                "$gte": semaine_debut,
                "$lte": semaine_fin
            },
            "tenant_id": tenant.id
        }).to_list(1000)
        
        nouvelles_assignations = []
        
        # MODE DÉMO AGRESSIF - REMPLIR AU MAXIMUM
        for type_garde in types_garde:
            for day_offset in range(7):
                current_date = datetime.strptime(semaine_debut, "%Y-%m-%d") + timedelta(days=day_offset)
                date_str = current_date.strftime("%Y-%m-%d")
                day_name = current_date.strftime("%A").lower()
                
                # CORRECTION CRITIQUE: Skip if type garde doesn't apply to this day
                jours_app = type_garde.get("jours_application", [])
                if jours_app and len(jours_app) > 0 and day_name not in jours_app:
                    logging.debug(f"⏭️ [SKIP DAY DEMO] {type_garde['nom']} - {date_str} ({day_name}): Jour non applicable (limité à {jours_app})")
                    continue
                
                # Compter combien de personnel déjà assigné pour cette garde
                existing_for_garde = [a for a in existing_assignations 
                                    if a["date"] == date_str and a["type_garde_id"] == type_garde["id"]]
                
                personnel_deja_assigne = len(existing_for_garde)
                personnel_requis = type_garde.get("personnel_requis", 1)
                
                # Assigner jusqu'au maximum requis
                for i in range(personnel_requis - personnel_deja_assigne):
                    # Trouver utilisateurs disponibles
                    available_users = []
                    
                    for user in users:
                        # Skip si déjà assigné cette garde ce jour
                        if any(a["user_id"] == user["id"] and a["date"] == date_str and a["type_garde_id"] == type_garde["id"] 
                               for a in existing_assignations):
                            continue
                        
                        # Skip si déjà assigné autre garde ce jour (éviter conflits)
                        if any(a["user_id"] == user["id"] and a["date"] == date_str 
                               for a in existing_assignations):
                            continue
                        
                        # Vérifier disponibilités
                        user_dispos = await db.disponibilites.find({
                            "user_id": user["id"],
                            "date": date_str,
                            "type_garde_id": type_garde["id"],
                            "statut": "disponible"
                        }).to_list(10)
                        
                        if user_dispos:
                            available_users.append(user)
                    
                    if not available_users:
                        break  # Pas d'utilisateurs disponibles pour ce poste
                    
                    # MODE DÉMO : ASSOUPLIR CONTRAINTE OFFICIER
                    if type_garde.get("officier_obligatoire", False):
                        # Chercher officiers d'abord
                        officers = [u for u in available_users if u.get("grade", "") in ["Capitaine", "Lieutenant", "Directeur"]]
                        # Sinon pompiers avec fonction supérieur
                        if not officers:
                            officers = [u for u in available_users if u.get("fonction_superieur", False)]
                        # En dernier recours : tous pompiers (MODE DÉMO)
                        if not officers:
                            officers = available_users
                        
                        if officers:
                            selected_user = officers[0]
                        else:
                            continue
                    else:
                        selected_user = available_users[0]
                    
                    # Créer assignation
                    assignation_obj = Assignation(
                        user_id=selected_user["id"],
                        type_garde_id=type_garde["id"],
                        date=date_str,
                        assignation_type="auto_demo",
                        tenant_id=tenant.id
                    )
                    
                    await db.assignations.insert_one(assignation_obj.dict())
                    nouvelles_assignations.append(assignation_obj.dict())
                    existing_assignations.append(assignation_obj.dict())
        
        return {
            "message": "Attribution DÉMO agressive effectuée avec succès",
            "assignations_creees": len(nouvelles_assignations),
            "algorithme": "Mode démo : Contraintes assouplies pour impression maximum",
            "semaine": f"{semaine_debut} - {semaine_fin}"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur attribution démo: {str(e)}")

# Vérification des assignations existantes pour une période

# GET check-periode
@router.get("/{tenant_slug}/planning/assignations/check-periode")
async def check_assignations_periode(
    tenant_slug: str, 
    debut: str, 
    fin: str, 
    current_user: User = Depends(get_current_user)
):
    """Vérifie s'il existe des assignations pour la période donnée"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    try:
        existing_count = await db.assignations.count_documents({
            "date": {
                "$gte": debut,
                "$lte": fin
            },
            "tenant_id": tenant.id
        })
        
        return {
            "existing_count": existing_count,
            "periode": f"{debut} au {fin}"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur vérification période: {str(e)}")


# GET progress/{task_id}
@router.get("/{tenant_slug}/planning/attribution-auto/progress/{task_id}")
async def attribution_progress_stream(
    tenant_slug: str,
    task_id: str
):
    """Stream SSE pour suivre la progression de l'attribution automatique
    
    Note: Pas d'authentification JWT car EventSource ne peut pas envoyer de headers.
    La sécurité est assurée par le task_id unique et éphémère.
    """
    # Vérifier que le task_id existe (sécurité basique)
    if task_id not in attribution_progress_store and task_id != "test":
        # Attendre un peu que la tâche soit créée
        await asyncio.sleep(1)
        if task_id not in attribution_progress_store:
            raise HTTPException(status_code=404, detail="Task ID non trouvé")
    
    return StreamingResponse(
        progress_event_generator(task_id),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"  # Nginx buffering disabled
        }
    )

# Attribution automatique intelligente avec rotation équitable et ancienneté

# POST attribution-auto
@router.post("/{tenant_slug}/planning/attribution-auto")
async def attribution_automatique(
    tenant_slug: str, 
    semaine_debut: str, 
    semaine_fin: str = None,
    reset: bool = False,  # Nouveau paramètre pour réinitialiser
    current_user: User = Depends(get_current_user)
):
    """Attribution automatique pour une ou plusieurs semaines avec progression temps réel
    
    Args:
        reset: Si True, supprime d'abord toutes les assignations AUTO de la période
        
    Returns:
        task_id: Identifiant pour suivre la progression via SSE
    """
    logging.info(f"🔥 [ENDPOINT] Attribution auto appelé par {current_user.email}")
    logging.info(f"🔥 [ENDPOINT] Paramètres reçus: tenant={tenant_slug}, debut={semaine_debut}, fin={semaine_fin}, reset={reset}")
    
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    # Vérifier le tenant
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Générer un task_id unique
    task_id = str(uuid.uuid4())
    
    # Lancer la tâche en arrière-plan
    asyncio.create_task(
        process_attribution_auto_async(
            task_id, tenant, semaine_debut, semaine_fin, reset
        )
    )
    
    # Retourner immédiatement le task_id
    return {
        "task_id": task_id,
        "message": "Attribution automatique lancée en arrière-plan",
        "stream_url": f"/api/{tenant_slug}/planning/attribution-auto/progress/{task_id}"
    }

async def process_attribution_auto_async(
    task_id: str,
    tenant,
    semaine_debut: str,
    semaine_fin: str = None,
    reset: bool = False
):
    """Traite l'attribution automatique de manière asynchrone avec suivi de progression"""
    progress = AttributionProgress(task_id)
    
    try:
        start_time = time.time()
        logging.info(f"⏱️ [PERF] Attribution auto démarrée - Task ID: {task_id}")
        logging.info(f"🔍 [DEBUG] reset={reset}, type={type(reset)}, tenant_id={tenant.id}")
        logging.info(f"🔍 [DEBUG] Période: {semaine_debut} → {semaine_fin}")
        
        # Si pas de semaine_fin fournie, calculer pour une seule semaine
        if not semaine_fin:
            semaine_fin = (datetime.strptime(semaine_debut, "%Y-%m-%d") + timedelta(days=6)).strftime("%Y-%m-%d")
        
        progress.update("Initialisation...", 5)
        
        # Si reset=True, supprimer d'abord toutes les assignations AUTO de la période
        assignations_supprimees = 0
        if reset:
            logging.info(f"🔍 [DEBUG] RESET MODE ACTIVÉ - Tentative de suppression...")
            
            # Vérifier combien d'assignations existent
            count_before = await db.assignations.count_documents({
                "tenant_id": tenant.id,
                "date": {"$gte": semaine_debut, "$lte": semaine_fin}
            })
            logging.info(f"🔍 [DEBUG] Assignations totales dans période: {count_before}")
            
            # Vérifier les types d'assignation existants
            distinct_types = await db.assignations.distinct("assignation_type", {
                "tenant_id": tenant.id,
                "date": {"$gte": semaine_debut, "$lte": semaine_fin}
            })
            logging.info(f"🔍 [DEBUG] Types d'assignation existants: {distinct_types}")
            
            progress.update("Suppression des assignations existantes...", 10)
            # CORRECTION: Supprimer TOUTES les assignations AUTO + celles sans type (anciennes)
            result = await db.assignations.delete_many({
                "tenant_id": tenant.id,
                "date": {
                    "$gte": semaine_debut,
                    "$lte": semaine_fin
                },
                "$or": [
                    {"assignation_type": {"$in": ["auto", "automatique"]}},
                    {"assignation_type": {"$exists": False}},  # Assignations sans type (anciennes)
                    {"assignation_type": None}  # Assignations avec type null
                ]
            })
            assignations_supprimees = result.deleted_count
            logging.info(f"⏱️ [PERF] ✅ {assignations_supprimees} assignations supprimées (incluant anciennes sans type)")
        else:
            logging.info(f"🔍 [DEBUG] RESET MODE DÉSACTIVÉ - Pas de suppression")
        
        # Pour une période complète (mois), traiter semaine par semaine
        start_date = datetime.strptime(semaine_debut, "%Y-%m-%d")
        end_date = datetime.strptime(semaine_fin, "%Y-%m-%d")
        
        # Calculer le nombre total de semaines
        total_weeks = ((end_date - start_date).days // 7) + 1
        progress.total_gardes = total_weeks
        
        total_assignations_creees = 0
        current_week_start = start_date
        week_number = 0
        
        # Itérer sur toutes les semaines de la période
        while current_week_start <= end_date:
            week_number += 1
            current_week_end = current_week_start + timedelta(days=6)
            if current_week_end > end_date:
                current_week_end = end_date
            
            week_start_str = current_week_start.strftime("%Y-%m-%d")
            week_end_str = current_week_end.strftime("%Y-%m-%d")
            
            # Mise à jour progression
            progress_percent = 15 + int((week_number / total_weeks) * 80)
            progress.update(
                f"Traitement semaine {week_number}/{total_weeks} ({week_start_str})",
                progress_percent,
                gardes_traitees=week_number
            )
            
            week_start_time = time.time()
            
            # Traiter cette semaine
            assignations_cette_semaine = await traiter_semaine_attribution_auto(
                tenant, 
                week_start_str, 
                week_end_str,
                progress=progress  # Passer l'objet progress pour mises à jour granulaires
            )
            
            # Compter le nombre d'assignations créées (la fonction retourne une liste)
            nb_assignations_semaine = len(assignations_cette_semaine) if isinstance(assignations_cette_semaine, list) else assignations_cette_semaine
            
            week_elapsed = time.time() - week_start_time
            logging.info(f"⏱️ [PERF] Semaine {week_number} traitée en {week_elapsed:.2f}s - {nb_assignations_semaine} assignations")
            
            total_assignations_creees += nb_assignations_semaine
            progress.assignations_creees = total_assignations_creees
            
            # Passer à la semaine suivante
            current_week_start += timedelta(days=7)
        
        # Terminer
        total_elapsed = time.time() - start_time
        logging.info(f"⏱️ [PERF] Attribution auto terminée en {total_elapsed:.2f}s - Total: {total_assignations_creees} assignations")
        
        progress.complete(total_assignations_creees)
        
    except Exception as e:
        logging.error(f"❌ [ERROR] Attribution auto échouée: {str(e)}", exc_info=True)
        progress.error(str(e))

async def generer_justification_attribution(
    selected_user: Dict,
    all_candidates: List[Dict],
    type_garde: Dict,
    date_str: str,
    user_monthly_hours_internes: Dict,
    user_monthly_hours_externes: Dict,
    activer_heures_sup: bool,
    existing_assignations: List[Dict],
    disponibilites_evaluees: List[Dict] = None,
    dispos_lookup: Dict = None  # NOUVEAU: pour vérifier les disponibilités
) -> Dict[str, Any]:
    """
    Génère une justification détaillée pour une attribution automatique
    """
    # Utiliser le compteur approprié selon le type de garde
    user_monthly_hours = user_monthly_hours_externes if type_garde.get("est_garde_externe", False) else user_monthly_hours_internes
    
    # Calculer les scores pour l'utilisateur sélectionné
    heures_selectionnee = user_monthly_hours.get(selected_user["id"], 0)
    moyenne_equipe = sum(user_monthly_hours.values()) / len(user_monthly_hours) if user_monthly_hours else 0
    
    # Score d'équité (0-100) - Plus les heures sont basses, meilleur le score
    if moyenne_equipe > 0:
        ecart_ratio = (moyenne_equipe - heures_selectionnee) / moyenne_equipe
        score_equite = min(100, max(0, 50 + (ecart_ratio * 50)))
    else:
        score_equite = 50
    
    # Score d'ancienneté (0-100)
    try:
        date_embauche = selected_user.get("date_embauche", "1900-01-01")
        try:
            embauche_dt = datetime.strptime(date_embauche, "%Y-%m-%d")
        except:
            embauche_dt = datetime.strptime(date_embauche, "%d/%m/%Y")
        
        annees_service = (datetime.now() - embauche_dt).days / 365.25
        score_anciennete = min(100, annees_service * 5)  # 5 points par an, max 100
    except:
        annees_service = 0
        score_anciennete = 0
    
    # Score de disponibilité (0-100)
    if selected_user.get("type_emploi") in ("temps_partiel", "temporaire"):
        score_disponibilite = 100 if disponibilites_evaluees else 50
    else:
        score_disponibilite = 75  # Temps plein toujours disponible
    
    # Score de compétences (0-100) - basé sur le grade
    grade_scores = {
        "Directeur": 100,
        "Capitaine": 85,
        "Lieutenant": 70,
        "Pompier": 50
    }
    score_competences = grade_scores.get(selected_user.get("grade", "Pompier"), 50)
    
    # Score total
    score_total = score_equite + score_anciennete + score_disponibilite + score_competences
    
    # Détails de l'utilisateur sélectionné
    assigned_user_info = {
        "user_id": selected_user["id"],
        "nom_complet": f"{selected_user['prenom']} {selected_user['nom']}",
        "grade": selected_user.get("grade", "N/A"),
        "type_emploi": selected_user.get("type_emploi", "N/A"),
        "scores": {
            "equite": round(score_equite, 1),
            "anciennete": round(score_anciennete, 1),
            "disponibilite": round(score_disponibilite, 1),
            "competences": round(score_competences, 1),
            "total": round(score_total, 1)
        },
        "details": {
            "heures_ce_mois": heures_selectionnee,
            "moyenne_equipe": round(moyenne_equipe, 1),
            "annees_service": round(annees_service, 1),
            "disponibilite_declaree": selected_user.get("type_emploi") in ("temps_partiel", "temporaire") and bool(disponibilites_evaluees),
            "heures_max_autorisees": selected_user.get("heures_max_semaine", 40) if not activer_heures_sup else None
        }
    }
    
    # Évaluer les autres candidats
    other_candidates = []
    for candidate in all_candidates:
        if candidate["id"] == selected_user["id"]:
            continue  # Skip l'utilisateur sélectionné
        
        # Déterminer la raison d'exclusion
        raison_exclusion = None
        candidate_scores = None
        
        # Vérifier heures supplémentaires (seulement si désactivées)
        if not activer_heures_sup:
            # Calculer heures de la semaine pour ce candidat
            heures_semaine_candidate = 0
            for assignation in existing_assignations:
                if assignation["user_id"] == candidate["id"]:
                    heures_semaine_candidate += 8  # Simplification
            
            heures_max_user = candidate.get("heures_max_semaine", 40)
            
            if heures_semaine_candidate + type_garde.get("duree_heures", 8) > heures_max_user:
                raison_exclusion = f"Heures max atteintes ({heures_semaine_candidate}h/{heures_max_user}h)"
        
        # Vérifier disponibilité (temps partiel)
        if not raison_exclusion and candidate.get("type_emploi") in ("temps_partiel", "temporaire"):
            # Vérifier s'il a déclaré une disponibilité (même logique que l'attribution)
            has_dispo = False
            if dispos_lookup and candidate["id"] in dispos_lookup and date_str in dispos_lookup[candidate["id"]]:
                # Disponibilité spécifique pour ce type de garde
                if type_garde["id"] in dispos_lookup[candidate["id"]][date_str]:
                    has_dispo = True
                # OU disponibilité générale (type_garde_id = None)
                elif None in dispos_lookup[candidate["id"]][date_str]:
                    has_dispo = True
            
            if not has_dispo:
                raison_exclusion = "Disponibilité non déclarée"
        
        # Vérifier s'il est déjà assigné
        if not raison_exclusion:
            deja_assigne = any(
                a["user_id"] == candidate["id"] and 
                a["date"] == date_str and 
                a["type_garde_id"] == type_garde["id"]
                for a in existing_assignations
            )
            if deja_assigne:
                raison_exclusion = "Déjà assigné à cette garde"
        
        # Si pas exclu, calculer les scores
        if not raison_exclusion:
            heures_candidate = user_monthly_hours.get(candidate["id"], 0)
            
            # Scores similaires à l'utilisateur sélectionné
            if moyenne_equipe > 0:
                ecart_ratio = (moyenne_equipe - heures_candidate) / moyenne_equipe
                cand_score_equite = min(100, max(0, 50 + (ecart_ratio * 50)))
            else:
                cand_score_equite = 50
            
            try:
                date_emb = candidate.get("date_embauche", "1900-01-01")
                try:
                    emb_dt = datetime.strptime(date_emb, "%Y-%m-%d")
                except:
                    emb_dt = datetime.strptime(date_emb, "%d/%m/%Y")
                cand_annees = (datetime.now() - emb_dt).days / 365.25
                cand_score_anc = min(100, cand_annees * 5)
            except:
                cand_score_anc = 0
            
            cand_score_dispo = 100 if candidate.get("type_emploi") in ("temps_partiel", "temporaire") else 75
            cand_score_comp = grade_scores.get(candidate.get("grade", "Pompier"), 50)
            cand_total = cand_score_equite + cand_score_anc + cand_score_dispo + cand_score_comp
            
            candidate_scores = {
                "equite": round(cand_score_equite, 1),
                "anciennete": round(cand_score_anc, 1),
                "disponibilite": round(cand_score_dispo, 1),
                "competences": round(cand_score_comp, 1),
                "total": round(cand_total, 1)
            }
            
            raison_exclusion = f"Score inférieur (total: {round(cand_total, 1)} vs {round(score_total, 1)})"
        
        other_candidates.append({
            "user_id": candidate["id"],
            "nom_complet": f"{candidate['prenom']} {candidate['nom']}",
            "grade": candidate.get("grade", "N/A"),
            "excluded_reason": raison_exclusion,
            "scores": candidate_scores,
            "heures_ce_mois": user_monthly_hours.get(candidate["id"], 0)
        })
    
    # Trier les autres candidats par score décroissant (si scores disponibles)
    other_candidates.sort(key=lambda x: x["scores"]["total"] if x["scores"] else 0, reverse=True)
    
    return {
        "assigned_user": assigned_user_info,
        "other_candidates": other_candidates[:10],  # Limiter à 10 pour ne pas surcharger
        "total_candidates_evaluated": len(all_candidates),
        "date_attribution": datetime.now(timezone.utc).isoformat(),
        "type_garde_info": {
            "nom": type_garde.get("nom", "N/A"),
            "duree_heures": type_garde.get("duree_heures", 8),
            "personnel_requis": type_garde.get("personnel_requis", 1)
        }
    }


async def traiter_semaine_attribution_auto(tenant, semaine_debut: str, semaine_fin: str, progress: AttributionProgress = None):
    """Traite l'attribution automatique pour une seule semaine avec suivi de performance"""
    perf_start = time.time()
    
    try:
        # Get all available users and types de garde pour ce tenant
        users = await db.users.find({"statut": "Actif", "tenant_id": tenant.id}).to_list(1000)
        types_garde = await db.types_garde.find({"tenant_id": tenant.id}).to_list(1000)
        
        # Créer un dictionnaire pour lookup rapide des utilisateurs par ID
        users_dict = {u["id"]: u for u in users}
        
        # Récupérer les paramètres de remplacements (incluant gestion heures sup)
        parametres = await db.parametres_remplacements.find_one({"tenant_id": tenant.id})
        if not parametres:
            # Créer des paramètres par défaut
            default_params = ParametresRemplacements(tenant_id=tenant.id)
            await db.parametres_remplacements.insert_one(default_params.dict())
            parametres = default_params.dict()
        
        activer_heures_sup = parametres.get("activer_gestion_heures_sup", False)
        
        # Récupérer les paramètres des équipes de garde
        params_equipes_garde = await db.parametres_equipes_garde.find_one({"tenant_id": tenant.id})
        equipes_garde_actif = params_equipes_garde.get("actif", False) if params_equipes_garde else False
        privilegier_equipe_garde_tp = False
        config_temps_partiel = {}
        if equipes_garde_actif and params_equipes_garde:
            config_temps_partiel = params_equipes_garde.get("temps_partiel", {})
            privilegier_equipe_garde_tp = config_temps_partiel.get("privilegier_equipe_garde", False)
        
        logging.info(f"📊 [EQUIPE GARDE] Actif: {equipes_garde_actif}, Prioriser équipe de garde: {privilegier_equipe_garde_tp}")
        
        # CORRECTION CRITIQUE: Charger les paramètres des niveaux d'attribution
        niveaux_actifs = {
            "niveau_2": tenant.parametres.get("niveau_2_actif", True),
            "niveau_3": tenant.parametres.get("niveau_3_actif", True),
            "niveau_4": tenant.parametres.get("niveau_4_actif", True),
            "niveau_5": tenant.parametres.get("niveau_5_actif", True)
        }
        logging.info(f"📊 [NIVEAUX] Niveaux d'attribution actifs: {niveaux_actifs}")
        
        # Récupérer les grades pour vérifier les officiers
        grades = await db.grades.find({"tenant_id": tenant.id}).to_list(1000)
        grades_map = {g["nom"]: g for g in grades}
        
        # Récupérer les compétences pour la priorisation des gardes
        competences = await db.competences.find({"tenant_id": tenant.id}).to_list(1000)
        
        # Get existing assignations for the week
        # NOTE: Ne PAS écraser semaine_fin car il est passé correctement depuis la boucle appelante
        # (Bug précédent: la ligne suivante écrasait semaine_fin et limitait à 7 jours)
        if not semaine_fin:
            semaine_fin = (datetime.strptime(semaine_debut, "%Y-%m-%d") + timedelta(days=6)).strftime("%Y-%m-%d")
        existing_assignations = await db.assignations.find({
            "date": {
                "$gte": semaine_debut,
                "$lte": semaine_fin
            },
            "tenant_id": tenant.id
        }).to_list(1000)
        
        # Get monthly statistics for rotation équitable (current month)
        current_month_start = datetime.strptime(semaine_debut, "%Y-%m-%d").replace(day=1).strftime("%Y-%m-%d")
        current_month_end = (datetime.strptime(current_month_start, "%Y-%m-%d") + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        current_month_end = current_month_end.strftime("%Y-%m-%d")
        
        monthly_assignations = await db.assignations.find({
            "date": {
                "$gte": current_month_start,
                "$lte": current_month_end
            },
            "tenant_id": tenant.id
        }).to_list(1000)
        
        # ⚡ OPTIMIZATION: Précharger TOUTES les disponibilités de la semaine en UNE SEULE requête
        # Cela évite le problème N+1 (une requête par user/garde)
        # Note: Accepter plusieurs formats de statut (disponible, Disponible, dispo, etc.)
        all_disponibilites = await db.disponibilites.find({
            "date": {
                "$gte": semaine_debut,
                "$lte": semaine_fin
            },
            "statut": {"$regex": "^dispo", "$options": "i"},  # Accepte "disponible", "Disponible", "dispo"
            "tenant_id": tenant.id
        }).to_list(10000)
        
        logging.info(f"📅 [DISPOS] {len(all_disponibilites)} disponibilités trouvées pour la période {semaine_debut} - {semaine_fin}")
        
        # Créer un index/dictionnaire pour lookup rapide
        # Structure: {user_id: {date: {type_garde_id: [list of dispos with horaires]}}}
        # Note: type_garde_id peut être None pour les disponibilités générales (toutes gardes)
        dispos_lookup = {}
        for dispo in all_disponibilites:
            user_id = dispo.get("user_id")
            date = dispo.get("date")
            type_garde_id = dispo.get("type_garde_id")  # Peut être None = disponible pour toutes les gardes
            
            if user_id not in dispos_lookup:
                dispos_lookup[user_id] = {}
            if date not in dispos_lookup[user_id]:
                dispos_lookup[user_id][date] = {"_general": []}  # _general pour les dispos sans type spécifique
            
            # Stocker sous le type_garde_id spécifique OU sous _general si None
            key = type_garde_id if type_garde_id else "_general"
            if key not in dispos_lookup[user_id][date]:
                dispos_lookup[user_id][date][key] = []
            
            # Stocker la dispo complète avec ses horaires
            # IMPORTANT: S'assurer que les heures ont des valeurs par défaut si None
            heure_debut = dispo.get("heure_debut") or "00:00"
            heure_fin = dispo.get("heure_fin") or "23:59"
            dispos_lookup[user_id][date][key].append({
                "heure_debut": heure_debut,
                "heure_fin": heure_fin
            })
        
        logging.info(f"📅 [DISPOS] Lookup créé pour {len(dispos_lookup)} utilisateurs")
        
        # ⚡ OPTIMIZATION: Précharger TOUTES les indisponibilités de la semaine
        all_indisponibilites = await db.disponibilites.find({
            "date": {
                "$gte": semaine_debut,
                "$lte": semaine_fin
            },
            "statut": "indisponible",
            "tenant_id": tenant.id
        }).to_list(10000)
        
        # Créer un index pour les indisponibilités
        # Structure: {user_id: {date: True}}
        # PRIORITÉ: Les disponibilités manuelles ont priorité sur les indisponibilités auto-générées
        indispos_lookup = {}
        # Nouveau: Stocker aussi les détails horaires des indisponibilités
        # Structure: {user_id: {date: [{heure_debut, heure_fin}, ...]}}
        indispos_details_lookup = {}
        
        for indispo in all_indisponibilites:
            user_id = indispo.get("user_id")
            date = indispo.get("date")
            source = indispo.get("source", "manuel")  # Par défaut: manuel
            heure_debut = indispo.get("heure_debut", "00:00")
            heure_fin = indispo.get("heure_fin", "23:59")
            
            # Vérifier s'il existe une disponibilité manuelle pour ce user/date
            has_manual_dispo = any(
                d.get("user_id") == user_id and 
                d.get("date") == date and 
                d.get("source", "manuel") == "manuel"
                for d in all_disponibilites
            )
            
            # N'ajouter l'indisponibilité que si:
            # - C'est une indispo manuelle OU
            # - Il n'y a pas de dispo manuelle qui la contredit
            if source == "manuel" or not has_manual_dispo:
                if user_id not in indispos_lookup:
                    indispos_lookup[user_id] = {}
                indispos_lookup[user_id][date] = True
                
                # Stocker les détails horaires
                if user_id not in indispos_details_lookup:
                    indispos_details_lookup[user_id] = {}
                if date not in indispos_details_lookup[user_id]:
                    indispos_details_lookup[user_id][date] = []
                indispos_details_lookup[user_id][date].append({
                    "heure_debut": heure_debut,
                    "heure_fin": heure_fin
                })
            else:
                logging.info(f"✅ [CONFLIT RÉSOLU] Indispo auto-générée ignorée pour {user_id} le {date} (dispo manuelle trouvée)")
        
        # Récupérer les paramètres d'équité
        params_planning = await db.parametres_validation_planning.find_one({"tenant_id": tenant.id})
        periode_equite = params_planning.get("periode_equite", "mensuel") if params_planning else "mensuel"
        periode_equite_jours = params_planning.get("periode_equite_jours", 30) if params_planning else 30
        
        # Calculer la date de début de la période d'équité
        start_date = datetime.strptime(semaine_debut, "%Y-%m-%d")
        end_date = datetime.strptime(semaine_fin, "%Y-%m-%d")
        date_debut_periode = start_date
        if periode_equite == "hebdomadaire":
            # Début de la semaine (lundi)
            jours_depuis_lundi = date_debut_periode.weekday()
            date_debut_periode = date_debut_periode - timedelta(days=jours_depuis_lundi)
        elif periode_equite == "bi-hebdomadaire":
            # Début de la bi-semaine (14 jours glissants)
            date_debut_periode = start_date - timedelta(days=14)
        elif periode_equite == "mensuel":
            # Début du mois
            date_debut_periode = date_debut_periode.replace(day=1)
        elif periode_equite == "personnalise":
            # Période personnalisée en jours
            date_debut_periode = start_date - timedelta(days=periode_equite_jours)
        
        logging.info(f"📊 [ÉQUITÉ] Période: {periode_equite}, Début: {date_debut_periode}, Jours: {periode_equite_jours if periode_equite == 'personnalise' else 'N/A'}")
        
        # Récupérer les assignations de la période d'équité
        assignations_periode = await db.assignations.find({
            "tenant_id": tenant.id,
            "date": {
                "$gte": date_debut_periode.strftime("%Y-%m-%d"),
                "$lt": end_date.strftime("%Y-%m-%d")
            }
        }).to_list(length=None)
        
        logging.info(f"📊 [ÉQUITÉ] {len(assignations_periode)} assignations trouvées pour la période d'équité")
        
        # Calculate hours for each user based on equity period (séparé interne/externe)
        user_monthly_hours_internes = {}
        user_monthly_hours_externes = {}
        for user in users:
            user_hours_internes = 0
            user_hours_externes = 0
            for assignation in assignations_periode:
                if assignation["user_id"] == user["id"]:
                    # Find type garde to get duration
                    type_garde = next((t for t in types_garde if t["id"] == assignation["type_garde_id"]), None)
                    if type_garde:
                        duree = type_garde.get("duree_heures", 8)
                        # Séparer les heures selon le type de garde
                        if type_garde.get("est_garde_externe", False):
                            user_hours_externes += duree
                        else:
                            user_hours_internes += duree
            user_monthly_hours_internes[user["id"]] = user_hours_internes
            user_monthly_hours_externes[user["id"]] = user_hours_externes
        
        logging.info(f"📊 [ÉQUITÉ] Heures calculées pour {len(users)} utilisateurs sur la période")
        
        # ==================== RÉCUPÉRATION DES PARAMÈTRES ====================
        # Lire les paramètres de remplacement pour heures_supplementaires_activees
        params_remplacements = await db.parametres_remplacements.find_one({"tenant_id": tenant.id})
        heures_sup_from_params = False
        if params_remplacements:
            heures_sup_from_params = params_remplacements.get("heures_supplementaires_activees", False)
        
        # Charger les niveaux depuis les paramètres du TENANT
        tenant_params = tenant.parametres or {}
        logging.info(f"📊 [NIVEAUX] Paramètres tenant: {tenant_params}")
        niveaux_actifs = {
            "niveau_2": tenant_params.get("niveau_2_actif", True),  # Temps partiel DISPONIBLES
            "niveau_3": tenant_params.get("niveau_3_actif", True),  # Temps partiel STAND-BY
            "niveau_4": tenant_params.get("niveau_4_actif", True),  # Temps plein INCOMPLETS
            "niveau_5": tenant_params.get("niveau_5_actif", True)   # Temps plein COMPLETS (heures sup)
        }
        # Activer heures sup si:
        # - Le paramètre dans la requête est True OU
        # - Le paramètre dans parametres_remplacements est True
        autoriser_heures_sup = activer_heures_sup or heures_sup_from_params or tenant_params.get("autoriser_heures_supplementaires", False)
        
        # Si heures sup non autorisées, désactiver niveau 5
        if not autoriser_heures_sup:
            niveaux_actifs["niveau_5"] = False
        
        logging.info(f"📋 Niveaux actifs: {niveaux_actifs}, Heures sup: {autoriser_heures_sup} (from_params={heures_sup_from_params}, activer_heures_sup={activer_heures_sup})")
        
        # Paramètres de paie pour seuil hebdomadaire temps plein
        params_paie = await db.parametres_paie.find_one({"tenant_id": tenant.id})
        seuil_hebdo_temps_plein = 40  # Défaut
        if params_paie:
            seuil_hebdo_temps_plein = params_paie.get("seuil_hebdomadaire", 40)
        
        logging.info(f"📋 Seuil hebdo temps plein: {seuil_hebdo_temps_plein}h")
        
        # ==================== RÉCUPÉRATION DES GRADES ====================
        grades_list = await db.grades.find({"tenant_id": tenant.id}).to_list(100)
        grades_map = {g.get("nom"): g for g in grades_list}
        
        def est_officier(user):
            """Vérifie si l'utilisateur est un officier basé sur son grade"""
            grade_nom = user.get("grade", "")
            grade_info = grades_map.get(grade_nom, {})
            return grade_info.get("est_officier", False) == True
        
        def est_eligible_fonction_superieure(user):
            """Vérifie si l'utilisateur peut opérer en fonction supérieure (grade +1)"""
            return user.get("fonction_superieur", False) == True
        
        def get_niveau_grade(user):
            """Retourne le niveau hiérarchique du grade de l'utilisateur"""
            grade_nom = user.get("grade", "")
            grade_info = grades_map.get(grade_nom, {})
            return grade_info.get("niveau_hierarchique", 0) or 0
        
        def user_a_competences_requises(user, competences_requises):
            """Vérifie si l'utilisateur possède toutes les compétences requises"""
            if not competences_requises:
                return True  # Pas de compétences requises
            user_competences = set(user.get("competences", []) or [])
            return all(comp_id in user_competences for comp_id in competences_requises)
        
        # ==================== PRÉPARATION DES DONNÉES UTILISATEURS ====================
        # Créer un dictionnaire pour accès rapide aux infos utilisateurs
        users_map = {u["id"]: u for u in users}
        
        # Calculer les heures hebdomadaires pour chaque utilisateur
        # (pour déterminer si temps plein incomplet ou complet)
        def get_heures_semaine(user_id, date_str, type_garde_externe=False):
            """Calcule les heures dans la semaine contenant date_str pour un type de garde (interne ou externe)"""
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            # Trouver le lundi de cette semaine
            lundi = date_obj - timedelta(days=date_obj.weekday())
            dimanche = lundi + timedelta(days=6)
            
            heures = 0
            for a in existing_assignations:
                if a.get("user_id") != user_id:
                    continue
                try:
                    a_date = datetime.strptime(a.get("date", ""), "%Y-%m-%d")
                    if lundi <= a_date <= dimanche:
                        tg = next((t for t in types_garde if t["id"] == a.get("type_garde_id")), None)
                        if tg:
                            is_externe = tg.get("est_garde_externe", False)
                            if type_garde_externe == is_externe:  # Compter séparément
                                heures += tg.get("duree_heures", 8)
                except:
                    pass
            return heures
        
        def get_heures_travaillees_semaine(user_id, date_str):
            """
            Calcule les heures TRAVAILLÉES dans la semaine contenant date_str.
            IMPORTANT: Seules les gardes INTERNES comptent comme heures travaillées.
            Les gardes EXTERNES ne comptent PAS vers le max d'heures.
            """
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            # Trouver le lundi de cette semaine
            lundi = date_obj - timedelta(days=date_obj.weekday())
            dimanche = lundi + timedelta(days=6)
            
            heures = 0
            for a in existing_assignations:
                if a.get("user_id") != user_id:
                    continue
                try:
                    a_date = datetime.strptime(a.get("date", ""), "%Y-%m-%d")
                    if lundi <= a_date <= dimanche:
                        tg = next((t for t in types_garde if t["id"] == a.get("type_garde_id")), None)
                        if tg:
                            # IMPORTANT: Seules les gardes INTERNES comptent
                            is_externe = tg.get("est_garde_externe", False)
                            if not is_externe:
                                heures += tg.get("duree_heures", 8)
                except:
                    pass
            return heures
        
        def get_heures_max_semaine(user):
            """Retourne le max d'heures par semaine pour cet utilisateur"""
            # Utiliser la valeur définie dans la fiche employé si disponible
            user_max = user.get("heures_max_semaine")
            if user_max is not None and user_max > 0:
                return user_max
            
            # Sinon, utiliser la valeur par défaut selon le type d'emploi
            type_emploi = user.get("type_emploi", "temps_plein")
            if type_emploi in ["temps_partiel", "temporaire"]:
                return 20  # Défaut temps partiel
            else:
                return seuil_hebdo_temps_plein  # Défaut temps plein
        
        def plages_se_chevauchent(debut1, fin1, debut2, fin2):
            """
            Vérifie si deux plages horaires se chevauchent.
            Les heures sont au format 'HH:MM'.
            Gère les gardes qui passent minuit (ex: 18:00-06:00).
            Retourne True si les plages se chevauchent.
            """
            # Convertir en minutes pour faciliter la comparaison
            def to_minutes(time_str):
                if not time_str:
                    return 0
                parts = time_str.split(':')
                return int(parts[0]) * 60 + int(parts[1])
            
            start1 = to_minutes(debut1)
            end1 = to_minutes(fin1)
            start2 = to_minutes(debut2)
            end2 = to_minutes(fin2)
            
            # Gérer le cas où fin = 23:59 (toute la journée)
            if end1 == 23 * 60 + 59:
                end1 = 24 * 60
            if end2 == 23 * 60 + 59:
                end2 = 24 * 60
            
            # Gérer les gardes qui passent minuit (fin < debut)
            # Ex: 18:00-06:00 -> on considère que ça se termine à 06:00 le LENDEMAIN
            # Pour la comparaison sur une même journée, une garde de nuit (18:00-06:00)
            # ne chevauche PAS une garde de jour (06:00-18:00)
            
            # Si la plage 1 passe minuit (ex: 18:00-06:00)
            if end1 < start1:
                # Elle couvre [start1, 24:00] et [00:00, end1]
                # Pour simplifier sur une journée: [start1, 24:00]
                end1 = 24 * 60
            
            # Si la plage 2 passe minuit (ex: 18:00-06:00)
            if end2 < start2:
                # Elle couvre [start2, 24:00] et [00:00, end2]
                # Pour simplifier sur une journée: [start2, 24:00]
                end2 = 24 * 60
            
            # Les plages se chevauchent si: start1 < end2 AND start2 < end1
            return start1 < end2 and start2 < end1
        
        def est_disponible_pour_garde(user_id, date_str, garde_heure_debut, garde_heure_fin, type_garde_id):
            """
            Vérifie si l'utilisateur a une disponibilité qui couvre la plage horaire de la garde.
            Retourne True si l'utilisateur est disponible pour cette garde.
            
            LOGIQUE MÉTIER:
            - Si la dispo a le MÊME type_garde_id et les MÊMES heures → match parfait
            - Pour une garde de jour: la dispo doit couvrir toute la plage
            - Pour une garde de nuit: il suffit que la dispo couvre le DÉBUT de la garde
            """
            def to_minutes(time_str):
                if not time_str:
                    return 0
                parts = time_str.split(':')
                return int(parts[0]) * 60 + int(parts[1])
            
            # Récupérer toutes les dispos (spécifiques + générales) pour cette date
            user_dispos = dispos_lookup.get(user_id, {}).get(date_str, {})
            specific_dispos = user_dispos.get(type_garde_id, [])
            general_dispos = user_dispos.get("_general", [])
            all_dispos = specific_dispos + general_dispos
            
            garde_start = to_minutes(garde_heure_debut)
            garde_end = to_minutes(garde_heure_fin)
            
            # Gérer 23:59 comme fin de journée
            if garde_end == 23 * 60 + 59:
                garde_end = 24 * 60
            
            # Déterminer si c'est une garde de nuit (traverse minuit)
            est_garde_nuit = garde_end < garde_start
            
            for dispo in all_dispos:
                dispo_debut = dispo.get("heure_debut") or "00:00"
                dispo_fin = dispo.get("heure_fin") or "23:59"
                
                dispo_start = to_minutes(dispo_debut)
                dispo_end = to_minutes(dispo_fin)
                
                # Gérer 23:59 comme fin de journée (= minuit)
                if dispo_end == 23 * 60 + 59:
                    dispo_end = 24 * 60
                
                # CAS 1: Match parfait des horaires (ex: dispo 18:00-06:00 pour garde 18:00-06:00)
                # Peu importe si ça traverse minuit, si les heures sont identiques c'est OK
                if dispo_debut == garde_heure_debut and dispo_fin == garde_heure_fin:
                    return True
                
                # CAS 2: La dispo elle-même traverse minuit (ex: 18:00-06:00)
                dispo_traverse_minuit = dispo_end < dispo_start
                if dispo_traverse_minuit:
                    # La dispo couvre de dispo_start jusqu'à minuit, puis de minuit jusqu'à dispo_end
                    # Pour une garde de nuit qui commence à garde_start, vérifier si dispo_start <= garde_start
                    if dispo_start <= garde_start:
                        return True
                    continue
                
                # CAS 3: Logique normale
                if est_garde_nuit:
                    # GARDE DE NUIT: Il suffit que la dispo couvre le DÉBUT de la garde
                    # Ex: Garde 18:00->06:00, dispo 00:00->23:59 = OK car couvre 18:00
                    if dispo_start <= garde_start and dispo_end > garde_start:
                        return True
                else:
                    # GARDE DE JOUR: La dispo doit couvrir toute la plage horaire
                    if dispo_start <= garde_start and dispo_end >= garde_end:
                        return True
            
            return False
        
        def a_indisponibilite_bloquante(user_id, date_str, garde_heure_debut, garde_heure_fin):
            """
            Vérifie si l'utilisateur a une indisponibilité qui chevauche la plage horaire de la garde.
            Retourne True si l'utilisateur est BLOQUÉ par une indisponibilité.
            """
            user_indispos = indispos_details_lookup.get(user_id, {}).get(date_str, [])
            
            for indispo in user_indispos:
                indispo_debut = indispo.get("heure_debut", "00:00")
                indispo_fin = indispo.get("heure_fin", "23:59")
                
                # Si l'indisponibilité chevauche la garde, l'utilisateur est bloqué
                if plages_se_chevauchent(garde_heure_debut, garde_heure_fin, indispo_debut, indispo_fin):
                    return True
            
            return False
        
        def trier_candidats_equite_anciennete(candidats, type_garde_externe=False, prioriser_officiers=False, user_monthly_hours=None, equipe_garde_du_jour=None, prioriser_equipe_garde=False):
            """
            Trie les candidats selon:
            0. Si prioriser_equipe_garde: Membres de l'équipe de garde du jour d'abord (NON ABSOLU - juste priorité)
            1. Si prioriser_officiers: Officiers d'abord, puis éligibles, puis autres
            2. Équité (moins d'heures d'abord)
            3. Ancienneté (plus ancien d'abord)
            """
            if user_monthly_hours is None:
                user_monthly_hours = user_monthly_hours_internes
                
            def sort_key(user):
                user_id = user["id"]
                
                # Priorité équipe de garde (0=membre équipe garde, 1=autre)
                # C'est une priorité NON ABSOLUE - elle influence le tri mais ne bloque pas les autres
                if prioriser_equipe_garde and equipe_garde_du_jour and type_garde_externe:
                    # La priorité équipe de garde s'applique uniquement aux gardes EXTERNES
                    equipe_utilisateur = user.get("equipe_garde")
                    if equipe_utilisateur == equipe_garde_du_jour:
                        equipe_priority = 0  # Membre de l'équipe de garde
                    else:
                        equipe_priority = 1  # Pas membre
                else:
                    equipe_priority = 0  # Pas de tri par équipe de garde
                
                # Priorité officier (0=officier, 1=éligible, 2=autre)
                if prioriser_officiers:
                    if est_officier(user):
                        officier_priority = 0
                    elif est_eligible_fonction_superieure(user):
                        officier_priority = 1
                    else:
                        officier_priority = 2
                else:
                    officier_priority = 0  # Pas de tri par officier
                
                # Heures selon le type de garde (interne ou externe)
                if type_garde_externe:
                    heures = user_monthly_hours_externes.get(user_id, 0)
                else:
                    heures = user_monthly_hours.get(user_id, 0)
                
                # Ancienneté (date_embauche)
                date_embauche = user.get("date_embauche", "2099-12-31")
                
                return (equipe_priority, officier_priority, heures, date_embauche)
            
            return sorted(candidats, key=sort_key)
        
        # ==================== INITIALISATION ====================
        nouvelles_assignations = []
        
        # ==================== ATTRIBUTION AUTOMATIQUE À 5 NIVEAUX ====================
        current_date = datetime.strptime(semaine_debut, "%Y-%m-%d")
        end_date = datetime.strptime(semaine_fin, "%Y-%m-%d")
        
        while current_date <= end_date:
            date_str = current_date.strftime("%Y-%m-%d")
            day_name = current_date.strftime("%A").lower()
            
            if progress:
                progress.current_step = f"📅 Traitement du {date_str}..."
            
            # Calculer l'équipe de garde du jour si le système est actif
            equipe_garde_du_jour = None
            if privilegier_equipe_garde_tp and config_temps_partiel.get("rotation_active"):
                # Calculer quelle équipe est de garde ce jour
                equipe_garde_du_jour = get_equipe_garde_du_jour_sync(
                    type_rotation=config_temps_partiel.get("type_rotation", "hebdomadaire"),
                    date_reference=config_temps_partiel.get("date_reference", date_str),
                    date_cible=date_str,
                    nombre_equipes=config_temps_partiel.get("nombre_equipes", 2),
                    pattern_mode=config_temps_partiel.get("pattern_mode", "hebdomadaire"),
                    pattern_personnalise=config_temps_partiel.get("pattern_personnalise", []),
                    duree_cycle=config_temps_partiel.get("duree_cycle", 14),
                    jour_rotation=config_temps_partiel.get("jour_rotation") or "monday",
                    heure_rotation=config_temps_partiel.get("heure_rotation") or "18:00",
                    heure_actuelle="12:00"  # Milieu de journée pour la comparaison
                )
                logging.info(f"📊 [EQUIPE GARDE] {date_str}: Équipe {equipe_garde_du_jour} de garde")
            
            # N0: Priorisation des types de garde (par priorité intrinsèque)
            types_garde_tries = sorted(types_garde, key=lambda t: t.get("priorite", 99))
            
            for type_garde in types_garde_tries:
                type_garde_id = type_garde["id"]
                type_garde_nom = type_garde.get("nom", "Garde")
                personnel_requis = type_garde.get("personnel_requis", 1)
                est_externe = type_garde.get("est_garde_externe", False)
                duree_garde = type_garde.get("duree_heures", 8)
                competences_requises = type_garde.get("competences_requises", [])
                officier_obligatoire = type_garde.get("officier_obligatoire", False)
                # IMPORTANT: Extraire les heures de début et fin de la garde
                heure_debut = type_garde.get("heure_debut", "00:00")
                heure_fin = type_garde.get("heure_fin", "23:59")
                
                # Vérifier si ce type de garde s'applique ce jour
                jours_app = type_garde.get("jours_application", [])
                if jours_app and len(jours_app) > 0 and day_name not in jours_app:
                    continue
                
                # N1: Compter les assignations MANUELLES existantes (ne jamais écraser)
                existing_this_garde = [
                    a for a in existing_assignations
                    if a.get("date") == date_str and a.get("type_garde_id") == type_garde_id
                ]
                existing_count = len(existing_this_garde)
                
                places_restantes = personnel_requis - existing_count
                
                if places_restantes <= 0:
                    continue  # Garde complète
                
                # Vérifier si un officier est déjà assigné à cette garde
                officier_deja_assigne = False
                if officier_obligatoire:
                    for a in existing_this_garde:
                        assigned_user = users_map.get(a.get("user_id"))
                        if assigned_user and est_officier(assigned_user):
                            officier_deja_assigne = True
                            break
                
                # Utilisateurs déjà assignés à des gardes qui CHEVAUCHENT cette garde
                # Une personne peut faire garde de jour + garde de nuit le même jour si elles ne se chevauchent pas
                def garde_chevauche(assignation_existante):
                    """Vérifie si une assignation existante chevauche la garde actuelle"""
                    tg_existant = next((t for t in types_garde if t["id"] == assignation_existante.get("type_garde_id")), None)
                    if not tg_existant:
                        return False
                    
                    existing_debut = tg_existant.get("heure_debut", "00:00")
                    existing_fin = tg_existant.get("heure_fin", "23:59")
                    
                    # Debug pour le 12 février
                    if date_str == "2026-02-12" and "jour" in type_garde_nom.lower():
                        chevauche = plages_se_chevauchent(heure_debut, heure_fin, existing_debut, existing_fin)
                        logging.info(f"🔎 Comparaison: {type_garde_nom}({heure_debut}-{heure_fin}) vs {tg_existant.get('nom')}({existing_debut}-{existing_fin}) = {chevauche}")
                        return chevauche
                    
                    return plages_se_chevauchent(heure_debut, heure_fin, existing_debut, existing_fin)
                
                # Séparer les utilisateurs déjà assignés à cette même garde vs une autre garde
                users_assignes_cette_garde = set(
                    a.get("user_id") for a in existing_assignations
                    if a.get("date") == date_str and a.get("type_garde_id") == type_garde_id
                )
                
                users_assignes_autre_garde = set(
                    a.get("user_id") for a in existing_assignations
                    if a.get("date") == date_str and a.get("type_garde_id") != type_garde_id and garde_chevauche(a)
                )
                
                users_assignes_ce_jour = users_assignes_cette_garde | users_assignes_autre_garde
                
                # Debug log pour le 12 février garde de jour
                if date_str == "2026-02-12" and "jour" in type_garde_nom.lower():
                    logging.info(f"🔍 DEBUG {date_str} {type_garde_nom}: users_assignes_ce_jour = {len(users_assignes_ce_jour)}")
                    for uid in users_assignes_ce_jour:
                        u = users_map.get(uid)
                        if u:
                            logging.info(f"   - {u.get('prenom')} {u.get('nom')} bloqué")
                
                # Récupérer les disponibilités pour ce jour/type_garde
                # IMPORTANT: Inclure aussi les disponibilités générales (type_garde_id = None)
                def get_user_dispos(user_id):
                    user_dispos = dispos_lookup.get(user_id, {}).get(date_str, {})
                    # Disponibilités spécifiques à ce type de garde
                    specific_dispos = user_dispos.get(type_garde_id, [])
                    # Disponibilités générales (toute la journée, sans type_garde spécifique)
                    general_dispos = user_dispos.get(None, [])
                    return specific_dispos + general_dispos
                
                # Récupérer les indisponibilités pour ce jour
                def has_indisponibilite(user_id):
                    user_indispos = indispos_lookup.get(user_id, {})
                    return date_str in user_indispos
                
                # ==================== N0: FILTRE COMPÉTENCES ====================
                # Pré-filtrer les utilisateurs qui ont les compétences requises
                users_avec_competences = [
                    u for u in users 
                    if user_a_competences_requises(u, competences_requises)
                ]
                
                # Log si peu de candidats avec les compétences requises
                if competences_requises and len(users_avec_competences) < len(users) // 2:
                    logging.info(f"  ⚠️ {type_garde_nom}: seulement {len(users_avec_competences)}/{len(users)} ont les compétences requises")
                
                # ==================== N0-bis: FILTRE GARDES EXTERNES ====================
                # Si c'est une garde externe, ne garder que les utilisateurs qui acceptent les gardes externes
                # LOGIQUE OPT-OUT: Par défaut tous acceptent, sauf ceux qui ont explicitement refusé (False)
                if est_externe:
                    nb_avant = len(users_avec_competences)
                    users_avec_competences = [
                        u for u in users_avec_competences
                        if u.get("accepter_gardes_externes") != False  # Exclure SEULEMENT ceux qui ont explicitement refusé
                    ]
                    logging.info(f"  🏠 Garde externe: {len(users_avec_competences)}/{nb_avant} candidats acceptant les gardes externes")
                
                # ==================== LOGIQUE OFFICIER OBLIGATOIRE ====================
                # Si officier obligatoire et pas encore d'officier assigné
                besoin_officier = officier_obligatoire and not officier_deja_assigne
                officier_assigne_cette_iteration = False
                
                # ==================== NIVEAUX 2-5 ====================
                assignes_cette_garde = 0
                
                for niveau in [2, 3, 4, 5]:
                    if assignes_cette_garde >= places_restantes:
                        break
                    
                    if not niveaux_actifs.get(f"niveau_{niveau}", True):
                        continue  # Niveau désactivé
                    
                    candidats = []
                    candidats_rejetes = []  # Pour l'audit: stocker les candidats non sélectionnés avec leurs raisons
                    
                    for user in users_avec_competences:  # Utiliser la liste filtrée par compétences
                        user_id = user["id"]
                        user_name = f"{user.get('prenom', '')} {user.get('nom', '')}"
                        
                        # Debug pour Alva le 12 février garde de jour
                        is_debug = date_str == "2026-02-12" and "jour" in type_garde_nom.lower() and "Alva" in user_name
                        
                        # Ignorer si déjà assigné ce jour à une garde qui chevauche
                        if user_id in users_assignes_ce_jour:
                            if is_debug:
                                logging.info(f"🔴 DEBUG Alva bloqué: déjà assigné à garde chevauchante")
                            
                            # Si déjà assigné à CETTE MÊME garde, ne pas l'ajouter aux rejetés
                            # car il a été sélectionné pour un slot précédent
                            if user_id in users_assignes_cette_garde:
                                # Ne rien faire - il est déjà sur cette garde
                                pass
                            else:
                                # Trouver le nom de l'autre garde
                                autre_garde = next(
                                    (a for a in existing_assignations 
                                     if a.get("user_id") == user_id and a.get("date") == date_str and a.get("type_garde_id") != type_garde_id),
                                    None
                                )
                                if autre_garde:
                                    autre_tg = next((t for t in types_garde if t["id"] == autre_garde.get("type_garde_id")), None)
                                    autre_nom = autre_tg.get("nom", "Autre garde") if autre_tg else "Autre garde"
                                    raison = f"Déjà assigné à '{autre_nom}' ce jour (conflit d'horaire)"
                                else:
                                    raison = "Déjà assigné à une autre garde ce jour"
                                
                                candidats_rejetes.append({
                                    "nom_complet": user_name,
                                    "grade": user.get("grade", ""),
                                    "type_emploi": user.get("type_emploi", ""),
                                    "raison_rejet": raison
                                })
                            continue
                        
                        # Ignorer si statut inactif
                        if user.get("statut") != "Actif":
                            if is_debug:
                                logging.info(f"🔴 DEBUG Alva bloqué: statut inactif")
                            candidats_rejetes.append({
                                "nom_complet": user_name,
                                "grade": user.get("grade", ""),
                                "type_emploi": user.get("type_emploi", ""),
                                "raison_rejet": "Statut inactif"
                            })
                            continue
                        
                        type_emploi = user.get("type_emploi", "temps_plein")
                        
                        # Heures travaillées = UNIQUEMENT les gardes INTERNES
                        # Les gardes externes ne comptent PAS vers le max d'heures
                        heures_travaillees = get_heures_travaillees_semaine(user_id, date_str)
                        heures_max = get_heures_max_semaine(user)
                        
                        # Pour une garde INTERNE : vérifier si on dépasse le max
                        # Pour une garde EXTERNE : pas de vérification du max (ne compte pas comme travaillé)
                        if est_externe:
                            # Garde externe - ne compte pas vers le max, toujours OK niveau heures
                            depasserait_max = False
                        else:
                            # Garde interne - vérifier le max
                            depasserait_max = (heures_travaillees + duree_garde) > heures_max
                        
                        # Vérification des plages horaires
                        # 1. Vérifie si l'utilisateur a une DISPONIBILITÉ qui COUVRE la garde
                        has_dispo_valide = est_disponible_pour_garde(user_id, date_str, heure_debut, heure_fin, type_garde_id)
                        # 2. Vérifie si l'utilisateur a une INDISPONIBILITÉ qui CHEVAUCHE la garde
                        has_indispo_bloquante = a_indisponibilite_bloquante(user_id, date_str, heure_debut, heure_fin)
                        
                        if is_debug:
                            logging.info(f"🔵 DEBUG Alva: niveau={niveau}, type_emploi={type_emploi}, heures={heures_travaillees}/{heures_max}, depasserait={depasserait_max}, dispo={has_dispo_valide}, indispo={has_indispo_bloquante}")
                        
                        # RÈGLE PRIORITAIRE: Si indisponibilité chevauche la garde, l'utilisateur est BLOQUÉ
                        if has_indispo_bloquante:
                            if is_debug:
                                logging.info(f"🔴 DEBUG Alva bloqué: indisponibilité")
                            candidats_rejetes.append({
                                "nom_complet": user_name,
                                "grade": user.get("grade", ""),
                                "type_emploi": type_emploi,
                                "heures_ce_mois": user_monthly_hours_internes.get(user_id, 0),
                                "raison_rejet": "Indisponibilité sur cette plage horaire"
                            })
                            continue
                        
                        # Variables pour tracking des raisons de rejet par niveau
                        accepte = False
                        raison_rejet = ""
                        
                        # Construire une explication détaillée
                        explication_niveaux = []
                        
                        # DEBUG pour gardes externes de nuit uniquement
                        # N2: Temps partiel DISPONIBLES
                        if niveau == 2:
                            if type_emploi in ["temps_partiel", "temporaire"] and has_dispo_valide:
                                # Vérifier qu'il n'a pas atteint son max (sauf garde externe)
                                if not depasserait_max:
                                    candidats.append(user)
                                    accepte = True
                                else:
                                    raison_rejet = f"A une dispo mais dépasserait {heures_max}h max ({heures_travaillees}h + {duree_garde}h)"
                            elif type_emploi not in ["temps_partiel", "temporaire"]:
                                raison_rejet = f"Temps plein (N2 = temps partiel)"
                            else:
                                raison_rejet = f"Pas de disponibilité (N2 requiert dispo)"
                        
                        # N3: Temps partiel STAND-BY (ni dispo explicite ni indispo bloquante)
                        elif niveau == 3:
                            if type_emploi in ["temps_partiel", "temporaire"] and not has_dispo_valide:
                                # Pas de dispo explicite mais pas d'indispo bloquante non plus
                                if not depasserait_max:
                                    candidats.append(user)
                                    accepte = True
                                else:
                                    raison_rejet = f"Dépasserait {heures_max}h max ({heures_travaillees}h + {duree_garde}h)"
                            elif type_emploi not in ["temps_partiel", "temporaire"]:
                                raison_rejet = f"Temps plein (N3 = temps partiel)"
                            else:
                                # A une dispo - expliquer pourquoi pas éligible N2 aussi
                                if depasserait_max:
                                    raison_rejet = f"A une dispo mais dépasserait {heures_max}h max → éligible N5 si heures sup autorisées"
                                else:
                                    raison_rejet = f"A une dispo → devrait être au N2 (vérifier les assignations)"
                        
                        # N4: Temps plein INCOMPLETS (heures < max de l'employé)
                        elif niveau == 4:
                            if type_emploi == "temps_plein":
                                # Pas encore au max d'heures de l'employé
                                if not depasserait_max:
                                    candidats.append(user)
                                    accepte = True
                                else:
                                    raison_rejet = f"Dépasserait {heures_max}h max ({heures_travaillees}h + {duree_garde}h) → éligible N5"
                            else:
                                raison_rejet = f"Temps partiel (N4 = temps plein)"
                        
                        # N5: HEURES SUPPLÉMENTAIRES (tous types d'emploi si autorisé)
                        elif niveau == 5:
                            # L'employé DÉPASSERAIT son max avec cette garde = heures supplémentaires
                            if depasserait_max:
                                if type_emploi in ["temps_partiel", "temporaire"]:
                                    # Temps partiel : DOIT avoir une dispo pour faire des heures sup
                                    if has_dispo_valide:
                                        candidats.append(user)
                                        accepte = True
                                    else:
                                        raison_rejet = f"Temps partiel en heures sup MAIS sans disponibilité"
                                else:
                                    # Temps plein : PAS besoin de dispo pour heures sup
                                    candidats.append(user)
                                    accepte = True
                            else:
                                raison_rejet = f"Pas en heures sup ({heures_travaillees}h < {heures_max}h max)"
                        
                        # Si pas accepté, ajouter aux rejetés
                        if not accepte and raison_rejet:
                            candidats_rejetes.append({
                                "nom_complet": user_name,
                                "grade": user.get("grade", ""),
                                "type_emploi": type_emploi,
                                "heures_ce_mois": user_monthly_hours_internes.get(user_id, 0),
                                "heures_semaine": heures_travaillees,
                                "heures_max": heures_max,
                                "raison_rejet": raison_rejet
                            })
                    
                    # Trier par équité puis ancienneté, avec priorité officier si nécessaire
                    # et priorité équipe de garde pour les gardes EXTERNES
                    candidats_tries = trier_candidats_equite_anciennete(
                        candidats, 
                        est_externe, 
                        prioriser_officiers=(besoin_officier and not officier_assigne_cette_iteration),
                        user_monthly_hours=user_monthly_hours_internes,
                        equipe_garde_du_jour=equipe_garde_du_jour,
                        prioriser_equipe_garde=privilegier_equipe_garde_tp
                    )
                    
                    # Assigner les candidats
                    for user in candidats_tries:
                        if assignes_cette_garde >= places_restantes:
                            break
                        
                        user_id = user["id"]
                        user_type_emploi = user.get("type_emploi", "temps_plein")
                        user_heures_travaillees = get_heures_travaillees_semaine(user_id, date_str)
                        user_heures_max = get_heures_max_semaine(user)
                        user_has_dispo = est_disponible_pour_garde(user_id, date_str, heure_debut, heure_fin, type_garde_id)
                        
                        # Si on a besoin d'un officier et qu'on vient d'en assigner un
                        if besoin_officier and not officier_assigne_cette_iteration:
                            if est_officier(user) or est_eligible_fonction_superieure(user):
                                officier_assigne_cette_iteration = True
                        
                        # Construire la liste des autres candidats (ceux qui n'ont pas été sélectionnés)
                        other_candidates_list = []
                        for other in candidats_tries:
                            if other["id"] != user_id:
                                other_heures = user_monthly_hours_internes.get(other["id"], 0)
                                other_candidates_list.append({
                                    "nom_complet": f"{other.get('prenom', '')} {other.get('nom', '')}",
                                    "grade": other.get("grade", ""),
                                    "type_emploi": other.get("type_emploi", ""),
                                    "heures_ce_mois": other_heures,
                                    "raison_non_selection": "Moins prioritaire (équité/ancienneté)"
                                })
                        
                        # Ajouter les candidats rejetés
                        other_candidates_list.extend(candidats_rejetes[:10])  # Limiter à 10 pour l'espace
                        
                        # Créer l'assignation
                        assignation = {
                            "id": str(uuid.uuid4()),
                            "tenant_id": tenant.id,
                            "user_id": user_id,
                            "type_garde_id": type_garde_id,
                            "date": date_str,
                            "statut": "assigne",
                            "auto_attribue": True,
                            "assignation_type": "auto",
                            "niveau_attribution": niveau,
                            "created_at": datetime.now(timezone.utc).isoformat(),
                            # Données d'audit/justification pour traçabilité (format attendu par le frontend)
                            "justification": {
                                "assigned_user": {
                                    "nom_complet": f"{user.get('prenom', '')} {user.get('nom', '')}",
                                    "grade": user.get("grade", ""),
                                    "type_emploi": user_type_emploi,
                                    "details": {
                                        "heures_ce_mois": user_monthly_hours_internes.get(user_id, 0),
                                        "heures_externes_ce_mois": user_monthly_hours_externes.get(user_id, 0),
                                        "heures_semaine": user_heures_travaillees,
                                        "heures_max": user_heures_max,
                                        "est_officier": est_officier(user),
                                        "est_eligible": est_eligible_fonction_superieure(user),
                                        "had_disponibilite": user_has_dispo,
                                        "date_embauche": user.get("date_embauche", "")
                                    }
                                },
                                "type_garde_info": {
                                    "nom": type_garde_nom,
                                    "heure_debut": heure_debut,
                                    "heure_fin": heure_fin,
                                    "duree_heures": duree_garde,
                                    "est_externe": est_externe
                                },
                                "niveau": niveau,
                                "niveau_description": {
                                    2: "Temps partiel DISPONIBLE",
                                    3: "Temps partiel STAND-BY",
                                    4: "Temps plein (heures incomplètes)",
                                    5: "Heures supplémentaires"
                                }.get(niveau, f"Niveau {niveau}"),
                                "total_candidates_evaluated": len(users_avec_competences),
                                "candidates_acceptes": len(candidats),
                                "candidates_rejetes": len(candidats_rejetes),
                                "other_candidates": other_candidates_list,
                                "raison": f"Niveau {niveau} - {user_type_emploi} - {user_heures_travaillees}h travaillées/{user_heures_max}h max"
                            }
                        }
                        
                        # Si officier manquant, ajouter le flag
                        if besoin_officier and not officier_assigne_cette_iteration and assignes_cette_garde == 0:
                            # Ce sera le premier assigné et ce n'est pas un officier
                            if not est_officier(user) and not est_eligible_fonction_superieure(user):
                                assignation["officier_manquant"] = True
                        
                        # Insérer dans la DB
                        await db.assignations.insert_one(assignation)
                        nouvelles_assignations.append(assignation)
                        
                        # Mise à jour locale
                        existing_assignations.append(assignation)
                        users_assignes_ce_jour.add(user_id)
                        
                        # Mettre à jour les heures mensuelles
                        if est_externe:
                            user_monthly_hours_externes[user_id] = user_monthly_hours_externes.get(user_id, 0) + duree_garde
                        else:
                            user_monthly_hours_internes[user_id] = user_monthly_hours_internes.get(user_id, 0) + duree_garde
                        
                        assignes_cette_garde += 1
                        
                        # Log avec info officier
                        officier_tag = ""
                        if est_officier(user):
                            officier_tag = " [OFFICIER]"
                        elif est_eligible_fonction_superieure(user):
                            officier_tag = " [ÉLIGIBLE]"
                        
                        logging.info(f"✅ [N{niveau}]{officier_tag} {user.get('prenom', '')} {user.get('nom', '')} → {type_garde_nom} le {date_str}")
                
                # Log si garde non remplie après tous les niveaux
                if assignes_cette_garde == 0 and places_restantes > 0:
                    logging.warning(f"❌ [NON REMPLIE] {type_garde_nom} @ {date_str}: 0 assignation sur {places_restantes} places. Candidats avec compétences: {len(users_avec_competences)}")
            
            current_date += timedelta(days=1)
        
        logging.info(f"📊 [RÉSULTAT] {len(nouvelles_assignations)} nouvelles assignations créées")
        
        return nouvelles_assignations
    
    except Exception as e:
        logging.error(f"Erreur attribution automatique: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Erreur attribution: {str(e)}")


# GET rapport-audit
@router.get("/{tenant_slug}/planning/rapport-audit")
async def generer_rapport_audit_assignations(
    tenant_slug: str,
    mois: str,  # Format: YYYY-MM
    format: str = "pdf",  # pdf ou excel
    current_user: User = Depends(get_current_user)
):
    """Génère un rapport d'audit complet des assignations automatiques pour un mois donné"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé - Admin uniquement")
    
    # Vérifier le tenant
    tenant = await get_tenant_from_slug(tenant_slug)
    
    try:
        # Parser le mois
        annee, mois_num = map(int, mois.split('-'))
        date_debut = datetime(annee, mois_num, 1)
        
        # Calculer la date de fin du mois
        if mois_num == 12:
            date_fin = datetime(annee + 1, 1, 1) - timedelta(days=1)
        else:
            date_fin = datetime(annee, mois_num + 1, 1) - timedelta(days=1)
        
        date_debut_str = date_debut.strftime("%Y-%m-%d")
        date_fin_str = date_fin.strftime("%Y-%m-%d")
        
        # Récupérer toutes les assignations automatiques du mois
        assignations_auto = await db.assignations.find({
            "tenant_id": tenant.id,
            "date": {
                "$gte": date_debut_str,
                "$lte": date_fin_str
            },
            "assignation_type": "auto",
            "justification": {"$exists": True, "$ne": None}
        }).to_list(1000)
        
        if not assignations_auto:
            raise HTTPException(status_code=404, detail="Aucune assignation automatique trouvée pour ce mois")
        
        # Récupérer les infos complémentaires (users, types garde)
        users = await db.users.find({"tenant_id": tenant.id}).to_list(1000)
        types_garde = await db.types_garde.find({"tenant_id": tenant.id}).to_list(1000)
        
        # Mapper users et types garde
        user_map = {u["id"]: u for u in users}
        type_garde_map = {t["id"]: t for t in types_garde}
        
        # Générer le rapport selon le format
        if format == "pdf":
            return await generer_pdf_audit(assignations_auto, user_map, type_garde_map, tenant, mois)
        else:  # excel
            return await generer_excel_audit(assignations_auto, user_map, type_garde_map, tenant, mois)
            
    except ValueError:
        raise HTTPException(status_code=400, detail="Format de mois invalide. Utilisez YYYY-MM")
    except Exception as e:
        logging.error(f"Erreur génération rapport audit: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur génération rapport: {str(e)}")

async def generer_pdf_audit(assignations, user_map, type_garde_map, tenant, mois):
    """Génère un PDF du rapport d'audit"""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    
    # Utiliser la fonction helper pour créer un PDF brandé
    buffer, doc, elements = create_branded_pdf(tenant, pagesize=A4)
    styles = getSampleStyleSheet()
    
    # Style titre
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Titre
    titre = Paragraph(f"<b>Rapport d'Audit des Affectations Automatiques</b><br/>{tenant.nom}<br/>Période: {mois}", title_style)
    elements.append(titre)
    elements.append(Spacer(1, 0.3*inch))
    
    # Statistiques globales
    stats = Paragraph(f"<b>Total d'assignations automatiques: {len(assignations)}</b>", styles['Normal'])
    elements.append(stats)
    elements.append(Spacer(1, 0.2*inch))
    
    # Tableau pour chaque assignation
    for idx, assignation in enumerate(assignations[:50], 1):  # Limiter à 50 pour PDF
        user = user_map.get(assignation["user_id"], {})
        type_garde = type_garde_map.get(assignation["type_garde_id"], {})
        justif = assignation.get("justification", {})
        
        # Info assignation
        info_title = Paragraph(f"<b>{idx}. {user.get('prenom', 'N/A')} {user.get('nom', 'N/A')} - {type_garde.get('nom', 'N/A')} - {assignation['date']}</b>", styles['Heading3'])
        elements.append(info_title)
        
        # Scores
        assigned_user = justif.get("assigned_user", {})
        scores = assigned_user.get("scores", {})
        details = assigned_user.get("details", {})
        
        data_scores = [
            ["Critère", "Score", "Détail"],
            ["Équité", f"{scores.get('equite', 0)}/100", f"{details.get('heures_ce_mois', 0)}h (moy: {details.get('moyenne_equipe', 0)}h)"],
            ["Ancienneté", f"{scores.get('anciennete', 0)}/100", f"{details.get('annees_service', 0)} ans"],
            ["Disponibilité", f"{scores.get('disponibilite', 0)}/100", "Déclarée" if details.get('disponibilite_declaree') else "Temps plein"],
            ["Compétences", f"{scores.get('competences', 0)}/100", user.get('grade', 'N/A')],
            ["TOTAL", f"{scores.get('total', 0)}/400", ""]
        ]
        
        table_scores = Table(data_scores, colWidths=[2*inch, 1.5*inch, 2.5*inch])
        table_scores.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e5e7eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        
        elements.append(table_scores)
        elements.append(Spacer(1, 0.1*inch))
        
        # Notes admin
        notes = assignation.get("notes_admin")
        if notes:
            notes_para = Paragraph(f"<b>Notes admin:</b> {notes}", styles['Normal'])
            elements.append(notes_para)
        
        # Autres candidats (top 3)
        other_candidates = justif.get("other_candidates", [])[:3]
        if other_candidates:
            autres_title = Paragraph("<b>Autres candidats évalués:</b>", styles['Normal'])
            elements.append(autres_title)
            
            for cand in other_candidates:
                cand_text = f"• {cand.get('nom_complet', 'N/A')} - {cand.get('excluded_reason', 'N/A')}"
                cand_para = Paragraph(cand_text, styles['Normal'])
                elements.append(cand_para)
        
        elements.append(Spacer(1, 0.3*inch))
        
        # Page break tous les 5 pour éviter surcharge
        if idx % 5 == 0 and idx < len(assignations):
            elements.append(PageBreak())
    
    doc.build(elements)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=audit_affectations_{mois}.pdf"
        }
    )

async def generer_excel_audit(assignations, user_map, type_garde_map, tenant, mois):
    """Génère un fichier Excel du rapport d'audit"""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Affectations"
    
    # En-tête
    ws['A1'] = f"Rapport d'Audit - {tenant.nom}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f"Période: {mois}"
    ws['A3'] = f"Total d'assignations: {len(assignations)}"
    
    # Colonnes
    headers = ["Date", "Garde", "Pompier", "Grade", "Heures mois", "Score Équité", 
               "Score Ancienneté", "Score Dispo", "Score Compét", "Score Total", 
               "Candidats évalués", "Notes Admin"]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Données
    row_num = 6
    for assignation in assignations:
        user = user_map.get(assignation["user_id"], {})
        type_garde = type_garde_map.get(assignation["type_garde_id"], {})
        justif = assignation.get("justification", {})
        assigned_user = justif.get("assigned_user", {})
        scores = assigned_user.get("scores", {})
        details = assigned_user.get("details", {})
        
        ws.cell(row=row_num, column=1).value = assignation["date"]
        ws.cell(row=row_num, column=2).value = type_garde.get("nom", "N/A")
        ws.cell(row=row_num, column=3).value = f"{user.get('prenom', '')} {user.get('nom', '')}"
        ws.cell(row=row_num, column=4).value = user.get("grade", "N/A")
        ws.cell(row=row_num, column=5).value = details.get("heures_ce_mois", 0)
        ws.cell(row=row_num, column=6).value = scores.get("equite", 0)
        ws.cell(row=row_num, column=7).value = scores.get("anciennete", 0)
        ws.cell(row=row_num, column=8).value = scores.get("disponibilite", 0)
        ws.cell(row=row_num, column=9).value = scores.get("competences", 0)
        ws.cell(row=row_num, column=10).value = scores.get("total", 0)
        ws.cell(row=row_num, column=11).value = justif.get("total_candidates_evaluated", 0)
        ws.cell(row=row_num, column=12).value = assignation.get("notes_admin", "")
        
        row_num += 1
    
    # Ajuster les largeurs avec des valeurs fixes pour éviter les erreurs MergedCell
    column_widths = {
        'A': 12, 'B': 12, 'C': 15, 'D': 15, 'E': 12, 'F': 12,
        'G': 12, 'H': 12, 'I': 12, 'J': 10, 'K': 10, 'L': 25
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=audit_affectations_{mois}.xlsx"
        }
    )

# Endpoint pour obtenir les statistiques personnelles mensuelles
@router.get("/{tenant_slug}/users/{user_id}/stats-mensuelles")

# POST reinitialiser
@router.post("/planning/reinitialiser")
async def reinitialiser_planning(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        # Supprimer toutes les assignations
        result = await db.assignations.delete_many({})
        
        return {
            "message": "Planning réinitialisé avec succès",
            "assignations_supprimees": result.deleted_count
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur réinitialisation: {str(e)}")


# GET parametres/validation-planning
@router.get("/{tenant_slug}/parametres/validation-planning")
async def get_parametres_validation(tenant_slug: str, current_user: User = Depends(get_current_user)):
    """
    Récupérer les paramètres de validation du planning pour le tenant
    """
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        tenant = await get_tenant_from_slug(tenant_slug)
        
        # Récupérer les paramètres de validation ou retourner valeurs par défaut
        validation_params = tenant.parametres.get('validation_planning', {
            'frequence': 'mensuel',
            'jour_envoi': 25,  # 25 du mois
            'heure_envoi': '17:00',
            'periode_couverte': 'mois_suivant',
            'envoi_automatique': True,
            'derniere_notification': None
        })
        
        return validation_params
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur récupération paramètres: {str(e)}")


# PUT parametres/validation-planning
@router.put("/{tenant_slug}/parametres/validation-planning")
async def update_parametres_validation(tenant_slug: str, parametres: dict, current_user: User = Depends(get_current_user)):
    """
    Mettre à jour les paramètres de validation du planning
    """
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        tenant = await get_tenant_from_slug(tenant_slug)
        tenant_doc = await db.tenants.find_one({"id": tenant.id})
        
        if not tenant_doc:
            raise HTTPException(status_code=404, detail="Tenant non trouvé")
        
        # Mettre à jour les paramètres
        current_parametres = tenant_doc.get('parametres', {})
        current_parametres['validation_planning'] = parametres
        
        await db.tenants.update_one(
            {"id": tenant.id},
            {"$set": {"parametres": current_parametres}}
        )
        
        return {"message": "Paramètres mis à jour avec succès", "parametres": parametres}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur mise à jour paramètres: {str(e)}")


# POST envoyer-notifications
@router.post("/{tenant_slug}/planning/envoyer-notifications")
async def envoyer_notifications_planning(tenant_slug: str, periode_debut: str, periode_fin: str, current_user: User = Depends(get_current_user)):
    """
    Envoyer les notifications par email à tous les pompiers avec leurs gardes assignées
    
    Args:
        tenant_slug: slug de la caserne
        periode_debut: Date début (YYYY-MM-DD)
        periode_fin: Date fin (YYYY-MM-DD)
    """
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        tenant = await get_tenant_from_slug(tenant_slug)
        
        # Récupérer toutes les assignations de la période
        assignations_list = await db.assignations.find({
            "tenant_id": tenant.id,
            "date": {"$gte": periode_debut, "$lte": periode_fin}
        }).to_list(length=None)
        
        # Récupérer tous les users et types de garde
        users_list = await db.users.find({"tenant_id": tenant.id}).to_list(length=None)
        types_garde_list = await db.types_garde.find({"tenant_id": tenant.id}).to_list(length=None)
        
        # Créer des maps pour accès rapide
        users_map = {u['id']: u for u in users_list}
        types_garde_map = {t['id']: t for t in types_garde_list}
        
        # Grouper les assignations par user
        gardes_par_user = {}
        for assignation in assignations_list:
            user_id = assignation['user_id']
            if user_id not in gardes_par_user:
                gardes_par_user[user_id] = []
            
            type_garde = types_garde_map.get(assignation['type_garde_id'], {})
            
            # Trouver les collègues pour cette garde
            collegues = [
                f"{users_map[a['user_id']]['prenom']} {users_map[a['user_id']]['nom']}"
                for a in assignations_list
                if a['date'] == assignation['date'] and 
                   a['type_garde_id'] == assignation['type_garde_id'] and 
                   a['user_id'] != user_id and 
                   a['user_id'] in users_map
            ]
            
            # Formater la date
            from datetime import datetime as dt
            date_obj = dt.strptime(assignation['date'], '%Y-%m-%d')
            jour_fr = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche'][date_obj.weekday()]
            
            gardes_par_user[user_id].append({
                'date': assignation['date'],  # Format YYYY-MM-DD pour compatibilité
                'jour': jour_fr,
                'type_garde': type_garde.get('nom', 'Garde'),
                'horaire': f"{type_garde.get('heure_debut', '08:00')} - {type_garde.get('heure_fin', '08:00')}",
                'duree_heures': type_garde.get('duree_heures', 0),
                'est_externe': type_garde.get('est_garde_externe', False),
                'collegues': collegues
            })
        
        # Envoyer les emails
        emails_envoyes = 0
        emails_echoues = 0
        
        for user_id, gardes in gardes_par_user.items():
            user = users_map.get(user_id)
            if not user or not user.get('email'):
                continue
            
            # Vérifier les préférences de notification
            preferences = user.get("preferences_notifications", {})
            if not preferences.get("email_actif", True):
                logger.info(f"📧 Email désactivé pour {user.get('prenom')} - préférences utilisateur")
                continue
            
            # Calculer les statistiques pour cet utilisateur
            stats = {
                "par_type": {},
                "heures_internes": 0,
                "heures_externes": 0,
                "total_gardes": len(gardes)
            }
            
            for garde in gardes:
                type_nom = garde.get("type_garde", "Garde")
                duree = garde.get("duree_heures", 0) or 0
                
                if type_nom not in stats["par_type"]:
                    stats["par_type"][type_nom] = 0
                stats["par_type"][type_nom] += 1
                
                if garde.get("est_externe", False):
                    stats["heures_externes"] += duree
                else:
                    stats["heures_internes"] += duree
            
            user_name = f"{user['prenom']} {user['nom']}"
            email_sent = send_planning_notification_email(
                user_email=user['email'],
                user_name=user_name,
                gardes_list=gardes,
                tenant_slug=tenant_slug,
                periode=f"{periode_debut} au {periode_fin}",
                tenant_nom=tenant.nom,
                stats=stats
            )
            
            if email_sent:
                emails_envoyes += 1
            else:
                emails_echoues += 1
        
        # Mettre à jour la date de dernière notification
        tenant_doc = await db.tenants.find_one({"id": tenant.id})
        current_parametres = tenant_doc.get('parametres', {})
        if 'validation_planning' not in current_parametres:
            current_parametres['validation_planning'] = {}
        current_parametres['validation_planning']['derniere_notification'] = datetime.now(timezone.utc).isoformat()
        
        await db.tenants.update_one(
            {"id": tenant.id},
            {"$set": {"parametres": current_parametres}}
        )
        
        return {
            "message": "Notifications envoyées",
            "emails_envoyes": emails_envoyes,
            "emails_echoues": emails_echoues,
            "total_pompiers": len(gardes_par_user)
        }
        
    except Exception as e:
        logger.error(f"Erreur envoi notifications planning: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Erreur envoi notifications: {str(e)}")



@router.put("/{tenant_slug}/assignations/{assignation_id}/notes")
async def update_assignation_notes(
    tenant_slug: str,
    assignation_id: str,
    data: dict,
    current_user: User = Depends(get_current_user)
):
    """Permet à un admin de mettre à jour les notes sur une assignation"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé - Admin uniquement")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Trouver l'assignation
    assignation = await db.assignations.find_one({
        "id": assignation_id,
        "tenant_id": tenant.id
    })
    
    if not assignation:
        raise HTTPException(status_code=404, detail="Assignation non trouvée")
    
    notes = data.get("notes", "")
    
    # Mettre à jour les notes
    await db.assignations.update_one(
        {"id": assignation_id},
        {"$set": {"notes_admin": notes}}
    )
    
    return {"message": "Notes mises à jour avec succès", "notes": notes}
