"""
Routes Audit du module Planning
================================
Extraits de planning.py pour maintenabilité.

Routes:
- GET /{tenant_slug}/planning/rapport-audit
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, Response
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
from io import BytesIO
import logging

from routes.dependencies import (
    db,
    get_current_user,
    get_tenant_from_slug,
    User,
    require_permission
)

from utils.pdf_helpers import (
    create_branded_pdf,
    get_modern_pdf_styles,
    create_pdf_footer_text
)

from services.email_builder import build_email

router = APIRouter(tags=["Planning Audit"])
logger = logging.getLogger(__name__)


@router.get("/{tenant_slug}/planning/rapport-audit")
async def generer_rapport_audit_assignations(
    tenant_slug: str,
    mois: str,  # Format: YYYY-MM
    format: str = "pdf",  # pdf ou excel
    current_user: User = Depends(get_current_user)
):
    """Génère un rapport d'audit complet des assignations automatiques pour un mois donné"""
    # Vérifier le tenant
    tenant = await get_tenant_from_slug(tenant_slug)
    await require_permission(tenant.id, current_user, "planning", "exporter")
    
    try:
        # Parser le mois
        annee, mois_num = map(int, mois.split('-'))
        date_debut = datetime(annee, mois_num, 1)
        
        # Calculer la date de fin du mois
        if mois_num == 12:
            date_fin = datetime(annee + 1, 1, 1) - timedelta(days=1)
        else:
            date_fin = datetime(annee, mois_num + 1, 1) - timedelta(days=1)
        
        date_debut_str = date_debut.strftime("%Y-%m-%d")
        date_fin_str = date_fin.strftime("%Y-%m-%d")
        
        # Récupérer toutes les assignations automatiques du mois
        assignations_auto = await db.assignations.find({
            "tenant_id": tenant.id,
            "date": {
                "$gte": date_debut_str,
                "$lte": date_fin_str
            },
            "assignation_type": "auto",
            "justification": {"$exists": True, "$ne": None}
        }).to_list(1000)
        
        if not assignations_auto:
            raise HTTPException(status_code=404, detail="Aucune assignation automatique trouvée pour ce mois")
        
        # Récupérer les infos complémentaires (users, types garde)
        users = await db.users.find({"tenant_id": tenant.id}).to_list(1000)
        types_garde = await db.types_garde.find({"tenant_id": tenant.id}).to_list(1000)
        
        # Mapper users et types garde
        user_map = {u["id"]: u for u in users}
        type_garde_map = {t["id"]: t for t in types_garde}
        
        # Générer le rapport selon le format
        if format == "pdf":
            return await generer_pdf_audit(assignations_auto, user_map, type_garde_map, tenant, mois)
        else:  # excel
            return await generer_excel_audit(assignations_auto, user_map, type_garde_map, tenant, mois)
            
    except ValueError:
        raise HTTPException(status_code=400, detail="Format de mois invalide. Utilisez YYYY-MM")
    except Exception as e:
        logging.error(f"Erreur génération rapport audit: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur génération rapport: {str(e)}")

async def generer_pdf_audit(assignations, user_map, type_garde_map, tenant, mois):
    """Génère un PDF du rapport d'audit"""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    
    # Utiliser la fonction helper pour créer un PDF brandé
    buffer, doc, elements = create_branded_pdf(tenant, pagesize=A4)
    styles = getSampleStyleSheet()
    
    # Style titre
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Titre
    titre = Paragraph(f"<b>Rapport d'Audit des Affectations Automatiques</b><br/>{tenant.nom}<br/>Période: {mois}", title_style)
    elements.append(titre)
    elements.append(Spacer(1, 0.3*inch))
    
    # Statistiques globales
    stats = Paragraph(f"<b>Total d'assignations automatiques: {len(assignations)}</b>", styles['Normal'])
    elements.append(stats)
    elements.append(Spacer(1, 0.2*inch))
    
    # Tableau pour chaque assignation
    for idx, assignation in enumerate(assignations[:50], 1):  # Limiter à 50 pour PDF
        user = user_map.get(assignation["user_id"], {})
        type_garde = type_garde_map.get(assignation["type_garde_id"], {})
        justif = assignation.get("justification", {})
        
        # Info assignation
        info_title = Paragraph(f"<b>{idx}. {user.get('prenom', 'N/A')} {user.get('nom', 'N/A')} - {type_garde.get('nom', 'N/A')} - {assignation['date']}</b>", styles['Heading3'])
        elements.append(info_title)
        
        # Scores
        assigned_user = justif.get("assigned_user", {})
        scores = assigned_user.get("scores", {})
        details = assigned_user.get("details", {})
        
        data_scores = [
            ["Critère", "Score", "Détail"],
            ["Équité", f"{scores.get('equite', 0)}/100", f"{details.get('heures_ce_mois', 0)}h (moy: {details.get('moyenne_equipe', 0)}h)"],
            ["Ancienneté", f"{scores.get('anciennete', 0)}/100", f"{details.get('annees_service', 0)} ans"],
            ["Disponibilité", f"{scores.get('disponibilite', 0)}/100", "Déclarée" if details.get('disponibilite_declaree') else "Temps plein"],
            ["Compétences", f"{scores.get('competences', 0)}/100", user.get('grade', 'N/A')],
            ["TOTAL", f"{scores.get('total', 0)}/400", ""]
        ]
        
        table_scores = Table(data_scores, colWidths=[2*inch, 1.5*inch, 2.5*inch])
        table_scores.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e5e7eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        
        elements.append(table_scores)
        elements.append(Spacer(1, 0.1*inch))
        
        # Notes admin
        notes = assignation.get("notes_admin")
        if notes:
            notes_para = Paragraph(f"<b>Notes admin:</b> {notes}", styles['Normal'])
            elements.append(notes_para)
        
        # Autres candidats (top 3)
        other_candidates = justif.get("other_candidates", [])[:3]
        if other_candidates:
            autres_title = Paragraph("<b>Autres candidats évalués:</b>", styles['Normal'])
            elements.append(autres_title)
            
            for cand in other_candidates:
                cand_text = f"• {cand.get('nom_complet', 'N/A')} - {cand.get('excluded_reason', 'N/A')}"
                cand_para = Paragraph(cand_text, styles['Normal'])
                elements.append(cand_para)
        
        elements.append(Spacer(1, 0.3*inch))
        
        # Page break tous les 5 pour éviter surcharge
        if idx % 5 == 0 and idx < len(assignations):
            elements.append(PageBreak())
    
    doc.build(elements)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=audit_affectations_{mois}.pdf"
        }
    )

async def generer_excel_audit(assignations, user_map, type_garde_map, tenant, mois):
    """Génère un fichier Excel du rapport d'audit"""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Affectations"
    
    # En-tête
    ws['A1'] = f"Rapport d'Audit - {tenant.nom}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f"Période: {mois}"
    ws['A3'] = f"Total d'assignations: {len(assignations)}"
    
    # Colonnes
    headers = ["Date", "Garde", "Pompier", "Grade", "Heures mois", "Score Équité", 
               "Score Ancienneté", "Score Dispo", "Score Compét", "Score Total", 
               "Candidats évalués", "Notes Admin"]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Données
    row_num = 6
    for assignation in assignations:
        user = user_map.get(assignation["user_id"], {})
        type_garde = type_garde_map.get(assignation["type_garde_id"], {})
        justif = assignation.get("justification", {})
        assigned_user = justif.get("assigned_user", {})
        scores = assigned_user.get("scores", {})
        details = assigned_user.get("details", {})
        
        ws.cell(row=row_num, column=1).value = assignation["date"]
        ws.cell(row=row_num, column=2).value = type_garde.get("nom", "N/A")
        ws.cell(row=row_num, column=3).value = f"{user.get('prenom', '')} {user.get('nom', '')}"
        ws.cell(row=row_num, column=4).value = user.get("grade", "N/A")
        ws.cell(row=row_num, column=5).value = details.get("heures_ce_mois", 0)
        ws.cell(row=row_num, column=6).value = scores.get("equite", 0)
        ws.cell(row=row_num, column=7).value = scores.get("anciennete", 0)
        ws.cell(row=row_num, column=8).value = scores.get("disponibilite", 0)
        ws.cell(row=row_num, column=9).value = scores.get("competences", 0)
        ws.cell(row=row_num, column=10).value = scores.get("total", 0)
        ws.cell(row=row_num, column=11).value = justif.get("total_candidates_evaluated", 0)
        ws.cell(row=row_num, column=12).value = assignation.get("notes_admin", "")
        
        row_num += 1
    
    # Ajuster les largeurs avec des valeurs fixes pour éviter les erreurs MergedCell
    column_widths = {
        'A': 12, 'B': 12, 'C': 15, 'D': 15, 'E': 12, 'F': 12,
        'G': 12, 'H': 12, 'I': 12, 'J': 10, 'K': 10, 'L': 25
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=audit_affectations_{mois}.xlsx"
        }
    )

