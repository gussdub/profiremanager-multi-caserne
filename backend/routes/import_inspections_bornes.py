"""
Routes API pour l'import des inspections de bornes sèches depuis des fichiers Excel
==================================================================================

Ce module gère l'import intelligent des données historiques d'inspections
depuis des fichiers Excel (ex: Google Forms exports).

Endpoint principal:
- POST /{tenant_slug}/import/inspections-bornes-seches - Import de fichiers Excel
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from typing import Optional
from datetime import datetime, timezone
import uuid
import logging
import re
import io

router = APIRouter(tags=["Import Inspections Bornes"])
logger = logging.getLogger(__name__)

from routes.dependencies import (
    db,
    get_current_user,
    get_tenant_from_slug,
    User
)

# Mapping des colonnes Excel vers les champs d'inspection
COLONNES_MAPPING = {
    "Horodateur": "horodateur",
    "Accessibilité de la borne?": "accessibilite",
    "Conditions atmosphériques lors du test?": "conditions_atmospheriques",
    "Température extérieure?": "temperature_exterieure",
    "Inspection visuelle [Joint présent]": "joint_present",
    "Inspection visuelle [Joint en bon état]": "joint_bon_etat",
    "Inspection visuelle [Site accessible]": "site_accessible",
    "Inspection visuelle [Site bien déneigé]": "site_deneige",
    "Inspection visuelle [Vanne de sortie Storz 4\"]": "vanne_storz_4",
    "Inspection visuelle [Vanne de sortie 6\" filetée]": "vanne_6_filetee",
    "Inspection visuelle [Vanne de sortie 4\" filetée]": "vanne_4_filetee",
    "Inspection visuelle [Niveau du plan d'eau]": "niveau_plan_eau",
    "Essai de pompage [Pompage en continu (5 minutes)]": "pompage_continu",
    "Essai de pompage [Cavitation durant le pompage]": "cavitation_pompage",
    "Temps d'amorçage (en secondes)?": "temps_amorcage",
    "Commentaire": "commentaire",
    "Date": "date_inspection",
    "Matricule du pompier ayant effectué l'inspection": "matricule_inspecteur"
}

# Valeurs acceptées et leur normalisation
VALEURS_CONFORME = ["Conforme", "conforme", "C", "Oui", "oui", "O"]
VALEURS_NON_CONFORME = ["Non conforme", "non conforme", "NC", "Non", "non", "N"]
VALEURS_NA = ["N/A", "n/a", "NA", "-", "", None]


def normaliser_valeur(valeur):
    """Normalise une valeur d'inspection vers le format standard"""
    if valeur is None:
        return "N/A"
    valeur_str = str(valeur).strip()
    if valeur_str in VALEURS_CONFORME:
        return "Conforme"
    if valeur_str in VALEURS_NON_CONFORME:
        return "Non conforme"
    if valeur_str in VALEURS_NA or valeur_str == "":
        return "N/A"
    return valeur_str


def extraire_nom_borne(filename: str) -> str:
    """Extrait le nom de la borne depuis le nom du fichier
    
    Exemples:
    - "Darby - Fiche technique borne seche (réponses).xlsx" -> "Darby"
    - "BS-001_inspections.xlsx" -> "BS-001"
    - "Lac_Vert_2024.xlsx" -> "Lac_Vert"
    """
    # Retirer l'extension
    name = filename.rsplit('.', 1)[0]
    
    # Patterns courants pour extraire le nom
    patterns = [
        r'^([^-]+)\s*-',          # "Nom - Suite..." -> "Nom"
        r'^([^_]+)_',              # "Nom_suite..." -> "Nom"
        r'^(.+?)[\s_-](?:fiche|inspection|borne)',  # Avant "fiche/inspection/borne"
    ]
    
    for pattern in patterns:
        match = re.search(pattern, name, re.IGNORECASE)
        if match:
            return match.group(1).strip()
    
    # Si aucun pattern ne match, retourner le nom complet sans extension
    return name.strip()


def parser_temps_amorcage(valeur) -> Optional[float]:
    """Parse le temps d'amorçage vers un nombre de secondes"""
    if valeur is None:
        return None
    valeur_str = str(valeur).strip().lower()
    
    # Retirer les unités
    valeur_str = re.sub(r'(sec|secondes?|s)', '', valeur_str).strip()
    
    try:
        return float(valeur_str)
    except ValueError:
        return None


def parser_date(valeur) -> Optional[str]:
    """Parse une date vers format ISO"""
    if valeur is None:
        return None
    
    if isinstance(valeur, datetime):
        return valeur.strftime('%Y-%m-%d')
    
    valeur_str = str(valeur).strip()
    
    # Différents formats de date
    formats = [
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M:%S.%f',
        '%Y-%m-%d',
        '%d/%m/%Y',
        '%d-%m-%Y',
    ]
    
    for fmt in formats:
        try:
            dt = datetime.strptime(valeur_str, fmt)
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            continue
    
    return valeur_str


@router.post("/{tenant_slug}/import/inspections-bornes-seches")
async def import_inspections_bornes_seches(
    tenant_slug: str,
    file: UploadFile = File(...),
    borne_id: Optional[str] = Form(None),
    current_user: User = Depends(get_current_user)
):
    """
    Importer des inspections de bornes sèches depuis un fichier Excel.
    
    Le nom de la borne est extrait du nom du fichier si borne_id n'est pas fourni.
    Les valeurs sont automatiquement mappées (Conforme, Non conforme, N/A).
    """
    if current_user.role not in ['admin', 'superviseur']:
        raise HTTPException(status_code=403, detail="Permission refusée - Admin/Superviseur requis")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Vérifier le type de fichier
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Format de fichier non supporté. Utilisez un fichier Excel (.xlsx ou .xls)"
        )
    
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Bibliothèque openpyxl non installée"
        )
    
    # Extraire le nom de la borne depuis le nom de fichier
    nom_borne_fichier = extraire_nom_borne(file.filename)
    logger.info(f"Nom de borne extrait du fichier: {nom_borne_fichier}")
    
    # Trouver la borne correspondante
    borne = None
    if borne_id:
        borne = await db.points_eau.find_one({
            "id": borne_id,
            "tenant_id": tenant.id
        })
    else:
        # Recherche par nom (insensible à la casse)
        borne = await db.points_eau.find_one({
            "tenant_id": tenant.id,
            "$or": [
                {"nom": {"$regex": f"^{re.escape(nom_borne_fichier)}$", "$options": "i"}},
                {"numero_identification": {"$regex": f"^{re.escape(nom_borne_fichier)}$", "$options": "i"}}
            ]
        })
    
    if not borne:
        # Lister les bornes disponibles pour aider l'utilisateur
        bornes_dispo = await db.points_eau.find(
            {"tenant_id": tenant.id},
            {"nom": 1, "numero_identification": 1, "_id": 0}
        ).to_list(length=20)
        
        noms_bornes = [b.get("nom") or b.get("numero_identification") for b in bornes_dispo]
        raise HTTPException(
            status_code=404,
            detail={
                "message": f"Borne '{nom_borne_fichier}' non trouvée",
                "nom_extrait": nom_borne_fichier,
                "bornes_disponibles": noms_bornes[:10],
                "suggestion": "Renommez le fichier avec le nom exact de la borne ou sélectionnez-la manuellement"
            }
        )
    
    # Lire le fichier Excel
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    # Récupérer les en-têtes
    headers = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            # Mapper vers notre format interne
            mapped_name = COLONNES_MAPPING.get(header, header.lower().replace(' ', '_'))
            headers[col] = mapped_name
    
    logger.info(f"Colonnes détectées: {list(headers.values())}")
    
    # Parser les lignes de données
    inspections_importees = []
    erreurs = []
    
    for row in range(2, ws.max_row + 1):
        try:
            row_data = {}
            for col, field_name in headers.items():
                value = ws.cell(row=row, column=col).value
                row_data[field_name] = value
            
            # Ignorer les lignes vides
            if not any(row_data.values()):
                continue
            
            # Construire l'objet d'inspection
            date_inspection = parser_date(row_data.get('date_inspection') or row_data.get('horodateur'))
            
            # Déterminer si l'inspection est conforme
            champs_inspection = [
                'joint_present', 'joint_bon_etat', 'site_accessible', 'site_deneige',
                'vanne_storz_4', 'vanne_6_filetee', 'vanne_4_filetee', 'niveau_plan_eau',
                'pompage_continu', 'cavitation_pompage'
            ]
            
            reponses = {}
            nb_non_conformes = 0
            
            for champ in champs_inspection:
                valeur = normaliser_valeur(row_data.get(champ))
                reponses[champ] = valeur
                if valeur == "Non conforme":
                    nb_non_conformes += 1
            
            # Ajouter les autres champs
            reponses['accessibilite'] = row_data.get('accessibilite', '')
            reponses['conditions_atmospheriques'] = row_data.get('conditions_atmospheriques', '')
            reponses['temperature_exterieure'] = row_data.get('temperature_exterieure')
            reponses['temps_amorcage'] = parser_temps_amorcage(row_data.get('temps_amorcage'))
            
            inspection = {
                "id": str(uuid.uuid4()),
                "tenant_id": tenant.id,
                "asset_id": borne["id"],
                "asset_type": "borne_seche",
                "borne_nom": borne.get("nom") or borne.get("numero_identification"),
                "date_inspection": date_inspection or datetime.now(timezone.utc).strftime('%Y-%m-%d'),
                "inspecteur_matricule": str(row_data.get('matricule_inspecteur', '')).replace('.0', ''),
                "inspecteur_nom": f"Matricule {row_data.get('matricule_inspecteur', 'inconnu')}",
                "reponses": reponses,
                "commentaire": row_data.get('commentaire', ''),
                "conforme": nb_non_conformes == 0,
                "nb_non_conformes": nb_non_conformes,
                "source": "import_excel",
                "fichier_source": file.filename,
                "imported_by": current_user.id,
                "imported_at": datetime.now(timezone.utc).isoformat(),
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            inspections_importees.append(inspection)
            
        except Exception as e:
            erreurs.append({
                "ligne": row,
                "erreur": str(e)
            })
            logger.error(f"Erreur ligne {row}: {e}")
    
    # Insérer les inspections en base
    if inspections_importees:
        await db.inspections_unifiees.insert_many(inspections_importees)
        
        # Mettre à jour la borne avec la dernière date d'inspection
        dates = [i["date_inspection"] for i in inspections_importees if i.get("date_inspection")]
        if dates:
            derniere_date = max(dates)
            await db.points_eau.update_one(
                {"id": borne["id"], "tenant_id": tenant.id},
                {
                    "$set": {
                        "derniere_inspection_date": derniere_date,
                        "nombre_inspections": (borne.get("nombre_inspections") or 0) + len(inspections_importees)
                    }
                }
            )
    
    logger.info(f"Import terminé: {len(inspections_importees)} inspections importées, {len(erreurs)} erreurs")
    
    return {
        "success": True,
        "message": f"Import terminé pour la borne '{borne.get('nom') or borne.get('numero_identification')}'",
        "borne": {
            "id": borne["id"],
            "nom": borne.get("nom") or borne.get("numero_identification")
        },
        "imported_count": len(inspections_importees),
        "errors_count": len(erreurs),
        "errors": erreurs[:10] if erreurs else []
    }


@router.get("/{tenant_slug}/import/inspections-bornes-seches/preview")
async def preview_import_file(
    tenant_slug: str,
    current_user: User = Depends(get_current_user)
):
    """
    Retourne les colonnes attendues pour l'import et des exemples.
    """
    return {
        "colonnes_attendues": list(COLONNES_MAPPING.keys()),
        "valeurs_acceptees": {
            "conforme": VALEURS_CONFORME,
            "non_conforme": VALEURS_NON_CONFORME,
            "non_applicable": VALEURS_NA
        },
        "format_date": "YYYY-MM-DD ou DD/MM/YYYY",
        "exemple_nom_fichier": "Darby - Fiche technique borne seche.xlsx",
        "note": "Le nom de la borne est extrait du début du nom de fichier (avant le premier tiret ou underscore)"
    }



from pydantic import BaseModel
from typing import List, Dict, Any

class InspectionMappedData(BaseModel):
    borne_nom: str = ""
    date_inspection: str = ""
    matricule_inspecteur: str = ""
    accessibilite: str = ""
    conditions_atmospheriques: str = ""
    temperature_exterieure: str = ""
    joint_present: str = ""
    joint_bon_etat: str = ""
    site_accessible: str = ""
    site_deneige: str = ""
    vanne_storz_4: str = ""
    vanne_6_filetee: str = ""
    vanne_4_filetee: str = ""
    niveau_plan_eau: str = ""
    pompage_continu: str = ""
    cavitation_pompage: str = ""
    temps_amorcage: str = ""
    commentaire: str = ""

class ImportCSVRequest(BaseModel):
    inspections: List[Dict[str, Any]]


@router.post("/{tenant_slug}/inspections-bornes/import-csv")
async def import_inspections_csv_mapped(
    tenant_slug: str,
    request: ImportCSVRequest,
    current_user: User = Depends(get_current_user)
):
    """
    Import des inspections de bornes sèches avec mapping pré-défini par le frontend.
    Les données arrivent déjà mappées aux bons champs.
    """
    tenant = await get_tenant_from_slug(tenant_slug)
    
    created = 0
    updated = 0
    erreurs = []
    
    # Cache des bornes pour éviter les requêtes répétées
    bornes_cache = {}
    
    for idx, insp_data in enumerate(request.inspections):
        try:
            # Trouver la borne
            borne_nom = str(insp_data.get('borne_nom', '')).strip()
            if not borne_nom:
                erreurs.append({"ligne": idx + 1, "erreur": "Nom de borne manquant"})
                continue
            
            # Chercher dans le cache ou en base
            if borne_nom not in bornes_cache:
                borne = await db.points_eau.find_one({
                    "tenant_id": tenant.id,
                    "type": "borne_seche",
                    "$or": [
                        {"nom": {"$regex": f"^{re.escape(borne_nom)}$", "$options": "i"}},
                        {"numero_identification": {"$regex": f"^{re.escape(borne_nom)}$", "$options": "i"}}
                    ]
                })
                bornes_cache[borne_nom] = borne
            else:
                borne = bornes_cache[borne_nom]
            
            if not borne:
                erreurs.append({"ligne": idx + 1, "erreur": f"Borne '{borne_nom}' non trouvée"})
                continue
            
            # Parser la date
            date_str = str(insp_data.get('date_inspection', '')).strip()
            date_inspection = None
            if date_str:
                # Essayer différents formats
                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']:
                    try:
                        date_inspection = datetime.strptime(date_str, fmt).strftime('%Y-%m-%d')
                        break
                    except:
                        continue
            
            if not date_inspection:
                date_inspection = datetime.now(timezone.utc).strftime('%Y-%m-%d')
            
            # Parser le temps d'amorçage
            temps_amorcage = None
            temps_str = str(insp_data.get('temps_amorcage', '')).strip()
            if temps_str:
                try:
                    # Extraire les chiffres
                    numbers = re.findall(r'\d+', temps_str)
                    if numbers:
                        temps_amorcage = int(numbers[0])
                except:
                    pass
            
            # Construire les réponses
            champs_inspection = [
                'joint_present', 'joint_bon_etat', 'site_accessible', 'site_deneige',
                'vanne_storz_4', 'vanne_6_filetee', 'vanne_4_filetee', 'niveau_plan_eau',
                'pompage_continu', 'cavitation_pompage'
            ]
            
            reponses = {}
            nb_non_conformes = 0
            
            for champ in champs_inspection:
                valeur = normaliser_valeur(insp_data.get(champ))
                reponses[champ] = valeur
                if valeur == "Non conforme":
                    nb_non_conformes += 1
            
            # Ajouter les autres champs
            reponses['accessibilite'] = insp_data.get('accessibilite', '')
            reponses['conditions_atmospheriques'] = insp_data.get('conditions_atmospheriques', '')
            reponses['temperature_exterieure'] = insp_data.get('temperature_exterieure')
            reponses['temps_amorcage'] = temps_amorcage
            
            # Créer l'inspection
            matricule = str(insp_data.get('matricule_inspecteur', '')).replace('.0', '').strip()
            
            inspection = {
                "id": str(uuid.uuid4()),
                "tenant_id": tenant.id,
                "asset_id": borne["id"],
                "asset_type": "borne_seche",
                "borne_nom": borne.get("nom") or borne.get("numero_identification"),
                "date_inspection": date_inspection,
                "inspecteur_matricule": matricule,
                "inspecteur_nom": f"Matricule {matricule}" if matricule else "Inconnu",
                "reponses": reponses,
                "commentaire": insp_data.get('commentaire', ''),
                "conforme": nb_non_conformes == 0,
                "nb_non_conformes": nb_non_conformes,
                "source": "import_csv_mapping",
                "imported_by": current_user.id,
                "imported_at": datetime.now(timezone.utc).isoformat(),
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            # Insérer en base
            await db.inspections_unifiees.insert_one(inspection)
            created += 1
            
            # Mettre à jour la borne
            await db.points_eau.update_one(
                {"id": borne["id"], "tenant_id": tenant.id},
                {
                    "$set": {"derniere_inspection_date": date_inspection},
                    "$inc": {"nombre_inspections": 1}
                }
            )
            
        except Exception as e:
            logger.error(f"Erreur import ligne {idx + 1}: {e}")
            erreurs.append({"ligne": idx + 1, "erreur": str(e)})
    
    logger.info(f"Import CSV terminé: {created} créées, {len(erreurs)} erreurs")
    
    return {
        "success": True,
        "created": created,
        "updated": updated,
        "errors": erreurs
    }
