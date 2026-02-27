"""
Routes API pour le module Remplacements
=======================================

STATUT: ACTIF
Ce module gère le système automatisé de demandes de remplacement entre pompiers.

Routes principales:
- POST   /{tenant_slug}/remplacements                        - Créer une demande de remplacement
- GET    /{tenant_slug}/remplacements                        - Liste des demandes
- GET    /{tenant_slug}/remplacements/propositions           - Propositions pour l'utilisateur
- GET    /{tenant_slug}/remplacements/export-pdf             - Export PDF des demandes
- GET    /{tenant_slug}/remplacements/export-excel           - Export Excel des demandes
- PUT    /{tenant_slug}/remplacements/{id}/accepter          - Accepter une demande
- PUT    /{tenant_slug}/remplacements/{id}/refuser           - Refuser une demande
- DELETE /{tenant_slug}/remplacements/{id}                   - Annuler une demande

Routes publiques (actions via email):
- GET    /remplacement-action/{token}/{action}               - Action via lien email

Paramètres remplacements:
- GET    /{tenant_slug}/parametres/remplacements             - Récupérer les paramètres
- PUT    /{tenant_slug}/parametres/remplacements             - Modifier les paramètres
"""

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import RedirectResponse, StreamingResponse
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
import asyncio
import uuid
import logging
import os

from routes.dependencies import (
    db,
    get_current_user,
    get_tenant_from_slug,
    clean_mongo_doc,
    User,
    creer_notification,
    creer_activite
)

# Import WebSocket pour synchronisation temps réel
from routes.websocket import broadcast_remplacement_update

router = APIRouter(tags=["Remplacements"])
logger = logging.getLogger(__name__)


# ==================== MODÈLES ====================

class TentativeRemplacement(BaseModel):
    """Historique des tentatives de remplacement"""
    user_id: str
    nom_complet: str
    date_contact: datetime
    statut: str  # contacted, accepted, refused, expired
    date_reponse: Optional[datetime] = None


class DemandeRemplacement(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    demandeur_id: str
    type_garde_id: str
    date: str  # Date de la garde à remplacer (format: YYYY-MM-DD)
    raison: str
    statut: str = "en_attente"  # en_attente, en_cours, accepte, expiree, annulee
    priorite: str = "normal"  # urgent (≤24h), normal (>24h) - calculé automatiquement
    remplacant_id: Optional[str] = None
    tentatives_historique: List[Dict[str, Any]] = []
    remplacants_contactes_ids: List[str] = []
    date_prochaine_tentative: Optional[datetime] = None
    nombre_tentatives: int = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class DemandeRemplacementCreate(BaseModel):
    type_garde_id: str
    date: str
    raison: str


class NotificationRemplacement(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    demande_remplacement_id: str
    destinataire_id: str
    message: str
    type_notification: str = "remplacement_disponible"
    statut: str = "envoye"  # envoye, lu, accepte, refuse
    date_envoi: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    date_reponse: Optional[datetime] = None
    ordre_priorite: Optional[int] = None


class ParametresRemplacements(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    mode_notification: str = "simultane"  # simultane, sequentiel, groupe_sequentiel
    taille_groupe: int = 3
    delai_attente_heures: int = 24  # Gardé pour compatibilité
    delai_attente_minutes: int = 1440  # Nouveau: en minutes
    max_contacts: int = 5
    priorite_grade: bool = True
    priorite_competences: bool = True
    activer_gestion_heures_sup: bool = False
    seuil_max_heures: int = 40
    periode_calcul_heures: str = "semaine"
    jours_periode_personnalisee: int = 7
    activer_regroupement_heures: bool = False
    duree_max_regroupement: int = 24


# ==================== FONCTIONS HELPER ====================

async def calculer_priorite_demande(date_garde: str) -> str:
    """
    Calcule la priorité d'une demande de remplacement
    - urgent: Si la garde est dans 24h ou moins
    - normal: Si la garde est dans plus de 24h
    """
    try:
        date_garde_obj = datetime.strptime(date_garde, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        maintenant = datetime.now(timezone.utc)
        delta = date_garde_obj - maintenant
        
        if delta.total_seconds() <= 86400:  # 24 heures en secondes
            return "urgent"
        return "normal"
    except Exception as e:
        logger.error(f"Erreur calcul priorité: {e}")
        return "normal"


async def trouver_remplacants_potentiels(
    tenant_id: str,
    type_garde_id: str,
    date_garde: str,
    demandeur_id: str,
    exclus_ids: List[str] = []
) -> List[Dict[str, Any]]:
    """
    Trouve les remplaçants potentiels selon la MÊME logique que le Planning:
    
    N0: Filtres absolus
        - Compétences requises
        - Pas d'assignation en conflit
        - Pas d'indisponibilité
        
    N1: Filtres secondaires
        - Statut actif
        
    Pour chaque niveau (N2, N3, N4, N5):
        Sous-tri par:
        1. Grade équivalent → équitabilité → équipe (si actif) → ancienneté
        2. Fonction supérieure → équitabilité → équipe (si actif) → ancienneté
        3. Autres → équitabilité → équipe (si actif) → ancienneté
    """
    try:
        # ==================== CHARGEMENT DES PARAMÈTRES ====================
        
        # Paramètres de remplacements
        parametres = await db.parametres_remplacements.find_one({"tenant_id": tenant_id})
        
        # Tenant pour les équipes de garde
        tenant = await db.tenants.find_one({"id": tenant_id})
        tenant_params = tenant.get("parametres", {}) if tenant else {}
        
        # IMPORTANT: Pour les remplacements, TOUS les niveaux sont TOUJOURS actifs
        # (contrairement au planning qui respecte les paramètres de tenant)
        niveau_2_actif = True  # Temps partiel DISPONIBLES
        niveau_3_actif = True  # Temps partiel STAND-BY
        niveau_4_actif = True  # Temps plein INCOMPLETS
        niveau_5_actif = True  # Heures supplémentaires
        
        # Paramètres équipe de garde
        params_equipes_garde = await db.parametres_equipes_garde.find_one({"tenant_id": tenant_id})
        equipes_garde_actif = params_equipes_garde.get("actif", False) if params_equipes_garde else False
        privilegier_equipe_garde = False
        if equipes_garde_actif and params_equipes_garde:
            config_temps_partiel = params_equipes_garde.get("temps_partiel", {})
            privilegier_equipe_garde = config_temps_partiel.get("privilegier_equipe_garde", False)
        
        logger.info(f"⚙️ Remplacements: Tous niveaux actifs (N2, N3, N4, N5 = True)")
        logger.info(f"⚙️ Équipe de garde - actif:{equipes_garde_actif}, prioriser:{privilegier_equipe_garde}")
        
        # Paramètres de remplacements
        competences_egales = parametres.get("competences_egales", False) if parametres else False
        
        # ==================== RÉCUPÉRATION DES DONNÉES ====================
        
        # Type de garde
        type_garde_data = await db.types_garde.find_one({"id": type_garde_id, "tenant_id": tenant_id})
        if not type_garde_data:
            logger.error(f"Type de garde non trouvé: {type_garde_id}")
            return []
        
        competences_requises = type_garde_data.get("competences_requises", [])
        duree_garde = type_garde_data.get("duree_heures", 8)
        est_garde_externe = type_garde_data.get("est_garde_externe", False)
        officier_obligatoire = type_garde_data.get("officier_obligatoire", False)
        heure_debut_garde = type_garde_data.get("heure_debut", "00:00")
        heure_fin_garde = type_garde_data.get("heure_fin", "23:59")
        
        # Fonction pour vérifier si deux plages horaires se chevauchent
        def plages_se_chevauchent(debut1: str, fin1: str, debut2: str, fin2: str) -> bool:
            """
            Vérifie si deux plages horaires se chevauchent.
            Format des heures: "HH:MM"
            Gère les gardes qui passent minuit (fin < début)
            """
            def heure_to_minutes(h: str) -> int:
                try:
                    parts = h.split(":")
                    return int(parts[0]) * 60 + int(parts[1])
                except:
                    return 0
            
            d1, f1 = heure_to_minutes(debut1), heure_to_minutes(fin1)
            d2, f2 = heure_to_minutes(debut2), heure_to_minutes(fin2)
            
            # Gestion des gardes qui passent minuit
            # Si fin < début, c'est une garde de nuit (ex: 22:00 - 06:00)
            if f1 < d1:
                f1 += 24 * 60  # Ajouter 24h
            if f2 < d2:
                f2 += 24 * 60
            
            # Deux plages se chevauchent si le début de l'une est avant la fin de l'autre
            return d1 < f2 and d2 < f1
        
        logger.info(f"🕐 Type de garde demandé: {type_garde_data.get('nom')} ({heure_debut_garde} - {heure_fin_garde})")
        
        # ==================== RÉCUPÉRATION DES GRADES ====================
        grades_list = await db.grades.find({"tenant_id": tenant_id}).to_list(100)
        grades_map = {g.get("nom"): g for g in grades_list}
        
        def est_officier_grade(grade_nom: str) -> bool:
            """Vérifie si un grade est considéré comme officier"""
            grade_info = grades_map.get(grade_nom, {})
            return grade_info.get("est_officier", False) == True
        
        def est_eligible_fonction_superieure(user_data: dict) -> bool:
            """Vérifie si l'utilisateur peut opérer en fonction supérieure"""
            return user_data.get("fonction_superieur", False) == True
        
        # Demandeur
        demandeur = await db.users.find_one({"id": demandeur_id, "tenant_id": tenant_id})
        demandeur_grade = demandeur.get("grade", "pompier") if demandeur else "pompier"
        demandeur_grade_lower = demandeur_grade.lower() if demandeur_grade else "pompier"
        demandeur_competences = set(demandeur.get("competences", [])) if demandeur else set()
        
        # ==================== RÈGLE OFFICIER ====================
        # La règle officier s'applique SEULEMENT si:
        # 1. Le type de garde nécessite un officier (officier_obligatoire = True)
        # 2. Le demandeur est officier
        # 3. ET il n'y a PAS d'autre officier déjà assigné à cette garde à cette date
        #
        # Si un autre officier est déjà assigné, la règle est respectée même si le remplaçant n'est pas officier
        
        demandeur_est_officier = est_officier_grade(demandeur_grade)
        besoin_officier_remplacement = False
        
        if officier_obligatoire and demandeur_est_officier:
            # Vérifier s'il y a d'autres officiers assignés à cette garde à cette date
            autres_assignations = await db.assignations.find({
                "tenant_id": tenant_id,
                "type_garde_id": type_garde_id,
                "date": date_garde,
                "user_id": {"$ne": demandeur_id}  # Exclure le demandeur
            }).to_list(100)
            
            # Vérifier si au moins un des autres assignés est officier
            autre_officier_present = False
            for assignation in autres_assignations:
                autre_user = await db.users.find_one({"id": assignation["user_id"], "tenant_id": tenant_id})
                if autre_user:
                    autre_grade = autre_user.get("grade", "")
                    if est_officier_grade(autre_grade):
                        autre_officier_present = True
                        logger.info(f"🎖️ Autre officier trouvé sur cette garde: {autre_user.get('prenom', '')} {autre_user.get('nom', '')} ({autre_grade})")
                        break
            
            # La règle s'applique SEULEMENT s'il n'y a pas d'autre officier
            besoin_officier_remplacement = not autre_officier_present
            
            if besoin_officier_remplacement:
                logger.info(f"🎖️ RÈGLE OFFICIER ACTIVE: Le remplaçant doit être officier ou éligible")
                logger.info(f"🎖️ Le demandeur {demandeur.get('prenom', '')} {demandeur.get('nom', '')} est le SEUL officier sur cette garde")
            else:
                logger.info(f"🎖️ Règle officier NON ACTIVE: Un autre officier est déjà présent sur cette garde")
        
        logger.info(f"🎖️ officier_obligatoire={officier_obligatoire}, demandeur_est_officier={demandeur_est_officier}, besoin_officier_remplacement={besoin_officier_remplacement}")
        
        # Hiérarchie des grades (fallback si pas en DB)
        grades_hierarchie = {
            "pompier": 1,
            "lieutenant": 2,
            "capitaine": 3,
            "chef": 4,
            "chef aux opérations": 5,
            "directeur": 6,
            "eligible": 2,
            "éligible": 2
        }
        
        def get_niveau_hierarchique(grade_nom: str) -> int:
            """Retourne le niveau hiérarchique d'un grade (depuis DB ou fallback)"""
            grade_info = grades_map.get(grade_nom, {})
            niveau_db = grade_info.get("niveau_hierarchique")
            if niveau_db is not None:
                return niveau_db
            return grades_hierarchie.get(grade_nom.lower() if grade_nom else "", 1)
        
        demandeur_grade_niveau = get_niveau_hierarchique(demandeur_grade)
        
        # Tous les utilisateurs actifs
        exclus_ids_set = set(exclus_ids + [demandeur_id])
        users_cursor = db.users.find({
            "tenant_id": tenant_id,
            "id": {"$nin": list(exclus_ids_set)},
            "statut": "Actif"
        })
        users_list = await users_cursor.to_list(length=None)
        
        logger.info(f"🔍 Recherche remplaçants pour date={date_garde}, type_garde={type_garde_id}")
        logger.info(f"🔍 {len(users_list)} employés actifs (excluant {len(exclus_ids)} déjà contactés + demandeur)")
        if exclus_ids:
            logger.info(f"🔍 IDs exclus (déjà contactés): {exclus_ids}")
        
        # Calculer les heures mensuelles pour l'équitabilité
        debut_mois = datetime.strptime(date_garde, "%Y-%m-%d").replace(day=1)
        fin_mois = (debut_mois + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        user_monthly_hours = {}
        for user in users_list:
            assignations_mois = await db.assignations.find({
                "user_id": user["id"],
                "tenant_id": tenant_id,
                "date": {
                    "$gte": debut_mois.strftime("%Y-%m-%d"),
                    "$lte": fin_mois.strftime("%Y-%m-%d")
                }
            }).to_list(1000)
            user_monthly_hours[user["id"]] = sum(8 for _ in assignations_mois)
        
        # Calculer l'équipe de garde du jour si actif
        equipe_garde_du_jour = None
        if privilegier_equipe_garde and params_equipes_garde:
            config_tp = params_equipes_garde.get("temps_partiel", {})
            if config_tp.get("rotation_active"):
                from routes.equipes_garde import get_equipe_garde_du_jour_sync
                equipe_garde_du_jour = get_equipe_garde_du_jour_sync(
                    type_rotation=config_tp.get("type_rotation", "hebdomadaire"),
                    date_reference=config_tp.get("date_reference", date_garde),
                    date_cible=date_garde,
                    nombre_equipes=config_tp.get("nombre_equipes", 2),
                    pattern_mode=config_tp.get("pattern_mode", "hebdomadaire"),
                    pattern_personnalise=config_tp.get("pattern_personnalise", []),
                    duree_cycle=config_tp.get("duree_cycle", 14),
                    jour_rotation=config_tp.get("jour_rotation") or "monday",
                    heure_rotation=config_tp.get("heure_rotation") or "18:00",
                    heure_actuelle="12:00"
                )
                logger.info(f"📊 Équipe de garde du jour: {equipe_garde_du_jour}")
        
        # ==================== CLASSIFICATION DES CANDIDATS ====================
        
        # Structure: niveau -> liste de candidats
        candidats_par_niveau = {2: [], 3: [], 4: [], 5: []}
        
        for user in users_list:
            user_name = f"{user.get('prenom', '')} {user.get('nom', '')}"
            user_id = user["id"]
            user_grade = user.get("grade", "pompier")
            user_grade_lower = user_grade.lower() if user_grade else "pompier"
            user_grade_niveau = get_niveau_hierarchique(user_grade)
            user_fonction_superieure = user.get("fonction_superieur", False)
            user_type_emploi = user.get("type_emploi", "temps_partiel")
            user_heures_max = user.get("heures_max_semaine", 40)
            date_embauche = user.get("date_embauche", "2999-12-31")
            user_equipe_garde = user.get("equipe_garde")
            
            # ========== N0: FILTRE COMPÉTENCES ==========
            # Normaliser les compétences (lowercase, strip) pour comparaison insensible à la casse
            def normalize_competences(comp_list):
                return set(c.lower().strip() for c in comp_list if c)
            
            if competences_egales:
                user_competences_norm = normalize_competences(user.get("competences", []))
                demandeur_competences_norm = normalize_competences(list(demandeur_competences))
                if demandeur_competences_norm and not demandeur_competences_norm.issubset(user_competences_norm):
                    manquantes = demandeur_competences_norm - user_competences_norm
                    logger.info(f"❌ {user_name} - N0: Compétences insuffisantes (manque: {manquantes})")
                    continue
            
            if competences_requises:
                user_competences_norm = normalize_competences(user.get("competences", []))
                competences_requises_norm = normalize_competences(competences_requises)
                if not competences_requises_norm.issubset(user_competences_norm):
                    manquantes = competences_requises_norm - user_competences_norm
                    logger.info(f"❌ {user_name} - N0: Compétences garde insuffisantes (requises: {list(competences_requises_norm)}, a: {list(user_competences_norm)}, manque: {list(manquantes)})")
                    continue
            
            # ========== N0: FILTRE INDISPONIBILITÉ ==========
            indispo = await db.disponibilites.find_one({
                "user_id": user_id,
                "tenant_id": tenant_id,
                "date": date_garde,
                "statut": "indisponible"
            })
            if indispo:
                logger.info(f"❌ {user_name} - N0: Indisponible déclaré pour {date_garde}")
                continue
            
            # ========== N0: FILTRE RÈGLE OFFICIER ==========
            # Si la règle officier est active, seuls les officiers ou éligibles peuvent remplacer
            if besoin_officier_remplacement:
                user_est_officier = est_officier_grade(user_grade)
                user_est_eligible = est_eligible_fonction_superieure(user)
                
                if not user_est_officier and not user_est_eligible:
                    logger.info(f"❌ {user_name} - N0: Règle officier - n'est pas officier ni éligible (grade={user_grade})")
                    continue
                else:
                    officier_tag = "OFFICIER" if user_est_officier else "ÉLIGIBLE"
                    logger.info(f"✅ {user_name} - Règle officier OK [{officier_tag}]")
            
            # ========== N1: FILTRE ASSIGNATION EN CONFLIT (avec chevauchement horaire) ==========
            # On vérifie si le candidat a une assignation qui CHEVAUCHE les horaires de la garde demandée
            assignations_ce_jour = await db.assignations.find({
                "user_id": user_id,
                "tenant_id": tenant_id,
                "date": date_garde
            }).to_list(10)
            
            conflit_horaire = False
            garde_en_conflit = None
            for assignation in assignations_ce_jour:
                # Récupérer le type de garde de cette assignation
                type_garde_assignation = await db.types_garde.find_one({
                    "id": assignation.get("type_garde_id"),
                    "tenant_id": tenant_id
                })
                if type_garde_assignation:
                    heure_debut_existante = type_garde_assignation.get("heure_debut", "00:00")
                    heure_fin_existante = type_garde_assignation.get("heure_fin", "23:59")
                    
                    # Vérifier si les horaires se chevauchent
                    if plages_se_chevauchent(heure_debut_garde, heure_fin_garde, heure_debut_existante, heure_fin_existante):
                        conflit_horaire = True
                        garde_en_conflit = f"{type_garde_assignation.get('nom', 'Inconnu')} ({heure_debut_existante}-{heure_fin_existante})"
                        break
                    else:
                        logger.info(f"ℹ️ {user_name} - Assigné à {type_garde_assignation.get('nom')} ({heure_debut_existante}-{heure_fin_existante}) mais PAS de chevauchement avec {heure_debut_garde}-{heure_fin_garde}")
            
            if conflit_horaire:
                logger.info(f"❌ {user_name} - N1: Conflit horaire avec {garde_en_conflit}")
                continue
            
            # ========== CALCULS COMMUNS ==========
            
            # Disponibilité déclarée
            dispo = await db.disponibilites.find_one({
                "user_id": user_id,
                "tenant_id": tenant_id,
                "date": date_garde,
                "statut": "disponible"
            })
            has_disponibilite = dispo is not None
            
            # Heures de la semaine
            semaine_debut = datetime.strptime(date_garde, "%Y-%m-%d")
            while semaine_debut.weekday() != 0:
                semaine_debut -= timedelta(days=1)
            semaine_fin = semaine_debut + timedelta(days=6)
            
            assignations_semaine = await db.assignations.find({
                "user_id": user_id,
                "tenant_id": tenant_id,
                "date": {
                    "$gte": semaine_debut.strftime("%Y-%m-%d"),
                    "$lte": semaine_fin.strftime("%Y-%m-%d")
                }
            }).to_list(1000)
            
            heures_travaillees = sum(8 for _ in assignations_semaine)
            depasserait_max = (heures_travaillees + duree_garde) > user_heures_max
            
            # Grade équivalent ou fonction supérieure
            est_grade_equivalent = user_grade_niveau == demandeur_grade_niveau
            est_fonction_superieure = user_fonction_superieure and user_grade_niveau == demandeur_grade_niveau - 1
            
            # Priorité grade: 1=équivalent, 2=fonction sup, 3=autre
            if est_grade_equivalent:
                grade_priorite = 1
            elif est_fonction_superieure:
                grade_priorite = 2
            else:
                grade_priorite = 3
            
            # Priorité équipe de garde: 0=membre équipe garde, 1=autre
            equipe_priorite = 1
            if privilegier_equipe_garde and equipe_garde_du_jour and est_garde_externe:
                if user_equipe_garde == equipe_garde_du_jour:
                    equipe_priorite = 0
            
            # Équitabilité (heures mensuelles)
            heures_mois = user_monthly_hours.get(user_id, 0)
            
            # Données du candidat
            candidat_data = {
                "user_id": user_id,
                "nom_complet": user_name,
                "email": user.get("email", ""),
                "telephone": user.get("telephone", ""),
                "grade": user_grade,
                "date_embauche": date_embauche,
                "has_disponibilite": has_disponibilite,
                "formations": list(set(user.get("competences", []))),
                "type_emploi": user_type_emploi,
                "heures_travaillees": heures_travaillees,
                "heures_max": user_heures_max,
                "heures_mois": heures_mois,
                "fonction_superieure": user_fonction_superieure,
                "equipe_garde": user_equipe_garde,
                # Clés de tri
                "grade_priorite": grade_priorite,
                "equipe_priorite": equipe_priorite
            }
            
            # ========== CLASSIFICATION PAR NIVEAU ==========
            
            # N2: Temps partiel DISPONIBLES
            if niveau_2_actif and user_type_emploi in ["temps_partiel", "temporaire"] and has_disponibilite and not depasserait_max:
                candidat_data["niveau"] = 2
                candidat_data["raison"] = f"N2 - TP Dispo (grade_prio:{grade_priorite}, équité:{heures_mois}h)"
                candidats_par_niveau[2].append(candidat_data.copy())
                logger.info(f"✅ {user_name} → N2")
                continue
            
            # N3: Temps partiel STAND-BY
            if niveau_3_actif and user_type_emploi in ["temps_partiel", "temporaire"] and not has_disponibilite and not depasserait_max:
                candidat_data["niveau"] = 3
                candidat_data["raison"] = f"N3 - TP Stand-by (grade_prio:{grade_priorite}, équité:{heures_mois}h)"
                candidats_par_niveau[3].append(candidat_data.copy())
                logger.info(f"✅ {user_name} → N3")
                continue
            
            # N4: Temps plein INCOMPLETS
            if niveau_4_actif and user_type_emploi == "temps_plein" and not depasserait_max:
                candidat_data["niveau"] = 4
                candidat_data["raison"] = f"N4 - TP Incomplet (grade_prio:{grade_priorite}, équité:{heures_mois}h)"
                candidats_par_niveau[4].append(candidat_data.copy())
                logger.info(f"✅ {user_name} → N4")
                continue
            
            # N5: Heures supplémentaires
            if niveau_5_actif and depasserait_max:
                # Temps partiel doit avoir une dispo pour heures sup
                if user_type_emploi in ["temps_partiel", "temporaire"]:
                    if not has_disponibilite:
                        logger.debug(f"❌ {user_name} - N5: TP sans dispo")
                        continue
                
                candidat_data["niveau"] = 5
                candidat_data["raison"] = f"N5 - Heures sup (grade_prio:{grade_priorite}, équité:{heures_mois}h)"
                candidats_par_niveau[5].append(candidat_data.copy())
                logger.info(f"✅ {user_name} → N5")
                continue
            
            logger.info(f"⚠️ {user_name} exclu - type:{user_type_emploi}, dispo:{has_disponibilite}, depasseMax:{depasserait_max}, N2:{niveau_2_actif}, N3:{niveau_3_actif}, N4:{niveau_4_actif}, N5:{niveau_5_actif}")
        
        # ==================== TRI ET ASSEMBLAGE FINAL ====================
        # Ordre de tri:
        # 1. Grade priorité (1=équivalent, 2=fonction sup, 3=autre)
        # 2. Équipe de garde (0=membre équipe, 1=autre) - privilégié mais pas restrictif
        # 3. Équitabilité (heures_mois - moins d'heures = priorité)
        # 4. Ancienneté (date_embauche - plus ancien = priorité)
        
        def sort_key(candidat):
            return (
                candidat.get("grade_priorite", 3),      # 1=grade equiv, 2=fonction sup, 3=autre
                candidat.get("equipe_priorite", 1),     # 0=membre équipe garde, 1=autre
                candidat.get("heures_mois", 999),       # Équitabilité: moins d'heures = priorité
                candidat.get("date_embauche", "2999-12-31")  # Ancienneté: plus ancien = priorité
            )
        
        remplacants_potentiels = []
        
        # Log de résumé par niveau
        logger.info(f"📊 Résumé par niveau: N2={len(candidats_par_niveau[2])}, N3={len(candidats_par_niveau[3])}, N4={len(candidats_par_niveau[4])}, N5={len(candidats_par_niveau[5])}")
        
        for niveau in [2, 3, 4, 5]:
            candidats = candidats_par_niveau[niveau]
            candidats_tries = sorted(candidats, key=sort_key)
            remplacants_potentiels.extend(candidats_tries)
        
        logger.info(f"✅ Trouvé {len(remplacants_potentiels)} remplaçants potentiels (tri: grade → équipe → équité → ancienneté):")
        for i, c in enumerate(remplacants_potentiels[:15]):
            equipe_info = f", équipe:{c.get('equipe_garde')}" if c.get('equipe_garde') else ""
            logger.info(f"   {i+1}. {c['nom_complet']} - N{c.get('niveau')} - grade_prio:{c.get('grade_priorite')}, {c.get('heures_mois')}h/mois{equipe_info}")
        
        return remplacants_potentiels
        
    except Exception as e:
        logger.error(f"❌ Erreur lors de la recherche de remplaçants: {e}", exc_info=True)
        return []


async def generer_token_remplacement(demande_id: str, remplacant_id: str, tenant_id: str) -> str:
    """Génère un token unique et temporaire pour accepter/refuser un remplacement par email"""
    token = str(uuid.uuid4())
    expiration = datetime.now(timezone.utc) + timedelta(hours=48)
    
    await db.tokens_remplacement.insert_one({
        "token": token,
        "demande_id": demande_id,
        "remplacant_id": remplacant_id,
        "tenant_id": tenant_id,
        "expiration": expiration.isoformat(),
        "utilise": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return token


async def envoyer_email_remplacement(
    demande_data: dict,
    remplacant: dict,
    demandeur: dict,
    type_garde: dict,
    tenant_id: str,
    token: str
):
    """Envoie un email au remplaçant potentiel avec les boutons Accepter/Refuser"""
    try:
        import resend
        
        resend_api_key = os.environ.get('RESEND_API_KEY')
        if not resend_api_key:
            logger.warning(f"RESEND_API_KEY non configurée - Email non envoyé")
            return False
        
        resend.api_key = resend_api_key
        
        remplacant_user = await db.users.find_one({"id": remplacant["user_id"]})
        if not remplacant_user or not remplacant_user.get("email"):
            logger.warning(f"Email non trouvé pour remplaçant {remplacant['user_id']}")
            return False
        
        # Vérifier les préférences de notification
        preferences = remplacant_user.get("preferences_notifications", {})
        if not preferences.get("email_actif", True):  # Par défaut activé
            logger.info(f"📧 Email désactivé pour {remplacant_user.get('prenom')} - préférences utilisateur")
            return False
        
        remplacant_email = remplacant_user["email"]
        remplacant_prenom = remplacant_user.get("prenom", "")
        
        frontend_url = os.environ.get('FRONTEND_URL', 'https://www.profiremanager.ca')
        backend_url = os.environ.get('REACT_APP_BACKEND_URL', frontend_url)
        
        lien_accepter = f"{backend_url}/api/remplacement-action/{token}/accepter"
        lien_refuser = f"{backend_url}/api/remplacement-action/{token}/refuser"
        
        demandeur_nom = f"{demandeur.get('prenom', '')} {demandeur.get('nom', '')}"
        type_garde_nom = type_garde.get("nom", "Garde")
        date_garde = demande_data.get("date", "")
        heure_debut = type_garde.get("heure_debut", "")
        heure_fin = type_garde.get("heure_fin", "")
        raison = demande_data.get("raison", "")
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
        </head>
        <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px; background-color: #f5f5f5;">
            <div style="background-color: white; border-radius: 12px; padding: 30px; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
                <div style="text-align: center; margin-bottom: 30px;">
                    <h1 style="color: #dc2626; margin: 0;">🚨 Demande de Remplacement</h1>
                </div>
                
                <p style="font-size: 16px; color: #333;">Bonjour {remplacant_prenom},</p>
                
                <p style="font-size: 16px; color: #333;">
                    <strong>{demandeur_nom}</strong> recherche un remplaçant et vous avez été identifié comme disponible.
                </p>
                
                <div style="background-color: #f8f9fa; border-radius: 8px; padding: 20px; margin: 20px 0;">
                    <h3 style="margin-top: 0; color: #1e3a5f;">📋 Détails de la garde</h3>
                    <table style="width: 100%; border-collapse: collapse;">
                        <tr>
                            <td style="padding: 8px 0; color: #666; width: 40%;">Type de garde:</td>
                            <td style="padding: 8px 0; color: #333; font-weight: bold;">{type_garde_nom}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px 0; color: #666;">Date:</td>
                            <td style="padding: 8px 0; color: #333; font-weight: bold;">{date_garde}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px 0; color: #666;">Horaire:</td>
                            <td style="padding: 8px 0; color: #333; font-weight: bold;">{heure_debut} - {heure_fin}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px 0; color: #666;">Demandeur:</td>
                            <td style="padding: 8px 0; color: #333; font-weight: bold;">{demandeur_nom}</td>
                        </tr>
                        {f'<tr><td style="padding: 8px 0; color: #666;">Raison:</td><td style="padding: 8px 0; color: #333;">{raison}</td></tr>' if raison else ''}
                    </table>
                </div>
                
                <p style="font-size: 16px; color: #333; text-align: center; margin: 30px 0 20px;">
                    <strong>Pouvez-vous effectuer ce remplacement ?</strong>
                </p>
                
                <div style="text-align: center; margin: 30px 0;">
                    <a href="{lien_accepter}" 
                       style="display: inline-block; background-color: #22c55e; color: white; padding: 15px 40px; 
                              text-decoration: none; border-radius: 8px; font-weight: bold; font-size: 16px; 
                              margin: 0 10px; box-shadow: 0 2px 5px rgba(34,197,94,0.3);">
                        ✅ J'accepte
                    </a>
                    <a href="{lien_refuser}" 
                       style="display: inline-block; background-color: #ef4444; color: white; padding: 15px 40px; 
                              text-decoration: none; border-radius: 8px; font-weight: bold; font-size: 16px; 
                              margin: 0 10px; box-shadow: 0 2px 5px rgba(239,68,68,0.3);">
                        ❌ Je refuse
                    </a>
                </div>
                
                <p style="font-size: 14px; color: #666; text-align: center; margin-top: 30px;">
                    Vous pouvez également répondre directement dans l'application ProFireManager.
                </p>
                
                <hr style="border: none; border-top: 1px solid #eee; margin: 30px 0;">
                
                <p style="font-size: 12px; color: #999; text-align: center;">
                    Ce lien est valide pendant 48 heures.<br>
                    Si vous n'êtes pas concerné par cette demande, veuillez ignorer cet email.
                </p>
            </div>
        </body>
        </html>
        """
        
        params = {
            "from": "ProFireManager <remplacement@profiremanager.ca>",
            "to": [remplacant_email],
            "subject": f"🚨 Demande de remplacement - {type_garde_nom} le {date_garde}",
            "html": html_content
        }
        
        response = resend.Emails.send(params)
        logger.info(f"✅ Email de remplacement envoyé à {remplacant_email} (ID: {response.get('id', 'N/A')})")
        return True
        
    except Exception as e:
        logger.error(f"❌ Erreur envoi email remplacement: {e}", exc_info=True)
        return False


async def envoyer_email_remplacement_trouve(
    demande_data: dict,
    remplacant: dict,
    demandeur: dict,
    type_garde: dict,
    tenant: dict
):
    """Envoie un email au demandeur pour l'informer qu'un remplaçant a été trouvé"""
    try:
        import resend
        from routes.emails_history import log_email_sent
        
        resend_api_key = os.environ.get('RESEND_API_KEY')
        if not resend_api_key:
            logger.warning(f"RESEND_API_KEY non configurée - Email non envoyé")
            return False
        
        resend.api_key = resend_api_key
        
        demandeur_email = demandeur.get("email")
        if not demandeur_email:
            logger.warning(f"Email non trouvé pour demandeur {demandeur.get('id')}")
            return False
        
        # Vérifier les préférences de notification
        preferences = demandeur.get("preferences_notifications", {})
        if not preferences.get("email_actif", True):  # Par défaut activé
            logger.info(f"📧 Email désactivé pour {demandeur.get('prenom')} - préférences utilisateur")
            return False
        
        demandeur_prenom = demandeur.get("prenom", "")
        remplacant_nom = f"{remplacant.get('prenom', '')} {remplacant.get('nom', '')}"
        type_garde_nom = type_garde.get("nom", "Garde") if type_garde else "Garde"
        date_garde = demande_data.get("date", "")
        tenant_nom = tenant.get("nom", "ProFireManager")
        
        sender_email = os.environ.get('SENDER_EMAIL', 'noreply@profiremanager.ca')
        frontend_url = os.environ.get('FRONTEND_URL', 'https://www.profiremanager.ca')
        
        # Construire l'email HTML
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; margin: 0; padding: 0; }}
                .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                .header {{ background: linear-gradient(135deg, #10B981 0%, #059669 100%); color: white; padding: 30px; text-align: center; border-radius: 8px 8px 0 0; }}
                .header h1 {{ margin: 0; font-size: 24px; }}
                .content {{ background-color: #f8f9fa; padding: 30px; border-radius: 0 0 8px 8px; }}
                .success-box {{ background-color: #D1FAE5; border: 1px solid #10B981; border-radius: 8px; padding: 20px; margin: 20px 0; text-align: center; }}
                .success-box .icon {{ font-size: 48px; margin-bottom: 10px; }}
                .details {{ background-color: white; border-radius: 8px; padding: 20px; margin: 20px 0; }}
                .details-row {{ display: flex; justify-content: space-between; padding: 10px 0; border-bottom: 1px solid #e5e7eb; }}
                .details-row:last-child {{ border-bottom: none; }}
                .details-label {{ color: #6B7280; }}
                .details-value {{ font-weight: 600; color: #111827; }}
                .btn {{ display: inline-block; background-color: #10B981; color: white; padding: 14px 30px; text-decoration: none; border-radius: 8px; margin-top: 20px; font-weight: 600; }}
                .footer {{ text-align: center; padding: 20px; color: #666; font-size: 12px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>✅ Remplaçant trouvé!</h1>
                </div>
                <div class="content">
                    <p>Bonjour {demandeur_prenom},</p>
                    
                    <div class="success-box">
                        <div class="icon">🎉</div>
                        <strong style="font-size: 18px; color: #059669;">Bonne nouvelle!</strong>
                        <p style="margin: 10px 0 0 0; color: #065F46;">
                            <strong>{remplacant_nom}</strong> a accepté de vous remplacer.
                        </p>
                    </div>
                    
                    <div class="details">
                        <div class="details-row">
                            <span class="details-label">📅 Date:</span>
                            <span class="details-value">{date_garde}</span>
                        </div>
                        <div class="details-row">
                            <span class="details-label">🚒 Type de garde:</span>
                            <span class="details-value">{type_garde_nom}</span>
                        </div>
                        <div class="details-row">
                            <span class="details-label">👤 Remplaçant:</span>
                            <span class="details-value">{remplacant_nom}</span>
                        </div>
                    </div>
                    
                    <p>Votre demande de remplacement a été traitée avec succès. Le planning a été mis à jour automatiquement.</p>
                    
                    <center>
                        <a href="{frontend_url}/remplacements" class="btn">Voir mes remplacements</a>
                    </center>
                    
                    <p style="margin-top: 30px;">Cordialement,<br>L'équipe {tenant_nom}</p>
                </div>
                <div class="footer">
                    <p>Ceci est un message automatique. Merci de ne pas y répondre.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        params = {
            "from": f"{tenant_nom} <{sender_email}>",
            "to": [demandeur_email],
            "subject": f"✅ Remplaçant trouvé pour le {date_garde}",
            "html": html_content
        }
        
        response = resend.Emails.send(params)
        logger.info(f"✅ Email 'remplacement trouvé' envoyé à {demandeur_email} (ID: {response.get('id', 'N/A')})")
        
        # Logger l'email dans l'historique
        try:
            await log_email_sent(
                tenant_id=tenant.get("id"),
                destinataire=demandeur_email,
                sujet=f"✅ Remplaçant trouvé pour le {date_garde}",
                type_email="remplacement_trouve",
                statut="sent",
                metadata={
                    "demande_id": demande_data.get("id"),
                    "remplacant_id": remplacant.get("id"),
                    "demandeur_id": demandeur.get("id"),
                    "date_garde": date_garde
                }
            )
        except Exception as log_error:
            logger.warning(f"⚠️ Erreur log email: {log_error}")
        
        return True
        
    except Exception as e:
        logger.error(f"❌ Erreur envoi email 'remplacement trouvé': {e}", exc_info=True)
        return False


async def envoyer_email_remplacement_non_trouve(
    demande_data: dict,
    demandeur: dict,
    type_garde: dict,
    tenant: dict
):
    """Envoie un email au demandeur pour l'informer qu'aucun remplaçant n'a été trouvé"""
    try:
        import resend
        from routes.emails_history import log_email_sent
        
        resend_api_key = os.environ.get('RESEND_API_KEY')
        if not resend_api_key:
            logger.warning(f"RESEND_API_KEY non configurée - Email non envoyé")
            return False
        
        resend.api_key = resend_api_key
        
        demandeur_email = demandeur.get("email")
        if not demandeur_email:
            logger.warning(f"Email non trouvé pour demandeur {demandeur.get('id')}")
            return False
        
        # Vérifier les préférences de notification
        preferences = demandeur.get("preferences_notifications", {})
        if not preferences.get("email_actif", True):  # Par défaut activé
            logger.info(f"📧 Email désactivé pour {demandeur.get('prenom')} - préférences utilisateur")
            return False
        
        demandeur_prenom = demandeur.get("prenom", "")
        type_garde_nom = type_garde.get("nom", "Garde") if type_garde else "Garde"
        date_garde = demande_data.get("date", "")
        tenant_nom = tenant.get("nom", "ProFireManager")
        
        sender_email = os.environ.get('SENDER_EMAIL', 'noreply@profiremanager.ca')
        frontend_url = os.environ.get('FRONTEND_URL', 'https://www.profiremanager.ca')
        
        # Construire l'email HTML
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; margin: 0; padding: 0; }}
                .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                .header {{ background: linear-gradient(135deg, #EF4444 0%, #DC2626 100%); color: white; padding: 30px; text-align: center; border-radius: 8px 8px 0 0; }}
                .header h1 {{ margin: 0; font-size: 24px; }}
                .content {{ background-color: #f8f9fa; padding: 30px; border-radius: 0 0 8px 8px; }}
                .alert-box {{ background-color: #FEE2E2; border: 1px solid #EF4444; border-radius: 8px; padding: 20px; margin: 20px 0; text-align: center; }}
                .alert-box .icon {{ font-size: 48px; margin-bottom: 10px; }}
                .details {{ background-color: white; border-radius: 8px; padding: 20px; margin: 20px 0; }}
                .details-row {{ display: flex; justify-content: space-between; padding: 10px 0; border-bottom: 1px solid #e5e7eb; }}
                .details-row:last-child {{ border-bottom: none; }}
                .details-label {{ color: #6B7280; }}
                .details-value {{ font-weight: 600; color: #111827; }}
                .action-box {{ background-color: #FEF3C7; border: 1px solid #F59E0B; border-radius: 8px; padding: 15px; margin: 20px 0; }}
                .btn {{ display: inline-block; background-color: #DC2626; color: white; padding: 14px 30px; text-decoration: none; border-radius: 8px; margin-top: 20px; font-weight: 600; }}
                .footer {{ text-align: center; padding: 20px; color: #666; font-size: 12px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>❌ Aucun remplaçant trouvé</h1>
                </div>
                <div class="content">
                    <p>Bonjour {demandeur_prenom},</p>
                    
                    <div class="alert-box">
                        <div class="icon">😔</div>
                        <strong style="font-size: 18px; color: #DC2626;">Demande expirée</strong>
                        <p style="margin: 10px 0 0 0; color: #991B1B;">
                            Malheureusement, aucun remplaçant n'a pu être trouvé pour votre demande.
                        </p>
                    </div>
                    
                    <div class="details">
                        <div class="details-row">
                            <span class="details-label">📅 Date:</span>
                            <span class="details-value">{date_garde}</span>
                        </div>
                        <div class="details-row">
                            <span class="details-label">🚒 Type de garde:</span>
                            <span class="details-value">{type_garde_nom}</span>
                        </div>
                        <div class="details-row">
                            <span class="details-label">📊 Statut:</span>
                            <span class="details-value" style="color: #DC2626;">Expirée - Aucun remplaçant</span>
                        </div>
                    </div>
                    
                    <div class="action-box">
                        <strong>⚠️ Action requise</strong>
                        <p style="margin: 10px 0 0 0;">
                            Veuillez contacter votre superviseur pour trouver une solution alternative.
                        </p>
                    </div>
                    
                    <center>
                        <a href="{frontend_url}/remplacements" class="btn">Voir mes demandes</a>
                    </center>
                    
                    <p style="margin-top: 30px;">Cordialement,<br>L'équipe {tenant_nom}</p>
                </div>
                <div class="footer">
                    <p>Ceci est un message automatique. Merci de ne pas y répondre.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        params = {
            "from": f"{tenant_nom} <{sender_email}>",
            "to": [demandeur_email],
            "subject": f"❌ Aucun remplaçant trouvé pour le {date_garde}",
            "html": html_content
        }
        
        response = resend.Emails.send(params)
        logger.info(f"✅ Email 'remplacement non trouvé' envoyé à {demandeur_email} (ID: {response.get('id', 'N/A')})")
        
        # Logger l'email dans l'historique
        try:
            await log_email_sent(
                tenant_id=tenant.get("id"),
                destinataire=demandeur_email,
                sujet=f"❌ Aucun remplaçant trouvé pour le {date_garde}",
                type_email="remplacement_non_trouve",
                statut="sent",
                metadata={
                    "demande_id": demande_data.get("id"),
                    "demandeur_id": demandeur.get("id"),
                    "date_garde": date_garde
                }
            )
        except Exception as log_error:
            logger.warning(f"⚠️ Erreur log email: {log_error}")
        
        return True
        
    except Exception as e:
        logger.error(f"❌ Erreur envoi email 'remplacement non trouvé': {e}", exc_info=True)
        return False


async def envoyer_sms_remplacement(
    remplacant: Dict[str, Any],
    demande_data: Dict[str, Any],
    demandeur: Dict[str, Any],
    type_garde: Dict[str, Any],
    tenant_id: str,
    token: str
) -> bool:
    """Envoie un SMS au remplaçant potentiel avec un lien pour accepter/refuser"""
    try:
        from twilio.rest import Client
        
        # Récupérer les credentials Twilio depuis les variables d'environnement
        account_sid = os.environ.get('TWILIO_ACCOUNT_SID')
        auth_token = os.environ.get('TWILIO_AUTH_TOKEN')
        twilio_phone = os.environ.get('TWILIO_PHONE_NUMBER')
        
        if not all([account_sid, auth_token, twilio_phone]):
            logger.warning("⚠️ Configuration Twilio incomplète - SMS non envoyé")
            return False
        
        # Récupérer les infos du remplaçant
        remplacant_user = await db.users.find_one({"id": remplacant["user_id"]})
        if not remplacant_user:
            logger.warning(f"Utilisateur non trouvé: {remplacant['user_id']}")
            return False
        
        # Vérifier les préférences de notification
        preferences = remplacant_user.get("preferences_notifications", {})
        if not preferences.get("sms_actif", True):  # Par défaut activé
            logger.info(f"📵 SMS désactivé pour {remplacant_user.get('prenom')} - préférences utilisateur")
            return False
        
        # Récupérer et formater le numéro de téléphone
        telephone = remplacant_user.get("telephone", "")
        if not telephone:
            logger.warning(f"Pas de téléphone pour {remplacant_user.get('prenom')} {remplacant_user.get('nom')}")
            return False
        
        # Formater le numéro au format E.164 si nécessaire
        telephone_formate = formater_numero_telephone(telephone)
        if not telephone_formate:
            logger.warning(f"Numéro de téléphone invalide: {telephone}")
            return False
        
        # Préparer le message
        demandeur_nom = f"{demandeur.get('prenom', '')} {demandeur.get('nom', '')}"
        type_garde_nom = type_garde.get("nom", "Garde")
        date_garde = demande_data.get("date", "")
        heure_debut = type_garde.get("heure_debut", "")
        heure_fin = type_garde.get("heure_fin", "")
        
        frontend_url = os.environ.get('FRONTEND_URL', 'https://www.profiremanager.ca')
        backend_url = os.environ.get('REACT_APP_BACKEND_URL', frontend_url)
        lien_reponse = f"{backend_url}/api/remplacement-action/{token}/choix"
        
        message = (
            f"🚨 ProFireManager: {demandeur_nom} cherche un remplaçant le {date_garde} "
            f"({type_garde_nom} {heure_debut}-{heure_fin}). "
            f"Répondez ici: {lien_reponse}"
        )
        
        # Envoyer le SMS via Twilio
        client = Client(account_sid, auth_token)
        
        sms = client.messages.create(
            body=message,
            from_=twilio_phone,
            to=telephone_formate
        )
        
        logger.info(f"✅ SMS envoyé à {telephone_formate} (SID: {sms.sid})")
        return True
        
    except Exception as e:
        logger.error(f"❌ Erreur envoi SMS remplacement: {e}", exc_info=True)
        return False


def formater_numero_telephone(numero: str) -> str:
    """
    Formate un numéro de téléphone au format E.164 (+1XXXXXXXXXX pour l'Amérique du Nord)
    """
    if not numero:
        return ""
    
    # Nettoyer le numéro (garder uniquement les chiffres et le +)
    numero_clean = ''.join(c for c in numero if c.isdigit() or c == '+')
    
    # Si déjà au format E.164
    if numero_clean.startswith('+'):
        return numero_clean
    
    # Enlever le 1 au début si présent (indicatif Amérique du Nord)
    if numero_clean.startswith('1') and len(numero_clean) == 11:
        numero_clean = numero_clean[1:]
    
    # Si 10 chiffres, ajouter +1 (Amérique du Nord - Canada/USA)
    if len(numero_clean) == 10:
        return f"+1{numero_clean}"
    
    # Si 11 chiffres commençant par 1
    if len(numero_clean) == 11 and numero_clean.startswith('1'):
        return f"+{numero_clean}"
    
    # Sinon, retourner avec + si valide
    if len(numero_clean) >= 10:
        return f"+{numero_clean}"
    
    return ""


# Import direct de send_push_notification_to_users depuis notifications.py
from routes.notifications import send_push_notification_to_users as push_notification_func

async def get_send_push_notification():
    """Retourne la fonction send_push_notification_to_users"""
    return push_notification_func


async def lancer_recherche_remplacant(demande_id: str, tenant_id: str):
    """Lance la recherche de remplaçant pour une demande"""
    try:
        send_push_notification_to_users = await get_send_push_notification()
        
        demande_data = await db.demandes_remplacement.find_one({"id": demande_id, "tenant_id": tenant_id})
        if not demande_data:
            logger.error(f"Demande de remplacement non trouvée: {demande_id}")
            return
        
        # IMPORTANT: Lire depuis parametres_remplacements (pas parametres) 
        # C'est là que le frontend sauvegarde les paramètres
        parametres_data = await db.parametres_remplacements.find_one({"tenant_id": tenant_id})
        if not parametres_data:
            mode_notification = "un_par_un"
            delai_attente_minutes = 120  # 2 heures par défaut
            nombre_simultane = 1
            max_contacts = 5
            logger.info(f"⚙️ Paramètres remplacements: Aucun trouvé, utilisation des valeurs par défaut")
        else:
            mode_notification = parametres_data.get("mode_notification", "un_par_un")
            # Utiliser delai_attente_minutes si disponible, sinon convertir delai_attente_heures
            delai_attente_minutes = parametres_data.get("delai_attente_minutes") or (parametres_data.get("delai_attente_heures", 2) * 60)
            nombre_simultane = parametres_data.get("nombre_simultane", 3)
            max_contacts = parametres_data.get("max_contacts", 5)
            logger.info(f"⚙️ Paramètres remplacements chargés: délai={delai_attente_minutes}min, mode={mode_notification}, max_contacts={max_contacts}")
        
        exclus_ids = [t.get("user_id") for t in demande_data.get("tentatives_historique", [])]
        
        remplacants = await trouver_remplacants_potentiels(
            tenant_id=tenant_id,
            type_garde_id=demande_data["type_garde_id"],
            date_garde=demande_data["date"],
            demandeur_id=demande_data["demandeur_id"],
            exclus_ids=exclus_ids
        )
        
        logger.info(f"📋 Résultat recherche remplaçants: {len(remplacants)} trouvé(s)")
        for r in remplacants[:5]:
            logger.info(f"  - {r.get('nom_complet')} (priorité {r.get('priorite')}, {r.get('raison')})")
        
        if not remplacants:
            logger.warning(f"⚠️ Aucun remplaçant trouvé pour la demande {demande_id}")
            await db.demandes_remplacement.update_one(
                {"id": demande_id},
                {
                    "$set": {
                        "statut": "expiree",
                        "updated_at": datetime.now(timezone.utc)
                    }
                }
            )
            
            # Notifier superviseurs
            superviseurs = await db.users.find({
                "tenant_id": tenant_id,
                "role": {"$in": ["superviseur", "admin"]}
            }).to_list(100)
            
            superviseur_ids = [s["id"] for s in superviseurs]
            if superviseur_ids:
                demandeur = await db.users.find_one({"id": demande_data["demandeur_id"]})
                demandeur_nom = f"{demandeur.get('prenom', '')} {demandeur.get('nom', '')}" if demandeur else "Un employé"
                
                # Push notification
                await send_push_notification_to_users(
                    user_ids=superviseur_ids,
                    title="❌ Aucun remplaçant trouvé",
                    body=f"Aucun remplaçant disponible pour {demandeur_nom} le {demande_data['date']}",
                    data={
                        "type": "remplacement_expiree",
                        "demande_id": demande_id
                    }
                )
                
                # Notification in-app + email pour chaque superviseur
                for sup in superviseurs:
                    await creer_notification(
                        tenant_id=tenant_id,
                        destinataire_id=sup["id"],
                        type="remplacement_expiree",
                        titre="❌ Aucun remplaçant trouvé",
                        message=f"Aucun remplaçant disponible pour {demandeur_nom} le {demande_data['date']}. Une intervention manuelle est requise.",
                        lien="/remplacements",
                        data={"demande_id": demande_id},
                        envoyer_email=True  # Email aux superviseurs quand demande expirée
                    )
            
            # Notifier le demandeur
            demandeur_id = demande_data.get("demandeur_id")
            if demandeur_id:
                await send_push_notification_to_users(
                    user_ids=[demandeur_id],
                    title="❌ Demande de remplacement expirée",
                    body=f"Aucun remplaçant n'a été trouvé pour votre demande du {demande_data['date']}. Contactez votre superviseur.",
                    data={
                        "type": "remplacement_expiree",
                        "demande_id": demande_id
                    }
                )
                await db.notifications.insert_one({
                    "id": str(uuid.uuid4()),
                    "tenant_id": tenant_id,
                    "user_id": demandeur_id,
                    "type": "remplacement_expiree",
                    "titre": "❌ Demande de remplacement expirée",
                    "message": f"Aucun remplaçant n'a été trouvé pour votre demande du {demande_data['date']}.",
                    "lu": False,
                    "data": {"demande_id": demande_id},
                    "created_at": datetime.now(timezone.utc).isoformat()
                })
                
                # Envoyer un email au demandeur pour l'informer qu'aucun remplaçant n'a été trouvé
                try:
                    tenant = await db.tenants.find_one({"id": tenant_id})
                    type_garde = await db.types_garde.find_one({"id": demande_data["type_garde_id"], "tenant_id": tenant_id})
                    await envoyer_email_remplacement_non_trouve(
                        demande_data=demande_data,
                        demandeur=demandeur,
                        type_garde=type_garde,
                        tenant=tenant
                    )
                except Exception as email_error:
                    logger.warning(f"⚠️ Erreur envoi email remplacement non trouvé: {email_error}")
                    
            return
        
        if mode_notification == "multiple":
            nombre_a_contacter = min(nombre_simultane, len(remplacants))
        else:
            nombre_a_contacter = 1
        
        remplacants_a_contacter = remplacants[:nombre_a_contacter]
        
        remplacant_ids = []
        maintenant = datetime.now(timezone.utc)
        
        for remplacant in remplacants_a_contacter:
            tentative = {
                "user_id": remplacant["user_id"],
                "nom_complet": remplacant["nom_complet"],
                "date_contact": maintenant.isoformat(),
                "statut": "contacted",
                "date_reponse": None
            }
            
            await db.demandes_remplacement.update_one(
                {"id": demande_id},
                {
                    "$push": {"tentatives_historique": tentative},
                    "$addToSet": {"remplacants_contactes_ids": remplacant["user_id"]}
                }
            )
            
            remplacant_ids.append(remplacant["user_id"])
            logger.info(f"📤 Contact remplaçant {remplacant['nom_complet']} pour demande {demande_id}")
        
        date_prochaine = maintenant + timedelta(minutes=delai_attente_minutes)
        
        await db.demandes_remplacement.update_one(
            {"id": demande_id},
            {
                "$set": {
                    "statut": "en_cours",
                    "date_prochaine_tentative": date_prochaine,
                    "updated_at": maintenant
                },
                "$inc": {"nombre_tentatives": 1}
            }
        )
        
        demandeur = await db.users.find_one({"id": demande_data["demandeur_id"]})
        type_garde = await db.types_garde.find_one({"id": demande_data["type_garde_id"]})
        
        await send_push_notification_to_users(
            user_ids=remplacant_ids,
            title="🚨 Demande de remplacement",
            body=f"{demandeur.get('prenom', '')} {demandeur.get('nom', '')} cherche un remplaçant pour {type_garde.get('nom', 'une garde')} le {demande_data['date']}",
            data={
                "type": "remplacement_proposition",
                "demande_id": demande_id,
                "lien": "/remplacements",
                "sound": "urgent"
            }
        )
        
        demandeur_nom = f"{demandeur.get('prenom', '')} {demandeur.get('nom', '')}"
        type_garde_nom = type_garde.get("nom", "une garde") if type_garde else "une garde"
        
        for remplacant_id in remplacant_ids:
            await db.notifications.insert_one({
                "id": str(uuid.uuid4()),
                "tenant_id": tenant_id,
                "user_id": remplacant_id,
                "type": "remplacement_proposition",
                "titre": "🚨 Demande de remplacement urgente",
                "message": f"{demandeur_nom} cherche un remplaçant pour {type_garde_nom} le {demande_data['date']}. Répondez rapidement !",
                "lu": False,
                "urgent": True,
                "data": {
                    "demande_id": demande_id,
                    "lien": "/remplacements"
                },
                "created_at": datetime.now(timezone.utc).isoformat()
            })
        
        for remplacant in remplacants_a_contacter:
            try:
                logger.info(f"📧 Envoi email à {remplacant.get('nom_complet', 'N/A')} ({remplacant.get('email', 'N/A')})")
                token = await generer_token_remplacement(demande_id, remplacant["user_id"], tenant_id)
                
                # Envoyer l'email
                email_result = await envoyer_email_remplacement(
                    demande_data=demande_data,
                    remplacant=remplacant,
                    demandeur=demandeur,
                    type_garde=type_garde,
                    tenant_id=tenant_id,
                    token=token
                )
                logger.info(f"📧 Résultat email: {email_result}")
                
                # Envoyer le SMS
                sms_result = await envoyer_sms_remplacement(
                    remplacant=remplacant,
                    demande_data=demande_data,
                    demandeur=demandeur,
                    type_garde=type_garde,
                    tenant_id=tenant_id,
                    token=token
                )
                logger.info(f"📱 Résultat SMS: {sms_result}")
                
            except Exception as notif_error:
                logger.error(f"Erreur envoi notification à {remplacant.get('nom_complet', 'N/A')}: {notif_error}", exc_info=True)
        
        logger.info(f"✅ Recherche lancée pour demande {demande_id}: {nombre_a_contacter} remplaçant(s) contacté(s)")
        
    except Exception as e:
        logger.error(f"❌ Erreur lors du lancement de la recherche de remplaçant: {e}", exc_info=True)


async def accepter_remplacement(demande_id: str, remplacant_id: str, tenant_id: str, tenant_slug: str = None):
    """Traite l'acceptation d'un remplacement par un remplaçant"""
    try:
        send_push_notification_to_users = await get_send_push_notification()
        
        demande_data = await db.demandes_remplacement.find_one({"id": demande_id, "tenant_id": tenant_id})
        if not demande_data:
            raise HTTPException(status_code=404, detail="Demande non trouvée")
        
        if demande_data["statut"] != "en_cours":
            raise HTTPException(status_code=400, detail="Cette demande n'est plus disponible")
        
        if remplacant_id not in demande_data.get("remplacants_contactes_ids", []):
            raise HTTPException(status_code=403, detail="Vous n'êtes pas autorisé à accepter cette demande")
        
        remplacant = await db.users.find_one({"id": remplacant_id, "tenant_id": tenant_id})
        if not remplacant:
            raise HTTPException(status_code=404, detail="Remplaçant non trouvé")
        
        maintenant = datetime.now(timezone.utc)
        await db.demandes_remplacement.update_one(
            {"id": demande_id},
            {
                "$set": {
                    "statut": "accepte",
                    "remplacant_id": remplacant_id,
                    "updated_at": maintenant
                }
            }
        )
        
        await db.demandes_remplacement.update_one(
            {
                "id": demande_id,
                "tentatives_historique.user_id": remplacant_id
            },
            {
                "$set": {
                    "tentatives_historique.$.statut": "accepted",
                    "tentatives_historique.$.date_reponse": maintenant.isoformat()
                }
            }
        )
        
        assignation = await db.assignations.find_one({
            "tenant_id": tenant_id,
            "user_id": demande_data["demandeur_id"],
            "date": demande_data["date"],
            "type_garde_id": demande_data["type_garde_id"]
        })
        
        if assignation:
            await db.assignations.update_one(
                {"id": assignation["id"]},
                {
                    "$set": {
                        "user_id": remplacant_id,
                        "est_remplacement": True,
                        "demandeur_original_id": demande_data["demandeur_id"],
                        "updated_at": maintenant
                    }
                }
            )
            logger.info(f"✅ Planning mis à jour: {remplacant['prenom']} {remplacant['nom']} remplace assignation {assignation['id']}")
        else:
            logger.warning(f"⚠️ Aucune assignation trouvée pour le demandeur {demande_data['demandeur_id']} le {demande_data['date']}")
        
        demandeur = await db.users.find_one({"id": demande_data["demandeur_id"]})
        
        # Notifications au demandeur (non bloquantes)
        try:
            await send_push_notification_to_users(
                user_ids=[demande_data["demandeur_id"]],
                title="✅ Remplacement trouvé!",
                body=f"{remplacant.get('prenom', '')} {remplacant.get('nom', '')} a accepté de vous remplacer le {demande_data['date']}",
                data={
                    "type": "remplacement_accepte",
                    "demande_id": demande_id,
                    "remplacant_id": remplacant_id
                }
            )
        except Exception as notif_error:
            logger.warning(f"⚠️ Erreur notification push demandeur: {notif_error}")
        
        try:
            await db.notifications.insert_one({
                "id": str(uuid.uuid4()),
                "tenant_id": tenant_id,
                "user_id": demande_data["demandeur_id"],
                "type": "remplacement_accepte",
                "titre": "✅ Remplacement trouvé!",
                "message": f"{remplacant.get('prenom', '')} {remplacant.get('nom', '')} a accepté de vous remplacer le {demande_data['date']}.",
                "lu": False,
                "data": {"demande_id": demande_id, "remplacant_id": remplacant_id},
                "created_at": datetime.now(timezone.utc).isoformat()
            })
        except Exception as notif_error:
            logger.warning(f"⚠️ Erreur insertion notification demandeur: {notif_error}")
        
        # Envoyer un email au demandeur pour l'informer que son remplacement a été trouvé
        try:
            tenant = await db.tenants.find_one({"id": tenant_id})
            type_garde = await db.types_garde.find_one({"id": demande_data["type_garde_id"], "tenant_id": tenant_id})
            await envoyer_email_remplacement_trouve(
                demande_data=demande_data,
                remplacant=remplacant,
                demandeur=demandeur,
                type_garde=type_garde,
                tenant=tenant
            )
        except Exception as email_error:
            logger.warning(f"⚠️ Erreur envoi email remplacement trouvé: {email_error}")
        
        superviseurs = await db.users.find({
            "tenant_id": tenant_id,
            "role": {"$in": ["superviseur", "admin"]}
        }).to_list(100)
        
        superviseur_ids = [s["id"] for s in superviseurs]
        remplacant_nom = f"{remplacant.get('prenom', '')} {remplacant.get('nom', '')}"
        demandeur_nom = f"{demandeur.get('prenom', '')} {demandeur.get('nom', '')}"
        
        if superviseur_ids:
            # Push notification (non bloquant)
            try:
                await send_push_notification_to_users(
                    user_ids=superviseur_ids,
                    title="✅ Remplacement confirmé",
                    body=f"{remplacant_nom} remplace {demandeur_nom} le {demande_data['date']}",
                    data={
                        "type": "remplacement_accepte",
                        "demande_id": demande_id
                    }
                )
            except Exception as notif_error:
                logger.warning(f"⚠️ Erreur notification push superviseurs: {notif_error}")
            
            # Notification in-app + email pour chaque superviseur (non bloquant)
            for sup in superviseurs:
                try:
                    await creer_notification(
                        tenant_id=tenant_id,
                        destinataire_id=sup["id"],
                        type="remplacement_accepte",
                        titre="✅ Remplacement confirmé",
                        message=f"{remplacant_nom} remplace {demandeur_nom} le {demande_data['date']}",
                        lien="/remplacements",
                        data={"demande_id": demande_id},
                        envoyer_email=True  # Email aux superviseurs quand remplaçant trouvé
                    )
                except Exception as notif_error:
                    logger.warning(f"⚠️ Erreur notification superviseur {sup.get('email', '')}: {notif_error}")
        
        autres_remplacants_ids = [
            rid for rid in demande_data.get("remplacants_contactes_ids", [])
            if rid != remplacant_id
        ]
        
        if autres_remplacants_ids:
            try:
                await send_push_notification_to_users(
                    user_ids=autres_remplacants_ids,
                    title="Remplacement pourvu",
                    body=f"Le remplacement du {demande_data['date']} a été pourvu par un autre pompier",
                    data={
                        "type": "remplacement_pourvu",
                        "demande_id": demande_id
                    }
                )
            except Exception as notif_error:
                logger.warning(f"⚠️ Erreur notification autres remplaçants: {notif_error}")
        
        type_garde = await db.types_garde.find_one({"id": demande_data["type_garde_id"], "tenant_id": tenant_id})
        garde_nom = type_garde['nom'] if type_garde else 'garde'
        
        # Créer l'activité (non bloquant)
        try:
            await creer_activite(
                tenant_id=tenant_id,
                type_activite="remplacement_accepte",
                description=f"✅ {remplacant.get('prenom', '')} {remplacant.get('nom', '')} a accepté de remplacer {demandeur.get('prenom', '')} {demandeur.get('nom', '')} pour la {garde_nom} du {demande_data['date']}",
                user_id=remplacant_id,
                user_nom=f"{remplacant.get('prenom', '')} {remplacant.get('nom', '')}"
            )
        except Exception as act_error:
            logger.warning(f"⚠️ Erreur création activité: {act_error}")
        
        # Broadcaster la mise à jour à tous les clients (non bloquant)
        asyncio.create_task(broadcast_remplacement_update(tenant_slug or tenant_id, "accepte", {
            "demande_id": demande_id,
            "remplacant_id": remplacant_id,
            "remplacant_nom": f"{remplacant.get('prenom', '')} {remplacant.get('nom', '')}",
            "demandeur_id": demande_data["demandeur_id"],
            "date": demande_data["date"]
        }))
        
        logger.info(f"✅ Remplacement accepté: demande {demande_id}, remplaçant {remplacant.get('nom', '')}")
        return True
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erreur lors de l'acceptation du remplacement: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Erreur lors de l'acceptation du remplacement")


async def refuser_remplacement(demande_id: str, remplacant_id: str, tenant_id: str):
    """Traite le refus d'un remplacement par un remplaçant"""
    try:
        demande_data = await db.demandes_remplacement.find_one({"id": demande_id, "tenant_id": tenant_id})
        if not demande_data:
            raise HTTPException(status_code=404, detail="Demande non trouvée")
        
        if remplacant_id not in demande_data.get("remplacants_contactes_ids", []):
            raise HTTPException(status_code=403, detail="Vous n'êtes pas autorisé à refuser cette demande")
        
        maintenant = datetime.now(timezone.utc)
        await db.demandes_remplacement.update_one(
            {
                "id": demande_id,
                "tentatives_historique.user_id": remplacant_id
            },
            {
                "$set": {
                    "tentatives_historique.$.statut": "refused",
                    "tentatives_historique.$.date_reponse": maintenant.isoformat()
                }
            }
        )
        
        await db.demandes_remplacement.update_one(
            {"id": demande_id},
            {
                "$pull": {"remplacants_contactes_ids": remplacant_id},
                "$set": {"updated_at": maintenant}
            }
        )
        
        demande_updated = await db.demandes_remplacement.find_one({"id": demande_id})
        if not demande_updated.get("remplacants_contactes_ids"):
            logger.info(f"🔄 Tous les remplaçants ont refusé, relance de la recherche pour demande {demande_id}")
            # Relancer la recherche en arrière-plan pour ne pas bloquer la réponse
            asyncio.create_task(lancer_recherche_remplacant(demande_id, tenant_id))
        
        logger.info(f"❌ Remplacement refusé par remplaçant {remplacant_id} pour demande {demande_id}")
        return True
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erreur lors du refus du remplacement: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Erreur lors du refus du remplacement")


async def verifier_et_traiter_timeouts():
    """Fonction appelée périodiquement pour vérifier les demandes en timeout"""
    try:
        maintenant = datetime.now(timezone.utc)
        
        demandes_cursor = db.demandes_remplacement.find({
            "statut": "en_cours",
            "date_prochaine_tentative": {"$lte": maintenant}
        })
        
        demandes_timeout = await demandes_cursor.to_list(length=None)
        
        for demande in demandes_timeout:
            logger.info(f"⏱️ Timeout atteint pour demande {demande['id']}, relance de la recherche")
            
            for remplacant_id in demande.get("remplacants_contactes_ids", []):
                await db.demandes_remplacement.update_one(
                    {
                        "id": demande["id"],
                        "tentatives_historique.user_id": remplacant_id,
                        "tentatives_historique.statut": "contacted"
                    },
                    {
                        "$set": {
                            "tentatives_historique.$.statut": "expired",
                            "tentatives_historique.$.date_reponse": maintenant.isoformat()
                        }
                    }
                )
            
            await db.demandes_remplacement.update_one(
                {"id": demande["id"]},
                {
                    "$set": {
                        "remplacants_contactes_ids": [],
                        "updated_at": maintenant
                    }
                }
            )
            
            await lancer_recherche_remplacant(demande["id"], demande["tenant_id"])
        
        if demandes_timeout:
            logger.info(f"✅ Traité {len(demandes_timeout)} demande(s) en timeout")
        
    except Exception as e:
        logger.error(f"❌ Erreur lors de la vérification des timeouts: {e}", exc_info=True)


# ==================== ROUTES API ====================

@router.post("/{tenant_slug}/remplacements", response_model=DemandeRemplacement)
async def create_demande_remplacement(tenant_slug: str, demande: DemandeRemplacementCreate, current_user: User = Depends(get_current_user)):
    """Créer une demande de remplacement et lancer automatiquement la recherche"""
    try:
        send_push_notification_to_users = await get_send_push_notification()
        
        tenant = await get_tenant_from_slug(tenant_slug)
        
        # ==================== VALIDATION: Vérifier que le demandeur est bien planifié ====================
        assignation_existante = await db.assignations.find_one({
            "tenant_id": tenant.id,
            "user_id": current_user.id,
            "date": demande.date,
            "type_garde_id": demande.type_garde_id
        })
        
        if not assignation_existante:
            # Récupérer le nom du type de garde pour un message clair
            type_garde = await db.types_garde.find_one({"id": demande.type_garde_id, "tenant_id": tenant.id})
            type_garde_nom = type_garde.get("nom", "ce type de garde") if type_garde else "ce type de garde"
            
            raise HTTPException(
                status_code=400, 
                detail=f"Vous n'êtes pas planifié sur '{type_garde_nom}' le {demande.date}. Veuillez vérifier votre planning."
            )
        # ==================== FIN VALIDATION ====================
        
        priorite = await calculer_priorite_demande(demande.date)
        
        demande_dict = demande.dict()
        demande_dict["tenant_id"] = tenant.id
        demande_dict["demandeur_id"] = current_user.id
        demande_dict["priorite"] = priorite
        demande_dict["statut"] = "en_attente"
        
        demande_obj = DemandeRemplacement(**demande_dict)
        await db.demandes_remplacement.insert_one(demande_obj.dict())
        
        logger.info(f"✅ Demande de remplacement créée: {demande_obj.id} (priorité: {priorite})")
        
        superviseurs_admins = await db.users.find({
            "tenant_id": tenant.id,
            "role": {"$in": ["superviseur", "admin"]}
        }).to_list(100)
        
        superviseur_ids = []
        for user in superviseurs_admins:
            await creer_notification(
                tenant_id=tenant.id,
                destinataire_id=user["id"],
                type="remplacement_demande",
                titre=f"{'🚨 ' if priorite == 'urgent' else ''}Recherche de remplacement en cours",
                message=f"{current_user.prenom} {current_user.nom} cherche un remplaçant pour le {demande.date}",
                lien="/remplacements",
                data={"demande_id": demande_obj.id},
                envoyer_email=False  # Pas d'email aux superviseurs, juste push
            )
            superviseur_ids.append(user["id"])
        
        if superviseur_ids:
            await send_push_notification_to_users(
                user_ids=superviseur_ids,
                title=f"{'🚨 ' if priorite == 'urgent' else ''}Recherche de remplacement",
                body=f"{current_user.prenom} {current_user.nom} cherche un remplaçant pour le {demande.date}",
                data={
                    "type": "remplacement_demande",
                    "demande_id": demande_obj.id,
                    "lien": "/remplacements"
                }
            )
        
        # Lancer la recherche de remplaçant EN ARRIÈRE-PLAN pour une réponse plus rapide
        asyncio.create_task(lancer_recherche_remplacant(demande_obj.id, tenant.id))
        
        type_garde = await db.types_garde.find_one({"id": demande.type_garde_id, "tenant_id": tenant.id})
        garde_nom = type_garde['nom'] if type_garde else 'garde'
        
        # Créer l'activité aussi en arrière-plan
        asyncio.create_task(creer_activite(
            tenant_id=tenant.id,
            type_activite="remplacement_demande",
            description=f"🔄 {current_user.prenom} {current_user.nom} cherche un remplaçant pour la {garde_nom} du {demande.date}",
            user_id=current_user.id,
            user_nom=f"{current_user.prenom} {current_user.nom}"
        ))
        
        cleaned_demande = clean_mongo_doc(demande_obj.dict())
        return DemandeRemplacement(**cleaned_demande)
        
    except HTTPException:
        raise  # Re-raise les HTTPException (comme notre validation)
    except Exception as e:
        logger.error(f"❌ Erreur lors de la création de la demande de remplacement: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Erreur lors de la création de la demande")


@router.get("/{tenant_slug}/remplacements/export-pdf")
async def export_remplacements_pdf(
    tenant_slug: str,
    user_id: str = None,
    current_user: User = Depends(get_current_user)
):
    """Export des demandes de remplacement en PDF"""
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        from io import BytesIO
        
        # Import des helpers PDF depuis server
        import server
        create_branded_pdf = server.create_branded_pdf
        get_modern_pdf_styles = server.get_modern_pdf_styles
        
        tenant = await get_tenant_from_slug(tenant_slug)
        
        if user_id:
            demandes_list = await db.demandes_remplacement.find({
                "tenant_id": tenant.id,
                "demandeur_id": user_id
            }).to_list(length=None)
        else:
            if current_user.role == "employe":
                demandes_list = await db.demandes_remplacement.find({
                    "tenant_id": tenant.id,
                    "demandeur_id": current_user.id
                }).to_list(length=None)
            else:
                demandes_list = await db.demandes_remplacement.find({
                    "tenant_id": tenant.id
                }).to_list(length=None)
        
        users_list = await db.users.find({"tenant_id": tenant.id}).to_list(length=None)
        types_garde_list = await db.types_garde.find({"tenant_id": tenant.id}).to_list(length=None)
        
        users_map = {u['id']: u for u in users_list}
        types_map = {t['id']: t for t in types_garde_list}
        
        buffer, doc, elements = create_branded_pdf(tenant, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        modern_styles = get_modern_pdf_styles(styles)
        
        titre = "Demandes de Remplacement"
        if user_id and user_id in users_map:
            titre = f"Demandes de {users_map[user_id]['prenom']} {users_map[user_id]['nom']}"
        
        elements.append(Paragraph(titre, modern_styles['title']))
        elements.append(Spacer(1, 0.3*inch))
        
        table_data = [['Date', 'Type Garde', 'Demandeur', 'Statut', 'Priorité', 'Remplaçant', 'Notes']]
        
        for demande in sorted(demandes_list, key=lambda x: x.get('date', ''), reverse=True):
            demandeur = users_map.get(demande['demandeur_id'], {})
            demandeur_nom = f"{demandeur.get('prenom', '')} {demandeur.get('nom', '')}" if demandeur else "N/A"
            
            type_garde = types_map.get(demande.get('type_garde_id', ''), {})
            type_nom = type_garde.get('nom', 'N/A')
            
            statut_labels = {
                'en_attente': 'En attente',
                'en_cours': 'En cours',
                'accepte': 'Accepté',
                'expiree': 'Expirée',
                'annulee': 'Annulée',
                'approuve_manuellement': 'Approuvé'
            }
            statut = statut_labels.get(demande.get('statut', ''), demande.get('statut', ''))
            
            priorite = 'Urgent' if demande.get('priorite') == 'urgent' else 'Normal'
            
            remplacant = users_map.get(demande.get('remplacant_id', ''), {})
            remplacant_nom = f"{remplacant.get('prenom', '')} {remplacant.get('nom', '')}" if remplacant else "-"
            
            table_data.append([
                demande.get('date', ''),
                type_nom,
                demandeur_nom,
                statut,
                priorite,
                remplacant_nom,
                demande.get('raison', '')[:30] + '...' if len(demande.get('raison', '')) > 30 else demande.get('raison', '')
            ])
        
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"remplacements_{tenant_slug}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Erreur export PDF remplacements: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Erreur export PDF: {str(e)}")


@router.get("/{tenant_slug}/remplacements/export-excel")
async def export_remplacements_excel(
    tenant_slug: str,
    user_id: str = None,
    current_user: User = Depends(get_current_user)
):
    """Export des demandes de remplacement en Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        tenant = await get_tenant_from_slug(tenant_slug)
        
        if user_id:
            demandes_list = await db.demandes_remplacement.find({
                "tenant_id": tenant.id,
                "demandeur_id": user_id
            }).to_list(length=None)
        else:
            if current_user.role == "employe":
                demandes_list = await db.demandes_remplacement.find({
                    "tenant_id": tenant.id,
                    "demandeur_id": current_user.id
                }).to_list(length=None)
            else:
                demandes_list = await db.demandes_remplacement.find({
                    "tenant_id": tenant.id
                }).to_list(length=None)
        
        users_list = await db.users.find({"tenant_id": tenant.id}).to_list(length=None)
        types_garde_list = await db.types_garde.find({"tenant_id": tenant.id}).to_list(length=None)
        
        users_map = {u['id']: u for u in users_list}
        types_map = {t['id']: t for t in types_garde_list}
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Remplacements"
        
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = ['Date', 'Type Garde', 'Demandeur', 'Statut', 'Priorité', 'Remplaçant', 'Raison', 'Créé le']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        for row, demande in enumerate(sorted(demandes_list, key=lambda x: x.get('date', ''), reverse=True), 2):
            demandeur = users_map.get(demande['demandeur_id'], {})
            type_garde = types_map.get(demande.get('type_garde_id', ''), {})
            remplacant = users_map.get(demande.get('remplacant_id', ''), {})
            
            statut_labels = {
                'en_attente': 'En attente',
                'en_cours': 'En cours',
                'accepte': 'Accepté',
                'expiree': 'Expirée',
                'annulee': 'Annulée',
                'approuve_manuellement': 'Approuvé'
            }
            
            data = [
                demande.get('date', ''),
                type_garde.get('nom', 'N/A'),
                f"{demandeur.get('prenom', '')} {demandeur.get('nom', '')}",
                statut_labels.get(demande.get('statut', ''), demande.get('statut', '')),
                'Urgent' if demande.get('priorite') == 'urgent' else 'Normal',
                f"{remplacant.get('prenom', '')} {remplacant.get('nom', '')}" if remplacant else "-",
                demande.get('raison', ''),
                str(demande.get('created_at', ''))[:10]
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
        
        for column_cells in ws.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"remplacements_{tenant_slug}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Erreur export Excel remplacements: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Erreur export Excel: {str(e)}")


@router.get("/{tenant_slug}/remplacements", response_model=List[DemandeRemplacement])
async def get_demandes_remplacement(tenant_slug: str, current_user: User = Depends(get_current_user)):
    """Liste des demandes de remplacement avec les noms des demandeurs"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if current_user.role == "employe":
        demandes = await db.demandes_remplacement.find({
            "tenant_id": tenant.id,
            "demandeur_id": current_user.id
        }).to_list(1000)
    else:
        demandes = await db.demandes_remplacement.find({"tenant_id": tenant.id}).to_list(1000)
    
    # Enrichir les demandes avec le nom du demandeur
    enriched_demandes = []
    for demande in demandes:
        cleaned = clean_mongo_doc(demande)
        
        # Récupérer le nom du demandeur si non présent
        if not cleaned.get("demandeur_nom") and cleaned.get("demandeur_id"):
            demandeur = await db.users.find_one({"id": cleaned["demandeur_id"]}, {"prenom": 1, "nom": 1})
            if demandeur:
                cleaned["demandeur_nom"] = f"{demandeur.get('prenom', '')} {demandeur.get('nom', '')}".strip()
        
        # Récupérer le nom du remplaçant si présent
        if cleaned.get("remplacant_id") and not cleaned.get("remplacant_nom"):
            remplacant = await db.users.find_one({"id": cleaned["remplacant_id"]}, {"prenom": 1, "nom": 1})
            if remplacant:
                cleaned["remplacant_nom"] = f"{remplacant.get('prenom', '')} {remplacant.get('nom', '')}".strip()
        
        enriched_demandes.append(DemandeRemplacement(**cleaned))
    
    return enriched_demandes


@router.get("/{tenant_slug}/remplacements/propositions")
async def get_propositions_remplacement(tenant_slug: str, current_user: User = Depends(get_current_user)):
    """Récupère les propositions de remplacement pour l'utilisateur connecté"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    demandes = await db.demandes_remplacement.find({
        "tenant_id": tenant.id,
        "statut": "en_cours",
        "remplacants_contactes_ids": current_user.id
    }).to_list(1000)
    
    propositions = []
    for demande in demandes:
        demandeur = await db.users.find_one({"id": demande["demandeur_id"]})
        type_garde = await db.types_garde.find_one({"id": demande["type_garde_id"]})
        
        demande["demandeur"] = {
            "nom": demandeur.get("nom", ""),
            "prenom": demandeur.get("prenom", ""),
            "email": demandeur.get("email", "")
        } if demandeur else None
        
        demande["type_garde"] = {
            "nom": type_garde.get("nom", ""),
            "heure_debut": type_garde.get("heure_debut", ""),
            "heure_fin": type_garde.get("heure_fin", "")
        } if type_garde else None
        
        propositions.append(clean_mongo_doc(demande))
    
    return propositions


@router.get("/{tenant_slug}/remplacements/{demande_id}/suivi")
async def get_suivi_remplacement(
    tenant_slug: str,
    demande_id: str,
    current_user: User = Depends(get_current_user)
):
    """
    Récupère le suivi détaillé d'une demande de remplacement
    Accessible par le demandeur, les admins et les superviseurs
    """
    tenant = await get_tenant_from_slug(tenant_slug)
    
    demande = await db.demandes_remplacement.find_one({
        "id": demande_id,
        "tenant_id": tenant.id
    })
    
    if not demande:
        raise HTTPException(status_code=404, detail="Demande non trouvée")
    
    # Vérifier les permissions : demandeur ou admin/superviseur
    if current_user.role == "employe" and demande.get("demandeur_id") != current_user.id:
        raise HTTPException(status_code=403, detail="Accès non autorisé")
    
    # Récupérer les informations de notifications envoyées
    tentatives = demande.get("tentatives_historique", [])
    
    # Enrichir chaque tentative avec les informations de l'utilisateur
    tentatives_enrichies = []
    for tentative in tentatives:
        user_id = tentative.get("user_id")
        user_info = await db.users.find_one({"id": user_id}, {"prenom": 1, "nom": 1, "email": 1, "telephone": 1})
        
        tentative_enrichie = {
            **tentative,
            "nom_complet": tentative.get("nom_complet") or (f"{user_info.get('prenom', '')} {user_info.get('nom', '')}" if user_info else "Inconnu"),
            "email": user_info.get("email") if user_info else None,
            "telephone": user_info.get("telephone") if user_info else None,
            # Par défaut, on considère que tous les canaux ont été utilisés
            # (à améliorer plus tard avec un vrai tracking des envois)
            "email_envoye": True,
            "sms_envoye": True,
            "push_envoye": True
        }
        tentatives_enrichies.append(tentative_enrichie)
    
    # Calculer les statistiques
    nb_tentatives = len(tentatives_enrichies)
    nb_acceptes = sum(1 for t in tentatives if t.get("statut") == "accepted")
    nb_refuses = sum(1 for t in tentatives if t.get("statut") == "refused")
    nb_en_attente = sum(1 for t in tentatives if t.get("statut") == "contacted")
    
    # Récupérer les informations du demandeur et du remplaçant
    demandeur = await db.users.find_one({"id": demande.get("demandeur_id")}, {"prenom": 1, "nom": 1})
    remplacant = None
    if demande.get("remplacant_id"):
        remplacant = await db.users.find_one({"id": demande.get("remplacant_id")}, {"prenom": 1, "nom": 1})
    
    type_garde = await db.types_garde.find_one({"id": demande.get("type_garde_id"), "tenant_id": tenant.id})
    
    suivi = {
        "demande_id": demande_id,
        "statut": demande.get("statut"),
        "date_garde": demande.get("date"),
        "raison": demande.get("raison"),
        "priorite": demande.get("priorite"),
        "type_garde": type_garde.get("nom") if type_garde else "Inconnu",
        "demandeur_nom": f"{demandeur.get('prenom', '')} {demandeur.get('nom', '')}" if demandeur else "Inconnu",
        "remplacant_nom": f"{remplacant.get('prenom', '')} {remplacant.get('nom', '')}" if remplacant else None,
        "created_at": demande.get("created_at"),
        "updated_at": demande.get("updated_at"),
        "tentatives": tentatives_enrichies,
        "notifications_envoyees": {
            "email": nb_tentatives,
            "sms": nb_tentatives,  # Approximation - à améliorer
            "push": nb_tentatives  # Approximation - à améliorer
        },
        "statistiques": {
            "total_contactes": nb_tentatives,
            "acceptes": nb_acceptes,
            "refuses": nb_refuses,
            "en_attente": nb_en_attente,
            "sans_reponse": nb_tentatives - nb_acceptes - nb_refuses - nb_en_attente
        }
    }
    
    logger.info(f"📋 Suivi demande {demande_id}: {nb_tentatives} tentatives, statut={demande.get('statut')}")
    
    return suivi


@router.put("/{tenant_slug}/remplacements/{demande_id}/accepter")
async def accepter_demande_remplacement(
    tenant_slug: str,
    demande_id: str,
    current_user: User = Depends(get_current_user)
):
    """Accepter/Approuver manuellement une demande de remplacement"""
    send_push_notification_to_users = await get_send_push_notification()
    tenant = await get_tenant_from_slug(tenant_slug)
    
    demande_data = await db.demandes_remplacement.find_one({"id": demande_id, "tenant_id": tenant.id})
    if not demande_data:
        raise HTTPException(status_code=404, detail="Demande non trouvée")
    
    if current_user.role in ["admin", "superviseur"]:
        maintenant = datetime.now(timezone.utc)
        await db.demandes_remplacement.update_one(
            {"id": demande_id},
            {
                "$set": {
                    "statut": "approuve_manuellement",
                    "approuve_par_id": current_user.id,
                    "date_approbation": maintenant.isoformat(),
                    "updated_at": maintenant
                }
            }
        )
        
        demandeur_id = demande_data.get("demandeur_id")
        if demandeur_id:
            await send_push_notification_to_users(
                user_ids=[demandeur_id],
                title="✅ Demande approuvée",
                body=f"Votre demande de remplacement du {demande_data['date']} a été approuvée par un superviseur.",
                data={
                    "type": "remplacement_approuve",
                    "demande_id": demande_id
                }
            )
            await db.notifications.insert_one({
                "id": str(uuid.uuid4()),
                "tenant_id": tenant.id,
                "user_id": demandeur_id,
                "type": "remplacement_approuve",
                "titre": "✅ Demande approuvée",
                "message": f"Votre demande de remplacement du {demande_data['date']} a été approuvée.",
                "lu": False,
                "data": {"demande_id": demande_id},
                "created_at": maintenant.isoformat()
            })
        
        return {
            "message": "Demande approuvée avec succès",
            "demande_id": demande_id
        }
    else:
        await accepter_remplacement(demande_id, current_user.id, tenant.id, tenant_slug)
        
        return {
            "message": "Remplacement accepté avec succès",
            "demande_id": demande_id
        }


@router.put("/{tenant_slug}/remplacements/{demande_id}/refuser")
async def refuser_demande_remplacement(
    tenant_slug: str,
    demande_id: str,
    current_user: User = Depends(get_current_user)
):
    """Refuser/Annuler une demande de remplacement"""
    send_push_notification_to_users = await get_send_push_notification()
    tenant = await get_tenant_from_slug(tenant_slug)
    
    demande_data = await db.demandes_remplacement.find_one({"id": demande_id, "tenant_id": tenant.id})
    if not demande_data:
        raise HTTPException(status_code=404, detail="Demande non trouvée")
    
    if current_user.role in ["admin", "superviseur"]:
        maintenant = datetime.now(timezone.utc)
        await db.demandes_remplacement.update_one(
            {"id": demande_id},
            {
                "$set": {
                    "statut": "annulee",
                    "annule_par_id": current_user.id,
                    "date_annulation": maintenant.isoformat(),
                    "updated_at": maintenant
                }
            }
        )
        
        demandeur_id = demande_data.get("demandeur_id")
        if demandeur_id:
            await send_push_notification_to_users(
                user_ids=[demandeur_id],
                title="❌ Demande annulée",
                body=f"Votre demande de remplacement du {demande_data['date']} a été annulée par un superviseur.",
                data={
                    "type": "remplacement_annulee",
                    "demande_id": demande_id
                }
            )
            await db.notifications.insert_one({
                "id": str(uuid.uuid4()),
                "tenant_id": tenant.id,
                "user_id": demandeur_id,
                "type": "remplacement_annulee",
                "titre": "❌ Demande annulée",
                "message": f"Votre demande de remplacement du {demande_data['date']} a été annulée.",
                "lu": False,
                "data": {"demande_id": demande_id},
                "created_at": maintenant.isoformat()
            })
        
        return {
            "message": "Demande annulée avec succès",
            "demande_id": demande_id
        }
    else:
        await refuser_remplacement(demande_id, current_user.id, tenant.id)
        
        return {
            "message": "Remplacement refusé",
            "demande_id": demande_id
        }


@router.get("/remplacement-action/{token}/{action}")
async def action_remplacement_via_email(token: str, action: str):
    """Traite une action de remplacement via le lien email"""
    frontend_url = os.environ.get('FRONTEND_URL', 'https://www.profiremanager.ca')
    
    try:
        token_data = await db.tokens_remplacement.find_one({"token": token})
        
        if not token_data:
            return RedirectResponse(
                url=f"{frontend_url}/remplacement-resultat?status=erreur&message=Lien invalide ou expiré",
                status_code=302
            )
        
        expiration = datetime.fromisoformat(token_data["expiration"].replace('Z', '+00:00'))
        if datetime.now(timezone.utc) > expiration:
            return RedirectResponse(
                url=f"{frontend_url}/remplacement-resultat?status=erreur&message=Ce lien a expiré",
                status_code=302
            )
        
        if token_data.get("utilise"):
            return RedirectResponse(
                url=f"{frontend_url}/remplacement-resultat?status=info&message=Cette action a déjà été traitée",
                status_code=302
            )
        
        demande_id = token_data["demande_id"]
        remplacant_id = token_data["remplacant_id"]
        tenant_id = token_data["tenant_id"]
        
        demande_data = await db.demandes_remplacement.find_one({"id": demande_id, "tenant_id": tenant_id})
        if not demande_data:
            return RedirectResponse(
                url=f"{frontend_url}/remplacement-resultat?status=erreur&message=Demande non trouvée",
                status_code=302
            )
        
        if demande_data["statut"] not in ["en_cours", "en_attente"]:
            status_label = {
                "accepte": "déjà acceptée",
                "expiree": "expirée",
                "annulee": "annulée",
                "approuve_manuellement": "déjà approuvée"
            }.get(demande_data["statut"], demande_data["statut"])
            return RedirectResponse(
                url=f"{frontend_url}/remplacement-resultat?status=info&message=Cette demande est {status_label}",
                status_code=302
            )
        
        # Si action = "choix", afficher une page avec les deux boutons
        if action == "choix":
            # Rediriger vers une page frontend qui affiche les deux options
            return RedirectResponse(
                url=f"{frontend_url}/remplacement-choix?token={token}",
                status_code=302
            )
        
        # Valider l'action avant de marquer le token comme utilisé
        if action not in ["accepter", "refuser"]:
            return RedirectResponse(
                url=f"{frontend_url}/remplacement-resultat?status=erreur&message=Action non reconnue",
                status_code=302
            )
        
        await db.tokens_remplacement.update_one(
            {"token": token},
            {"$set": {"utilise": True, "action": action, "date_utilisation": datetime.now(timezone.utc).isoformat()}}
        )
        
        if action == "accepter":
            try:
                await accepter_remplacement(demande_id, remplacant_id, tenant_id)
                
                demandeur = await db.users.find_one({"id": demande_data["demandeur_id"]})
                demandeur_nom = f"{demandeur.get('prenom', '')} {demandeur.get('nom', '')}" if demandeur else "le demandeur"
                
                return RedirectResponse(
                    url=f"{frontend_url}/remplacement-resultat?status=succes&message=Vous avez accepté le remplacement de {demandeur_nom} le {demande_data['date']}",
                    status_code=302
                )
            except Exception as e:
                logger.error(f"Erreur acceptation via email: {e}")
                return RedirectResponse(
                    url=f"{frontend_url}/remplacement-resultat?status=erreur&message=Erreur lors de l'acceptation",
                    status_code=302
                )
        
        elif action == "refuser":
            try:
                await refuser_remplacement(demande_id, remplacant_id, tenant_id)
                
                return RedirectResponse(
                    url=f"{frontend_url}/remplacement-resultat?status=info&message=Vous avez refusé cette demande de remplacement",
                    status_code=302
                )
            except Exception as e:
                logger.error(f"Erreur refus via email: {e}")
                return RedirectResponse(
                    url=f"{frontend_url}/remplacement-resultat?status=erreur&message=Erreur lors du refus",
                    status_code=302
                )
            
    except Exception as e:
        logger.error(f"Erreur traitement action email: {e}", exc_info=True)
        return RedirectResponse(
            url=f"{frontend_url}/remplacement-resultat?status=erreur&message=Une erreur est survenue",
            status_code=302
        )


@router.delete("/{tenant_slug}/remplacements/{demande_id}")
async def annuler_demande_remplacement(
    tenant_slug: str,
    demande_id: str,
    current_user: User = Depends(get_current_user)
):
    """Annuler une demande de remplacement (seulement par le demandeur)"""
    send_push_notification_to_users = await get_send_push_notification()
    tenant = await get_tenant_from_slug(tenant_slug)
    
    demande = await db.demandes_remplacement.find_one({"id": demande_id, "tenant_id": tenant.id})
    if not demande:
        raise HTTPException(status_code=404, detail="Demande non trouvée")
    
    if demande["demandeur_id"] != current_user.id:
        raise HTTPException(status_code=403, detail="Seul le demandeur peut annuler la demande")
    
    if demande["statut"] == "accepte":
        raise HTTPException(status_code=400, detail="Impossible d'annuler une demande déjà acceptée")
    
    await db.demandes_remplacement.update_one(
        {"id": demande_id},
        {
            "$set": {
                "statut": "annulee",
                "updated_at": datetime.now(timezone.utc)
            }
        }
    )
    
    if demande.get("remplacants_contactes_ids"):
        await send_push_notification_to_users(
            user_ids=demande["remplacants_contactes_ids"],
            title="Demande annulée",
            body=f"La demande de remplacement du {demande['date']} a été annulée",
            data={
                "type": "remplacement_annulee",
                "demande_id": demande_id
            }
        )
    
    logger.info(f"✅ Demande de remplacement annulée: {demande_id}")
    
    return {
        "message": "Demande annulée avec succès",
        "demande_id": demande_id
    }


# ==================== PARAMÈTRES REMPLACEMENTS ====================

@router.get("/{tenant_slug}/parametres/remplacements")
async def get_parametres_remplacements(tenant_slug: str, current_user: User = Depends(get_current_user)):
    """Récupère les paramètres de remplacements"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    parametres = await db.parametres_remplacements.find_one({"tenant_id": tenant.id})
    
    if not parametres:
        logger.info(f"📋 Création des paramètres de remplacement par défaut pour tenant {tenant.slug}")
        default_params = ParametresRemplacements(tenant_id=tenant.id)
        await db.parametres_remplacements.insert_one(default_params.dict())
        return default_params
    
    cleaned_params = clean_mongo_doc(parametres)
    logger.info(f"📋 Paramètres remplacements chargés pour {tenant.slug}: delai={cleaned_params.get('delai_attente_minutes')}min, max_contacts={cleaned_params.get('max_contacts')}")
    return cleaned_params


@router.put("/{tenant_slug}/parametres/remplacements")
async def update_parametres_remplacements(
    tenant_slug: str,
    parametres_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Met à jour les paramètres de remplacements"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    logger.info(f"💾 Sauvegarde paramètres remplacements pour {tenant.slug}: {parametres_data}")
    
    existing = await db.parametres_remplacements.find_one({"tenant_id": tenant.id})
    
    parametres_data["tenant_id"] = tenant.id
    
    if existing:
        await db.parametres_remplacements.update_one(
            {"tenant_id": tenant.id},
            {"$set": parametres_data}
        )
        logger.info(f"✅ Paramètres remplacements MIS À JOUR pour {tenant.slug}")
    else:
        if "id" not in parametres_data:
            parametres_data["id"] = str(uuid.uuid4())
        await db.parametres_remplacements.insert_one(parametres_data)
        logger.info(f"✅ Paramètres remplacements CRÉÉS pour {tenant.slug}")
    
    return {"message": "Paramètres mis à jour avec succès"}



@router.get("/{tenant_slug}/remplacements/debug/{demande_id}")
async def debug_recherche_remplacant(
    tenant_slug: str,
    demande_id: str,
    current_user: User = Depends(get_current_user)
):
    """
    Endpoint de diagnostic pour comprendre pourquoi aucun remplaçant n'est trouvé.
    Retourne des détails sur chaque utilisateur et pourquoi il est éligible ou non.
    """
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès réservé aux admins")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Récupérer la demande
    demande = await db.demandes_remplacement.find_one({"id": demande_id, "tenant_id": tenant.id})
    if not demande:
        raise HTTPException(status_code=404, detail="Demande non trouvée")
    
    date_garde = demande.get("date")
    type_garde_id = demande.get("type_garde_id")
    demandeur_id = demande.get("demandeur_id")
    exclus_ids = demande.get("remplacants_contactes_ids", [])
    
    # Type de garde
    type_garde = await db.types_garde.find_one({"id": type_garde_id, "tenant_id": tenant.id})
    competences_requises = type_garde.get("competences_requises", []) if type_garde else []
    officier_obligatoire = type_garde.get("officier_obligatoire", False) if type_garde else False
    heure_debut_garde = type_garde.get("heure_debut", "00:00") if type_garde else "00:00"
    heure_fin_garde = type_garde.get("heure_fin", "23:59") if type_garde else "23:59"
    
    # Fonction pour vérifier si deux plages horaires se chevauchent
    def plages_se_chevauchent(debut1: str, fin1: str, debut2: str, fin2: str) -> bool:
        def heure_to_minutes(h: str) -> int:
            try:
                parts = h.split(":")
                return int(parts[0]) * 60 + int(parts[1])
            except:
                return 0
        
        d1, f1 = heure_to_minutes(debut1), heure_to_minutes(fin1)
        d2, f2 = heure_to_minutes(debut2), heure_to_minutes(fin2)
        
        if f1 < d1:
            f1 += 24 * 60
        if f2 < d2:
            f2 += 24 * 60
        
        return d1 < f2 and d2 < f1
    
    # Récupérer les grades pour déterminer qui est officier
    grades_list = await db.grades.find({"tenant_id": tenant.id}).to_list(100)
    grades_map = {g.get("nom"): g for g in grades_list}
    
    def est_officier_grade(grade_nom: str) -> bool:
        """Vérifie si un grade est considéré comme officier"""
        grade_info = grades_map.get(grade_nom, {})
        return grade_info.get("est_officier", False) == True
    
    # Demandeur
    demandeur = await db.users.find_one({"id": demandeur_id})
    demandeur_grade = demandeur.get("grade", "") if demandeur else ""
    demandeur_est_officier = est_officier_grade(demandeur_grade)
    
    # Vérifier la règle officier
    besoin_officier_remplacement = False
    autre_officier_present = False
    
    if officier_obligatoire and demandeur_est_officier:
        # Vérifier s'il y a d'autres officiers assignés à cette garde à cette date
        autres_assignations = await db.assignations.find({
            "tenant_id": tenant.id,
            "type_garde_id": type_garde_id,
            "date": date_garde,
            "user_id": {"$ne": demandeur_id}
        }).to_list(100)
        
        for assignation in autres_assignations:
            autre_user = await db.users.find_one({"id": assignation["user_id"], "tenant_id": tenant.id})
            if autre_user:
                autre_grade = autre_user.get("grade", "")
                if est_officier_grade(autre_grade):
                    autre_officier_present = True
                    break
        
        besoin_officier_remplacement = not autre_officier_present
    
    # Tous les utilisateurs actifs
    users = await db.users.find({"tenant_id": tenant.id, "statut": "Actif"}).to_list(1000)
    
    resultats = []
    
    for user in users:
        user_id = user["id"]
        user_name = f"{user.get('prenom', '')} {user.get('nom', '')}"
        user_competences = user.get("competences", [])
        user_type_emploi = user.get("type_emploi", "")
        
        resultat = {
            "nom": user_name,
            "id": user_id,
            "type_emploi": user_type_emploi,
            "competences": user_competences,
            "eligible": True,
            "raisons_exclusion": []
        }
        
        # Check 1: Est-ce le demandeur ?
        if user_id == demandeur_id:
            resultat["eligible"] = False
            resultat["raisons_exclusion"].append("C'est le demandeur lui-même")
            resultats.append(resultat)
            continue
        
        # Check 2: Déjà contacté ?
        if user_id in exclus_ids:
            resultat["eligible"] = False
            resultat["raisons_exclusion"].append("Déjà contacté précédemment")
            resultats.append(resultat)
            continue
        
        # Check 3: Compétences requises
        if competences_requises:
            # Normaliser pour comparaison insensible à la casse
            competences_requises_norm = set(c.lower().strip() for c in competences_requises if c)
            user_competences_norm = set(c.lower().strip() for c in user_competences if c)
            manquantes = competences_requises_norm - user_competences_norm
            if manquantes:
                resultat["eligible"] = False
                resultat["raisons_exclusion"].append(f"Compétences manquantes: {list(manquantes)}")
            resultat["competences_requises_norm"] = list(competences_requises_norm)
            resultat["competences_user_norm"] = list(user_competences_norm)
        
        # Check 4: Indisponibilité déclarée
        indispo = await db.disponibilites.find_one({
            "user_id": user_id,
            "tenant_id": tenant.id,
            "date": date_garde,
            "statut": "indisponible"
        })
        if indispo:
            resultat["eligible"] = False
            resultat["raisons_exclusion"].append("Indisponibilité déclarée pour cette date")
        
        # Check 4b: Règle officier
        if besoin_officier_remplacement:
            user_grade = user.get("grade", "")
            user_est_officier = est_officier_grade(user_grade)
            user_est_eligible = user.get("fonction_superieur", False) == True
            resultat["est_officier"] = user_est_officier
            resultat["est_eligible_fonction_sup"] = user_est_eligible
            
            if not user_est_officier and not user_est_eligible:
                resultat["eligible"] = False
                resultat["raisons_exclusion"].append(f"Règle officier: Le demandeur est le seul officier, le remplaçant doit être officier ou éligible (grade={user_grade})")
        
        # Check 5: Conflit horaire (assignation existante qui chevauche)
        assignations_ce_jour = await db.assignations.find({
            "user_id": user_id,
            "tenant_id": tenant.id,
            "date": date_garde
        }).to_list(10)
        
        for assignation in assignations_ce_jour:
            type_garde_assignation = await db.types_garde.find_one({"id": assignation.get("type_garde_id")})
            if type_garde_assignation:
                heure_debut_existante = type_garde_assignation.get("heure_debut", "00:00")
                heure_fin_existante = type_garde_assignation.get("heure_fin", "23:59")
                nom_garde_existante = type_garde_assignation.get("nom", "Inconnu")
                
                if plages_se_chevauchent(heure_debut_garde, heure_fin_garde, heure_debut_existante, heure_fin_existante):
                    resultat["eligible"] = False
                    resultat["raisons_exclusion"].append(f"Conflit horaire: Déjà assigné à {nom_garde_existante} ({heure_debut_existante}-{heure_fin_existante}) qui chevauche {heure_debut_garde}-{heure_fin_garde}")
                else:
                    # Info: assigné mais pas de chevauchement
                    if "assignations_sans_conflit" not in resultat:
                        resultat["assignations_sans_conflit"] = []
                    resultat["assignations_sans_conflit"].append(f"{nom_garde_existante} ({heure_debut_existante}-{heure_fin_existante})")
        
        # Check 6: Disponibilité déclarée
        dispo = await db.disponibilites.find_one({
            "user_id": user_id,
            "tenant_id": tenant.id,
            "date": date_garde,
            "statut": "disponible"
        })
        resultat["a_disponibilite"] = dispo is not None
        
        # Classification niveau
        if resultat["eligible"]:
            if user_type_emploi in ["temps_partiel", "temporaire"]:
                if dispo:
                    resultat["niveau"] = "N2 (TP Disponible)"
                else:
                    resultat["niveau"] = "N3 (TP Stand-by)"
            elif user_type_emploi == "temps_plein":
                resultat["niveau"] = "N4 (Temps plein)"
            else:
                resultat["niveau"] = "N5 ou non classé"
        
        resultats.append(resultat)
    
    # Compter les éligibles
    eligibles = [r for r in resultats if r["eligible"]]
    non_eligibles = [r for r in resultats if not r["eligible"]]
    
    return {
        "demande": {
            "id": demande_id,
            "date": date_garde,
            "type_garde": type_garde.get("nom") if type_garde else "Inconnu",
            "horaires": f"{heure_debut_garde} - {heure_fin_garde}",
            "competences_requises": competences_requises,
            "demandeur": f"{demandeur.get('prenom', '')} {demandeur.get('nom', '')}" if demandeur else "Inconnu",
            "demandeur_grade": demandeur_grade,
            "demandeur_est_officier": demandeur_est_officier
        },
        "regle_officier": {
            "officier_obligatoire": officier_obligatoire,
            "demandeur_est_officier": demandeur_est_officier,
            "autre_officier_present_sur_garde": autre_officier_present,
            "regle_active": besoin_officier_remplacement,
            "explication": (
                "Le remplaçant DOIT être officier ou éligible car le demandeur est le seul officier sur cette garde"
                if besoin_officier_remplacement else
                "Pas de contrainte officier (soit type de garde sans officier obligatoire, soit le demandeur n'est pas officier, soit un autre officier est déjà présent)"
            )
        },
        "resume": {
            "total_utilisateurs": len(resultats),
            "eligibles": len(eligibles),
            "non_eligibles": len(non_eligibles)
        },
        "eligibles": eligibles,
        "non_eligibles": non_eligibles
    }
