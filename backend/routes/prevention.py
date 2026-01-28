"""
Routes API pour le module Prévention
====================================

STATUT: ACTIF
Ce module gère la prévention incendie : bâtiments, inspections, grilles, secteurs.

Routes Bâtiments:
- GET    /{tenant_slug}/prevention/batiments                    - Liste bâtiments
- GET    /{tenant_slug}/prevention/batiments/search             - Recherche bâtiments
- GET    /{tenant_slug}/prevention/batiments/{id}               - Détail bâtiment
- POST   /{tenant_slug}/prevention/batiments                    - Créer bâtiment
- PUT    /{tenant_slug}/prevention/batiments/{id}               - Modifier bâtiment
- DELETE /{tenant_slug}/prevention/batiments/{id}               - Supprimer bâtiment
- POST   /{tenant_slug}/prevention/batiments/import-csv         - Importer CSV

Routes Inspections:
- GET    /{tenant_slug}/prevention/inspections                  - Liste inspections
- GET    /{tenant_slug}/prevention/inspections/{id}             - Détail inspection
- POST   /{tenant_slug}/prevention/inspections                  - Créer inspection
- PUT    /{tenant_slug}/prevention/inspections/{id}             - Modifier inspection

Routes Grilles:
- GET    /{tenant_slug}/prevention/grilles-inspection           - Liste grilles
- GET    /{tenant_slug}/prevention/grilles-inspection/{id}      - Détail grille
- POST   /{tenant_slug}/prevention/grilles-inspection           - Créer grille
- PUT    /{tenant_slug}/prevention/grilles-inspection/{id}      - Modifier grille
- DELETE /{tenant_slug}/prevention/grilles-inspection/{id}      - Supprimer grille
- POST   /{tenant_slug}/prevention/grilles-inspection/{id}/dupliquer - Dupliquer

Routes Secteurs:
- GET    /{tenant_slug}/prevention/secteurs                     - Liste secteurs
- GET    /{tenant_slug}/prevention/secteurs/{id}                - Détail secteur
- POST   /{tenant_slug}/prevention/secteurs                     - Créer secteur
- PUT    /{tenant_slug}/prevention/secteurs/{id}                - Modifier secteur
- DELETE /{tenant_slug}/prevention/secteurs/{id}                - Supprimer secteur

Routes Références:
- GET    /{tenant_slug}/prevention/references                   - Catégories et risques
- GET    /{tenant_slug}/prevention/meta/niveaux-risque          - Niveaux de risque
- GET    /{tenant_slug}/prevention/meta/categories-batiments    - Catégories bâtiments
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Body
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone
import uuid
import logging
import csv
import base64
import json
from io import StringIO, BytesIO
from datetime import timedelta
from starlette.responses import StreamingResponse, Response

from routes.dependencies import (
    db,
    get_current_user,
    get_tenant_from_slug,
    clean_mongo_doc,
    User
)

# Import des helpers PDF partagés
from utils.pdf_helpers import (
    create_branded_pdf,
    create_pdf_header_elements,
    create_pdf_footer_text,
    get_modern_pdf_styles,
    BrandedDocTemplate
)

router = APIRouter(tags=["Prévention"])
logger = logging.getLogger(__name__)


# ==================== MODÈLES ====================

class Batiment(BaseModel):
    """Fiche d'établissement/bâtiment pour les inspections de prévention"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    nom_etablissement: str = ""
    adresse_civique: str = ""
    ville: str = ""
    province: str = "QC"
    code_postal: str = ""
    cadastre_matricule: str = ""
    valeur_fonciere: Optional[str] = ""
    type_batiment: str = ""
    sous_type_batiment: str = ""
    annee_construction: str = ""
    nombre_etages: str = ""
    superficie_totale_m2: str = ""
    proprietaire_nom: str = ""
    proprietaire_prenom: str = ""
    proprietaire_telephone: str = ""
    proprietaire_courriel: str = ""
    gerant_nom: str = ""
    gerant_telephone: str = ""
    gerant_courriel: str = ""
    gestionnaire_nom: str = ""
    gestionnaire_prenom: str = ""
    gestionnaire_telephone: str = ""
    gestionnaire_courriel: str = ""
    locataire_nom: str = ""
    locataire_prenom: str = ""
    locataire_telephone: str = ""
    locataire_courriel: str = ""
    responsable_securite_nom: str = ""
    responsable_securite_telephone: str = ""
    responsable_securite_courriel: str = ""
    groupe_occupation: str = ""
    sous_groupe: str = ""
    description_activite: str = ""
    niveau_risque: str = ""
    risques: List[str] = []
    risques_identifies: List[str] = []
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    photo_url: Optional[str] = ""
    statut: str = "actif"
    notes_generales: str = ""
    notes: str = ""
    preventionniste_assigne_id: Optional[str] = None
    historique_assignations: List[Dict[str, Any]] = []
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class BatimentCreate(BaseModel):
    nom_etablissement: str = ""
    adresse_civique: str = ""
    ville: str = ""
    province: str = "QC"
    code_postal: str = ""
    cadastre_matricule: str = ""
    valeur_fonciere: Optional[str] = ""
    type_batiment: str = ""
    sous_type_batiment: str = ""
    annee_construction: str = ""
    nombre_etages: str = ""
    superficie_totale_m2: str = ""
    proprietaire_nom: str = ""
    proprietaire_prenom: str = ""
    proprietaire_telephone: str = ""
    proprietaire_courriel: str = ""
    gerant_nom: str = ""
    gerant_telephone: str = ""
    gerant_courriel: str = ""
    gestionnaire_nom: str = ""
    gestionnaire_prenom: str = ""
    gestionnaire_telephone: str = ""
    gestionnaire_courriel: str = ""
    locataire_nom: str = ""
    locataire_prenom: str = ""
    locataire_telephone: str = ""
    locataire_courriel: str = ""
    responsable_securite_nom: str = ""
    responsable_securite_telephone: str = ""
    responsable_securite_courriel: str = ""
    groupe_occupation: str = ""
    sous_groupe: str = ""
    description_activite: str = ""
    niveau_risque: str = ""
    risques: List[str] = []
    risques_identifies: List[str] = []
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    photo_url: Optional[str] = ""
    statut: str = "actif"
    notes_generales: str = ""
    notes: str = ""
    preventionniste_assigne_id: Optional[str] = None


class SecteurGeographique(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    nom: str
    description: str = ""
    couleur: str = "#3B82F6"
    coordonnees: List[Dict[str, float]] = []
    preventionniste_assigne_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class SecteurGeographiqueCreate(BaseModel):
    nom: str
    description: str = ""
    couleur: str = "#3B82F6"
    coordonnees: List[Dict[str, float]] = []
    preventionniste_assigne_id: Optional[str] = None


class GrilleInspection(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    nom: str
    description: str = ""
    sections: List[Dict[str, Any]] = []
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class GrilleInspectionCreate(BaseModel):
    nom: str
    description: str = ""
    sections: List[Dict[str, Any]] = []


class Inspection(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    batiment_id: str
    grille_id: str = ""
    preventionniste_id: str
    preventionniste_nom: str = ""
    date_inspection: str
    type_inspection: str = "routine"
    statut: str = "planifiee"
    resultats: List[Dict[str, Any]] = []
    photos: List[str] = []
    anomalies: List[Dict[str, Any]] = []
    recommandations: str = ""
    notes: str = ""
    signature_preventionniste: str = ""
    signature_responsable: str = ""
    date_signature_responsable: str = ""
    prochaine_inspection: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class InspectionCreate(BaseModel):
    batiment_id: str
    grille_id: str = ""
    preventionniste_id: str
    preventionniste_nom: str = ""
    date_inspection: str
    type_inspection: str = "routine"
    statut: str = "planifiee"
    resultats: List[Dict[str, Any]] = []
    photos: List[str] = []
    anomalies: List[Dict[str, Any]] = []
    recommandations: str = ""
    notes: str = ""
    signature_preventionniste: str = ""
    signature_responsable: str = ""
    date_signature_responsable: str = ""
    prochaine_inspection: str = ""


# ==================== DONNÉES DE RÉFÉRENCE ====================

CATEGORIES_NR24_27 = {
    "A": {
        "nom": "Établissements de réunion",
        "sous_groupes": {
            "A-1": "Destinés à la production et à la présentation d'arts du spectacle",
            "A-2": "Qui ne figurent dans aucune autre division du groupe A",
            "A-3": "De type aréna",
            "A-4": "Où les occupants sont rassemblés en plein air"
        }
    },
    "B": {
        "nom": "Établissements de détention, soins, traitement et habitations supervisées",
        "sous_groupes": {
            "B-1": "Établissements de détention",
            "B-2": "Établissements de traitement",
            "B-3": "Établissements de soins",
            "B-4": "Établissements de soins de type résidentiel"
        }
    },
    "C": {"nom": "Habitations", "sous_groupes": {}},
    "D": {"nom": "Établissements d'affaires", "sous_groupes": {}},
    "E": {"nom": "Établissements commerciaux", "sous_groupes": {}},
    "F": {
        "nom": "Établissements industriels",
        "sous_groupes": {
            "F-1": "À risques très élevés",
            "F-2": "À risques moyens",
            "F-3": "À risques faibles"
        }
    },
    "G": {
        "nom": "Établissements agricoles",
        "sous_groupes": {
            "G-1": "À risques très élevés",
            "G-2": "Qui ne figurent dans aucune autre division du groupe G",
            "G-3": "Abritant des serres",
            "G-4": "Sans occupation humaine"
        }
    }
}

RISQUES_GUIDE_PLANIFICATION = {
    "incendies": ["Incendie", "Conflagration"],
    "matieres": [
        "Matières dangereuses", "Matières hautement inflammables",
        "Matières très toxiques", "Matières combustibles", "Produits chimiques"
    ],
    "infrastructure": [
        "Explosion", "Défaillances des systèmes de sécurité incendie",
        "Matériaux de construction douteux"
    ],
    "naturels": ["Inondation", "Tremblement de terre", "Tempête de verglas"],
    "humains": ["Négligence", "Acte criminel", "Occupation illégale"]
}

NIVEAUX_RISQUE = [
    {"id": "faible", "nom": "Faible", "couleur": "#22C55E", "description": "Risque minimal"},
    {"id": "moyen", "nom": "Moyen", "couleur": "#F59E0B", "description": "Risque modéré"},
    {"id": "eleve", "nom": "Élevé", "couleur": "#EF4444", "description": "Risque important"},
    {"id": "tres_eleve", "nom": "Très élevé", "couleur": "#7C3AED", "description": "Risque critique"}
]


# ==================== ROUTES RÉFÉRENCES ====================

@router.get("/{tenant_slug}/prevention/references")
async def get_prevention_references(tenant_slug: str):
    """Récupère les catégories officielles (NR24-27) et risques pour le module prévention"""
    return {
        "categories_nr24_27": CATEGORIES_NR24_27,
        "risques_guide_planification": RISQUES_GUIDE_PLANIFICATION,
        "niveaux_risque": NIVEAUX_RISQUE
    }


@router.get("/{tenant_slug}/prevention/meta/niveaux-risque")
async def get_niveaux_risque(tenant_slug: str):
    """Récupère les niveaux de risque disponibles"""
    return NIVEAUX_RISQUE


@router.get("/{tenant_slug}/prevention/meta/categories-batiments")
async def get_categories_batiments(tenant_slug: str):
    """Récupère les catégories de bâtiments selon le Code national"""
    return CATEGORIES_NR24_27


# ==================== ROUTES BÂTIMENTS ====================

@router.get("/{tenant_slug}/prevention/batiments")
async def get_batiments(
    tenant_slug: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer tous les bâtiments"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    batiments = await db.batiments.find(
        {"tenant_id": tenant.id},
        {"_id": 0}
    ).sort("nom_etablissement", 1).to_list(10000)
    
    return batiments


@router.get("/{tenant_slug}/prevention/batiments/search")
async def search_batiments(
    tenant_slug: str,
    q: str = "",
    niveau_risque: str = "",
    groupe_occupation: str = "",
    current_user: User = Depends(get_current_user)
):
    """Recherche de bâtiments avec filtres"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    query = {"tenant_id": tenant.id}
    
    if q:
        query["$or"] = [
            {"nom_etablissement": {"$regex": q, "$options": "i"}},
            {"adresse_civique": {"$regex": q, "$options": "i"}},
            {"ville": {"$regex": q, "$options": "i"}}
        ]
    
    if niveau_risque:
        query["niveau_risque"] = niveau_risque
    
    if groupe_occupation:
        query["groupe_occupation"] = groupe_occupation
    
    batiments = await db.batiments.find(query, {"_id": 0}).sort("nom_etablissement", 1).to_list(1000)
    
    return batiments


@router.get("/{tenant_slug}/prevention/batiments/{batiment_id}")
async def get_batiment(
    tenant_slug: str,
    batiment_id: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer un bâtiment spécifique"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    batiment = await db.batiments.find_one(
        {"id": batiment_id, "tenant_id": tenant.id},
        {"_id": 0}
    )
    
    if not batiment:
        raise HTTPException(status_code=404, detail="Bâtiment non trouvé")
    
    return batiment


@router.post("/{tenant_slug}/prevention/batiments")
async def create_batiment(
    tenant_slug: str,
    batiment: BatimentCreate,
    current_user: User = Depends(get_current_user)
):
    """Créer un nouveau bâtiment"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    batiment_obj = Batiment(tenant_id=tenant.id, **batiment.dict())
    
    await db.batiments.insert_one(batiment_obj.dict())
    
    return clean_mongo_doc(batiment_obj.dict())


@router.put("/{tenant_slug}/prevention/batiments/{batiment_id}")
async def update_batiment(
    tenant_slug: str,
    batiment_id: str,
    batiment_data: BatimentCreate,
    current_user: User = Depends(get_current_user)
):
    """Mettre à jour un bâtiment"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    existing = await db.batiments.find_one({"id": batiment_id, "tenant_id": tenant.id})
    if not existing:
        raise HTTPException(status_code=404, detail="Bâtiment non trouvé")
    
    update_data = batiment_data.dict()
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.batiments.update_one(
        {"id": batiment_id, "tenant_id": tenant.id},
        {"$set": update_data}
    )
    
    updated = await db.batiments.find_one({"id": batiment_id, "tenant_id": tenant.id}, {"_id": 0})
    return updated


@router.delete("/{tenant_slug}/prevention/batiments/{batiment_id}")
async def delete_batiment(
    tenant_slug: str,
    batiment_id: str,
    current_user: User = Depends(get_current_user)
):
    """Supprimer un bâtiment"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    existing = await db.batiments.find_one({"id": batiment_id, "tenant_id": tenant.id})
    if not existing:
        raise HTTPException(status_code=404, detail="Bâtiment non trouvé")
    
    await db.batiments.delete_one({"id": batiment_id, "tenant_id": tenant.id})
    await db.inspections.delete_many({"batiment_id": batiment_id})
    
    return {"message": "Bâtiment supprimé avec succès"}


@router.post("/{tenant_slug}/prevention/batiments/import-csv")
async def import_batiments_csv(
    tenant_slug: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Importer des bâtiments depuis un fichier CSV"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    contents = await file.read()
    try:
        csv_text = contents.decode('utf-8')
    except UnicodeDecodeError:
        csv_text = contents.decode('latin-1')
    
    csv_reader = csv.DictReader(StringIO(csv_text))
    
    imported = 0
    errors = 0
    error_details = []
    
    for row in csv_reader:
        try:
            batiment = Batiment(
                tenant_id=tenant.id,
                nom_etablissement=row.get('nom_etablissement', row.get('nom', '')),
                adresse_civique=row.get('adresse_civique', row.get('adresse', '')),
                ville=row.get('ville', ''),
                code_postal=row.get('code_postal', ''),
                type_batiment=row.get('type_batiment', ''),
                niveau_risque=row.get('niveau_risque', ''),
                groupe_occupation=row.get('groupe_occupation', ''),
                notes=row.get('notes', '')
            )
            await db.batiments.insert_one(batiment.dict())
            imported += 1
        except Exception as e:
            errors += 1
            error_details.append(str(e))
    
    return {
        "imported": imported,
        "errors": errors,
        "error_details": error_details[:10],
        "message": f"{imported} bâtiment(s) importé(s), {errors} erreur(s)"
    }


# ==================== ROUTES INSPECTIONS ====================

@router.get("/{tenant_slug}/prevention/inspections")
async def get_inspections(
    tenant_slug: str,
    batiment_id: Optional[str] = None,
    preventionniste_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Récupérer toutes les inspections avec filtres optionnels"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    query = {"tenant_id": tenant.id}
    if batiment_id:
        query["batiment_id"] = batiment_id
    if preventionniste_id:
        query["preventionniste_id"] = preventionniste_id
    
    inspections = await db.inspections.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return inspections


@router.get("/{tenant_slug}/prevention/inspections/{inspection_id}")
async def get_inspection(
    tenant_slug: str,
    inspection_id: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer une inspection spécifique"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    inspection = await db.inspections.find_one(
        {"id": inspection_id, "tenant_id": tenant.id},
        {"_id": 0}
    )
    
    if not inspection:
        raise HTTPException(status_code=404, detail="Inspection non trouvée")
    
    return inspection


@router.post("/{tenant_slug}/prevention/inspections")
async def create_inspection(
    tenant_slug: str,
    inspection: InspectionCreate,
    current_user: User = Depends(get_current_user)
):
    """Créer une nouvelle inspection"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    inspection_obj = Inspection(tenant_id=tenant.id, **inspection.dict())
    
    await db.inspections.insert_one(inspection_obj.dict())
    
    return clean_mongo_doc(inspection_obj.dict())


@router.put("/{tenant_slug}/prevention/inspections/{inspection_id}")
async def update_inspection(
    tenant_slug: str,
    inspection_id: str,
    inspection_data: InspectionCreate,
    current_user: User = Depends(get_current_user)
):
    """Mettre à jour une inspection"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    existing = await db.inspections.find_one({"id": inspection_id, "tenant_id": tenant.id})
    if not existing:
        raise HTTPException(status_code=404, detail="Inspection non trouvée")
    
    update_data = inspection_data.dict()
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.inspections.update_one(
        {"id": inspection_id, "tenant_id": tenant.id},
        {"$set": update_data}
    )
    
    updated = await db.inspections.find_one({"id": inspection_id, "tenant_id": tenant.id}, {"_id": 0})
    return updated


# ==================== ROUTES GRILLES ====================

@router.get("/{tenant_slug}/prevention/grilles-inspection")
async def get_grilles_inspection(
    tenant_slug: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer toutes les grilles d'inspection"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    grilles = await db.grilles_inspection.find(
        {"tenant_id": tenant.id},
        {"_id": 0}
    ).sort("nom", 1).to_list(1000)
    
    return grilles


@router.get("/{tenant_slug}/prevention/grilles-inspection/{grille_id}")
async def get_grille_inspection(
    tenant_slug: str,
    grille_id: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer une grille d'inspection spécifique"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    grille = await db.grilles_inspection.find_one(
        {"id": grille_id, "tenant_id": tenant.id},
        {"_id": 0}
    )
    
    if not grille:
        raise HTTPException(status_code=404, detail="Grille non trouvée")
    
    return grille


@router.post("/{tenant_slug}/prevention/grilles-inspection")
async def create_grille_inspection(
    tenant_slug: str,
    grille: GrilleInspectionCreate,
    current_user: User = Depends(get_current_user)
):
    """Créer une nouvelle grille d'inspection"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    grille_obj = GrilleInspection(tenant_id=tenant.id, **grille.dict())
    
    await db.grilles_inspection.insert_one(grille_obj.dict())
    
    return clean_mongo_doc(grille_obj.dict())


@router.put("/{tenant_slug}/prevention/grilles-inspection/{grille_id}")
async def update_grille_inspection(
    tenant_slug: str,
    grille_id: str,
    grille_data: GrilleInspectionCreate,
    current_user: User = Depends(get_current_user)
):
    """Mettre à jour une grille d'inspection"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    existing = await db.grilles_inspection.find_one({"id": grille_id, "tenant_id": tenant.id})
    if not existing:
        raise HTTPException(status_code=404, detail="Grille non trouvée")
    
    update_data = grille_data.dict()
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.grilles_inspection.update_one(
        {"id": grille_id, "tenant_id": tenant.id},
        {"$set": update_data}
    )
    
    updated = await db.grilles_inspection.find_one({"id": grille_id, "tenant_id": tenant.id}, {"_id": 0})
    return updated


@router.delete("/{tenant_slug}/prevention/grilles-inspection/{grille_id}")
async def delete_grille_inspection(
    tenant_slug: str,
    grille_id: str,
    current_user: User = Depends(get_current_user)
):
    """Supprimer une grille d'inspection"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    existing = await db.grilles_inspection.find_one({"id": grille_id, "tenant_id": tenant.id})
    if not existing:
        raise HTTPException(status_code=404, detail="Grille non trouvée")
    
    await db.grilles_inspection.delete_one({"id": grille_id, "tenant_id": tenant.id})
    
    return {"message": "Grille supprimée avec succès"}


@router.post("/{tenant_slug}/prevention/grilles-inspection/{grille_id}/dupliquer")
async def dupliquer_grille_inspection(
    tenant_slug: str,
    grille_id: str,
    current_user: User = Depends(get_current_user)
):
    """Dupliquer une grille d'inspection"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    grille = await db.grilles_inspection.find_one({"id": grille_id, "tenant_id": tenant.id})
    if not grille:
        raise HTTPException(status_code=404, detail="Grille non trouvée")
    
    new_grille = GrilleInspection(
        tenant_id=tenant.id,
        nom=f"{grille['nom']} (copie)",
        description=grille.get('description', ''),
        sections=grille.get('sections', [])
    )
    
    await db.grilles_inspection.insert_one(new_grille.dict())
    
    return clean_mongo_doc(new_grille.dict())


# ==================== ROUTES SECTEURS ====================

@router.get("/{tenant_slug}/prevention/secteurs")
async def get_secteurs(
    tenant_slug: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer tous les secteurs géographiques"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    secteurs = await db.secteurs_geographiques.find(
        {"tenant_id": tenant.id},
        {"_id": 0}
    ).sort("nom", 1).to_list(1000)
    
    return secteurs


@router.get("/{tenant_slug}/prevention/secteurs/{secteur_id}")
async def get_secteur(
    tenant_slug: str,
    secteur_id: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer un secteur spécifique"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    secteur = await db.secteurs_geographiques.find_one(
        {"id": secteur_id, "tenant_id": tenant.id},
        {"_id": 0}
    )
    
    if not secteur:
        raise HTTPException(status_code=404, detail="Secteur non trouvé")
    
    return secteur


@router.post("/{tenant_slug}/prevention/secteurs")
async def create_secteur(
    tenant_slug: str,
    secteur: SecteurGeographiqueCreate,
    current_user: User = Depends(get_current_user)
):
    """Créer un nouveau secteur géographique"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    secteur_obj = SecteurGeographique(tenant_id=tenant.id, **secteur.dict())
    
    await db.secteurs_geographiques.insert_one(secteur_obj.dict())
    
    return clean_mongo_doc(secteur_obj.dict())


@router.put("/{tenant_slug}/prevention/secteurs/{secteur_id}")
async def update_secteur(
    tenant_slug: str,
    secteur_id: str,
    secteur_data: SecteurGeographiqueCreate,
    current_user: User = Depends(get_current_user)
):
    """Mettre à jour un secteur"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    existing = await db.secteurs_geographiques.find_one({"id": secteur_id, "tenant_id": tenant.id})
    if not existing:
        raise HTTPException(status_code=404, detail="Secteur non trouvé")
    
    update_data = secteur_data.dict()
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.secteurs_geographiques.update_one(
        {"id": secteur_id, "tenant_id": tenant.id},
        {"$set": update_data}
    )
    
    updated = await db.secteurs_geographiques.find_one({"id": secteur_id, "tenant_id": tenant.id}, {"_id": 0})
    return updated


@router.delete("/{tenant_slug}/prevention/secteurs/{secteur_id}")
async def delete_secteur(
    tenant_slug: str,
    secteur_id: str,
    current_user: User = Depends(get_current_user)
):
    """Supprimer un secteur"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    existing = await db.secteurs_geographiques.find_one({"id": secteur_id, "tenant_id": tenant.id})
    if not existing:
        raise HTTPException(status_code=404, detail="Secteur non trouvé")
    
    await db.secteurs_geographiques.delete_one({"id": secteur_id, "tenant_id": tenant.id})
    
    return {"message": "Secteur supprimé avec succès"}


# ==================== MODÈLES ÉTENDUS MIGRÉS DE SERVER.PY ====================

class BatimentPhotoUpload(BaseModel):
    """Modèle pour l'upload de photo de bâtiment en base64"""
    photo_base64: str  # Data URL base64 (ex: data:image/jpeg;base64,...)

class SecteurGeographique(BaseModel):
    """Secteur géographique pour l'assignation des préventionnistes"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    
    # Informations du secteur
    nom: str  # Ex: "Secteur Nord", "Zone industrielle Est"
    description: str = ""
    couleur: str = "#3b82f6"  # Couleur d'affichage sur la carte (hex)
    
    # Géométrie (polygone GeoJSON)
    geometry: Dict[str, Any]  # Format GeoJSON: {"type": "Polygon", "coordinates": [[[lng, lat], ...]]}
    
    # Assignation
    preventionniste_assigne_id: Optional[str] = None  # ID de l'employé préventionniste
    
    # Métadonnées
    actif: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SecteurGeographiqueCreate(BaseModel):
    """Modèle pour la création d'un secteur géographique"""
    nom: str
    description: str = ""
    couleur: str = "#3b82f6"
    geometry: Dict[str, Any]
    preventionniste_assigne_id: Optional[str] = None
    actif: bool = True
    
    class Config:
        extra = "ignore"

class SymbolePersonnalise(BaseModel):
    """Symbole personnalisé pour les plans d'intervention"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    nom: str  # Ex: "Borne-fontaine personnalisée"
    categorie: str = "Personnalisé"  # Catégorie du symbole
    image_base64: str  # Image en base64
    couleur: str = "#3b82f6"  # Couleur de bordure dans la palette
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str  # ID de l'utilisateur qui a créé

class SymbolePersonnaliseCreate(BaseModel):
    """Modèle pour la création d'un symbole personnalisé"""
    nom: str
    categorie: str = "Personnalisé"
    image_base64: str
    couleur: str = "#3b82f6"
    
    class Config:
        extra = "ignore"


class GrilleInspection(BaseModel):
    """Template de grille d'inspection selon le groupe d'occupation"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    nom: str  # Ex: "Grille Groupe C - Résidentiel"
    groupe_occupation: str  # C, E, F, I, etc.
    sections: List[Dict[str, Any]] = []  # Structure JSON des sections et questions
    actif: bool = True
    version: str = "1.0"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class GrilleInspectionCreate(BaseModel):
    nom: str
    groupe_occupation: str
    sections: List[Dict[str, Any]] = []
    actif: bool = True
    version: str = "1.0"

class Inspection(BaseModel):
    """Inspection réalisée sur un bâtiment"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    batiment_id: str
    grille_inspection_id: str
    preventionniste_id: str  # ID de l'employé qui a fait l'inspection
    
    # Métadonnées inspection
    date_inspection: str = ""  # YYYY-MM-DD
    heure_debut: str = ""
    heure_fin: str = ""
    type_inspection: str = "reguliere"  # reguliere, suivi, urgence, plainte
    
    # Résultats
    resultats: Dict[str, Any] = {}  # Réponses JSON de la grille
    statut_global: str = "conforme"  # conforme, non_conforme, partiellement_conforme
    score_conformite: float = 100.0  # Pourcentage de conformité
    
    # Documentation
    photos: List[str] = []  # URLs des photos
    notes_inspection: str = ""
    recommandations: str = ""
    
    # Signature et validation
    signature_proprietaire: Optional[str] = None  # Signature numérique base64
    nom_representant: str = ""
    rapport_pdf_url: Optional[str] = None
    
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class InspectionCreate(BaseModel):
    batiment_id: str
    grille_inspection_id: str
    preventionniste_id: str
    date_inspection: str
    heure_debut: str = ""
    heure_fin: str = ""
    type_inspection: str = "reguliere"
    resultats: Dict[str, Any] = {}
    statut_global: str = "conforme"
    score_conformite: float = 100.0
    photos: List[str] = []
    notes_inspection: str = ""
    recommandations: str = ""
    signature_proprietaire: Optional[str] = None
    nom_representant: str = ""

class NonConformite(BaseModel):
    """Non-conformité identifiée lors d'une inspection"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    inspection_id: str
    batiment_id: str
    
    # Description de la non-conformité
    titre: str = ""
    description: str = ""
    section_grille: str = ""  # Section de la grille où elle a été identifiée
    gravite: str = "moyen"  # faible, moyen, eleve, critique
    article_code: str = ""  # Article du code de sécurité
    
    # Suivi
    statut: str = "ouverte"  # ouverte, en_cours, corrigee, fermee
    delai_correction: Optional[str] = None  # Date limite YYYY-MM-DD
    date_correction: Optional[str] = None
    notes_correction: str = ""
    
    # Documentation
    photos_avant: List[str] = []
    photos_apres: List[str] = []
    
    # Responsabilité
    responsable_correction: str = ""  # Propriétaire/Gestionnaire
    preventionniste_suivi_id: Optional[str] = None
    
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class NonConformiteCreate(BaseModel):
    inspection_id: str
    batiment_id: str
    titre: str
    description: str = ""
    section_grille: str = ""
    gravite: str = "moyen"
    article_code: str = ""
    delai_correction: Optional[str] = None
    photos_avant: List[str] = []
    responsable_correction: str = ""
    preventionniste_suivi_id: Optional[str] = None


# ==================== MODÈLES ÉTENDUS POUR INSPECTIONS VISUELLES ====================

class PhotoInspection(BaseModel):
    """Photo prise lors d'une inspection"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    url: str  # URL de stockage de la photo
    categorie: str = ""  # Ex: "Preuve accroche porte", "Adresse non visible", "Matières dangereuses"
    secteur: Optional[str] = None  # Secteur 1, 2, 3, 4, 5 selon schéma
    cadran: Optional[str] = None  # Cadran A, B, C, D (subdivision du Secteur 1)
    description: str = ""
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ParticipantInspection(BaseModel):
    """Participant à une inspection (pompier ou préventionniste)"""
    user_id: str
    nom_complet: str
    role: str  # "pompier" ou "preventionniste"
    est_principal: bool = False  # Le pompier connecté qui crée l'inspection

class InspectionVisuelle(BaseModel):
    """Inspection visuelle complète pour tablette/mobile"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    batiment_id: str
    
    # Participants
    participants: List[ParticipantInspection] = []
    
    # Timing
    date_inspection: str = ""  # YYYY-MM-DD
    heure_debut: Optional[str] = None
    heure_fin: Optional[str] = None
    duree_minutes: Optional[int] = None
    
    # Géolocalisation (capture automatique)
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    
    # Photos catégorisées
    photos: List[PhotoInspection] = []
    
    # Non-conformités détaillées
    non_conformites_ids: List[str] = []  # Références aux NonConformite
    
    # Checklist dynamique selon type de bâtiment
    checklist_reponses: Dict[str, Any] = {}
    
    # Statuts
    statut: str = "en_cours"  # en_cours, validee, non_conforme, suivi_requis
    statut_conformite: str = "conforme"  # conforme, non_conforme, partiellement_conforme
    
    # Plan d'intervention
    plan_intervention_url: Optional[str] = None  # URL du PDF du plan
    
    # Notes
    notes_terrain: str = ""
    recommandations: str = ""
    
    # Validation (modifiable en tout temps)
    validee_par_id: Optional[str] = None
    date_validation: Optional[datetime] = None
    
    # Mode hors-ligne
    sync_status: str = "synced"  # synced, pending, offline
    
    # Métadonnées
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class InspectionVisuelleCreate(BaseModel):
    batiment_id: str
    participants: List[ParticipantInspection]
    date_inspection: str
    heure_debut: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    notes_terrain: str = ""

class InspectionVisuelleUpdate(BaseModel):
    participants: Optional[List[ParticipantInspection]] = None
    heure_fin: Optional[str] = None
    photos: Optional[List[PhotoInspection]] = None
    checklist_reponses: Optional[Dict[str, Any]] = None
    statut: Optional[str] = None
    statut_conformite: Optional[str] = None
    notes_terrain: Optional[str] = None
    recommandations: Optional[str] = None
    validee_par_id: Optional[str] = None

class NonConformiteVisuelle(BaseModel):
    """Non-conformité avec photos et gravité détaillée"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    inspection_id: str
    batiment_id: str
    
    # Description
    titre: str
    description: str = ""
    gravite: str = "mineur"  # mineur, majeur, critique
    
    # Articles et délais
    article_municipal: str = ""  # Ex: "Article 45.2"
    delai_correction_jours: Optional[int] = None
    date_limite: Optional[str] = None  # YYYY-MM-DD
    
    # Photos
    photos_nc: List[PhotoInspection] = []  # Photos de la non-conformité
    photos_resolution: List[PhotoInspection] = []  # Photos après correction
    
    # Statut
    statut: str = "nouvelle"  # nouvelle, en_cours, resolue
    date_resolution: Optional[datetime] = None
    notes_resolution: str = ""
    
    # Suivi
    responsable_correction: str = ""  # Nom propriétaire/gestionnaire
    preventionniste_suivi_id: Optional[str] = None
    
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class NonConformiteVisuelleCreate(BaseModel):
    inspection_id: str
    batiment_id: str
    titre: str
    description: str = ""
    gravite: str = "mineur"
    article_municipal: str = ""
    delai_correction_jours: Optional[int] = None
    photos_nc: List[PhotoInspection] = []
    responsable_correction: str = ""

class BatimentMapView(BaseModel):
    """Vue simplifiée pour affichage sur carte"""
    id: str
    nom_etablissement: str
    adresse_civique: str
    ville: str
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    niveau_risque: str
    statut_inspection: str  # "fait_conforme", "a_faire", "non_conforme", "en_cours"
    derniere_inspection: Optional[str] = None  # Date ISO
    groupe_occupation: str
    sous_groupe: str

class GeocodeRequest(BaseModel):
    """Requête de géocodage d'adresse"""
    adresse_complete: str

class GeocodeResponse(BaseModel):
    """Réponse de géocodage"""
    latitude: float
    longitude: float
    adresse_formatee: str
    precision: str  # "building", "street", "city"


# ==================== MODÈLES PLANS D'INTERVENTION ====================

class ElementPlanBase(BaseModel):
    """Classe de base pour tous les éléments d'un plan d'intervention"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    type_element: str  # hydrant, sortie, matiere_dangereuse, generatrice, gaz_naturel, reservoir_propane, vehicule
    latitude: float
    longitude: float
    numero: Optional[str] = None  # Ex: H1, S1, MD1 (auto-généré)
    notes: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class HydrantElement(ElementPlanBase):
    """Hydrant sur le plan"""
    type_element: str = "hydrant"
    type_hydrant: str  # borne_fontaine, borne_seche, aspiration
    debit: float  # Débit
    unite_debit: str = "gal/min"  # gal/min ou L/min
    couleur_indicateur: Optional[str] = None  # Rouge, jaune, vert selon débit

class SortieElement(ElementPlanBase):
    """Sortie d'urgence sur le plan"""
    type_element: str = "sortie"
    type_sortie: str  # urgence, principale, secondaire
    largeur_m: Optional[float] = None
    acces_fauteuil: bool = False
    eclairage_secours: bool = False

class MatiereDangereuse(ElementPlanBase):
    """Matière dangereuse présente"""
    type_element: str = "matiere_dangereuse"
    nom_produit: str
    pictogramme_simdut: str  # URL ou code du pictogramme
    quantite: Optional[float] = None
    unite_quantite: str = "L"  # L, kg, m³
    classe_danger: str = ""  # Ex: "Inflammable", "Toxique", "Corrosif"

class GeneratriceElement(ElementPlanBase):
    """Génératrice d'urgence"""
    type_element: str = "generatrice"
    puissance_kw: Optional[float] = None
    emplacement_commutateur: str = ""
    type_carburant: str = ""  # diesel, essence, gaz naturel

class GazNaturelElement(ElementPlanBase):
    """Entrée de gaz naturel"""
    type_element: str = "gaz_naturel"
    emplacement_vanne_coupure: str
    accessible_exterieur: bool = True

class ReservoirPropaneElement(ElementPlanBase):
    """Réservoir de propane"""
    type_element: str = "reservoir_propane"
    capacite: float
    unite_capacite: str = "gallons"  # gallons ou litres
    emplacement_vanne: str
    type_reservoir: str = ""  # aerien, enterre

class VehiculeElement(ElementPlanBase):
    """Position recommandée pour véhicules d'intervention"""
    type_element: str = "vehicule"
    type_vehicule: str  # echelle, pompe, citerne
    position_recommandee: str  # Ex: "Face façade nord", "Cour arrière"
    notes_stationnement: str = ""

class RouteAcces(BaseModel):
    """Route d'accès au bâtiment"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nom: str = "Route principale"
    chemin_polyline: List[Dict[str, float]] = []  # Liste de {lat, lng}
    largeur_m: Optional[float] = None
    pente: Optional[str] = None  # faible, moyenne, forte
    notes: str = ""
    est_principale: bool = True

class ZoneDanger(BaseModel):
    """Zone de danger ou périmètre d'évacuation"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nom: str
    type_zone: str  # perimetre_evacuation, zone_chaude, zone_tiede, zone_froide
    polygone: List[Dict[str, float]] = []  # Liste de {lat, lng}
    couleur: str = "#ff0000"  # Hex color
    opacite: float = 0.3
    rayon_m: Optional[float] = None
    description: str = ""

class SecteurPlan(BaseModel):
    """Secteur du bâtiment (même système que photos inspection)"""
    numero: int  # 1, 2, 3, 4, 5
    cadran: Optional[str] = None  # A, B, C, D (subdivision secteur 1)
    description: str = ""
    elements_ids: List[str] = []  # IDs des éléments dans ce secteur

class PlanEtage(BaseModel):
    """Plan d'un étage intérieur"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    numero_etage: int  # -1 (sous-sol), 0 (RDC), 1, 2, 3...
    nom: str  # "Rez-de-chaussée", "1er étage", "Sous-sol"
    image_url: Optional[str] = None  # Image du plan d'étage
    annotations: List[Dict[str, Any]] = []  # Annotations sur le plan
    elements_interieurs: List[Dict[str, Any]] = []  # Escaliers, ascenseurs, etc.

class PhotoPlanIntervention(BaseModel):
    """Photo attachée au plan d'intervention"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    url: str
    latitude: float = 0.0
    longitude: float = 0.0
    titre: str = ""
    description: str = ""
    localisation: str = ""  # Localisation textuelle dans le bâtiment (ex: "Entrée principale", "2e étage - côté est")
    categorie: str = ""  # facade, entree, systeme_alarme, points_eau, risques, autre
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class IconePersonnalisee(BaseModel):
    """Icône personnalisée pour les plans d'intervention"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    nom: str
    image_base64: str  # Image encodée en base64
    categorie: str  # hydrants, sorties, matieres_dangereuses, generateurs, gaz_naturel, propane, vehicules, autre
    created_by_id: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class IconePersonnaliseeCreate(BaseModel):
    """Création d'une icône personnalisée"""
    nom: str
    image_base64: str
    categorie: str

class PlanIntervention(BaseModel):
    """Plan d'intervention complet"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    batiment_id: str
    
    # Identification
    numero_plan: str  # Ex: "PI-2025-001"
    nom: str = ""
    
    # Versioning
    version: str = "1.0"
    version_precedente_id: Optional[str] = None
    
    # Statut et workflow
    statut: str = "brouillon"  # brouillon, en_attente_validation, valide, archive, rejete
    created_by_id: str  # ID du préventionniste créateur
    validated_by_id: Optional[str] = None  # ID admin/superviseur qui valide
    date_validation: Optional[datetime] = None
    commentaires_validation: str = ""
    commentaires_rejet: str = ""
    
    # Éléments du plan
    hydrants: List[HydrantElement] = []
    sorties: List[SortieElement] = []
    matieres_dangereuses: List[MatiereDangereuse] = []
    generatrices: List[GeneratriceElement] = []
    gaz_naturel: List[GazNaturelElement] = []
    reservoirs_propane: List[ReservoirPropaneElement] = []
    vehicules: List[VehiculeElement] = []
    
    # Structure spatiale
    routes_acces: List[RouteAcces] = []
    zones_danger: List[ZoneDanger] = []
    secteurs: List[SecteurPlan] = []
    plans_etages: List[PlanEtage] = []
    photos: List[PhotoPlanIntervention] = []
    
    # Layers GeoJSON pour le plan interactif (depuis le builder)
    layers: List[Dict[str, Any]] = []
    
    # Vue aérienne
    centre_lat: float
    centre_lng: float
    zoom_level: int = 18
    vue_aerienne_url: Optional[str] = None  # Google Static Map URL
    carte_image: Optional[str] = None  # Capture d'écran de la carte en base64
    predefined_symbol_overrides: Dict[str, Any] = {}  # Modifications des icônes prédéfinies
    
    # Calculs automatiques
    distance_caserne_km: Optional[float] = None
    distance_caserne_unite: str = "km"  # km ou m
    temps_acces_minutes: Optional[int] = None
    
    # Documentation
    notes_generales: str = ""
    instructions_particulieres: str = ""
    
    # Export
    pdf_url: Optional[str] = None
    date_derniere_maj: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    
    # Métadonnées
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PlanInterventionCreate(BaseModel):
    batiment_id: str
    nom: str = ""
    centre_lat: float
    centre_lng: float
    notes_generales: str = ""
    
    # Champs optionnels pour permettre la sauvegarde depuis le builder
    layers: List[Dict[str, Any]] = Field(default_factory=list)
    hydrants: List[HydrantElement] = Field(default_factory=list)
    sorties: List[SortieElement] = Field(default_factory=list)
    matieres_dangereuses: List[MatiereDangereuse] = Field(default_factory=list)
    generatrices: List[GeneratriceElement] = Field(default_factory=list)
    gaz_naturel: List[GazNaturelElement] = Field(default_factory=list)
    reservoirs_propane: List[ReservoirPropaneElement] = Field(default_factory=list)
    vehicules: List[VehiculeElement] = Field(default_factory=list)
    routes_acces: List[RouteAcces] = Field(default_factory=list)
    zones_danger: List[ZoneDanger] = Field(default_factory=list)
    secteurs: List[SecteurPlan] = Field(default_factory=list)
    plans_etages: List[PlanEtage] = Field(default_factory=list)
    photos: List[PhotoPlanIntervention] = Field(default_factory=list)
    instructions_particulieres: str = ""
    carte_image: Optional[str] = None  # Capture d'écran de la carte en base64
    predefined_symbol_overrides: Dict[str, Any] = Field(default_factory=dict)  # Modifications des icônes prédéfinies

class PlanInterventionUpdate(BaseModel):
    nom: Optional[str] = None
    statut: Optional[str] = None  # Permettre la mise à jour du statut (pour repasser rejete -> brouillon)
    layers: Optional[List[Dict[str, Any]]] = None  # Layers GeoJSON du builder
    hydrants: Optional[List[HydrantElement]] = None
    sorties: Optional[List[SortieElement]] = None
    matieres_dangereuses: Optional[List[MatiereDangereuse]] = None
    generatrices: Optional[List[GeneratriceElement]] = None
    gaz_naturel: Optional[List[GazNaturelElement]] = None
    reservoirs_propane: Optional[List[ReservoirPropaneElement]] = None
    vehicules: Optional[List[VehiculeElement]] = None
    routes_acces: Optional[List[RouteAcces]] = None
    zones_danger: Optional[List[ZoneDanger]] = None
    secteurs: Optional[List[SecteurPlan]] = None
    plans_etages: Optional[List[PlanEtage]] = None
    photos: Optional[List[PhotoPlanIntervention]] = None
    notes_generales: Optional[str] = None
    instructions_particulieres: Optional[str] = None
    carte_image: Optional[str] = None  # Capture d'écran de la carte en base64
    predefined_symbol_overrides: Optional[Dict[str, Any]] = None  # Modifications des icônes prédéfinies

class TemplatePlanIntervention(BaseModel):
    """Template pré-défini de plan d'intervention"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    nom: str  # Ex: "Résidentiel unifamilial", "Commercial petit", "Industriel F-1"
    type_batiment: str  # residentiel, commercial, industriel
    groupe_occupation: str  # A, B, C, D, E, F, G, I
    sous_groupe: Optional[str] = None  # F-1, F-2, F-3, etc.
    
    # Éléments pré-configurés (positions relatives)
    hydrants_defaut: List[Dict[str, Any]] = []
    sorties_defaut: List[Dict[str, Any]] = []
    vehicules_defaut: List[Dict[str, Any]] = []
    
    # Instructions
    instructions_utilisation: str = ""
    
    actif: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ValidationRequest(BaseModel):
    """Requête de validation de plan"""
    commentaires: str = ""

class RejectionRequest(BaseModel):
    """Requête de rejet de plan"""
    commentaires_rejet: str





# ==================== ROUTES PRÉVENTION MIGRÉES DE SERVER.PY ====================

# ==================== PRÉVENTION ENDPOINTS ====================
















@router.post("/{tenant_slug}/prevention/initialiser")
async def initialiser_module_prevention(
    tenant_slug: str,
    current_user: User = Depends(get_current_user)
):
    """Initialiser le module prévention avec les 7 grilles d'inspection standards"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé - Admin uniquement")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Vérifier que le module prévention est activé
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Vérifier si des grilles existent déjà
    existing_count = await db.grilles_inspection.count_documents({"tenant_id": tenant.id})
    if existing_count > 0:
        raise HTTPException(status_code=400, detail=f"{existing_count} grille(s) existent déjà. Supprimez-les d'abord si vous voulez réinitialiser.")
    
    # Importer les grilles depuis insert_grilles.py
    import sys
    import os
    sys.path.insert(0, os.path.dirname(__file__))
    from insert_grilles import GRILLES
    
    # Créer les 7 grilles pour ce tenant
    grilles_creees = []
    for grille_template in GRILLES:
        grille = grille_template.copy()
        grille["id"] = str(uuid.uuid4())
        grille["tenant_id"] = tenant.id
        grille["actif"] = True
        grille["version"] = "1.0"
        
        await db.grilles_inspection.insert_one(grille)
        grilles_creees.append({"id": grille["id"], "nom": grille["nom"]})
    
    return {
        "message": f"{len(grilles_creees)} grilles d'inspection créées avec succès",
        "grilles": grilles_creees
    }



@router.post("/{tenant_slug}/prevention/batiments/{batiment_id}/photo")
async def upload_batiment_photo(
    tenant_slug: str,
    batiment_id: str,
    photo_data: BatimentPhotoUpload,
    current_user: User = Depends(get_current_user)
):
    """Uploader/Mettre à jour la photo d'un bâtiment (en base64)"""
    # Admins, superviseurs et préventionnistes peuvent uploader des photos
    if current_user.role not in ["admin", "superviseur"] and not current_user.est_preventionniste:
        raise HTTPException(status_code=403, detail="Accès refusé - Permission insuffisante")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Vérifier que le module prévention est activé
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Vérifier que le bâtiment existe
    existing = await db.batiments.find_one({"id": batiment_id, "tenant_id": tenant.id})
    if not existing:
        raise HTTPException(status_code=404, detail="Bâtiment non trouvé")
    
    # Vérifier que la photo est au bon format base64
    if not photo_data.photo_base64.startswith('data:image/'):
        raise HTTPException(status_code=400, detail="Format de photo invalide (doit être base64 data URL)")
    
    # Mettre à jour la photo
    await db.batiments.update_one(
        {"id": batiment_id, "tenant_id": tenant.id},
        {"$set": {"photo_url": photo_data.photo_base64, "updated_at": datetime.now(timezone.utc)}}
    )
    
    return {"message": "Photo mise à jour avec succès", "photo_url": photo_data.photo_base64}

@router.delete("/{tenant_slug}/prevention/batiments/{batiment_id}/photo")
async def delete_batiment_photo(
    tenant_slug: str,
    batiment_id: str,
    current_user: User = Depends(get_current_user)
):
    """Supprimer la photo d'un bâtiment"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé - Admin uniquement")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Vérifier que le module prévention est activé
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Mettre à jour pour retirer la photo
    result = await db.batiments.update_one(
        {"id": batiment_id, "tenant_id": tenant.id},
        {"$set": {"photo_url": "", "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Bâtiment non trouvé")
    
    return {"message": "Photo supprimée avec succès"}

# ==================== SYMBOLES PERSONNALISÉS ====================

@router.get("/{tenant_slug}/prevention/symboles-personnalises")
async def get_symboles_personnalises(
    tenant_slug: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer tous les symboles personnalisés du tenant"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé - Admin uniquement")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    symboles = await db.symboles_personnalises.find({"tenant_id": tenant.id}).to_list(1000)
    return [clean_mongo_doc(symbole) for symbole in symboles]

@router.post("/{tenant_slug}/prevention/symboles-personnalises")
async def create_symbole_personnalise(
    tenant_slug: str,
    symbole: SymbolePersonnaliseCreate,
    current_user: User = Depends(get_current_user)
):
    """Créer un nouveau symbole personnalisé"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé - Admin uniquement")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Vérifier que l'image est au bon format base64
    if not symbole.image_base64.startswith('data:image/'):
        raise HTTPException(status_code=400, detail="Format d'image invalide (doit être base64 data URL)")
    
    symbole_dict = symbole.dict()
    symbole_dict["tenant_id"] = tenant.id
    symbole_dict["id"] = str(uuid.uuid4())
    symbole_dict["created_at"] = datetime.now(timezone.utc)
    symbole_dict["created_by"] = current_user.id
    
    symbole_obj = SymbolePersonnalise(**symbole_dict)
    await db.symboles_personnalises.insert_one(symbole_obj.dict())
    
    return clean_mongo_doc(symbole_obj.dict())

@router.put("/{tenant_slug}/prevention/symboles-personnalises/{symbole_id}")
async def update_symbole_personnalise(
    tenant_slug: str,
    symbole_id: str,
    symbole: SymbolePersonnaliseCreate,
    current_user: User = Depends(get_current_user)
):
    """Modifier un symbole personnalisé"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé - Admin uniquement")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Vérifier que le symbole existe
    existing = await db.symboles_personnalises.find_one({"id": symbole_id, "tenant_id": tenant.id})
    if not existing:
        raise HTTPException(status_code=404, detail="Symbole non trouvé")
    
    # Vérifier que l'image est au bon format si elle est fournie
    if symbole.image_base64 and not symbole.image_base64.startswith('data:image/'):
        raise HTTPException(status_code=400, detail="Format d'image invalide (doit être base64 data URL)")
    
    update_dict = symbole.dict()
    update_dict["updated_at"] = datetime.now(timezone.utc)
    
    await db.symboles_personnalises.update_one(
        {"id": symbole_id, "tenant_id": tenant.id},
        {"$set": update_dict}
    )
    
    updated = await db.symboles_personnalises.find_one({"id": symbole_id, "tenant_id": tenant.id})
    return clean_mongo_doc(updated)

@router.delete("/{tenant_slug}/prevention/symboles-personnalises/{symbole_id}")
async def delete_symbole_personnalise(
    tenant_slug: str,
    symbole_id: str,
    current_user: User = Depends(get_current_user)
):
    """Supprimer un symbole personnalisé"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé - Admin uniquement")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Vérifier si le symbole existe
    symbole = await db.symboles_personnalises.find_one({"id": symbole_id, "tenant_id": tenant.id})
    if not symbole:
        raise HTTPException(status_code=404, detail="Symbole non trouvé")
    
    # Vérifier si le symbole est utilisé dans des plans d'intervention
    # Les plans peuvent stocker les symboles dans les layers (format GeoJSON)
    plans_utilisant = await db.plans_intervention.count_documents({
        "tenant_id": tenant.id,
        "$or": [
            {"layers.properties.symbolId": symbole_id},
            {"layers": {"$elemMatch": {"properties.symbolId": symbole_id}}}
        ]
    })
    
    if plans_utilisant > 0:
        raise HTTPException(
            status_code=409, 
            detail=f"Impossible de supprimer ce symbole. Il est utilisé dans {plans_utilisant} plan(s) d'intervention."
        )
    
    result = await db.symboles_personnalises.delete_one({"id": symbole_id, "tenant_id": tenant.id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Symbole non trouvé")
    
    return {"message": "Symbole supprimé avec succès"}



# ==================== SECTEURS GÉOGRAPHIQUES ====================











# ==================== GÉNÉRATION RAPPORT PDF ====================# ==================== GÉNÉRATION RAPPORT PDF ====================

async def generer_rapport_inspection_pdf(inspection_id: str, tenant_id: str) -> BytesIO:
    """Générer un rapport PDF pour une inspection"""
    # Récupérer le tenant
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    
    # Récupérer l'inspection
    inspection = await db.inspections.find_one({"id": inspection_id, "tenant_id": tenant_id})
    if not inspection:
        raise HTTPException(status_code=404, detail="Inspection non trouvée")
    
    # Récupérer le bâtiment
    batiment = await db.batiments.find_one({"id": inspection["batiment_id"], "tenant_id": tenant_id})
    
    # Récupérer la grille d'inspection
    grille = await db.grilles_inspection.find_one({"id": inspection["grille_inspection_id"], "tenant_id": tenant_id})
    
    # Récupérer le préventionniste
    preventionniste = await db.users.find_one({"id": inspection["preventionniste_id"]})
    
    # Récupérer les non-conformités
    non_conformites = await db.non_conformites.find({
        "inspection_id": inspection_id,
        "tenant_id": tenant_id
    }).to_list(1000)
    
    # Créer le PDF avec branding
    from types import SimpleNamespace
    tenant_obj = SimpleNamespace(**tenant) if tenant else None
    buffer, doc, story = create_branded_pdf(tenant_obj, pagesize=letter)
    styles = getSampleStyleSheet()
    
    # Style personnalisé
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#374151'),
        spaceAfter=12,
        spaceBefore=20
    )
    
    # Titre
    story.append(Paragraph("RAPPORT D'INSPECTION INCENDIE", title_style))
    story.append(Spacer(1, 0.3*inch))
    
    # Informations générales
    story.append(Paragraph("INFORMATIONS GÉNÉRALES", heading_style))
    
    info_data = [
        ["Date d'inspection:", inspection.get("date_inspection", "N/A")],
        ["Type:", inspection.get("type_inspection", "régulière").upper()],
        ["Préventionniste:", f"{preventionniste.get('prenom', '')} {preventionniste.get('nom', '')}" if preventionniste else "N/A"],
        ["Heure début:", inspection.get("heure_debut", "N/A")],
        ["Heure fin:", inspection.get("heure_fin", "N/A")],
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1f2937')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb'))
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Informations bâtiment
    if batiment:
        story.append(Paragraph("INFORMATIONS BÂTIMENT", heading_style))
        
        bat_data = [
            ["Nom établissement:", batiment.get("nom_etablissement", "N/A")],
            ["Adresse:", batiment.get("adresse_civique", "N/A")],
            ["Ville:", batiment.get("ville", "N/A")],
            ["Code postal:", batiment.get("code_postal", "N/A")],
            ["Groupe occupation:", batiment.get("groupe_occupation", "N/A")],
        ]
        
        bat_table = Table(bat_data, colWidths=[2*inch, 4*inch])
        bat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1f2937')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb'))
        ]))
        story.append(bat_table)
        story.append(Spacer(1, 0.3*inch))
    
    # Résultat global
    story.append(Paragraph("RÉSULTAT GLOBAL", heading_style))
    
    statut_color = colors.HexColor('#10b981') if inspection.get("statut_global") == "conforme" else colors.HexColor('#ef4444')
    statut_text = inspection.get("statut_global", "N/A").upper()
    score = inspection.get("score_conformite", 100)
    
    result_data = [
        ["Statut:", statut_text],
        ["Score de conformité:", f"{score}%"],
        ["Non-conformités:", str(len(non_conformites))],
    ]
    
    result_table = Table(result_data, colWidths=[2*inch, 4*inch])
    result_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('BACKGROUND', (1, 0), (1, 0), statut_color),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1f2937')),
        ('TEXTCOLOR', (1, 0), (1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb'))
    ]))
    story.append(result_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Non-conformités
    if non_conformites:
        story.append(Paragraph("NON-CONFORMITÉS IDENTIFIÉES", heading_style))
        
        for idx, nc in enumerate(non_conformites, 1):
            nc_data = [
                [f"#{idx}", ""],
                ["Titre:", nc.get("titre", "N/A")],
                ["Description:", nc.get("description", "N/A")],
                ["Gravité:", nc.get("gravite", "N/A").upper()],
                ["Statut:", nc.get("statut", "N/A")],
                ["Délai correction:", nc.get("delai_correction", "N/A")],
            ]
            
            nc_table = Table(nc_data, colWidths=[2*inch, 4*inch])
            nc_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fef2f2')),
                ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#f3f4f6')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1f2937')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                ('SPAN', (0, 0), (-1, 0)),
            ]))
            story.append(nc_table)
            story.append(Spacer(1, 0.2*inch))
    
    # Notes et recommandations
    if inspection.get("notes_inspection") or inspection.get("recommandations"):
        story.append(Paragraph("NOTES ET RECOMMANDATIONS", heading_style))
        
        if inspection.get("notes_inspection"):
            story.append(Paragraph(f"<b>Notes:</b> {inspection.get('notes_inspection')}", styles['Normal']))
            story.append(Spacer(1, 0.1*inch))
        
        if inspection.get("recommandations"):
            story.append(Paragraph(f"<b>Recommandations:</b> {inspection.get('recommandations')}", styles['Normal']))
            story.append(Spacer(1, 0.3*inch))
    
    # Signature
    story.append(Spacer(1, 0.5*inch))
    story.append(Paragraph("SIGNATURES", heading_style))
    
    sig_data = [
        ["Préventionniste:", "_" * 40],
        ["Date:", "_" * 40],
        ["", ""],
        ["Représentant bâtiment:", "_" * 40],
        ["Nom:", inspection.get("nom_representant", "_" * 40)],
        ["Date:", "_" * 40],
    ]
    
    sig_table = Table(sig_data, colWidths=[2*inch, 4*inch])
    sig_table.setStyle(TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1f2937')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(sig_table)
    
    # Générer le PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


@router.get("/{tenant_slug}/prevention/inspections/{inspection_id}/rapport-pdf")
async def get_inspection_rapport_pdf(
    tenant_slug: str,
    inspection_id: str,
    current_user: User = Depends(get_current_user)
):
    """Générer et télécharger le rapport PDF d'une inspection"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Vérifier que l'inspection existe
    inspection = await db.inspections.find_one({"id": inspection_id, "tenant_id": tenant.id})
    if not inspection:
        raise HTTPException(status_code=404, detail="Inspection non trouvée")
    
    # Générer le PDF
    pdf_buffer = await generer_rapport_inspection_pdf(inspection_id, tenant.id)
    
    # Retourner le PDF
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=rapport_inspection_{inspection_id}.pdf"
        }
    )


# ==================== INSPECTIONS ====================





@router.get("/{tenant_slug}/prevention/inspections-planifiees")
async def get_inspections_planifiees(
    tenant_slug: str,
    days: int = 7,
    current_user: User = Depends(get_current_user)
):
    """Récupérer les inspections planifiées pour les X prochains jours (pour mode offline)"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Calculer la plage de dates
    from datetime import datetime, timedelta, timezone
    today = datetime.now(timezone.utc).date()
    end_date = today + timedelta(days=days)
    
    # Convertir en format ISO string YYYY-MM-DD pour comparaison
    today_str = today.isoformat()
    end_date_str = end_date.isoformat()
    
    # Récupérer les inspections planifiées (date_inspection entre aujourd'hui et end_date)
    inspections = await db.inspections.find({
        "tenant_id": tenant.id,
        "date_inspection": {
            "$gte": today_str,
            "$lte": end_date_str
        },
        "statut": {"$in": ["planifiee", "en_cours", "a_faire", None]}  # Exclure les terminées
    }).sort("date_inspection", 1).to_list(100)
    
    # Enrichir avec les infos du bâtiment
    enriched = []
    for insp in inspections:
        batiment = await db.batiments.find_one(
            {"id": insp.get("batiment_id"), "tenant_id": tenant.id},
            {"_id": 0}
        )
        
        enriched.append({
            **clean_mongo_doc(insp),
            "batiment": batiment
        })
    
    return enriched

@router.delete("/{tenant_slug}/prevention/inspections/{inspection_id}")
async def delete_inspection(
    tenant_slug: str,
    inspection_id: str,
    current_user: User = Depends(get_current_user)
):
    """Supprimer une inspection"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé - Admin uniquement")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    result = await db.inspections.delete_one({"id": inspection_id, "tenant_id": tenant.id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Inspection non trouvée")
    
    # Supprimer aussi les non-conformités associées
    await db.non_conformites.delete_many({"inspection_id": inspection_id, "tenant_id": tenant.id})
    
    return {"message": "Inspection supprimée avec succès"}


# ==================== NON-CONFORMITÉS ====================

@router.post("/{tenant_slug}/prevention/non-conformites")
async def create_non_conformite(
    tenant_slug: str,
    non_conformite: NonConformiteCreate,
    current_user: User = Depends(get_current_user)
):
    """Créer une nouvelle non-conformité"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    nc_dict = non_conformite.dict()
    nc_dict["tenant_id"] = tenant.id
    nc_obj = NonConformite(**nc_dict)
    
    await db.non_conformites.insert_one(nc_obj.dict())
    
    return clean_mongo_doc(nc_obj.dict())

@router.get("/{tenant_slug}/prevention/non-conformites")
async def get_non_conformites(
    tenant_slug: str,
    inspection_id: Optional[str] = None,
    batiment_id: Optional[str] = None,
    statut: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Récupérer toutes les non-conformités avec filtres optionnels"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    query = {"tenant_id": tenant.id}
    if inspection_id:
        query["inspection_id"] = inspection_id
    if batiment_id:
        query["batiment_id"] = batiment_id
    if statut:
        query["statut"] = statut
    
    non_conformites = await db.non_conformites.find(query).sort("created_at", -1).to_list(1000)
    return [clean_mongo_doc(nc) for nc in non_conformites]

@router.get("/{tenant_slug}/prevention/non-conformites/{nc_id}")
async def get_non_conformite(
    tenant_slug: str,
    nc_id: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer une non-conformité spécifique"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    nc = await db.non_conformites.find_one({"id": nc_id, "tenant_id": tenant.id})
    
    if not nc:
        raise HTTPException(status_code=404, detail="Non-conformité non trouvée")
    
    return clean_mongo_doc(nc)

@router.put("/{tenant_slug}/prevention/non-conformites/{nc_id}")
async def update_non_conformite(
    tenant_slug: str,
    nc_id: str,
    nc_data: NonConformiteCreate,
    current_user: User = Depends(get_current_user)
):
    """Mettre à jour une non-conformité"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    update_dict = nc_data.dict()
    update_dict["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.non_conformites.update_one(
        {"id": nc_id, "tenant_id": tenant.id},
        {"$set": update_dict}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Non-conformité non trouvée")
    
    updated_nc = await db.non_conformites.find_one({"id": nc_id})
    return clean_mongo_doc(updated_nc)

@router.patch("/{tenant_slug}/prevention/non-conformites/{nc_id}/statut")
async def update_non_conformite_statut(
    tenant_slug: str,
    nc_id: str,
    statut: str = Body(..., embed=True),
    notes_correction: str = Body("", embed=True),
    current_user: User = Depends(get_current_user)
):
    """Mettre à jour le statut d'une non-conformité"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    update_data = {
        "statut": statut,
        "notes_correction": notes_correction,
        "updated_at": datetime.now(timezone.utc)
    }
    
    if statut == "corrigee" or statut == "fermee":
        update_data["date_correction"] = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    result = await db.non_conformites.update_one(
        {"id": nc_id, "tenant_id": tenant.id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Non-conformité non trouvée")
    
    updated_nc = await db.non_conformites.find_one({"id": nc_id})
    return clean_mongo_doc(updated_nc)

@router.delete("/{tenant_slug}/prevention/non-conformites/{nc_id}")
async def delete_non_conformite(
    tenant_slug: str,
    nc_id: str,
    current_user: User = Depends(get_current_user)
):
    """Supprimer une non-conformité"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé - Admin uniquement")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    result = await db.non_conformites.delete_one({"id": nc_id, "tenant_id": tenant.id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Non-conformité non trouvée")
    
    return {"message": "Non-conformité supprimée avec succès"}


# ==================== UPLOAD PHOTOS ====================

@router.post("/{tenant_slug}/prevention/upload-photo")
async def upload_photo(
    tenant_slug: str,
    photo_base64: str = Body(..., embed=True),
    filename: str = Body(..., embed=True),
    current_user: User = Depends(get_current_user)
):
    """Upload une photo en base64 et retourne l'URL"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    try:
        # Générer un ID unique pour la photo
        photo_id = str(uuid.uuid4())
        
        # Stocker la photo dans la collection photos
        photo_doc = {
            "id": photo_id,
            "tenant_id": tenant.id,
            "filename": filename,
            "data": photo_base64,
            "uploaded_by": current_user.id,
            "uploaded_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.photos_prevention.insert_one(photo_doc)
        
        # Retourner l'ID de la photo (qui servira d'URL)
        return {
            "photo_id": photo_id,
            "url": f"/api/{tenant_slug}/prevention/photos/{photo_id}"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur upload photo: {str(e)}")

@router.post("/{tenant_slug}/inventaires/upload-photo")
async def upload_inventaire_photo(
    tenant_slug: str,
    photo_base64: str = Body(..., embed=True),
    filename: str = Body(..., embed=True),
    current_user: User = Depends(get_current_user)
):
    """Upload une photo pour un inventaire véhicule et retourne l'URL"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    try:
        # Générer un ID unique pour la photo
        photo_id = str(uuid.uuid4())
        
        # Stocker la photo dans la collection photos
        photo_doc = {
            "id": photo_id,
            "tenant_id": tenant.id,
            "filename": filename,
            "data": photo_base64,
            "uploaded_by": current_user.id,
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "type": "inventaire_vehicule"
        }
        
        await db.photos_inventaires.insert_one(photo_doc)
        
        # Retourner l'URL de la photo
        return {
            "photo_id": photo_id,
            "url": f"/api/{tenant_slug}/inventaires/photos/{photo_id}"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur upload photo: {str(e)}")

@router.get("/{tenant_slug}/inventaires/photos/{photo_id}")
async def get_inventaire_photo(
    tenant_slug: str,
    photo_id: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer une photo d'inventaire par son ID"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    photo = await db.photos_inventaires.find_one({
        "id": photo_id,
        "tenant_id": tenant.id
    }, {"_id": 0})
    
    if not photo:
        raise HTTPException(status_code=404, detail="Photo non trouvée")
    
    # Retourner la photo en base64
    from fastapi.responses import Response
    import base64
    
    # Extraire les données de l'image
    if photo['data'].startswith('data:'):
        # Format: data:image/png;base64,xxxx
        header, data = photo['data'].split(',', 1)
        mime_type = header.split(':')[1].split(';')[0]
    else:
        data = photo['data']
        mime_type = 'image/png'
    
    image_data = base64.b64decode(data)
    
    return Response(content=image_data, media_type=mime_type)

@router.get("/{tenant_slug}/prevention/photos/{photo_id}")
async def get_photo(
    tenant_slug: str,
    photo_id: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer une photo par son ID"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    photo = await db.photos_prevention.find_one({"id": photo_id, "tenant_id": tenant.id})
    
    if not photo:
        raise HTTPException(status_code=404, detail="Photo non trouvée")
    
    return {
        "id": photo["id"],
        "filename": photo.get("filename", "photo.jpg"),
        "data": photo["data"],
        "uploaded_at": photo.get("uploaded_at")
    }

@router.delete("/{tenant_slug}/prevention/photos/{photo_id}")
async def delete_photo(
    tenant_slug: str,
    photo_id: str,
    current_user: User = Depends(get_current_user)
):
    """Supprimer une photo"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    result = await db.photos_prevention.delete_one({"id": photo_id, "tenant_id": tenant.id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Photo non trouvée")
    
    return {"message": "Photo supprimée avec succès"}


# ==================== ICÔNES PERSONNALISÉES ====================

@router.post("/{tenant_slug}/prevention/icones-personnalisees")
async def create_icone_personnalisee(
    tenant_slug: str,
    icone: IconePersonnaliseeCreate,
    current_user: User = Depends(get_current_user)
):
    """Créer une icône personnalisée"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    try:
        # Créer l'icône
        icone_dict = icone.dict()
        icone_dict["id"] = str(uuid.uuid4())
        icone_dict["tenant_id"] = tenant.id
        icone_dict["created_by_id"] = current_user.id
        icone_dict["created_at"] = datetime.now(timezone.utc).isoformat()
        
        await db.icones_personnalisees.insert_one(icone_dict)
        
        return clean_mongo_doc(icone_dict)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur création icône: {str(e)}")

@router.get("/{tenant_slug}/prevention/icones-personnalisees")
async def get_icones_personnalisees(
    tenant_slug: str,
    categorie: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Récupérer toutes les icônes personnalisées d'un tenant"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    query = {"tenant_id": tenant.id}
    if categorie:
        query["categorie"] = categorie
    
    icones = await db.icones_personnalisees.find(query).to_list(length=None)
    
    return [clean_mongo_doc(icone) for icone in icones]

@router.delete("/{tenant_slug}/prevention/icones-personnalisees/{icone_id}")
async def delete_icone_personnalisee(
    tenant_slug: str,
    icone_id: str,
    current_user: User = Depends(get_current_user)
):
    """Supprimer une icône personnalisée"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    result = await db.icones_personnalisees.delete_one({"id": icone_id, "tenant_id": tenant.id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Icône non trouvée")
    
    return {"message": "Icône supprimée avec succès"}


# ==================== INSPECTIONS VISUELLES (NOUVEAU SYSTÈME) ====================

@router.post("/{tenant_slug}/prevention/inspections-visuelles")
async def create_inspection_visuelle(
    tenant_slug: str,
    inspection: InspectionVisuelleCreate,
    current_user: User = Depends(get_current_user)
):
    """Créer une nouvelle inspection visuelle (pompiers + préventionnistes)"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Vérifier que le bâtiment existe
    batiment = await db.batiments.find_one({"id": inspection.batiment_id, "tenant_id": tenant.id})
    if not batiment:
        raise HTTPException(status_code=404, detail="Bâtiment non trouvé")
    
    inspection_dict = inspection.dict()
    inspection_dict["tenant_id"] = tenant.id
    
    # Créer l'objet InspectionVisuelle complet
    inspection_obj = InspectionVisuelle(**inspection_dict)
    
    await db.inspections_visuelles.insert_one(inspection_obj.dict())
    
    return clean_mongo_doc(inspection_obj.dict())

@router.get("/{tenant_slug}/prevention/inspections-visuelles")
async def get_inspections_visuelles(
    tenant_slug: str,
    batiment_id: Optional[str] = None,
    statut: Optional[str] = None,
    date_debut: Optional[str] = None,
    date_fin: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Récupérer la liste des inspections visuelles avec filtres"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    query = {"tenant_id": tenant.id}
    
    if batiment_id:
        query["batiment_id"] = batiment_id
    
    if statut:
        query["statut"] = statut
    
    if date_debut and date_fin:
        query["date_inspection"] = {"$gte": date_debut, "$lte": date_fin}
    
    inspections = await db.inspections_visuelles.find(query).sort("created_at", -1).to_list(length=None)
    
    return [clean_mongo_doc(insp) for insp in inspections]

@router.get("/{tenant_slug}/prevention/inspections-visuelles/{inspection_id}")
async def get_inspection_visuelle(
    tenant_slug: str,
    inspection_id: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer une inspection visuelle spécifique"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    inspection = await db.inspections_visuelles.find_one({"id": inspection_id, "tenant_id": tenant.id})
    
    if not inspection:
        raise HTTPException(status_code=404, detail="Inspection non trouvée")
    
    return clean_mongo_doc(inspection)

@router.put("/{tenant_slug}/prevention/inspections-visuelles/{inspection_id}")
async def update_inspection_visuelle(
    tenant_slug: str,
    inspection_id: str,
    inspection_update: InspectionVisuelleUpdate,
    current_user: User = Depends(get_current_user)
):
    """Mettre à jour une inspection visuelle (toujours modifiable)"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Récupérer l'inspection existante
    existing = await db.inspections_visuelles.find_one({"id": inspection_id, "tenant_id": tenant.id})
    if not existing:
        raise HTTPException(status_code=404, detail="Inspection non trouvée")
    
    # Mettre à jour uniquement les champs fournis
    update_dict = {k: v for k, v in inspection_update.dict(exclude_unset=True).items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc)
    
    # Calculer la durée si heure_fin est fournie
    if "heure_fin" in update_dict and existing.get("heure_debut"):
        try:
            debut = datetime.fromisoformat(f"{existing['date_inspection']}T{existing['heure_debut']}")
            fin = datetime.fromisoformat(f"{existing['date_inspection']}T{update_dict['heure_fin']}")
            update_dict["duree_minutes"] = int((fin - debut).total_seconds() / 60)
        except:
            pass
    
    result = await db.inspections_visuelles.update_one(
        {"id": inspection_id, "tenant_id": tenant.id},
        {"$set": update_dict}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Inspection non trouvée")
    
    updated = await db.inspections_visuelles.find_one({"id": inspection_id})
    return clean_mongo_doc(updated)

@router.delete("/{tenant_slug}/prevention/inspections-visuelles/{inspection_id}")
async def delete_inspection_visuelle(
    tenant_slug: str,
    inspection_id: str,
    current_user: User = Depends(get_current_user)
):
    """Supprimer une inspection visuelle (préventionnistes uniquement)"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Vérifier que l'utilisateur est préventionniste ou admin
    if current_user.role not in ["admin", "superviseur"] and current_user.type_emploi != "preventionniste":
        raise HTTPException(status_code=403, detail="Seuls les préventionnistes peuvent supprimer des inspections")
    
    result = await db.inspections_visuelles.delete_one({"id": inspection_id, "tenant_id": tenant.id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Inspection non trouvée")
    
    # Supprimer aussi les non-conformités associées
    await db.non_conformites_visuelles.delete_many({"inspection_id": inspection_id, "tenant_id": tenant.id})
    
    return {"message": "Inspection supprimée avec succès"}


# ==================== NON-CONFORMITÉS VISUELLES ====================

@router.post("/{tenant_slug}/prevention/non-conformites-visuelles")
async def create_non_conformite_visuelle(
    tenant_slug: str,
    nc: NonConformiteVisuelleCreate,
    current_user: User = Depends(get_current_user)
):
    """Créer une nouvelle non-conformité visuelle"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    nc_dict = nc.dict()
    nc_dict["tenant_id"] = tenant.id
    
    # Calculer la date limite si délai fourni
    if nc_dict.get("delai_correction_jours"):
        from datetime import timedelta
        inspection = await db.inspections_visuelles.find_one({"id": nc.inspection_id})
        if inspection:
            date_insp = datetime.fromisoformat(inspection["date_inspection"])
            date_limite = date_insp + timedelta(days=nc_dict["delai_correction_jours"])
            nc_dict["date_limite"] = date_limite.strftime("%Y-%m-%d")
    
    nc_obj = NonConformiteVisuelle(**nc_dict)
    
    await db.non_conformites_visuelles.insert_one(nc_obj.dict())
    
    # Ajouter l'ID de la NC à l'inspection
    await db.inspections_visuelles.update_one(
        {"id": nc.inspection_id, "tenant_id": tenant.id},
        {"$push": {"non_conformites_ids": nc_obj.id}}
    )
    
    return clean_mongo_doc(nc_obj.dict())

@router.get("/{tenant_slug}/prevention/non-conformites-visuelles")
async def get_non_conformites_visuelles(
    tenant_slug: str,
    inspection_id: Optional[str] = None,
    batiment_id: Optional[str] = None,
    statut: Optional[str] = None,
    gravite: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Récupérer les non-conformités visuelles avec filtres"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    query = {"tenant_id": tenant.id}
    
    if inspection_id:
        query["inspection_id"] = inspection_id
    if batiment_id:
        query["batiment_id"] = batiment_id
    if statut:
        query["statut"] = statut
    if gravite:
        query["gravite"] = gravite
    
    ncs = await db.non_conformites_visuelles.find(query).sort("created_at", -1).to_list(length=None)
    
    return [clean_mongo_doc(nc) for nc in ncs]

@router.put("/{tenant_slug}/prevention/non-conformites-visuelles/{nc_id}")
async def update_non_conformite_visuelle(
    tenant_slug: str,
    nc_id: str,
    statut: Optional[str] = Body(None),
    photos_resolution: Optional[List[PhotoInspection]] = Body(None),
    notes_resolution: Optional[str] = Body(None),
    current_user: User = Depends(get_current_user)
):
    """Mettre à jour le statut d'une non-conformité"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    update_dict = {"updated_at": datetime.now(timezone.utc)}
    
    if statut:
        update_dict["statut"] = statut
        if statut == "resolue":
            update_dict["date_resolution"] = datetime.now(timezone.utc)
    
    if photos_resolution:
        update_dict["photos_resolution"] = [p.dict() for p in photos_resolution]
    
    if notes_resolution:
        update_dict["notes_resolution"] = notes_resolution
    
    result = await db.non_conformites_visuelles.update_one(
        {"id": nc_id, "tenant_id": tenant.id},
        {"$set": update_dict}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Non-conformité non trouvée")
    
    updated = await db.non_conformites_visuelles.find_one({"id": nc_id})
    return clean_mongo_doc(updated)


# ==================== CARTE INTERACTIVE & GÉOCODAGE ====================

@router.get("/{tenant_slug}/prevention/batiments/map")
async def get_batiments_for_map(
    tenant_slug: str,
    niveau_risque: Optional[str] = None,
    statut_inspection: Optional[str] = None,
    secteur: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Récupérer les bâtiments formatés pour affichage sur carte"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Récupérer tous les bâtiments
    query = {"tenant_id": tenant.id, "statut": "actif"}
    
    if niveau_risque:
        query["niveau_risque"] = niveau_risque
    
    batiments = await db.batiments.find(query).to_list(length=None)
    
    # Pour chaque bâtiment, déterminer le statut d'inspection
    map_data = []
    for bat in batiments:
        # Chercher la dernière inspection
        derniere_inspection = await db.inspections_visuelles.find_one(
            {"batiment_id": bat["id"], "tenant_id": tenant.id},
            sort=[("date_inspection", -1)]
        )
        
        # Déterminer le statut
        if not derniere_inspection:
            statut_insp = "a_faire"
        elif derniere_inspection["statut"] == "en_cours":
            statut_insp = "en_cours"
        elif derniere_inspection["statut_conformite"] == "non_conforme":
            statut_insp = "non_conforme"
        else:
            statut_insp = "fait_conforme"
        
        # Filtrer par statut si demandé
        if statut_inspection and statut_insp != statut_inspection:
            continue
        
        # Géocoder l'adresse si pas déjà fait (latitude/longitude manquants)
        latitude = bat.get("latitude")
        longitude = bat.get("longitude")
        
        map_item = BatimentMapView(
            id=bat["id"],
            nom_etablissement=bat.get("nom_etablissement", ""),
            adresse_civique=bat.get("adresse_civique", ""),
            ville=bat.get("ville", ""),
            latitude=latitude,
            longitude=longitude,
            niveau_risque=bat.get("niveau_risque", ""),
            statut_inspection=statut_insp,
            derniere_inspection=derniere_inspection["date_inspection"] if derniere_inspection else None,
            groupe_occupation=bat.get("groupe_occupation", ""),
            sous_groupe=bat.get("sous_groupe", "")
        )
        
        map_data.append(map_item.dict())
    
    return map_data

@router.post("/{tenant_slug}/prevention/geocode")
async def geocode_address(
    tenant_slug: str,
    request: GeocodeRequest,
    current_user: User = Depends(get_current_user)
):
    """Géocoder une adresse en latitude/longitude avec Google Maps API"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    try:
        import requests
        import os
        
        api_key = os.getenv("GOOGLE_MAPS_API_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="Clé API Google Maps non configurée")
        
        # Appeler l'API Google Geocoding
        url = "https://maps.googleapis.com/maps/api/geocode/json"
        params = {
            "address": request.adresse_complete,
            "key": api_key
        }
        
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        
        data = response.json()
        
        if data["status"] == "OK" and len(data["results"]) > 0:
            result = data["results"][0]
            location = result["geometry"]["location"]
            
            # Déterminer la précision
            location_type = result["geometry"]["location_type"]
            if location_type == "ROOFTOP":
                precision = "building"
            elif location_type in ["RANGE_INTERPOLATED", "GEOMETRIC_CENTER"]:
                precision = "street"
            else:
                precision = "city"
            
            return GeocodeResponse(
                latitude=location["lat"],
                longitude=location["lng"],
                adresse_formatee=result["formatted_address"],
                precision=precision
            )
        else:
            raise HTTPException(status_code=404, detail="Adresse non trouvée")
    
    except requests.exceptions.RequestException as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors du géocodage: {str(e)}")

@router.put("/{tenant_slug}/prevention/batiments/{batiment_id}/coordinates")
async def update_batiment_coordinates(
    tenant_slug: str,
    batiment_id: str,
    latitude: float = Body(...),
    longitude: float = Body(...),
    current_user: User = Depends(get_current_user)
):
    """Mettre à jour les coordonnées GPS d'un bâtiment"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    result = await db.batiments.update_one(
        {"id": batiment_id, "tenant_id": tenant.id},
        {"$set": {
            "latitude": latitude,
            "longitude": longitude,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Bâtiment non trouvé")
    
    return {"message": "Coordonnées mises à jour avec succès"}


# ==================== GESTION DES PRÉVENTIONNISTES ====================

@router.put("/{tenant_slug}/users/{user_id}/toggle-preventionniste")
async def toggle_preventionniste(
    tenant_slug: str,
    user_id: str,
    current_user: User = Depends(get_current_user)
):
    """Activer/désactiver le statut de préventionniste pour un utilisateur (admin uniquement)"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Seuls les admins et superviseurs peuvent modifier ce statut
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Permissions insuffisantes")
    
    # Récupérer l'utilisateur
    user = await db.users.find_one({"id": user_id, "tenant_id": tenant.id})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    # Toggle le statut
    new_status = not user.get('est_preventionniste', False)
    
    await db.users.update_one(
        {"id": user_id, "tenant_id": tenant.id},
        {"$set": {"est_preventionniste": new_status}}
    )
    
    return {
        "message": "Statut de préventionniste mis à jour",
        "user_id": user_id,
        "est_preventionniste": new_status
    }


@router.get("/{tenant_slug}/prevention/preventionnistes")
async def get_preventionnistes(
    tenant_slug: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer la liste de tous les préventionnistes actifs"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Récupérer tous les utilisateurs avec est_preventionniste = true et statut actif
    preventionnistes_cursor = db.users.find({
        "tenant_id": tenant.id,
        "est_preventionniste": True,
        "statut": "Actif"
    })
    
    preventionnistes = await preventionnistes_cursor.to_list(length=None)
    
    # Pour chaque préventionniste, ajouter des statistiques
    result = []
    for prev in preventionnistes:
        # Compter les bâtiments assignés
        nb_batiments = await db.batiments.count_documents({
            "tenant_id": tenant.id,
            "preventionniste_assigne_id": prev["id"]
        })
        
        # Compter les secteurs assignés
        nb_secteurs = await db.secteurs_geographiques.count_documents({
            "tenant_id": tenant.id,
            "preventionniste_assigne_id": prev["id"]
        })
        
        # Compter les inspections ce mois
        start_of_month = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        nb_inspections_mois = await db.inspections.count_documents({
            "tenant_id": tenant.id,
            "preventionniste_id": prev["id"],
            "date_inspection": {"$gte": start_of_month.isoformat()}
        })
        
        result.append({
            "id": prev["id"],
            "nom": prev["nom"],
            "prenom": prev["prenom"],
            "email": prev["email"],
            "telephone": prev.get("telephone", ""),
            "grade": prev.get("grade", ""),
            "nb_batiments": nb_batiments,
            "nb_secteurs": nb_secteurs,
            "nb_inspections_mois": nb_inspections_mois
        })
    
    return result


@router.get("/{tenant_slug}/prevention/preventionnistes/{preventionniste_id}/stats")
async def get_preventionniste_stats(
    tenant_slug: str,
    preventionniste_id: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer les statistiques détaillées d'un préventionniste"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Vérifier que le préventionniste existe
    preventionniste = await db.users.find_one({
        "id": preventionniste_id,
        "tenant_id": tenant.id,
        "est_preventionniste": True
    })
    
    if not preventionniste:
        raise HTTPException(status_code=404, detail="Préventionniste non trouvé")
    
    # Statistiques globales
    nb_batiments = await db.batiments.count_documents({
        "tenant_id": tenant.id,
        "preventionniste_assigne_id": preventionniste_id
    })
    
    nb_secteurs = await db.secteurs_geographiques.count_documents({
        "tenant_id": tenant.id,
        "preventionniste_assigne_id": preventionniste_id
    })
    
    # Inspections par période
    start_of_month = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    start_of_year = datetime.now(timezone.utc).replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    
    nb_inspections_mois = await db.inspections.count_documents({
        "tenant_id": tenant.id,
        "preventionniste_id": preventionniste_id,
        "date_inspection": {"$gte": start_of_month.isoformat()}
    })
    
    nb_inspections_annee = await db.inspections.count_documents({
        "tenant_id": tenant.id,
        "preventionniste_id": preventionniste_id,
        "date_inspection": {"$gte": start_of_year.isoformat()}
    })
    
    # Plans d'intervention créés
    nb_plans = await db.plans_intervention.count_documents({
        "tenant_id": tenant.id,
        "created_by": preventionniste_id
    })
    
    return {
        "preventionniste": {
            "id": preventionniste["id"],
            "nom": preventionniste["nom"],
            "prenom": preventionniste["prenom"],
            "email": preventionniste["email"],
            "telephone": preventionniste.get("telephone", ""),
            "grade": preventionniste.get("grade", "")
        },
        "stats": {
            "nb_batiments": nb_batiments,
            "nb_secteurs": nb_secteurs,
            "nb_inspections_mois": nb_inspections_mois,
            "nb_inspections_annee": nb_inspections_annee,
            "nb_plans": nb_plans
        }
    }


@router.get("/{tenant_slug}/prevention/preventionnistes/{preventionniste_id}/batiments")
async def get_preventionniste_batiments(
    tenant_slug: str,
    preventionniste_id: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer tous les bâtiments assignés à un préventionniste"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    batiments_cursor = db.batiments.find({
        "tenant_id": tenant.id,
        "preventionniste_assigne_id": preventionniste_id
    })
    
    batiments = await batiments_cursor.to_list(length=None)
    
    # Nettoyer les ObjectIds pour sérialisation JSON
    for batiment in batiments:
        if "_id" in batiment:
            del batiment["_id"]
    
    return batiments


@router.get("/{tenant_slug}/prevention/preventionnistes/{preventionniste_id}/secteurs")
async def get_preventionniste_secteurs(
    tenant_slug: str,
    preventionniste_id: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer tous les secteurs assignés à un préventionniste"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    secteurs_cursor = db.secteurs_geographiques.find({
        "tenant_id": tenant.id,
        "preventionniste_assigne_id": preventionniste_id
    })
    
    secteurs = await secteurs_cursor.to_list(length=None)
    
    # Nettoyer les ObjectIds pour sérialisation JSON
    for secteur in secteurs:
        if "_id" in secteur:
            del secteur["_id"]
    
    return secteurs


@router.put("/{tenant_slug}/prevention/batiments/{batiment_id}/assigner")
async def assigner_batiment_preventionniste(
    tenant_slug: str,
    batiment_id: str,
    preventionniste_id: Optional[str] = Body(None),
    raison: Optional[str] = Body(""),
    current_user: User = Depends(get_current_user)
):
    """Assigner un préventionniste à un bâtiment (avec historique)"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Vérifier permissions
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Permissions insuffisantes")
    
    # Récupérer le bâtiment
    batiment = await db.batiments.find_one({"id": batiment_id, "tenant_id": tenant.id})
    if not batiment:
        raise HTTPException(status_code=404, detail="Bâtiment non trouvé")
    
    # Si preventionniste_id fourni, vérifier qu'il existe et est actif
    if preventionniste_id:
        preventionniste = await db.users.find_one({
            "id": preventionniste_id,
            "tenant_id": tenant.id,
            "est_preventionniste": True,
            "statut": "Actif"
        })
        if not preventionniste:
            raise HTTPException(status_code=404, detail="Préventionniste non trouvé ou inactif")
    
    # Créer l'entrée d'historique
    ancien_preventionniste_id = batiment.get("preventionniste_assigne_id")
    historique_entry = {
        "date": datetime.now(timezone.utc).isoformat(),
        "ancien_preventionniste_id": ancien_preventionniste_id,
        "nouveau_preventionniste_id": preventionniste_id,
        "modifie_par": current_user.id,
        "modifie_par_nom": f"{current_user.prenom} {current_user.nom}",
        "raison": raison
    }
    
    # Mettre à jour le bâtiment
    await db.batiments.update_one(
        {"id": batiment_id, "tenant_id": tenant.id},
        {
            "$set": {
                "preventionniste_assigne_id": preventionniste_id,
                "updated_at": datetime.now(timezone.utc)
            },
            "$push": {"historique_assignations": historique_entry}
        }
    )
    
    # Créer notification pour le préventionniste
    if preventionniste_id:
        notification = {
            "id": str(uuid.uuid4()),
            "tenant_id": tenant.id,
            "user_id": preventionniste_id,
            "type": "assignation_batiment",
            "titre": "Nouveau bâtiment assigné",
            "message": f"Le bâtiment '{batiment.get('nom_etablissement') or batiment.get('adresse_civique')}' vous a été assigné.",
            "lue": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "data": {
                "batiment_id": batiment_id,
                "batiment_nom": batiment.get('nom_etablissement') or batiment.get('adresse_civique')
            }
        }
        await db.notifications.insert_one(notification)
    
    return {
        "message": "Bâtiment assigné avec succès",
        "batiment_id": batiment_id,
        "preventionniste_id": preventionniste_id
    }


@router.put("/{tenant_slug}/prevention/secteurs/{secteur_id}/assigner")
async def assigner_secteur_preventionniste(
    tenant_slug: str,
    secteur_id: str,
    preventionniste_id: Optional[str] = Body(None),
    assigner_batiments: bool = Body(True),
    current_user: User = Depends(get_current_user)
):
    """Assigner un préventionniste à un secteur (et optionnellement tous ses bâtiments)"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Vérifier permissions
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Permissions insuffisantes")
    
    # Récupérer le secteur
    secteur = await db.secteurs_geographiques.find_one({"id": secteur_id, "tenant_id": tenant.id})
    if not secteur:
        raise HTTPException(status_code=404, detail="Secteur non trouvé")
    
    # Si preventionniste_id fourni, vérifier qu'il existe
    if preventionniste_id:
        preventionniste = await db.users.find_one({
            "id": preventionniste_id,
            "tenant_id": tenant.id,
            "est_preventionniste": True,
            "statut": "Actif"
        })
        if not preventionniste:
            raise HTTPException(status_code=404, detail="Préventionniste non trouvé ou inactif")
    
    # Mettre à jour le secteur
    await db.secteurs_geographiques.update_one(
        {"id": secteur_id, "tenant_id": tenant.id},
        {
            "$set": {
                "preventionniste_assigne_id": preventionniste_id,
                "updated_at": datetime.now(timezone.utc)
            }
        }
    )
    
    nb_batiments_assignes = 0
    
    # Si demandé, assigner tous les bâtiments du secteur
    if assigner_batiments:
        # Trouver tous les bâtiments dans ce secteur (géométriquement)
        # Pour simplifier, on va assigner tous les bâtiments sans preventionniste ou avec autre preventionniste
        batiments_cursor = db.batiments.find({
            "tenant_id": tenant.id,
            "latitude": {"$ne": None},
            "longitude": {"$ne": None}
        })
        
        batiments = await batiments_cursor.to_list(length=None)
        
        # Pour chaque bâtiment, vérifier s'il est dans le polygone du secteur
        from shapely.geometry import Point, shape
        
        secteur_polygon = shape(secteur["geometry"])
        
        for batiment in batiments:
            if batiment.get("latitude") and batiment.get("longitude"):
                point = Point(batiment["longitude"], batiment["latitude"])
                
                if secteur_polygon.contains(point):
                    # Créer l'entrée d'historique
                    ancien_preventionniste_id = batiment.get("preventionniste_assigne_id")
                    historique_entry = {
                        "date": datetime.now(timezone.utc).isoformat(),
                        "ancien_preventionniste_id": ancien_preventionniste_id,
                        "nouveau_preventionniste_id": preventionniste_id,
                        "modifie_par": current_user.id,
                        "modifie_par_nom": f"{current_user.prenom} {current_user.nom}",
                        "raison": f"Assignation automatique via secteur '{secteur['nom']}'"
                    }
                    
                    # Mettre à jour le bâtiment
                    await db.batiments.update_one(
                        {"id": batiment["id"]},
                        {
                            "$set": {
                                "preventionniste_assigne_id": preventionniste_id,
                                "updated_at": datetime.now(timezone.utc)
                            },
                            "$push": {"historique_assignations": historique_entry}
                        }
                    )
                    nb_batiments_assignes += 1
    
    # Créer notification pour le préventionniste
    if preventionniste_id:
        notification = {
            "id": str(uuid.uuid4()),
            "tenant_id": tenant.id,
            "user_id": preventionniste_id,
            "type": "assignation_secteur",
            "titre": "Nouveau secteur assigné",
            "message": f"Le secteur '{secteur['nom']}' vous a été assigné" + (f" avec {nb_batiments_assignes} bâtiments." if assigner_batiments else "."),
            "lue": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "data": {
                "secteur_id": secteur_id,
                "secteur_nom": secteur['nom'],
                "nb_batiments": nb_batiments_assignes
            }
        }
        await db.notifications.insert_one(notification)
    
    return {
        "message": "Secteur assigné avec succès",
        "secteur_id": secteur_id,
        "preventionniste_id": preventionniste_id,
        "nb_batiments_assignes": nb_batiments_assignes
    }


# ==================== PARAMÈTRES PRÉVENTION ====================

@router.put("/{tenant_slug}/prevention/parametres")
async def update_parametres_prevention(
    tenant_slug: str,
    recurrence_inspections: int = Body(...),
    nombre_visites_requises: int = Body(...),
    superviseur_prevention_id: Optional[str] = Body(None),
    current_user: User = Depends(get_current_user)
):
    """Mettre à jour les paramètres de prévention (admin uniquement)"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Vérifier permissions (admin seulement)
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Seuls les administrateurs peuvent modifier les paramètres")
    
    # Valider les valeurs
    if recurrence_inspections not in [1, 2, 3, 4, 5]:
        raise HTTPException(status_code=400, detail="La récurrence doit être entre 1 et 5 ans")
    
    if nombre_visites_requises not in [1, 2, 3]:
        raise HTTPException(status_code=400, detail="Le nombre de visites doit être entre 1 et 3")
    
    # Si superviseur fourni, vérifier qu'il existe
    if superviseur_prevention_id:
        superviseur = await db.users.find_one({
            "id": superviseur_prevention_id,
            "tenant_id": tenant.id
        })
        if not superviseur:
            raise HTTPException(status_code=404, detail="Superviseur non trouvé")
    
    # Mettre à jour les paramètres du tenant
    parametres_update = {
        "parametres.recurrence_inspections": recurrence_inspections,
        "parametres.nombre_visites_requises": nombre_visites_requises,
        "parametres.superviseur_prevention_id": superviseur_prevention_id,
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.tenants.update_one(
        {"id": tenant.id},
        {"$set": parametres_update}
    )
    
    logging.info(f"Paramètres prévention mis à jour pour {tenant_slug} par {current_user.prenom} {current_user.nom}")
    
    return {
        "message": "Paramètres mis à jour avec succès",
        "parametres": {
            "recurrence_inspections": recurrence_inspections,
            "nombre_visites_requises": nombre_visites_requises,
            "superviseur_prevention_id": superviseur_prevention_id
        }
    }


# ==================== PLANS D'INTERVENTION ====================

@router.post("/{tenant_slug}/prevention/plans-intervention")
async def create_plan_intervention(
    tenant_slug: str,
    plan: PlanInterventionCreate,
    current_user: User = Depends(get_current_user)
):
    """Créer un nouveau plan d'intervention (préventionnistes uniquement)"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Vérifier que l'utilisateur est préventionniste ou admin
    if current_user.role not in ["admin", "superviseur"] and current_user.type_emploi != "preventionniste":
        raise HTTPException(status_code=403, detail="Seuls les préventionnistes peuvent créer des plans")
    
    # Vérifier que le bâtiment existe
    batiment = await db.batiments.find_one({"id": plan.batiment_id, "tenant_id": tenant.id})
    if not batiment:
        raise HTTPException(status_code=404, detail="Bâtiment non trouvé")
    
    # Générer le numéro de plan unique
    current_year = datetime.now().year
    count = await db.plans_intervention.count_documents({"tenant_id": tenant.id})
    numero_plan = f"PI-{current_year}-{str(count + 1).zfill(3)}"
    
    plan_dict = plan.dict()
    
    # Debug logs
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"📥 Layers reçus dans plan_dict: {len(plan_dict.get('layers', []))} layers")
    logger.info(f"📥 Détails layers reçus: {plan_dict.get('layers', [])}")
    
    plan_dict["tenant_id"] = tenant.id
    plan_dict["numero_plan"] = numero_plan
    plan_dict["created_by_id"] = current_user.id
    plan_dict["statut"] = "brouillon"
    
    plan_obj = PlanIntervention(**plan_dict)
    
    logger.info(f"📤 Layers dans plan_obj après conversion: {len(plan_obj.layers)} layers")
    
    plan_to_insert = plan_obj.dict()
    logger.info(f"💾 Layers avant insertion MongoDB: {len(plan_to_insert.get('layers', []))} layers")
    
    await db.plans_intervention.insert_one(plan_to_insert)
    
    # Créer une activité
    batiment_nom = batiment.get('nom') or batiment.get('nom_batiment') or batiment.get('adresse_civique') or 'Bâtiment'
    await creer_activite(
        tenant_id=tenant.id,
        type_activite="prevention_plan_creation",
        description=f"🏢 {current_user.prenom} {current_user.nom} a créé le plan d'intervention #{numero_plan} pour '{batiment_nom}'",
        user_id=current_user.id,
        user_nom=f"{current_user.prenom} {current_user.nom}"
    )
    
    result = clean_mongo_doc(plan_to_insert)
    logger.info(f"✅ Layers dans résultat final: {len(result.get('layers', []))} layers")
    
    return result

@router.get("/{tenant_slug}/prevention/plans-intervention")
async def get_plans_intervention(
    tenant_slug: str,
    batiment_id: Optional[str] = None,
    statut: Optional[str] = None,
    created_by_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Récupérer la liste des plans d'intervention avec filtres"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    query = {"tenant_id": tenant.id}
    
    if batiment_id:
        query["batiment_id"] = batiment_id
    if statut:
        query["statut"] = statut
    if created_by_id:
        query["created_by_id"] = created_by_id
    
    plans = await db.plans_intervention.find(query).sort("created_at", -1).to_list(length=None)
    
    return [clean_mongo_doc(plan) for plan in plans]

@router.get("/{tenant_slug}/prevention/plans-intervention/{plan_id}")
async def get_plan_intervention(
    tenant_slug: str,
    plan_id: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer un plan d'intervention spécifique"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    plan = await db.plans_intervention.find_one({"id": plan_id, "tenant_id": tenant.id})
    
    if not plan:
        raise HTTPException(status_code=404, detail="Plan d'intervention non trouvé")
    
    return clean_mongo_doc(plan)

@router.put("/{tenant_slug}/prevention/plans-intervention/{plan_id}")
async def update_plan_intervention(
    tenant_slug: str,
    plan_id: str,
    plan_update: PlanInterventionUpdate,
    current_user: User = Depends(get_current_user)
):
    """Mettre à jour un plan d'intervention (seulement si brouillon ou en_attente)"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Récupérer le plan existant
    existing = await db.plans_intervention.find_one({"id": plan_id, "tenant_id": tenant.id})
    if not existing:
        raise HTTPException(status_code=404, detail="Plan non trouvé")
    
    # Vérifier que le plan est modifiable
    if existing["statut"] not in ["brouillon", "en_attente_validation", "rejete"]:
        raise HTTPException(status_code=403, detail="Plan validé non modifiable - créer une nouvelle version")
    
    # Vérifier que l'utilisateur est le créateur ou admin
    if existing["created_by_id"] != current_user.id and current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Seul le créateur ou un admin peut modifier ce plan")
    
    # Mettre à jour les champs fournis
    update_dict = {k: v for k, v in plan_update.dict(exclude_unset=True).items() if v is not None}
    
    # Debug logs
    import logging
    logger = logging.getLogger(__name__)
    if 'layers' in update_dict:
        logger.info(f"📥 UPDATE - Layers reçus: {len(update_dict.get('layers', []))} layers")
        logger.info(f"📥 UPDATE - Détails layers: {update_dict.get('layers', [])}")
    
    update_dict["updated_at"] = datetime.now(timezone.utc)
    update_dict["date_derniere_maj"] = datetime.now(timezone.utc)
    
    result = await db.plans_intervention.update_one(
        {"id": plan_id, "tenant_id": tenant.id},
        {"$set": update_dict}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Plan non trouvé")
    
    updated = await db.plans_intervention.find_one({"id": plan_id})
    
    if updated and 'layers' in updated:
        logger.info(f"✅ UPDATE - Layers dans résultat: {len(updated.get('layers', []))} layers")
    
    return clean_mongo_doc(updated)

@router.delete("/{tenant_slug}/prevention/plans-intervention/{plan_id}")
async def delete_plan_intervention(
    tenant_slug: str,
    plan_id: str,
    current_user: User = Depends(get_current_user)
):
    """Supprimer un plan d'intervention (admin uniquement)"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé - Admin uniquement")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    result = await db.plans_intervention.delete_one({"id": plan_id, "tenant_id": tenant.id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Plan non trouvé")
    
    return {"message": "Plan d'intervention supprimé avec succès"}

@router.post("/{tenant_slug}/prevention/plans-intervention/{plan_id}/valider")
async def soumettre_plan_validation(
    tenant_slug: str,
    plan_id: str,
    request: ValidationRequest,
    current_user: User = Depends(get_current_user)
):
    """Soumettre un plan pour validation (préventionniste créateur)"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    plan = await db.plans_intervention.find_one({"id": plan_id, "tenant_id": tenant.id})
    if not plan:
        raise HTTPException(status_code=404, detail="Plan non trouvé")
    
    # Vérifier que l'utilisateur est le créateur
    if plan["created_by_id"] != current_user.id and current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Seul le créateur peut soumettre le plan")
    
    # Vérifier que le plan est en brouillon
    if plan["statut"] != "brouillon":
        raise HTTPException(status_code=400, detail="Le plan n'est pas en brouillon")
    
    result = await db.plans_intervention.update_one(
        {"id": plan_id, "tenant_id": tenant.id},
        {"$set": {
            "statut": "en_attente_validation",
            "commentaires_validation": request.commentaires,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"message": "Plan soumis pour validation"}

@router.post("/{tenant_slug}/prevention/plans-intervention/{plan_id}/approuver")
async def approuver_plan_intervention(
    tenant_slug: str,
    plan_id: str,
    request: ValidationRequest,
    current_user: User = Depends(get_current_user)
):
    """Approuver un plan d'intervention (admin/superviseur uniquement)"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Seuls les admin/superviseurs peuvent approuver")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    plan = await db.plans_intervention.find_one({"id": plan_id, "tenant_id": tenant.id})
    if not plan:
        raise HTTPException(status_code=404, detail="Plan non trouvé")
    
    if plan["statut"] != "en_attente_validation":
        raise HTTPException(status_code=400, detail="Le plan n'est pas en attente de validation")
    
    result = await db.plans_intervention.update_one(
        {"id": plan_id, "tenant_id": tenant.id},
        {"$set": {
            "statut": "valide",
            "validated_by_id": current_user.id,
            "date_validation": datetime.now(timezone.utc),
            "commentaires_validation": request.commentaires,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"message": "Plan d'intervention approuvé"}

@router.post("/{tenant_slug}/prevention/plans-intervention/{plan_id}/rejeter")
async def rejeter_plan_intervention(
    tenant_slug: str,
    plan_id: str,
    request: RejectionRequest,
    current_user: User = Depends(get_current_user)
):
    """Rejeter un plan d'intervention (admin/superviseur uniquement)"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Seuls les admin/superviseurs peuvent rejeter")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    plan = await db.plans_intervention.find_one({"id": plan_id, "tenant_id": tenant.id})
    if not plan:
        raise HTTPException(status_code=404, detail="Plan non trouvé")
    
    if plan["statut"] != "en_attente_validation":
        raise HTTPException(status_code=400, detail="Le plan n'est pas en attente de validation")
    
    result = await db.plans_intervention.update_one(
        {"id": plan_id, "tenant_id": tenant.id},
        {"$set": {
            "statut": "rejete",
            "validated_by_id": current_user.id,
            "commentaires_rejet": request.commentaires_rejet,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"message": "Plan d'intervention rejeté"}

@router.post("/{tenant_slug}/prevention/plans-intervention/{plan_id}/nouvelle-version")
async def creer_nouvelle_version_plan(
    tenant_slug: str,
    plan_id: str,
    current_user: User = Depends(get_current_user)
):
    """Créer une nouvelle version d'un plan validé"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Vérifier permissions
    if current_user.role not in ["admin", "superviseur"] and current_user.type_emploi != "preventionniste":
        raise HTTPException(status_code=403, detail="Seuls les préventionnistes peuvent créer des versions")
    
    # Récupérer le plan existant
    plan_actuel = await db.plans_intervention.find_one({"id": plan_id, "tenant_id": tenant.id})
    if not plan_actuel:
        raise HTTPException(status_code=404, detail="Plan non trouvé")
    
    if plan_actuel["statut"] != "valide":
        raise HTTPException(status_code=400, detail="Seul un plan validé peut avoir une nouvelle version")
    
    # Archiver l'ancien plan
    await db.plans_intervention.update_one(
        {"id": plan_id, "tenant_id": tenant.id},
        {"$set": {"statut": "archive"}}
    )
    
    # Créer la nouvelle version
    nouveau_plan = plan_actuel.copy()
    nouveau_plan["id"] = str(uuid.uuid4())
    nouveau_plan["version_precedente_id"] = plan_id
    nouveau_plan["statut"] = "brouillon"
    nouveau_plan["created_by_id"] = current_user.id
    nouveau_plan["validated_by_id"] = None
    nouveau_plan["date_validation"] = None
    nouveau_plan["commentaires_validation"] = ""
    nouveau_plan["commentaires_rejet"] = ""
    nouveau_plan["created_at"] = datetime.now(timezone.utc)
    nouveau_plan["updated_at"] = datetime.now(timezone.utc)
    
    # Incrémenter la version
    version_parts = nouveau_plan["version"].split(".")
    version_parts[-1] = str(int(version_parts[-1]) + 1)
    nouveau_plan["version"] = ".".join(version_parts)
    
    # Supprimer _id MongoDB du dict avant insertion
    if "_id" in nouveau_plan:
        del nouveau_plan["_id"]
    
    await db.plans_intervention.insert_one(nouveau_plan)
    
    return clean_mongo_doc(nouveau_plan)


@router.get("/{tenant_slug}/prevention/plans-intervention/{plan_id}/export-pdf")
async def export_plan_intervention_pdf(
    tenant_slug: str,
    plan_id: str,
    current_user: User = Depends(get_current_user)
):
    """Exporter un plan d'intervention en PDF"""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from io import BytesIO
    import base64
    from PIL import Image as PILImage
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Récupérer le plan
    plan = await db.plans_intervention.find_one({"id": plan_id, "tenant_id": tenant.id})
    if not plan:
        raise HTTPException(status_code=404, detail="Plan non trouvé")
    
    # Récupérer le bâtiment associé
    batiment = None
    if plan.get("batiment_id"):
        batiment = await db.batiments.find_one({"id": plan["batiment_id"], "tenant_id": tenant.id})
    
    # Créer le buffer PDF avec branding
    buffer, doc, elements = create_branded_pdf(
        tenant, 
        pagesize=A4, 
        rightMargin=40, 
        leftMargin=40, 
        topMargin=60, 
        bottomMargin=40
    )
    
    # Styles
    styles = getSampleStyleSheet()
    modern_styles = get_modern_pdf_styles(styles)
    title_style = modern_styles['title']
    heading_style = modern_styles['heading']
    normal_style = styles['Normal']
    
    # En-tête avec titre du plan
    elements.append(Paragraph(f"🔥 Plan d'Intervention", modern_styles['title']))
    plan_title = plan.get('titre') or plan.get('nom_plan') or f"Plan - {batiment.get('nom_etablissement', 'Sans titre') if batiment else 'Sans titre'}"
    elements.append(Paragraph(f"<b>{plan_title}</b>", heading_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Informations générales
    info_data = [
        ['Numéro de plan:', plan.get('numero_plan', 'N/A')],
        ['Statut:', plan.get('statut', 'brouillon').replace('_', ' ').capitalize()],
        ['Date de création:', plan.get('created_at').strftime('%Y-%m-%d') if plan.get('created_at') and hasattr(plan.get('created_at'), 'strftime') else (str(plan.get('created_at', 'N/A'))[:10] if plan.get('created_at') else 'N/A')],
    ]
    
    if plan.get('date_validation'):
        date_val = plan['date_validation']
        if hasattr(date_val, 'strftime'):
            info_data.append(['Date de validation:', date_val.strftime('%Y-%m-%d')])
        elif isinstance(date_val, str):
            info_data.append(['Date de validation:', date_val[:10]])
        else:
            info_data.append(['Date de validation:', 'N/A'])
    
    if batiment:
        info_data.append(['Bâtiment:', f"{batiment.get('nom_etablissement', 'N/A')}"])
        info_data.append(['Adresse complète:', f"{batiment.get('adresse_civique', '')} {batiment.get('ville', '')} {batiment.get('province', '')} {batiment.get('code_postal', '')}".strip() or 'N/A'])
        
        if batiment.get('type_batiment'):
            info_data.append(['Type de bâtiment:', batiment['type_batiment']])
        if batiment.get('sous_type_batiment'):
            info_data.append(['Sous-type:', batiment['sous_type_batiment']])
        if batiment.get('groupe_occupation'):
            info_data.append(['Groupe d\'occupation:', batiment['groupe_occupation']])
        if batiment.get('sous_groupe'):
            info_data.append(['Sous-groupe:', batiment['sous_groupe']])
        if batiment.get('niveau_risque'):
            info_data.append(['Niveau de risque:', batiment['niveau_risque']])
        if batiment.get('annee_construction'):
            info_data.append(['Année construction:', batiment['annee_construction']])
        if batiment.get('nombre_etages'):
            info_data.append(['Nombre d\'étages:', batiment['nombre_etages']])
        if batiment.get('superficie_totale_m2'):
            info_data.append(['Superficie totale:', f"{batiment['superficie_totale_m2']} m²"])
        if batiment.get('cadastre_matricule'):
            info_data.append(['Cadastre/Matricule:', batiment['cadastre_matricule']])
        if batiment.get('description_activite'):
            info_data.append(['Description activité:', batiment['description_activite']])
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F3F4F6')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Section Contacts (si bâtiment disponible)
    if batiment:
        has_contacts = False
        contact_data = []
        
        # Propriétaire
        if batiment.get('proprietaire_nom') or batiment.get('proprietaire_prenom'):
            contact_data.append(['👤 Propriétaire:', f"{batiment.get('proprietaire_prenom', '')} {batiment.get('proprietaire_nom', '')}".strip()])
            if batiment.get('proprietaire_telephone'):
                contact_data.append(['Téléphone:', batiment.get('proprietaire_telephone')])
            if batiment.get('proprietaire_courriel'):
                contact_data.append(['Courriel:', batiment.get('proprietaire_courriel')])
            has_contacts = True
        
        # Gestionnaire
        if batiment.get('gestionnaire_nom') or batiment.get('gestionnaire_prenom') or batiment.get('gerant_nom'):
            nom = batiment.get('gestionnaire_nom') or batiment.get('gerant_nom', '')
            prenom = batiment.get('gestionnaire_prenom', '')
            contact_data.append(['👨‍💼 Gestionnaire:', f"{prenom} {nom}".strip()])
            tel = batiment.get('gestionnaire_telephone') or batiment.get('gerant_telephone')
            if tel:
                contact_data.append(['Téléphone:', tel])
            email = batiment.get('gestionnaire_courriel') or batiment.get('gerant_courriel')
            if email:
                contact_data.append(['Courriel:', email])
            has_contacts = True
        
        # Locataire
        if batiment.get('locataire_nom') or batiment.get('locataire_prenom') or batiment.get('localaire_nom'):
            nom = batiment.get('locataire_nom') or batiment.get('localaire_nom', '')
            prenom = batiment.get('locataire_prenom') or batiment.get('localaire_prenom', '')
            contact_data.append(['🏠 Locataire:', f"{prenom} {nom}".strip()])
            tel = batiment.get('locataire_telephone') or batiment.get('localaire_telephone')
            if tel:
                contact_data.append(['Téléphone:', tel])
            email = batiment.get('locataire_courriel') or batiment.get('localaire_courriel')
            if email:
                contact_data.append(['Courriel:', email])
            has_contacts = True
        
        # Responsable sécurité
        if batiment.get('responsable_securite_nom'):
            contact_data.append(['🔒 Responsable sécurité:', batiment.get('responsable_securite_nom')])
            if batiment.get('responsable_securite_telephone'):
                contact_data.append(['Téléphone:', batiment.get('responsable_securite_telephone')])
            if batiment.get('responsable_securite_courriel'):
                contact_data.append(['Courriel:', batiment.get('responsable_securite_courriel')])
            has_contacts = True
        
        if has_contacts:
            elements.append(Paragraph("<b>📞 Contacts</b>", heading_style))
            contact_table = Table(contact_data, colWidths=[2*inch, 4*inch])
            contact_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F3F4F6')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(contact_table)
            elements.append(Spacer(1, 0.3*inch))
    
    # Description
    description = plan.get('description') or plan.get('notes_generales')
    if description:
        elements.append(Paragraph("<b>📋 Description</b>", heading_style))
        elements.append(Paragraph(description, normal_style))
        elements.append(Spacer(1, 0.2*inch))
    
    # Notes Tactiques
    notes_tactiques = plan.get('notes_tactiques') or plan.get('instructions_particulieres')
    if notes_tactiques:
        elements.append(Paragraph("<b>⚠️ Notes Tactiques</b>", heading_style))
        elements.append(Paragraph(notes_tactiques, normal_style))
        elements.append(Spacer(1, 0.2*inch))
    
    # Section Carte et Légendes (affichage visuel des symboles)
    layers = plan.get('layers', [])
    if layers and len(layers) > 0:
        elements.append(PageBreak())
        elements.append(Paragraph("<b>🗺️ Carte et Légendes</b>", title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Informations de la carte
        carte_info = f"""
        <b>Centre de la carte:</b><br/>
        Latitude: {plan.get('centre_lat', 'N/A')}<br/>
        Longitude: {plan.get('centre_lng', 'N/A')}<br/>
        <br/>
        <b>Éléments placés sur la carte:</b> {len(layers)} symbole(s)
        """
        elements.append(Paragraph(carte_info, normal_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Image de la carte (si disponible)
        carte_image = plan.get('carte_image')
        if carte_image:
            try:
                # Décoder l'image base64
                if carte_image.startswith('data:image'):
                    image_data = carte_image.split(',')[1]
                    image_bytes = base64.b64decode(image_data)
                    
                    # Créer une image PIL pour optimiser
                    img = PILImage.open(BytesIO(image_bytes))
                    if img.mode in ('RGBA', 'LA', 'P'):
                        img = img.convert('RGB')
                    
                    # Redimensionner pour le PDF (largeur max 6.5 inches)
                    max_width = 6.5 * inch
                    aspect_ratio = img.height / img.width
                    img_width = max_width
                    img_height = img_width * aspect_ratio
                    
                    # Limiter la hauteur maximale à 4 inches
                    if img_height > 4 * inch:
                        img_height = 4 * inch
                        img_width = img_height / aspect_ratio
                    
                    # Compresser l'image
                    img_buffer = BytesIO()
                    img.save(img_buffer, format='JPEG', quality=80, optimize=True)
                    img_buffer.seek(0)
                    
                    # Créer l'image ReportLab
                    carte_rl_image = RLImage(img_buffer, width=img_width, height=img_height)
                    
                    # Centrer l'image
                    elements.append(Paragraph("<b>📍 Vue de la Carte</b>", heading_style))
                    elements.append(Spacer(1, 0.1*inch))
                    elements.append(carte_rl_image)
                    elements.append(Spacer(1, 0.3*inch))
            except Exception as e:
                print(f"Erreur lors de l'ajout de l'image de la carte: {e}")
                elements.append(Paragraph("<i>Erreur lors du chargement de l'image de la carte</i>", normal_style))
                elements.append(Spacer(1, 0.2*inch))
        else:
            elements.append(Paragraph("<i>💡 Astuce: L'image de la carte sera capturée automatiquement à la prochaine sauvegarde du plan.</i>", normal_style))
            elements.append(Spacer(1, 0.3*inch))
        
        # Légendes des symboles avec icônes
        elements.append(Paragraph("<b>📌 Légende des Symboles</b>", heading_style))
        elements.append(Spacer(1, 0.1*inch))
        
        # Grouper les symboles par type pour éviter les répétitions
        symbol_types = {}
        for layer in layers:
            if layer.get('type') == 'symbol':
                props = layer.get('properties', {})
                label = props.get('label', 'Symbole')
                if label not in symbol_types:
                    symbol_types[label] = {
                        'symbol': props.get('symbol', '📍'),
                        'image': props.get('image'),
                        'color': props.get('color', '#6B7280'),
                        'count': 0
                    }
                symbol_types[label]['count'] += 1
        
        # Créer un tableau de légendes avec icônes
        legend_data = [['Icône', 'Type', 'Quantité']]
        for label, info in symbol_types.items():
            # Si c'est une image, essayer de l'afficher
            if info['image']:
                try:
                    # Décoder l'image base64
                    if info['image'].startswith('data:image'):
                        image_data = info['image'].split(',')[1]
                        image_bytes = base64.b64decode(image_data)
                        
                        # Créer une image PIL
                        img = PILImage.open(BytesIO(image_bytes))
                        if img.mode in ('RGBA', 'LA', 'P'):
                            img = img.convert('RGB')
                        
                        # Redimensionner pour la légende
                        img.thumbnail((24, 24))
                        img_buffer = BytesIO()
                        img.save(img_buffer, format='PNG')
                        img_buffer.seek(0)
                        
                        # Créer une image ReportLab
                        icon_display = RLImage(img_buffer, width=24, height=24)
                        legend_data.append([icon_display, label, f"{info['count']}x"])
                    else:
                        legend_data.append([Paragraph(info['symbol'], normal_style), label, f"{info['count']}x"])
                except Exception as e:
                    print(f"Erreur affichage icône: {e}")
                    legend_data.append([Paragraph(info['symbol'], normal_style), label, f"{info['count']}x"])
            else:
                # Afficher l'emoji
                legend_data.append([Paragraph(info['symbol'], normal_style), label, f"{info['count']}x"])
        
        legend_table = Table(legend_data, colWidths=[0.8*inch, 3*inch, 1*inch])
        legend_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')])
        ]))
        elements.append(legend_table)
        elements.append(Spacer(1, 0.4*inch))
    
    # Points d'accès
    points_acces = plan.get('points_acces', [])
    if points_acces and len(points_acces) > 0:
        elements.append(Paragraph(f"<b>📍 Points d'Accès ({len(points_acces)})</b>", heading_style))
        for idx, point in enumerate(points_acces, 1):
            elements.append(Paragraph(f"{idx}. {point.get('description', 'N/A')}", normal_style))
        elements.append(Spacer(1, 0.2*inch))
    else:
        elements.append(Paragraph("<b>📍 Points d'Accès</b>", heading_style))
        elements.append(Paragraph("Aucun point d'accès défini", normal_style))
        elements.append(Spacer(1, 0.2*inch))
    
    # Zones dangereuses
    zones_danger = plan.get('zones_dangereuses', []) or plan.get('zones_danger', [])
    if zones_danger and len(zones_danger) > 0:
        elements.append(Paragraph(f"<b>⚠️ Zones Dangereuses ({len(zones_danger)})</b>", heading_style))
        for idx, zone in enumerate(zones_danger, 1):
            elements.append(Paragraph(f"{idx}. {zone.get('description', 'N/A')}", normal_style))
        elements.append(Spacer(1, 0.2*inch))
    else:
        elements.append(Paragraph("<b>⚠️ Zones Dangereuses</b>", heading_style))
        elements.append(Paragraph("Aucune zone dangereuse identifiée", normal_style))
        elements.append(Spacer(1, 0.2*inch))
    
    # Équipements
    equipements = plan.get('equipements', [])
    if equipements and len(equipements) > 0:
        elements.append(Paragraph(f"<b>🔧 Équipements ({len(equipements)})</b>", heading_style))
        for idx, equip in enumerate(equipements, 1):
            elements.append(Paragraph(f"{idx}. {equip.get('description', 'N/A')}", normal_style))
        elements.append(Spacer(1, 0.2*inch))
    else:
        elements.append(Paragraph("<b>🔧 Équipements</b>", heading_style))
        elements.append(Paragraph("Aucun équipement spécifique", normal_style))
        elements.append(Spacer(1, 0.2*inch))
    
    # Risques identifiés
    risques = plan.get('risques_identifies', [])
    if risques and len(risques) > 0:
        elements.append(Paragraph(f"<b>🔥 Risques Identifiés ({len(risques)})</b>", heading_style))
        for idx, risque in enumerate(risques, 1):
            elements.append(Paragraph(f"{idx}. {risque.get('description', 'N/A')}", normal_style))
        elements.append(Spacer(1, 0.2*inch))
    
    # Commentaires de validation
    if plan.get('commentaires_validation'):
        elements.append(Paragraph("<b>✅ Commentaires de Validation</b>", heading_style))
        elements.append(Paragraph(plan['commentaires_validation'], normal_style))
        elements.append(Spacer(1, 0.2*inch))
    
    # Galerie Photos (nouvelle section)
    photos = plan.get('photos', [])
    if photos and len(photos) > 0:
        try:
            elements.append(PageBreak())
            elements.append(Paragraph(f"<b>📷 Galerie Photos ({len(photos)})</b>", title_style))
            elements.append(Spacer(1, 0.3*inch))
        except Exception as e:
            print(f"Erreur lors de l'ajout du titre galerie photos: {e}")
        
        for idx, photo in enumerate(photos, 1):
            try:
                # En-tête de la photo
                photo_title = photo.get('titre', f'Photo {idx}')
                elements.append(Paragraph(f"<b>{idx}. {photo_title}</b>", heading_style))
                
                # Informations de la photo
                photo_info = []
                if photo.get('categorie'):
                    categorie_labels = {
                        'facade': '🏢 Façade',
                        'entree': '🚪 Entrée',
                        'systeme_alarme': "🚨 Système d'alarme",
                        'points_eau': "💧 Points d'eau",
                        'risques': '⚠️ Risques',
                        'autre': '📷 Autre'
                    }
                    categorie = categorie_labels.get(photo.get('categorie'), photo.get('categorie'))
                    photo_info.append(f"<b>Catégorie:</b> {categorie}")
                
                if photo.get('localisation'):
                    photo_info.append(f"<b>Localisation:</b> {photo.get('localisation')}")
                
                if photo.get('description'):
                    photo_info.append(f"<b>Description:</b> {photo.get('description')}")
                
                if photo_info:
                    info_text = ' | '.join(photo_info)
                    elements.append(Paragraph(info_text, normal_style))
                    elements.append(Spacer(1, 0.1*inch))
                
                # Image (si disponible)
                if photo.get('url'):
                    try:
                        # Gérer les images base64
                        if photo['url'].startswith('data:image'):
                            # Extraire les données base64
                            image_data = photo['url'].split(',')[1]
                            image_bytes = base64.b64decode(image_data)
                            
                            # Ouvrir l'image avec PIL pour la compresser
                            img = PILImage.open(BytesIO(image_bytes))
                            
                            # Convertir en RGB si nécessaire
                            if img.mode in ('RGBA', 'LA', 'P'):
                                img = img.convert('RGB')
                            
                            # Compresser l'image (qualité optimisée 70-80%)
                            img_buffer = BytesIO()
                            img.save(img_buffer, format='JPEG', quality=75, optimize=True)
                            img_buffer.seek(0)
                            
                            # Créer l'image ReportLab avec une largeur maximale de 5 inches
                            rl_image = RLImage(img_buffer, width=5*inch, height=3.5*inch, kind='proportional')
                            elements.append(rl_image)
                    except Exception as img_error:
                        print(f"Erreur chargement image {idx}: {img_error}")
                        elements.append(Paragraph(f"<i>Image non disponible</i>", normal_style))
                
                elements.append(Spacer(1, 0.4*inch))
                
                # Saut de page après 2 photos pour éviter la surcharge
                if idx % 2 == 0 and idx < len(photos):
                    elements.append(PageBreak())
                    
            except Exception as e:
                print(f"Erreur traitement photo {idx}: {e}")
                elements.append(Paragraph(f"<i>Erreur lors du traitement de la photo {idx}</i>", normal_style))
                elements.append(Spacer(1, 0.2*inch))
    
    # Pied de page
    elements.append(Spacer(1, 0.5*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    elements.append(Paragraph(f"<i>Document généré le {datetime.now(timezone.utc).strftime('%d/%m/%Y à %H:%M')} - {tenant.nom}</i>", footer_style))
    
    # Construire le PDF
    doc.build(elements)
    
    # Retourner le PDF
    buffer.seek(0)
    
    # Générer un nom de fichier avec le nom du bâtiment ou l'adresse
    print(f"DEBUG - batiment: {batiment}")
    if batiment:
        print(f"DEBUG - batiment keys: {batiment.keys()}")
        batiment_info = batiment.get('nom_etablissement') or batiment.get('nom') or batiment.get('nom_batiment') or batiment.get('adresse_civique') or batiment.get('adresse') or 'batiment'
    else:
        batiment_info = 'batiment'
    
    print(f"DEBUG - batiment_info: {batiment_info}")
    
    # Nettoyer le nom pour le rendre compatible avec les noms de fichiers
    batiment_safe = batiment_info.replace(' ', '_').replace('/', '-').replace('\\', '-').replace(',', '')
    numero_plan = plan.get('numero_plan', plan_id[:8])
    filename = f"plan_intervention_{numero_plan}_{batiment_safe}.pdf"
    
    print(f"DEBUG - filename généré: {filename}")
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@router.get("/{tenant_slug}/prevention/plans-intervention/{plan_id}/versions")
async def get_versions_plan(
    tenant_slug: str,
    plan_id: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer l'historique des versions d'un plan"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Trouver toutes les versions liées
    versions = []
    
    # Chercher la version actuelle
    plan_actuel = await db.plans_intervention.find_one({"id": plan_id, "tenant_id": tenant.id})
    if not plan_actuel:
        raise HTTPException(status_code=404, detail="Plan non trouvé")
    
    versions.append(clean_mongo_doc(plan_actuel))
    
    # Chercher les versions précédentes
    version_precedente_id = plan_actuel.get("version_precedente_id")
    while version_precedente_id:
        plan_prec = await db.plans_intervention.find_one({"id": version_precedente_id, "tenant_id": tenant.id})
        if plan_prec:
            versions.append(clean_mongo_doc(plan_prec))
            version_precedente_id = plan_prec.get("version_precedente_id")
        else:
            break
    
    # Chercher les versions suivantes
    versions_suivantes = await db.plans_intervention.find({
        "version_precedente_id": plan_id,
        "tenant_id": tenant.id
    }).to_list(length=None)
    
    for v in versions_suivantes:
        versions.append(clean_mongo_doc(v))
    
    return sorted(versions, key=lambda x: x["version"], reverse=True)

@router.post("/{tenant_slug}/prevention/plans-intervention/{plan_id}/calculer-distance")
async def calculer_distance_caserne(
    tenant_slug: str,
    plan_id: str,
    caserne_lat: float = Body(...),
    caserne_lng: float = Body(...),
    current_user: User = Depends(get_current_user)
):
    """Calculer la distance entre la caserne et le bâtiment"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    plan = await db.plans_intervention.find_one({"id": plan_id, "tenant_id": tenant.id})
    if not plan:
        raise HTTPException(status_code=404, detail="Plan non trouvé")
    
    # Calculer la distance en utilisant l'API Google Distance Matrix
    try:
        import requests
        import os
        
        api_key = os.getenv("GOOGLE_MAPS_API_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="Clé API Google Maps non configurée")
        
        url = "https://maps.googleapis.com/maps/api/distancematrix/json"
        params = {
            "origins": f"{caserne_lat},{caserne_lng}",
            "destinations": f"{plan['centre_lat']},{plan['centre_lng']}",
            "key": api_key,
            "mode": "driving"
        }
        
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        
        data = response.json()
        
        if data["status"] == "OK" and len(data["rows"]) > 0:
            element = data["rows"][0]["elements"][0]
            if element["status"] == "OK":
                distance_m = element["distance"]["value"]
                duree_s = element["duration"]["value"]
                
                distance_km = distance_m / 1000.0
                temps_minutes = duree_s // 60
                
                # Mettre à jour le plan
                await db.plans_intervention.update_one(
                    {"id": plan_id, "tenant_id": tenant.id},
                    {"$set": {
                        "distance_caserne_km": distance_km,
                        "distance_caserne_unite": "km",
                        "temps_acces_minutes": temps_minutes,
                        "updated_at": datetime.now(timezone.utc)
                    }}
                )
                
                return {
                    "distance_km": distance_km,
                    "distance_m": distance_m,
                    "temps_acces_minutes": temps_minutes,
                    "message": "Distance calculée avec succès"
                }
        
        raise HTTPException(status_code=404, detail="Impossible de calculer la distance")
    
    except requests.exceptions.RequestException as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors du calcul de distance: {str(e)}")

@router.post("/{tenant_slug}/prevention/plans-intervention/{plan_id}/generer-pdf")
async def generer_pdf_plan(
    tenant_slug: str,
    plan_id: str,
    current_user: User = Depends(get_current_user)
):
    """Générer le PDF d'un plan d'intervention"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    plan = await db.plans_intervention.find_one({"id": plan_id, "tenant_id": tenant.id})
    if not plan:
        raise HTTPException(status_code=404, detail="Plan non trouvé")
    
    # TODO: Implémenter la génération PDF complète avec ReportLab/WeasyPrint
    # Pour l'instant, retourner un placeholder
    
    pdf_url = f"/api/{tenant_slug}/prevention/plans-intervention/{plan_id}/pdf"
    
    await db.plans_intervention.update_one(
        {"id": plan_id, "tenant_id": tenant.id},
        {"$set": {
            "pdf_url": pdf_url,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {
        "pdf_url": pdf_url,
        "message": "Génération PDF programmée (fonctionnalité à compléter)"
    }


# ==================== TEMPLATES PLANS D'INTERVENTION ====================

@router.get("/{tenant_slug}/prevention/plans-intervention/templates")
async def get_templates_plans(
    tenant_slug: str,
    type_batiment: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Récupérer les templates de plans d'intervention"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    query = {"tenant_id": tenant.id, "actif": True}
    
    if type_batiment:
        query["type_batiment"] = type_batiment
    
    templates = await db.templates_plans_intervention.find(query).to_list(length=None)
    
    return [clean_mongo_doc(t) for t in templates]

@router.post("/{tenant_slug}/prevention/plans-intervention/from-template/{template_id}")
async def creer_plan_depuis_template(
    tenant_slug: str,
    template_id: str,
    batiment_id: str = Body(...),
    current_user: User = Depends(get_current_user)
):
    """Créer un nouveau plan à partir d'un template"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Vérifier permissions
    if current_user.role not in ["admin", "superviseur"] and current_user.type_emploi != "preventionniste":
        raise HTTPException(status_code=403, detail="Seuls les préventionnistes peuvent créer des plans")
    
    # Récupérer le template
    template = await db.templates_plans_intervention.find_one({"id": template_id, "tenant_id": tenant.id})
    if not template:
        raise HTTPException(status_code=404, detail="Template non trouvé")
    
    # Vérifier que le bâtiment existe
    batiment = await db.batiments.find_one({"id": batiment_id, "tenant_id": tenant.id})
    if not batiment:
        raise HTTPException(status_code=404, detail="Bâtiment non trouvé")
    
    # Créer le plan basé sur le template
    current_year = datetime.now().year
    count = await db.plans_intervention.count_documents({"tenant_id": tenant.id})
    numero_plan = f"PI-{current_year}-{str(count + 1).zfill(3)}"
    
    # Utiliser les coordonnées du bâtiment si disponibles
    centre_lat = batiment.get("latitude", 45.5017)  # Default Montreal
    centre_lng = batiment.get("longitude", -73.5673)
    
    nouveau_plan = PlanIntervention(
        tenant_id=tenant.id,
        batiment_id=batiment_id,
        numero_plan=numero_plan,
        nom=f"Plan {batiment.get('nom_etablissement', '')}",
        created_by_id=current_user.id,
        centre_lat=centre_lat,
        centre_lng=centre_lng,
        notes_generales=template.get("instructions_utilisation", "")
    )
    
    # Appliquer les éléments par défaut du template
    # TODO: Adapter les positions relatives du template aux coordonnées du bâtiment
    
    await db.plans_intervention.insert_one(nouveau_plan.dict())
    
    return clean_mongo_doc(nouveau_plan.dict())


# ==================== STATISTIQUES PRÉVENTION ====================

@router.get("/{tenant_slug}/prevention/statistiques")
async def get_prevention_statistics(
    tenant_slug: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer les statistiques du module prévention"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Compter les bâtiments
    total_batiments = await db.batiments.count_documents({"tenant_id": tenant.id})
    batiments_avec_preventionniste = await db.batiments.count_documents({
        "tenant_id": tenant.id,
        "preventionniste_assigne_id": {"$exists": True, "$ne": None}
    })
    
    # Compter les inspections
    total_inspections = await db.inspections.count_documents({"tenant_id": tenant.id})
    inspections_conformes = await db.inspections.count_documents({
        "tenant_id": tenant.id,
        "statut_global": "conforme"
    })
    
    # Compter les non-conformités
    total_non_conformites = await db.non_conformites.count_documents({"tenant_id": tenant.id})
    nc_ouvertes = await db.non_conformites.count_documents({
        "tenant_id": tenant.id,
        "statut": {"$in": ["ouverte", "en_cours"]}
    })
    nc_corrigees = await db.non_conformites.count_documents({
        "tenant_id": tenant.id,
        "statut": {"$in": ["corrigee", "fermee"]}
    })
    
    # Récupérer les préventionnistes actifs
    preventionnistes = await db.users.find({
        "tenant_slug": tenant.slug,
        "role": {"$in": ["admin", "superviseur"]}
    }).to_list(100)
    
    preventionnistes_stats = []
    for prev in preventionnistes:
        batiments_assignes = await db.batiments.count_documents({
            "tenant_id": tenant.id,
            "preventionniste_assigne_id": prev["id"]
        })
        inspections_realisees = await db.inspections.count_documents({
            "tenant_id": tenant.id,
            "preventionniste_id": prev["id"]
        })
        
        preventionnistes_stats.append({
            "id": prev["id"],
            "nom": f"{prev.get('prenom', '')} {prev.get('nom', '')}",
            "batiments_assignes": batiments_assignes,
            "inspections_realisees": inspections_realisees
        })
    
    return {
        "batiments": {
            "total": total_batiments,
            "avec_preventionniste": batiments_avec_preventionniste,
            "sans_preventionniste": total_batiments - batiments_avec_preventionniste
        },
        "inspections": {
            "total": total_inspections,
            "conformes": inspections_conformes,
            "non_conformes": total_inspections - inspections_conformes,
            "taux_conformite": round((inspections_conformes / total_inspections * 100) if total_inspections > 0 else 100, 1)
        },
        "non_conformites": {
            "total": total_non_conformites,
            "ouvertes": nc_ouvertes,
            "corrigees": nc_corrigees,
            "taux_resolution": round((nc_corrigees / total_non_conformites * 100) if total_non_conformites > 0 else 100, 1)
        },
        "preventionnistes": preventionnistes_stats
    }


# ==================== RAPPORT BÂTIMENT PDF ====================

@router.get("/{tenant_slug}/prevention/batiments/{batiment_id}/rapport-pdf")
async def export_rapport_batiment_pdf(
    tenant_slug: str,
    batiment_id: str,
    current_user: User = Depends(get_current_user)
):
    """Générer un rapport complet PDF pour un bâtiment"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Récupérer le bâtiment
    batiment = await db.batiments.find_one({"id": batiment_id, "tenant_id": tenant.id})
    if not batiment:
        raise HTTPException(status_code=404, detail="Bâtiment non trouvé")
    
    # Récupérer les inspections du bâtiment
    inspections_cursor = db.inspections.find({
        "tenant_id": tenant.id,
        "batiment_id": batiment_id
    }).sort("date_inspection", -1)
    inspections = await inspections_cursor.to_list(length=None)
    
    # Récupérer le plan d'intervention validé
    plan = await db.plans_intervention.find_one({
        "tenant_id": tenant.id,
        "batiment_id": batiment_id,
        "statut": "valide"
    })
    
    # Récupérer le préventionniste assigné
    preventionniste = None
    if batiment.get("preventionniste_assigne_id"):
        preventionniste = await db.users.find_one({
            "id": batiment["preventionniste_assigne_id"],
            "tenant_id": tenant.id
        })
    
    # Créer le PDF
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    import io
    from PIL import Image as PILImage
    import base64
    
    buffer, doc, story = create_branded_pdf(
        tenant, 
        pagesize=letter, 
        topMargin=0.5*inch, 
        bottomMargin=0.5*inch
    )
    styles = getSampleStyleSheet()
    
    # Style personnalisé
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#2563eb'),
        spaceAfter=10,
        spaceBefore=15
    )
    
    # Titre principal
    story.append(Paragraph(f"RAPPORT DE PRÉVENTION", title_style))
    story.append(Paragraph(f"{batiment.get('nom_etablissement') or batiment.get('adresse_civique')}", styles['Heading2']))
    story.append(Spacer(1, 0.3*inch))
    
    # Section A : Informations du Bâtiment
    story.append(Paragraph("📋 INFORMATIONS DU BÂTIMENT", heading_style))
    
    info_data = [
        ["Adresse", f"{batiment.get('adresse_civique', '')}, {batiment.get('ville', '')}, {batiment.get('province', 'QC')}"],
        ["Type de bâtiment", batiment.get('type_batiment', 'N/A')],
        ["Catégorie", batiment.get('categorie', 'N/A')],
        ["Niveau de risque", batiment.get('niveau_risque', 'N/A')],
        ["Nombre d'occupants", str(batiment.get('nombre_occupants', 'N/A'))],
        ["Valeur foncière", f"{batiment.get('valeur_fonciere', 0):,.2f} $" if batiment.get('valeur_fonciere') else 'N/A'],
        ["Préventionniste assigné", f"{preventionniste['prenom']} {preventionniste['nom']}" if preventionniste else "Non assigné"]
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4.5*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.2*inch))
    
    # Photo du bâtiment si disponible
    if batiment.get('photo_url'):
        try:
            photo_data = batiment['photo_url']
            if photo_data.startswith('data:image'):
                photo_data = photo_data.split(',')[1]
            
            img_data = base64.b64decode(photo_data)
            img = PILImage.open(io.BytesIO(img_data))
            
            # Redimensionner
            max_width = 4 * inch
            max_height = 3 * inch
            img.thumbnail((int(max_width * 2), int(max_height * 2)), PILImage.Resampling.LANCZOS)
            
            img_buffer = io.BytesIO()
            img.save(img_buffer, format='JPEG', quality=85)
            img_buffer.seek(0)
            
            rl_img = RLImage(img_buffer, width=max_width, height=max_height)
            story.append(rl_img)
            story.append(Spacer(1, 0.2*inch))
        except Exception as e:
            print(f"Erreur chargement photo: {e}")
    
    # Section B : Historique des Inspections
    story.append(Paragraph("📜 HISTORIQUE DES INSPECTIONS", heading_style))
    
    if inspections:
        insp_data = [["Date", "Statut", "Non-conformités", "Inspecteur"]]
        
        for insp in inspections[:10]:  # Limiter à 10 dernières
            date_str = insp.get('date_inspection', 'N/A')
            if isinstance(date_str, str):
                try:
                    date_obj = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                    date_str = date_obj.strftime('%Y-%m-%d')
                except:
                    pass
            
            statut = insp.get('statut_conformite', 'N/A')
            nb_nc = len(insp.get('non_conformites', []))
            inspecteur = insp.get('inspecteur_nom', 'N/A')
            
            insp_data.append([date_str, statut, str(nb_nc), inspecteur])
        
        insp_table = Table(insp_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 2*inch])
        insp_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')])
        ]))
        story.append(insp_table)
        
        # Statistiques
        story.append(Spacer(1, 0.15*inch))
        conformes = len([i for i in inspections if i.get('statut_conformite') == 'Conforme'])
        taux = (conformes / len(inspections) * 100) if inspections else 0
        
        stats_text = f"<b>Statistiques :</b> {len(inspections)} inspections | Taux de conformité : {taux:.1f}%"
        story.append(Paragraph(stats_text, styles['Normal']))
    else:
        story.append(Paragraph("Aucune inspection enregistrée pour ce bâtiment.", styles['Normal']))
    
    story.append(Spacer(1, 0.2*inch))
    
    # Section C : Plan d'Intervention
    story.append(Paragraph("🗺️ PLAN D'INTERVENTION", heading_style))
    
    if plan:
        plan_text = f"Plan validé : <b>{plan.get('numero_plan', 'N/A')}</b><br/>"
        plan_text += f"Points d'accès : {len(plan.get('points_acces', []))}<br/>"
        plan_text += f"Zones dangereuses : {len(plan.get('zones_dangereuses', []))}<br/>"
        plan_text += f"Équipements : {len(plan.get('equipements_disponibles', []))}"
        story.append(Paragraph(plan_text, styles['Normal']))
    else:
        story.append(Paragraph("Aucun plan d'intervention validé.", styles['Normal']))
    
    story.append(Spacer(1, 0.2*inch))
    
    # Section D : Recommandations
    story.append(Paragraph("💡 RECOMMANDATIONS", heading_style))
    
    recommandations = []
    
    # Analyse des dernières inspections
    if inspections:
        derniere_insp = inspections[0]
        date_derniere = derniere_insp.get('date_inspection')
        if date_derniere:
            try:
                date_obj = datetime.fromisoformat(date_derniere.replace('Z', '+00:00'))
                jours_depuis = (datetime.now(timezone.utc) - date_obj).days
                
                if jours_depuis > 365:
                    recommandations.append(f"⚠️ Dernière inspection il y a {jours_depuis} jours - Prévoir une nouvelle inspection")
            except:
                pass
        
        if derniere_insp.get('statut_conformite') == 'Non conforme':
            nb_nc = len(derniere_insp.get('non_conformites', []))
            recommandations.append(f"🔴 {nb_nc} non-conformité(s) à corriger en priorité")
    
    if not plan:
        recommandations.append("📋 Créer un plan d'intervention pour ce bâtiment")
    
    if batiment.get('niveau_risque') in ['Élevé', 'Très élevé'] and not preventionniste:
        recommandations.append("👤 Assigner un préventionniste pour le suivi régulier")
    
    if not recommandations:
        recommandations.append("✅ Bâtiment en bon état, poursuivre le suivi régulier")
    
    for reco in recommandations:
        story.append(Paragraph(f"• {reco}", styles['Normal']))
    
    # Footer
    story.append(Spacer(1, 0.3*inch))
    footer_text = f"<i>Rapport généré le {datetime.now().strftime('%Y-%m-%d %H:%M')} par {current_user.prenom} {current_user.nom}</i>"
    story.append(Paragraph(footer_text, styles['Normal']))
    
    # Construire le PDF
    doc.build(story)
    buffer.seek(0)
    
    # Nom du fichier
    filename = f"rapport_{batiment.get('nom_etablissement', 'batiment').replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )


# ==================== EXPORT EXCEL ====================

@router.get("/{tenant_slug}/prevention/export-excel")
async def export_excel_prevention(
    tenant_slug: str,
    type_export: str = "inspections",  # inspections, batiments, non_conformites
    current_user: User = Depends(get_current_user)
):
    """Exporter les données en Excel"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    try:
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if type_export == "inspections":
            ws.title = "Inspections"
            
            # En-têtes
            headers = ["Date", "Bâtiment", "Préventionniste", "Type", "Statut", "Score (%)", "Non-conformités"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Données
            inspections = await db.inspections.find({"tenant_id": tenant.id}).to_list(10000)
            batiments = await db.batiments.find({"tenant_id": tenant.id}).to_list(10000)
            users = await db.users.find({"tenant_slug": tenant.slug}).to_list(10000)
            
            batiments_dict = {b["id"]: b for b in batiments}
            users_dict = {u["id"]: u for u in users}
            
            for idx, insp in enumerate(inspections, 2):
                batiment = batiments_dict.get(insp.get("batiment_id"), {})
                preventionniste = users_dict.get(insp.get("preventionniste_id"), {})
                
                nc_count = await db.non_conformites.count_documents({
                    "inspection_id": insp["id"],
                    "tenant_id": tenant.id
                })
                
                ws.cell(row=idx, column=1, value=insp.get("date_inspection", ""))
                ws.cell(row=idx, column=2, value=batiment.get("nom_etablissement", ""))
                ws.cell(row=idx, column=3, value=f"{preventionniste.get('prenom', '')} {preventionniste.get('nom', '')}")
                ws.cell(row=idx, column=4, value=insp.get("type_inspection", ""))
                ws.cell(row=idx, column=5, value=insp.get("statut_global", ""))
                ws.cell(row=idx, column=6, value=insp.get("score_conformite", 0))
                ws.cell(row=idx, column=7, value=nc_count)
        
        elif type_export == "batiments":
            ws.title = "Bâtiments"
            
            headers = ["Nom", "Adresse", "Ville", "Code Postal", "Groupe Occ.", "Préventionniste", "Nb Inspections"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            batiments = await db.batiments.find({"tenant_id": tenant.id}).to_list(10000)
            users = await db.users.find({"tenant_slug": tenant.slug}).to_list(10000)
            users_dict = {u["id"]: u for u in users}
            
            for idx, bat in enumerate(batiments, 2):
                preventionniste = users_dict.get(bat.get("preventionniste_assigne_id"), {})
                insp_count = await db.inspections.count_documents({
                    "batiment_id": bat["id"],
                    "tenant_id": tenant.id
                })
                
                ws.cell(row=idx, column=1, value=bat.get("nom_etablissement", ""))
                ws.cell(row=idx, column=2, value=bat.get("adresse_civique", ""))
                ws.cell(row=idx, column=3, value=bat.get("ville", ""))
                ws.cell(row=idx, column=4, value=bat.get("code_postal", ""))
                ws.cell(row=idx, column=5, value=bat.get("groupe_occupation", ""))
                ws.cell(row=idx, column=6, value=f"{preventionniste.get('prenom', '')} {preventionniste.get('nom', '')}")
                ws.cell(row=idx, column=7, value=insp_count)
        
        elif type_export == "non_conformites":
            ws.title = "Non-Conformités"
            
            headers = ["Date Détection", "Bâtiment", "Titre", "Description", "Gravité", "Statut", "Délai Correction"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            non_conformites = await db.non_conformites.find({"tenant_id": tenant.id}).to_list(10000)
            batiments = await db.batiments.find({"tenant_id": tenant.id}).to_list(10000)
            batiments_dict = {b["id"]: b for b in batiments}
            
            for idx, nc in enumerate(non_conformites, 2):
                batiment = batiments_dict.get(nc.get("batiment_id"), {})
                
                ws.cell(row=idx, column=1, value=nc.get("created_at", "")[:10])
                ws.cell(row=idx, column=2, value=batiment.get("nom_etablissement", ""))
                ws.cell(row=idx, column=3, value=nc.get("titre", ""))
                ws.cell(row=idx, column=4, value=nc.get("description", ""))
                ws.cell(row=idx, column=5, value=nc.get("gravite", ""))
                ws.cell(row=idx, column=6, value=nc.get("statut", ""))
                ws.cell(row=idx, column=7, value=nc.get("delai_correction", ""))
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=export_{type_export}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur export: {str(e)}")


# ==================== NOTIFICATIONS ====================

@router.get("/{tenant_slug}/prevention/notifications")
async def get_notifications(
    tenant_slug: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer les notifications pour l'utilisateur"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    notifications = []
    today = datetime.now(timezone.utc).date()
    
    # 1. Non-conformités en retard
    non_conformites = await db.non_conformites.find({
        "tenant_id": tenant.id,
        "statut": {"$in": ["ouverte", "en_cours"]}
    }).to_list(1000)
    
    batiments = await db.batiments.find({"tenant_id": tenant.id}).to_list(10000)
    batiments_dict = {b["id"]: b for b in batiments}
    
    for nc in non_conformites:
        if nc.get("delai_correction"):
            try:
                delai_date = datetime.strptime(nc["delai_correction"], "%Y-%m-%d").date()
                days_remaining = (delai_date - today).days
                
                if days_remaining < 0:
                    notifications.append({
                        "id": f"nc_late_{nc['id']}",
                        "type": "nc_retard",
                        "priority": "urgent",
                        "titre": f"Non-conformité en retard",
                        "description": f"{nc.get('titre', 'NC')} au {batiments_dict.get(nc.get('batiment_id'), {}).get('nom_etablissement', 'bâtiment')}",
                        "jours_retard": abs(days_remaining),
                        "link": f"/prevention/non-conformites/{nc['id']}",
                        "date": nc.get("created_at", "")
                    })
                elif days_remaining <= 7:
                    notifications.append({
                        "id": f"nc_soon_{nc['id']}",
                        "type": "nc_echeance_proche",
                        "priority": "high",
                        "titre": f"Échéance proche ({days_remaining}j)",
                        "description": f"{nc.get('titre', 'NC')} au {batiments_dict.get(nc.get('batiment_id'), {}).get('nom_etablissement', 'bâtiment')}",
                        "jours_restants": days_remaining,
                        "link": f"/prevention/non-conformites/{nc['id']}",
                        "date": nc.get("created_at", "")
                    })
            except:
                pass
    
    # 2. Bâtiments sans inspection depuis 6 mois
    six_months_ago = (datetime.now(timezone.utc) - timedelta(days=180)).date().isoformat()
    
    for batiment in batiments:
        last_inspection = await db.inspections.find_one(
            {"batiment_id": batiment["id"], "tenant_id": tenant.id},
            sort=[("date_inspection", -1)]
        )
        
        # Gérer le cas où date_inspection peut être datetime ou string
        last_inspection_date = None
        if last_inspection:
            date_insp = last_inspection.get("date_inspection")
            if isinstance(date_insp, datetime):
                last_inspection_date = date_insp.date().isoformat()
            elif isinstance(date_insp, str):
                last_inspection_date = date_insp
        
        if not last_inspection_date or last_inspection_date < six_months_ago:
            notifications.append({
                "id": f"bat_inspection_{batiment['id']}",
                "type": "inspection_requise",
                "priority": "medium",
                "titre": "Inspection requise",
                "description": f"{batiment.get('nom_etablissement', 'Bâtiment')} - Dernière inspection il y a >6 mois",
                "link": f"/prevention/batiments/{batiment['id']}",
                "date": last_inspection.get("date_inspection", "") if last_inspection else None
            })
    
    # 3. Inspections non-conformes récentes (< 30 jours)
    thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).date().isoformat()
    
    recent_non_conformes = await db.inspections.find({
        "tenant_id": tenant.id,
        "statut_global": {"$ne": "conforme"},
        "date_inspection": {"$gte": thirty_days_ago}
    }).to_list(100)
    
    for inspection in recent_non_conformes:
        nc_count = await db.non_conformites.count_documents({
            "inspection_id": inspection["id"],
            "statut": {"$in": ["ouverte", "en_cours"]}
        })
        
        if nc_count > 0:
            notifications.append({
                "id": f"insp_nc_{inspection['id']}",
                "type": "inspection_nc",
                "priority": "medium",
                "titre": f"{nc_count} NC non résolues",
                "description": f"Inspection du {inspection.get('date_inspection', '')} au {batiments_dict.get(inspection.get('batiment_id'), {}).get('nom_etablissement', 'bâtiment')}",
                "link": f"/prevention/inspections/{inspection['id']}",
                "date": inspection.get("date_inspection", "")
            })
    
    # Trier par priorité
    priority_order = {"urgent": 0, "high": 1, "medium": 2, "low": 3}
    notifications.sort(key=lambda x: priority_order.get(x["priority"], 999))
    
    return {
        "notifications": notifications,
        "count": len(notifications),
        "urgent_count": len([n for n in notifications if n["priority"] == "urgent"]),
        "high_count": len([n for n in notifications if n["priority"] == "high"])
    }


# ==================== RAPPORTS AVANCÉS ====================

@router.get("/{tenant_slug}/prevention/rapports/tendances")
async def get_tendances(
    tenant_slug: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer les tendances sur les 6 derniers mois"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Calculer les 6 derniers mois
    today = datetime.now(timezone.utc)
    months_data = []
    
    for i in range(6):
        month_date = today - timedelta(days=30 * i)
        month_start = month_date.replace(day=1).strftime("%Y-%m-%d")
        
        if month_date.month == 12:
            next_month = month_date.replace(year=month_date.year + 1, month=1, day=1)
        else:
            next_month = month_date.replace(month=month_date.month + 1, day=1)
        month_end = next_month.strftime("%Y-%m-%d")
        
        # Inspections du mois
        inspections_count = await db.inspections.count_documents({
            "tenant_id": tenant.id,
            "date_inspection": {"$gte": month_start, "$lt": month_end}
        })
        
        conformes_count = await db.inspections.count_documents({
            "tenant_id": tenant.id,
            "date_inspection": {"$gte": month_start, "$lt": month_end},
            "statut_global": "conforme"
        })
        
        # Non-conformités du mois
        nc_ouvertes = await db.non_conformites.count_documents({
            "tenant_id": tenant.id,
            "created_at": {"$gte": month_start, "$lt": month_end}
        })
        
        months_data.append({
            "mois": month_date.strftime("%B %Y"),
            "inspections_total": inspections_count,
            "inspections_conformes": conformes_count,
            "taux_conformite": round((conformes_count / inspections_count * 100) if inspections_count > 0 else 0, 1),
            "non_conformites_nouvelles": nc_ouvertes
        })
    
    return {
        "tendances": list(reversed(months_data))
    }
