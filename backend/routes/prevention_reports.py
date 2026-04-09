"""
Routes API pour les Rapports, Statistiques et Exports (Module Prévention)
=========================================================================
Fichier extrait de prevention.py pour améliorer la maintenabilité.

Routes:
- GET  /{tenant_slug}/prevention/inspections/{inspection_id}/rapport-pdf  - PDF inspection
- GET  /{tenant_slug}/prevention/statistiques                            - Statistiques module
- GET  /{tenant_slug}/prevention/batiments/{batiment_id}/rapport-pdf     - PDF bâtiment
- GET  /{tenant_slug}/prevention/export-excel                            - Export Excel
- GET  /{tenant_slug}/prevention/notifications                           - Notifications prévention
- GET  /{tenant_slug}/prevention/rapports/tendances                      - Tendances 6 mois
"""

from fastapi import APIRouter, Depends, HTTPException
from typing import Optional
from datetime import datetime, timezone, timedelta
from io import BytesIO
from starlette.responses import StreamingResponse, Response
import logging
import httpx

from routes.dependencies import (
    db,
    get_current_user,
    get_tenant_from_slug,
    clean_mongo_doc,
    User,
)

from utils.pdf_helpers import (
    create_branded_pdf,
    get_modern_pdf_styles,
)

router = APIRouter(tags=["Prévention - Rapports"])
logger = logging.getLogger(__name__)


# ==================== GÉNÉRATION RAPPORT PDF INSPECTION ====================

async def generer_rapport_inspection_pdf(inspection_id: str, tenant_id: str) -> BytesIO:
    """Générer un rapport PDF pour une inspection"""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    
    # Récupérer le tenant
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    
    # Récupérer l'inspection
    inspection = await db.inspections.find_one({"id": inspection_id, "tenant_id": tenant_id})
    if not inspection:
        raise HTTPException(status_code=404, detail="Inspection non trouvée")
    
    # Récupérer le bâtiment
    batiment = await db.batiments.find_one({"id": inspection["batiment_id"], "tenant_id": tenant_id})
    
    # Récupérer la grille d'inspection
    grille = await db.grilles_inspection.find_one({"id": inspection["grille_inspection_id"], "tenant_id": tenant_id})
    
    # Récupérer le préventionniste
    preventionniste = await db.users.find_one({"id": inspection["preventionniste_id"]})
    
    # Récupérer les non-conformités
    non_conformites = await db.non_conformites.find({
        "inspection_id": inspection_id,
        "tenant_id": tenant_id
    }).to_list(1000)
    
    # Créer le PDF avec branding
    from types import SimpleNamespace
    tenant_obj = SimpleNamespace(**tenant) if tenant else None
    buffer, doc, story = create_branded_pdf(tenant_obj, pagesize=letter)
    styles = getSampleStyleSheet()
    
    # Style personnalisé
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#374151'),
        spaceAfter=12,
        spaceBefore=20
    )
    
    # Titre
    story.append(Paragraph("RAPPORT D'INSPECTION INCENDIE", title_style))
    story.append(Spacer(1, 0.3*inch))
    
    # Informations générales
    story.append(Paragraph("INFORMATIONS GÉNÉRALES", heading_style))
    
    info_data = [
        ["Date d'inspection:", inspection.get("date_inspection", "N/A")],
        ["Type:", inspection.get("type_inspection", "régulière").upper()],
        ["Préventionniste:", f"{preventionniste.get('prenom', '')} {preventionniste.get('nom', '')}" if preventionniste else "N/A"],
        ["Heure début:", inspection.get("heure_debut", "N/A")],
        ["Heure fin:", inspection.get("heure_fin", "N/A")],
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1f2937')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb'))
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Informations bâtiment
    if batiment:
        story.append(Paragraph("INFORMATIONS BÂTIMENT", heading_style))
        
        bat_data = [
            ["Nom établissement:", batiment.get("nom_etablissement", "N/A")],
            ["Adresse:", batiment.get("adresse_civique", "N/A")],
            ["Ville:", batiment.get("ville", "N/A")],
            ["Code postal:", batiment.get("code_postal", "N/A")],
            ["Groupe occupation:", batiment.get("groupe_occupation", "N/A")],
        ]
        
        bat_table = Table(bat_data, colWidths=[2*inch, 4*inch])
        bat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1f2937')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb'))
        ]))
        story.append(bat_table)
        story.append(Spacer(1, 0.3*inch))
    
    # Résultat global
    story.append(Paragraph("RÉSULTAT GLOBAL", heading_style))
    
    statut_color = colors.HexColor('#10b981') if inspection.get("statut_global") == "conforme" else colors.HexColor('#ef4444')
    statut_text = inspection.get("statut_global", "N/A").upper()
    score = inspection.get("score_conformite", 100)
    
    result_data = [
        ["Statut:", statut_text],
        ["Score de conformité:", f"{score}%"],
        ["Non-conformités:", str(len(non_conformites))],
    ]
    
    result_table = Table(result_data, colWidths=[2*inch, 4*inch])
    result_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('BACKGROUND', (1, 0), (1, 0), statut_color),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1f2937')),
        ('TEXTCOLOR', (1, 0), (1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb'))
    ]))
    story.append(result_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Non-conformités
    if non_conformites:
        story.append(Paragraph("NON-CONFORMITÉS IDENTIFIÉES", heading_style))
        
        for idx, nc in enumerate(non_conformites, 1):
            nc_data = [
                [f"#{idx}", ""],
                ["Titre:", nc.get("titre", "N/A")],
                ["Description:", nc.get("description", "N/A")],
                ["Gravité:", nc.get("gravite", "N/A").upper()],
                ["Statut:", nc.get("statut", "N/A")],
                ["Délai correction:", nc.get("delai_correction", "N/A")],
            ]
            
            nc_table = Table(nc_data, colWidths=[2*inch, 4*inch])
            nc_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fef2f2')),
                ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#f3f4f6')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1f2937')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                ('SPAN', (0, 0), (-1, 0)),
            ]))
            story.append(nc_table)
            story.append(Spacer(1, 0.2*inch))
    
    # Notes et recommandations
    if inspection.get("notes_inspection") or inspection.get("recommandations"):
        story.append(Paragraph("NOTES ET RECOMMANDATIONS", heading_style))
        
        if inspection.get("notes_inspection"):
            story.append(Paragraph(f"<b>Notes:</b> {inspection.get('notes_inspection')}", styles['Normal']))
            story.append(Spacer(1, 0.1*inch))
        
        if inspection.get("recommandations"):
            story.append(Paragraph(f"<b>Recommandations:</b> {inspection.get('recommandations')}", styles['Normal']))
            story.append(Spacer(1, 0.3*inch))
    
    # Signature
    story.append(Spacer(1, 0.5*inch))
    story.append(Paragraph("SIGNATURES", heading_style))
    
    sig_data = [
        ["Préventionniste:", "_" * 40],
        ["Date:", "_" * 40],
        ["", ""],
        ["Représentant bâtiment:", "_" * 40],
        ["Nom:", inspection.get("nom_representant", "_" * 40)],
        ["Date:", "_" * 40],
    ]
    
    sig_table = Table(sig_data, colWidths=[2*inch, 4*inch])
    sig_table.setStyle(TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1f2937')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(sig_table)
    
    # Générer le PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


@router.get("/{tenant_slug}/prevention/inspections/{inspection_id}/rapport-pdf")
async def get_inspection_rapport_pdf(
    tenant_slug: str,
    inspection_id: str,
    current_user: User = Depends(get_current_user)
):
    """Générer et télécharger le rapport PDF d'une inspection"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Vérifier que l'inspection existe
    inspection = await db.inspections.find_one({"id": inspection_id, "tenant_id": tenant.id})
    if not inspection:
        raise HTTPException(status_code=404, detail="Inspection non trouvée")
    
    # Générer le PDF
    pdf_buffer = await generer_rapport_inspection_pdf(inspection_id, tenant.id)
    
    # Retourner le PDF
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=rapport_inspection_{inspection_id}.pdf"
        }
    )


# ==================== STATISTIQUES PRÉVENTION ====================

@router.get("/{tenant_slug}/prevention/statistiques")
async def get_prevention_statistics(
    tenant_slug: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer les statistiques du module prévention"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Compter les bâtiments
    total_batiments = await db.batiments.count_documents({"tenant_id": tenant.id})
    batiments_avec_preventionniste = await db.batiments.count_documents({
        "tenant_id": tenant.id,
        "preventionniste_assigne_id": {"$exists": True, "$ne": None}
    })
    
    # Compter les inspections
    total_inspections = await db.inspections.count_documents({"tenant_id": tenant.id})
    inspections_conformes = await db.inspections.count_documents({
        "tenant_id": tenant.id,
        "statut_global": "conforme"
    })
    
    # Compter les non-conformités
    total_non_conformites = await db.non_conformites.count_documents({"tenant_id": tenant.id})
    nc_ouvertes = await db.non_conformites.count_documents({
        "tenant_id": tenant.id,
        "statut": {"$in": ["ouverte", "en_cours"]}
    })
    nc_corrigees = await db.non_conformites.count_documents({
        "tenant_id": tenant.id,
        "statut": {"$in": ["corrigee", "fermee"]}
    })
    
    # Récupérer les préventionnistes actifs
    preventionnistes = await db.users.find({
        "tenant_slug": tenant.slug,
        "role": {"$in": ["admin", "superviseur"]}
    }).to_list(100)
    
    preventionnistes_stats = []
    for prev in preventionnistes:
        batiments_assignes = await db.batiments.count_documents({
            "tenant_id": tenant.id,
            "preventionniste_assigne_id": prev["id"]
        })
        inspections_realisees = await db.inspections.count_documents({
            "tenant_id": tenant.id,
            "preventionniste_id": prev["id"]
        })
        
        preventionnistes_stats.append({
            "id": prev["id"],
            "nom": f"{prev.get('prenom', '')} {prev.get('nom', '')}",
            "batiments_assignes": batiments_assignes,
            "inspections_realisees": inspections_realisees
        })
    
    return {
        "batiments": {
            "total": total_batiments,
            "avec_preventionniste": batiments_avec_preventionniste,
            "sans_preventionniste": total_batiments - batiments_avec_preventionniste
        },
        "inspections": {
            "total": total_inspections,
            "conformes": inspections_conformes,
            "non_conformes": total_inspections - inspections_conformes,
            "taux_conformite": round((inspections_conformes / total_inspections * 100) if total_inspections > 0 else 100, 1)
        },
        "non_conformites": {
            "total": total_non_conformites,
            "ouvertes": nc_ouvertes,
            "corrigees": nc_corrigees,
            "taux_resolution": round((nc_corrigees / total_non_conformites * 100) if total_non_conformites > 0 else 100, 1)
        },
        "preventionnistes": preventionnistes_stats
    }


# ==================== RAPPORT BÂTIMENT PDF ====================

@router.get("/{tenant_slug}/prevention/batiments/{batiment_id}/rapport-pdf")
async def export_rapport_batiment_pdf(
    tenant_slug: str,
    batiment_id: str,
    current_user: User = Depends(get_current_user)
):
    """Générer un rapport complet PDF pour un bâtiment"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Récupérer le bâtiment
    batiment = await db.batiments.find_one({"id": batiment_id, "tenant_id": tenant.id})
    if not batiment:
        raise HTTPException(status_code=404, detail="Bâtiment non trouvé")
    
    # Récupérer les inspections du bâtiment
    inspections_cursor = db.inspections.find({
        "tenant_id": tenant.id,
        "batiment_id": batiment_id
    }).sort("date_inspection", -1)
    inspections = await inspections_cursor.to_list(length=None)
    
    # Récupérer le plan d'intervention validé
    plan = await db.plans_intervention.find_one({
        "tenant_id": tenant.id,
        "batiment_id": batiment_id,
        "statut": "valide"
    })
    
    # Récupérer le préventionniste assigné
    preventionniste = None
    if batiment.get("preventionniste_assigne_id"):
        preventionniste = await db.users.find_one({
            "id": batiment["preventionniste_assigne_id"],
            "tenant_id": tenant.id
        })
    
    # Créer le PDF
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer, Image as RLImage, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    import io
    from PIL import Image as PILImage
    import base64
    
    buffer, doc, story = create_branded_pdf(
        tenant, 
        pagesize=letter, 
        topMargin=0.5*inch, 
        bottomMargin=0.5*inch
    )
    styles = getSampleStyleSheet()
    
    # Style personnalisé
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#2563eb'),
        spaceAfter=10,
        spaceBefore=15
    )
    
    # Titre principal
    story.append(Paragraph(f"RAPPORT DE PRÉVENTION", title_style))
    story.append(Paragraph(f"{batiment.get('nom_etablissement') or batiment.get('adresse_civique')}", styles['Heading2']))
    story.append(Spacer(1, 0.3*inch))
    
    # Section A : Informations du Bâtiment
    story.append(Paragraph("INFORMATIONS DU BÂTIMENT", heading_style))
    
    info_data = [
        ["Adresse", f"{batiment.get('adresse_civique', '')}, {batiment.get('ville', '')}, {batiment.get('province', 'QC')}"],
        ["Type de bâtiment", batiment.get('type_batiment', 'N/A')],
        ["Catégorie", batiment.get('categorie', 'N/A')],
        ["Niveau de risque", batiment.get('niveau_risque', 'N/A')],
        ["Nombre d'occupants", str(batiment.get('nombre_occupants', 'N/A'))],
        ["Valeur foncière", f"{batiment.get('valeur_fonciere', 0):,.2f} $" if batiment.get('valeur_fonciere') else 'N/A'],
        ["Préventionniste assigné", f"{preventionniste['prenom']} {preventionniste['nom']}" if preventionniste else "Non assigné"]
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4.5*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.2*inch))
    
    # Photo du bâtiment si disponible (Azure ou legacy base64)
    photo_loaded = False
    if batiment.get('photo_blob_name'):
        try:
            from services.azure_storage import get_object
            img_data, _ = get_object(batiment['photo_blob_name'])
            img = PILImage.open(io.BytesIO(img_data))
            
            max_width = 4 * inch
            max_height = 3 * inch
            img.thumbnail((int(max_width * 2), int(max_height * 2)), PILImage.Resampling.LANCZOS)
            
            img_buffer = io.BytesIO()
            img.save(img_buffer, format='JPEG', quality=85)
            img_buffer.seek(0)
            
            rl_img = RLImage(img_buffer, width=max_width, height=max_height)
            story.append(rl_img)
            story.append(Spacer(1, 0.2*inch))
            photo_loaded = True
        except Exception as e:
            logger.error(f"Erreur chargement photo Azure: {e}")
    
    if not photo_loaded and batiment.get('photo_url'):
        try:
            photo_data = batiment['photo_url']
            if photo_data.startswith('data:image'):
                photo_data = photo_data.split(',')[1]
            
            img_data = base64.b64decode(photo_data)
            img = PILImage.open(io.BytesIO(img_data))
            
            max_width = 4 * inch
            max_height = 3 * inch
            img.thumbnail((int(max_width * 2), int(max_height * 2)), PILImage.Resampling.LANCZOS)
            
            img_buffer = io.BytesIO()
            img.save(img_buffer, format='JPEG', quality=85)
            img_buffer.seek(0)
            
            rl_img = RLImage(img_buffer, width=max_width, height=max_height)
            story.append(rl_img)
            story.append(Spacer(1, 0.2*inch))
        except Exception as e:
            logger.error(f"Erreur chargement photo: {e}")
    
    # Section Galerie Photos (photos annotées)
    gallery_photos = batiment.get('photos', [])
    if gallery_photos:
        story.append(Paragraph("GALERIE PHOTOS", heading_style))
        
        caption_style = ParagraphStyle(
            'PhotoCaption',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#4b5563'),
            alignment=TA_CENTER,
            spaceAfter=15,
            spaceBefore=4,
            italic=True
        )
        
        photo_max_w = 4.5 * inch
        photo_max_h = 3.5 * inch
        
        async with httpx.AsyncClient(timeout=15.0) as client:
            for idx, photo in enumerate(gallery_photos):
                photo_url = photo.get('url', '')
                if not photo_url:
                    continue
                try:
                    # Télécharger l'image depuis l'URL
                    resp = await client.get(photo_url)
                    if resp.status_code != 200:
                        continue
                    
                    pil_img = PILImage.open(io.BytesIO(resp.content))
                    orig_w, orig_h = pil_img.size
                    
                    # Calculer les dimensions proportionnelles
                    ratio = min(photo_max_w / orig_w, photo_max_h / orig_h)
                    display_w = orig_w * ratio
                    display_h = orig_h * ratio
                    
                    img_buf = io.BytesIO()
                    pil_img.save(img_buf, format='JPEG', quality=80)
                    img_buf.seek(0)
                    
                    rl_photo = RLImage(img_buf, width=display_w, height=display_h)
                    rl_photo.hAlign = 'CENTER'
                    story.append(rl_photo)
                    
                    # Légende
                    legende = photo.get('legende', '').strip()
                    label = legende if legende else f"Photo {idx + 1}"
                    story.append(Paragraph(label, caption_style))
                    
                except Exception as e:
                    logger.warning(f"Impossible de charger la photo {photo.get('id', '?')}: {e}")
                    continue
        
        story.append(Spacer(1, 0.2*inch))
    
    # Section B : Historique des Inspections
    story.append(Paragraph("HISTORIQUE DES INSPECTIONS", heading_style))
    
    if inspections:
        insp_data = [["Date", "Statut", "Non-conformités", "Inspecteur"]]
        
        for insp in inspections[:10]:  # Limiter à 10 dernières
            date_str = insp.get('date_inspection', 'N/A')
            if isinstance(date_str, str):
                try:
                    date_obj = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                    date_str = date_obj.strftime('%Y-%m-%d')
                except Exception:
                    pass
            
            statut = insp.get('statut_conformite', 'N/A')
            nb_nc = len(insp.get('non_conformites', []))
            inspecteur = insp.get('inspecteur_nom', 'N/A')
            
            insp_data.append([date_str, statut, str(nb_nc), inspecteur])
        
        insp_table = Table(insp_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 2*inch])
        insp_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')])
        ]))
        story.append(insp_table)
        
        # Statistiques
        story.append(Spacer(1, 0.15*inch))
        conformes = len([i for i in inspections if i.get('statut_conformite') == 'Conforme'])
        taux = (conformes / len(inspections) * 100) if inspections else 0
        
        stats_text = f"<b>Statistiques :</b> {len(inspections)} inspections | Taux de conformité : {taux:.1f}%"
        story.append(Paragraph(stats_text, styles['Normal']))
    else:
        story.append(Paragraph("Aucune inspection enregistrée pour ce bâtiment.", styles['Normal']))
    
    story.append(Spacer(1, 0.2*inch))
    
    # Section C : Plan d'Intervention
    story.append(Paragraph("PLAN D'INTERVENTION", heading_style))
    
    if plan:
        plan_text = f"Plan validé : <b>{plan.get('numero_plan', 'N/A')}</b><br/>"
        plan_text += f"Points d'accès : {len(plan.get('points_acces', []))}<br/>"
        plan_text += f"Zones dangereuses : {len(plan.get('zones_dangereuses', []))}<br/>"
        plan_text += f"Équipements : {len(plan.get('equipements_disponibles', []))}"
        story.append(Paragraph(plan_text, styles['Normal']))
    else:
        story.append(Paragraph("Aucun plan d'intervention validé.", styles['Normal']))
    
    story.append(Spacer(1, 0.2*inch))
    
    # Section D : Recommandations
    story.append(Paragraph("RECOMMANDATIONS", heading_style))
    
    recommandations = []
    
    # Analyse des dernières inspections
    if inspections:
        derniere_insp = inspections[0]
        date_derniere = derniere_insp.get('date_inspection')
        if date_derniere:
            try:
                date_obj = datetime.fromisoformat(date_derniere.replace('Z', '+00:00'))
                jours_depuis = (datetime.now(timezone.utc) - date_obj).days
                
                if jours_depuis > 365:
                    recommandations.append(f"Dernière inspection il y a {jours_depuis} jours - Prévoir une nouvelle inspection")
            except Exception:
                pass
        
        if derniere_insp.get('statut_conformite') == 'Non conforme':
            nb_nc = len(derniere_insp.get('non_conformites', []))
            recommandations.append(f"{nb_nc} non-conformité(s) à corriger en priorité")
    
    if not plan:
        recommandations.append("Créer un plan d'intervention pour ce bâtiment")
    
    if batiment.get('niveau_risque') in ['Élevé', 'Très élevé'] and not preventionniste:
        recommandations.append("Assigner un préventionniste pour le suivi régulier")
    
    if not recommandations:
        recommandations.append("Bâtiment en bon état, poursuivre le suivi régulier")
    
    for reco in recommandations:
        story.append(Paragraph(f"- {reco}", styles['Normal']))
    
    # Footer
    story.append(Spacer(1, 0.3*inch))
    footer_text = f"<i>Rapport généré le {datetime.now().strftime('%Y-%m-%d %H:%M')} par {current_user.prenom} {current_user.nom}</i>"
    story.append(Paragraph(footer_text, styles['Normal']))
    
    # Construire le PDF
    doc.build(story)
    buffer.seek(0)
    
    # Nom du fichier
    filename = f"rapport_{batiment.get('nom_etablissement', 'batiment').replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )


# ==================== EXPORT EXCEL ====================

@router.get("/{tenant_slug}/prevention/export-excel")
async def export_excel_prevention(
    tenant_slug: str,
    type_export: str = "inspections",  # inspections, batiments, non_conformites
    current_user: User = Depends(get_current_user)
):
    """Exporter les données en Excel"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if type_export == "inspections":
            ws.title = "Inspections"
            
            # En-têtes
            headers = ["Date", "Bâtiment", "Préventionniste", "Type", "Statut", "Score (%)", "Non-conformités"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Données
            inspections = await db.inspections.find({"tenant_id": tenant.id}).to_list(10000)
            batiments = await db.batiments.find({"tenant_id": tenant.id}).to_list(10000)
            users = await db.users.find({"tenant_slug": tenant.slug}).to_list(10000)
            
            batiments_dict = {b["id"]: b for b in batiments}
            users_dict = {u["id"]: u for u in users}
            
            for idx, insp in enumerate(inspections, 2):
                bat = batiments_dict.get(insp.get("batiment_id"), {})
                prev = users_dict.get(insp.get("preventionniste_id"), {})
                
                nc_count = await db.non_conformites.count_documents({
                    "inspection_id": insp["id"],
                    "tenant_id": tenant.id
                })
                
                ws.cell(row=idx, column=1, value=insp.get("date_inspection", ""))
                ws.cell(row=idx, column=2, value=bat.get("nom_etablissement", ""))
                ws.cell(row=idx, column=3, value=f"{prev.get('prenom', '')} {prev.get('nom', '')}")
                ws.cell(row=idx, column=4, value=insp.get("type_inspection", ""))
                ws.cell(row=idx, column=5, value=insp.get("statut_global", ""))
                ws.cell(row=idx, column=6, value=insp.get("score_conformite", 0))
                ws.cell(row=idx, column=7, value=nc_count)
        
        elif type_export == "batiments":
            ws.title = "Bâtiments"
            
            headers = ["Nom", "Adresse", "Ville", "Code Postal", "Groupe Occ.", "Préventionniste", "Nb Inspections"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            batiments = await db.batiments.find({"tenant_id": tenant.id}).to_list(10000)
            users = await db.users.find({"tenant_slug": tenant.slug}).to_list(10000)
            users_dict = {u["id"]: u for u in users}
            
            for idx, bat in enumerate(batiments, 2):
                prev = users_dict.get(bat.get("preventionniste_assigne_id"), {})
                insp_count = await db.inspections.count_documents({
                    "batiment_id": bat["id"],
                    "tenant_id": tenant.id
                })
                
                ws.cell(row=idx, column=1, value=bat.get("nom_etablissement", ""))
                ws.cell(row=idx, column=2, value=bat.get("adresse_civique", ""))
                ws.cell(row=idx, column=3, value=bat.get("ville", ""))
                ws.cell(row=idx, column=4, value=bat.get("code_postal", ""))
                ws.cell(row=idx, column=5, value=bat.get("groupe_occupation", ""))
                ws.cell(row=idx, column=6, value=f"{prev.get('prenom', '')} {prev.get('nom', '')}")
                ws.cell(row=idx, column=7, value=insp_count)
        
        elif type_export == "non_conformites":
            ws.title = "Non-Conformités"
            
            headers = ["Date Détection", "Bâtiment", "Titre", "Description", "Gravité", "Statut", "Délai Correction"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            non_conformites = await db.non_conformites.find({"tenant_id": tenant.id}).to_list(10000)
            batiments = await db.batiments.find({"tenant_id": tenant.id}).to_list(10000)
            batiments_dict = {b["id"]: b for b in batiments}
            
            for idx, nc in enumerate(non_conformites, 2):
                bat = batiments_dict.get(nc.get("batiment_id"), {})
                
                ws.cell(row=idx, column=1, value=nc.get("created_at", "")[:10])
                ws.cell(row=idx, column=2, value=bat.get("nom_etablissement", ""))
                ws.cell(row=idx, column=3, value=nc.get("titre", ""))
                ws.cell(row=idx, column=4, value=nc.get("description", ""))
                ws.cell(row=idx, column=5, value=nc.get("gravite", ""))
                ws.cell(row=idx, column=6, value=nc.get("statut", ""))
                ws.cell(row=idx, column=7, value=nc.get("delai_correction", ""))
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=export_{type_export}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur export: {str(e)}")


# ==================== NOTIFICATIONS ====================

@router.get("/{tenant_slug}/prevention/notifications")
async def get_notifications(
    tenant_slug: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer les notifications pour l'utilisateur"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    notifications = []
    today = datetime.now(timezone.utc).date()
    
    # 1. Non-conformités en retard
    non_conformites = await db.non_conformites.find({
        "tenant_id": tenant.id,
        "statut": {"$in": ["ouverte", "en_cours"]}
    }).to_list(1000)
    
    batiments = await db.batiments.find({"tenant_id": tenant.id}).to_list(10000)
    batiments_dict = {b["id"]: b for b in batiments}
    
    for nc in non_conformites:
        if nc.get("delai_correction"):
            try:
                delai_date = datetime.strptime(nc["delai_correction"], "%Y-%m-%d").date()
                days_remaining = (delai_date - today).days
                
                if days_remaining < 0:
                    notifications.append({
                        "id": f"nc_late_{nc['id']}",
                        "type": "nc_retard",
                        "priority": "urgent",
                        "titre": f"Non-conformité en retard",
                        "description": f"{nc.get('titre', 'NC')} au {batiments_dict.get(nc.get('batiment_id'), {}).get('nom_etablissement', 'bâtiment')}",
                        "jours_retard": abs(days_remaining),
                        "link": f"/prevention/non-conformites/{nc['id']}",
                        "date": nc.get("created_at", "")
                    })
                elif days_remaining <= 7:
                    notifications.append({
                        "id": f"nc_soon_{nc['id']}",
                        "type": "nc_echeance_proche",
                        "priority": "high",
                        "titre": f"Échéance proche ({days_remaining}j)",
                        "description": f"{nc.get('titre', 'NC')} au {batiments_dict.get(nc.get('batiment_id'), {}).get('nom_etablissement', 'bâtiment')}",
                        "jours_restants": days_remaining,
                        "link": f"/prevention/non-conformites/{nc['id']}",
                        "date": nc.get("created_at", "")
                    })
            except Exception:
                pass
    
    # 2. Bâtiments sans inspection depuis 6 mois
    six_months_ago = (datetime.now(timezone.utc) - timedelta(days=180)).date().isoformat()
    
    for batiment in batiments:
        last_inspection = await db.inspections.find_one(
            {"batiment_id": batiment["id"], "tenant_id": tenant.id},
            sort=[("date_inspection", -1)]
        )
        
        # Gérer le cas où date_inspection peut être datetime ou string
        last_inspection_date = None
        if last_inspection:
            date_insp = last_inspection.get("date_inspection")
            if isinstance(date_insp, datetime):
                last_inspection_date = date_insp.date().isoformat()
            elif isinstance(date_insp, str):
                last_inspection_date = date_insp
        
        if not last_inspection_date or last_inspection_date < six_months_ago:
            notifications.append({
                "id": f"bat_inspection_{batiment['id']}",
                "type": "inspection_requise",
                "priority": "medium",
                "titre": "Inspection requise",
                "description": f"{batiment.get('nom_etablissement', 'Bâtiment')} - Dernière inspection il y a >6 mois",
                "link": f"/prevention/batiments/{batiment['id']}",
                "date": last_inspection.get("date_inspection", "") if last_inspection else None
            })
    
    # 3. Inspections non-conformes récentes (< 30 jours)
    thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).date().isoformat()
    
    recent_non_conformes = await db.inspections.find({
        "tenant_id": tenant.id,
        "statut_global": {"$ne": "conforme"},
        "date_inspection": {"$gte": thirty_days_ago}
    }).to_list(100)
    
    for inspection in recent_non_conformes:
        nc_count = await db.non_conformites.count_documents({
            "inspection_id": inspection["id"],
            "statut": {"$in": ["ouverte", "en_cours"]}
        })
        
        if nc_count > 0:
            notifications.append({
                "id": f"insp_nc_{inspection['id']}",
                "type": "inspection_nc",
                "priority": "medium",
                "titre": f"{nc_count} NC non résolues",
                "description": f"Inspection du {inspection.get('date_inspection', '')} au {batiments_dict.get(inspection.get('batiment_id'), {}).get('nom_etablissement', 'bâtiment')}",
                "link": f"/prevention/inspections/{inspection['id']}",
                "date": inspection.get("date_inspection", "")
            })
    
    # Trier par priorité
    priority_order = {"urgent": 0, "high": 1, "medium": 2, "low": 3}
    notifications.sort(key=lambda x: priority_order.get(x["priority"], 999))
    
    return {
        "notifications": notifications,
        "count": len(notifications),
        "urgent_count": len([n for n in notifications if n["priority"] == "urgent"]),
        "high_count": len([n for n in notifications if n["priority"] == "high"])
    }


# ==================== RAPPORTS AVANCÉS ====================

@router.get("/{tenant_slug}/prevention/rapports/tendances")
async def get_tendances(
    tenant_slug: str,
    current_user: User = Depends(get_current_user)
):
    """Récupérer les tendances sur les 6 derniers mois"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not tenant.parametres.get('module_prevention_active', False):
        raise HTTPException(status_code=403, detail="Module prévention non activé")
    
    # Calculer les 6 derniers mois
    today = datetime.now(timezone.utc)
    months_data = []
    
    for i in range(6):
        month_date = today - timedelta(days=30 * i)
        month_start = month_date.replace(day=1).strftime("%Y-%m-%d")
        
        if month_date.month == 12:
            next_month = month_date.replace(year=month_date.year + 1, month=1, day=1)
        else:
            next_month = month_date.replace(month=month_date.month + 1, day=1)
        month_end = next_month.strftime("%Y-%m-%d")
        
        # Inspections du mois
        inspections_count = await db.inspections.count_documents({
            "tenant_id": tenant.id,
            "date_inspection": {"$gte": month_start, "$lt": month_end}
        })
        
        conformes_count = await db.inspections.count_documents({
            "tenant_id": tenant.id,
            "date_inspection": {"$gte": month_start, "$lt": month_end},
            "statut_global": "conforme"
        })
        
        # Non-conformités du mois
        nc_ouvertes = await db.non_conformites.count_documents({
            "tenant_id": tenant.id,
            "created_at": {"$gte": month_start, "$lt": month_end}
        })
        
        months_data.append({
            "mois": month_date.strftime("%B %Y"),
            "inspections_total": inspections_count,
            "inspections_conformes": conformes_count,
            "taux_conformite": round((conformes_count / inspections_count * 100) if inspections_count > 0 else 0, 1),
            "non_conformites_nouvelles": nc_ouvertes
        })
    
    return {
        "tendances": list(reversed(months_data))
    }
