"""
Routes API pour le module Disponibilités
=========================================

STATUT: ACTIF
Ce module gère les disponibilités des pompiers.

Routes de base (les routes avancées restent dans server.py pour l'instant):
- GET    /{tenant_slug}/disponibilites/{user_id}  - Disponibilités d'un utilisateur
- POST   /{tenant_slug}/disponibilites            - Créer une disponibilité
- PUT    /{tenant_slug}/disponibilites/{user_id}  - Modifier disponibilités
- GET    /{tenant_slug}/disponibilites/statut-blocage - Statut de blocage
"""

from fastapi import APIRouter, Depends, HTTPException, Body
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, date
import uuid
import logging

# Import des dépendances partagées
from routes.dependencies import (
    db,
    get_current_user,
    get_tenant_from_slug,
    clean_mongo_doc,
    User
)

router = APIRouter(tags=["Disponibilités"])


# ==================== MODÈLES ====================

class Disponibilite(BaseModel):
    """Modèle disponibilité"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    tenant_id: str
    date: str  # Format YYYY-MM-DD
    type_garde_id: Optional[str] = None
    heure_debut: str = "00:00"
    heure_fin: str = "23:59"
    statut: str = "disponible"  # "disponible" ou "indisponible"
    est_disponible: bool = True  # Garder pour compatibilité
    commentaire: Optional[str] = None
    origine: Optional[str] = "manuelle"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    
    class Config:
        extra = "allow"


class DisponibiliteCreate(BaseModel):
    """Modèle pour la création d'une disponibilité"""
    user_id: Optional[str] = None  # Optionnel car peut être fourni par current_user
    date: str
    type_garde_id: Optional[str] = None
    heure_debut: str = "00:00"
    heure_fin: str = "23:59"
    statut: str = "disponible"  # "disponible" ou "indisponible"
    est_disponible: bool = True  # Garder pour compatibilité
    commentaire: Optional[str] = None
    origine: Optional[str] = "manuelle"  # "manuelle", "recurrence", etc.


class DisponibiliteReinitialiser(BaseModel):
    """Modèle pour la réinitialisation des disponibilités"""
    user_id: str
    periode: str  # "semaine", "mois", "annee", "personnalisee"
    mode: str  # "tout" ou "generees_seulement"
    type_entree: str = "les_deux"  # "disponibilites", "indisponibilites", "les_deux"
    date_debut: Optional[str] = None
    date_fin: Optional[str] = None


class ConflictResolution(BaseModel):
    """Historique des résolutions de conflits disponibilité/indisponibilité"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    user_id: str
    affected_user_id: str
    action: str
    type_created: str
    conflicts_deleted: List[Dict[str, Any]] = []
    created_item: Optional[Dict[str, Any]] = None
    resolved_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class ConflictDetail(BaseModel):
    """Détail d'un conflit"""
    type: str
    date: str
    id: str


class DisponibiliteImportCSV(BaseModel):
    """Modèle pour l'import CSV"""
    csv_content: str
    mois: str


# ==================== ROUTES ====================

@router.get("/{tenant_slug}/disponibilites/statut-blocage")
async def get_statut_blocage(
    tenant_slug: str,
    current_user: User = Depends(get_current_user)
):
    """Retourne le statut de blocage des soumissions"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if current_user.tenant_id != tenant.id:
        raise HTTPException(status_code=403, detail="Accès interdit à cette caserne")
    
    params = await db.parametres_disponibilites.find_one({"tenant_id": tenant.id})
    
    if not params:
        return {
            "bloque": False,
            "message": "Soumissions ouvertes",
            "deadline": None
        }
    
    today = date.today()
    jour_fin = params.get("jour_fin_soumission", 25)
    bloque = params.get("bloquer_apres_deadline", False) and today.day > jour_fin
    
    return {
        "bloque": bloque,
        "message": f"Deadline: jour {jour_fin} du mois",
        "deadline": jour_fin,
        "jour_actuel": today.day
    }


@router.get("/{tenant_slug}/disponibilites/{user_id}", response_model=List[Disponibilite])
async def get_disponibilites_user(
    tenant_slug: str,
    user_id: str,
    mois: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Récupère les disponibilités d'un utilisateur"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if current_user.tenant_id != tenant.id:
        raise HTTPException(status_code=403, detail="Accès interdit à cette caserne")
    
    # Vérifier permissions
    if current_user.role not in ["admin", "superviseur"] and current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    # Construire le filtre
    filter_query = {"user_id": user_id, "tenant_id": tenant.id}
    
    if mois:
        filter_query["date"] = {"$regex": f"^{mois}"}
    
    dispos = await db.disponibilites.find(filter_query).to_list(1000)
    cleaned_dispos = [clean_mongo_doc(d) for d in dispos]
    
    return [Disponibilite(**d) for d in cleaned_dispos]


@router.post("/{tenant_slug}/disponibilites", response_model=Disponibilite)
async def create_disponibilite(
    tenant_slug: str,
    dispo: DisponibiliteCreate,
    current_user: User = Depends(get_current_user)
):
    """Créer une nouvelle disponibilité"""
    import logging
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if current_user.tenant_id != tenant.id:
        raise HTTPException(status_code=403, detail="Accès interdit à cette caserne")
    
    # Déterminer le user_id cible
    target_user_id = dispo.user_id if dispo.user_id else current_user.id
    
    # Si l'admin/superviseur crée pour quelqu'un d'autre, vérifier les permissions
    if target_user_id != current_user.id:
        if current_user.role not in ["admin", "superviseur"]:
            raise HTTPException(
                status_code=403,
                detail="Vous ne pouvez pas créer des disponibilités pour un autre utilisateur"
            )
        # Vérifier que l'utilisateur cible existe dans le même tenant
        target_user = await db.users.find_one({"id": target_user_id, "tenant_id": tenant.id})
        if not target_user:
            raise HTTPException(status_code=404, detail="Utilisateur cible non trouvé")
    
    # Vérifier si l'utilisateur peut soumettre (deadline) - seulement pour les non-admins
    if current_user.role not in ["admin", "superviseur"]:
        params = await db.parametres_disponibilites.find_one({"tenant_id": tenant.id})
        if params and params.get("bloquer_apres_deadline"):
            today = date.today()
            jour_fin = params.get("jour_fin_soumission", 25)
            
            if today.day > jour_fin:
                raise HTTPException(
                    status_code=400,
                    detail=f"La période de soumission est terminée (deadline: jour {jour_fin})"
                )
    
    # Créer la disponibilité
    dispo_dict = dispo.dict()
    dispo_dict["id"] = str(uuid.uuid4())
    dispo_dict["user_id"] = target_user_id
    dispo_dict["tenant_id"] = tenant.id
    dispo_dict["created_at"] = datetime.now(timezone.utc)
    
    logging.info(f"📅 [DISPOS] POST - Création dispo pour user={target_user_id}, date={dispo.date}, statut={dispo.statut}")
    
    await db.disponibilites.insert_one(dispo_dict)
    
    # Supprimer _id ajouté par MongoDB avant de retourner
    dispo_dict.pop("_id", None)
    
    return Disponibilite(**dispo_dict)


@router.put("/{tenant_slug}/disponibilites/{user_id}")
async def update_disponibilites(
    tenant_slug: str,
    user_id: str,
    disponibilites: List[Dict] = Body(...),
    current_user: User = Depends(get_current_user)
):
    """Met à jour les disponibilités d'un utilisateur"""
    import logging
    logging.info(f"📅 [DISPOS] PUT reçu pour user_id={user_id}, {len(disponibilites)} disponibilités")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if current_user.tenant_id != tenant.id:
        raise HTTPException(status_code=403, detail="Accès interdit à cette caserne")
    
    # Vérification de sécurité côté serveur
    if current_user.id != user_id and current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(
            status_code=403,
            detail="Vous ne pouvez pas modifier les disponibilités d'un autre utilisateur"
        )
    
    # Vérifier que l'utilisateur modifié appartient au même tenant
    target_user = await db.users.find_one({"id": user_id, "tenant_id": tenant.id})
    if not target_user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    logging.info(f"📅 [DISPOS] Utilisateur cible: {target_user.get('prenom')} {target_user.get('nom')}")
    
    # Supprimer les anciennes disponibilités pour ce mois
    if disponibilites and len(disponibilites) > 0:
        mois = disponibilites[0].get("date", "")[:7]  # Format YYYY-MM
        delete_result = await db.disponibilites.delete_many({
            "user_id": user_id,
            "tenant_id": tenant.id,
            "date": {"$regex": f"^{mois}"}
        })
        logging.info(f"📅 [DISPOS] Supprimées {delete_result.deleted_count} anciennes dispos du mois {mois}")
    
    # Insérer les nouvelles
    inserted_count = 0
    for dispo in disponibilites:
        dispo["id"] = dispo.get("id") or str(uuid.uuid4())
        dispo["user_id"] = user_id
        dispo["tenant_id"] = tenant.id
        dispo["updated_at"] = datetime.now(timezone.utc)
        await db.disponibilites.insert_one(dispo)
        inserted_count += 1
    
    logging.info(f"📅 [DISPOS] Insérées {inserted_count} nouvelles dispos")
    
    return {"message": f"{len(disponibilites)} disponibilités mises à jour"}


# ==================== ROUTES MIGRÉES DE SERVER.PY ====================

# GET rapports/disponibilite
@router.get("/{tenant_slug}/rapports/disponibilite")
async def get_rapport_disponibilite(
    tenant_slug: str,
    date_debut: str,
    date_fin: str,
    current_user: User = Depends(get_current_user)
):
    """Rapport de disponibilité/indisponibilité des pompiers"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Convertir dates
    date_debut_dt = datetime.fromisoformat(date_debut)
    date_fin_dt = datetime.fromisoformat(date_fin)
    
    # Récupérer les données
    users = await db.users.find({"tenant_id": tenant.id}).to_list(1000)
    disponibilites = await db.disponibilites.find({"tenant_id": tenant.id}).to_list(10000)
    
    rapport = []
    total_jours_disponibles = 0
    total_jours_indisponibles = 0
    
    for user in users:
        # Filtrer disponibilités par période
        user_disponibilites = []
        for dispo in disponibilites:
            if dispo["user_id"] == user["id"] and "date" in dispo:
                try:
                    date_dispo = datetime.fromisoformat(dispo["date"]).date()
                    if date_debut_dt.date() <= date_dispo <= date_fin_dt.date():
                        user_disponibilites.append(dispo)
                except:
                    pass
        
        jours_disponibles = len([d for d in user_disponibilites if d.get("disponible") == True])
        jours_indisponibles = len([d for d in user_disponibilites if d.get("disponible") == False])
        
        # Analyser motifs d'indisponibilité
        motifs = {}
        for dispo in user_disponibilites:
            if not dispo.get("disponible"):
                motif = dispo.get("motif", "non_specifie")
                motifs[motif] = motifs.get(motif, 0) + 1
        
        total_jours = jours_disponibles + jours_indisponibles
        taux_disponibilite = round((jours_disponibles / total_jours * 100) if total_jours > 0 else 0, 1)
        
        total_jours_disponibles += jours_disponibles
        total_jours_indisponibles += jours_indisponibles
        
        rapport.append({
            "nom": f"{user.get('prenom', '')} {user.get('nom', '')}",
            "grade": user.get("grade", "N/A"),
            "jours_disponibles": jours_disponibles,
            "jours_indisponibles": jours_indisponibles,
            "taux_disponibilite": taux_disponibilite,
            "motifs_indisponibilite": motifs
        })
    
    # Calculer statistiques globales
    total_jours = total_jours_disponibles + total_jours_indisponibles
    taux_global = round((total_jours_disponibles / total_jours * 100) if total_jours > 0 else 0, 1)
    
    return {
        "periode": {"debut": date_debut, "fin": date_fin},
        "employes": rapport,
        "total_jours_disponibles": total_jours_disponibles,
        "total_jours_indisponibles": total_jours_indisponibles,
        "taux_disponibilite_global": taux_global,
        "nombre_employes": len(rapport)
    }


# GET disponibilites/export-pdf
@router.get("/{tenant_slug}/disponibilites/export-pdf")
async def export_disponibilites_pdf(
    tenant_slug: str,
    user_id: str = None,
    current_user: User = Depends(get_current_user)
):
    """Export des disponibilités en PDF"""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        from io import BytesIO
        
        tenant = await get_tenant_from_slug(tenant_slug)
        
        # Récupérer les disponibilités
        if user_id:
            disponibilites_list = await db.disponibilites.find({
                "tenant_id": tenant.id,
                "user_id": user_id
            }).to_list(length=None)
            users_list = [await db.users.find_one({"id": user_id, "tenant_id": tenant.id})]
        else:
            disponibilites_list = await db.disponibilites.find({
                "tenant_id": tenant.id
            }).to_list(length=None)
            users_list = await db.users.find({"tenant_id": tenant.id}).to_list(length=None)
        
        users_map = {u['id']: u for u in users_list}
        
        # Créer le PDF avec branding
        buffer, doc, elements = create_branded_pdf(tenant, pagesize=letter)
        styles = getSampleStyleSheet()
        modern_styles = get_modern_pdf_styles(styles)
        
        # Titre
        titre = "Disponibilités du Personnel Temps Partiel"
        if user_id and user_id in users_map:
            titre = f"Disponibilités de {users_map[user_id]['prenom']} {users_map[user_id]['nom']}"
        
        elements.append(Paragraph(titre, modern_styles['title']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Construire le tableau
        table_data = [['Date', 'Heure Début', 'Heure Fin', 'Statut', 'Type Garde', 'Pompier']]
        
        for dispo in sorted(disponibilites_list, key=lambda x: x.get('date', '')):
            user = users_map.get(dispo['user_id'], {})
            pompier_nom = f"{user.get('prenom', '')} {user.get('nom', '')}" if user else "N/A"
            
            statut_fr = {
                'disponible': 'Disponible',
                'indisponible': 'Indisponible',
                'conge': 'Congé'
            }.get(dispo.get('statut', ''), dispo.get('statut', ''))
            
            table_data.append([
                dispo.get('date', 'N/A'),
                dispo.get('heure_debut', 'N/A'),
                dispo.get('heure_fin', 'N/A'),
                statut_fr,
                dispo.get('type_garde_id', 'Tous') if dispo.get('type_garde_id') else 'Tous',
                pompier_nom if not user_id else ''
            ])
        
        # Si pas de user_id, afficher la colonne pompier, sinon la cacher
        if user_id:
            table_data = [[row[i] for i in range(5)] for row in table_data]
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), modern_styles['primary_color']),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, modern_styles['grid']),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, modern_styles['bg_light']])
        ]))
        
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"disponibilites_{user_id if user_id else 'tous'}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur export PDF: {str(e)}")


# GET disponibilites/export-excel
@router.get("/{tenant_slug}/disponibilites/export-excel")
async def export_disponibilites_excel(
    tenant_slug: str,
    user_id: str = None,
    current_user: User = Depends(get_current_user)
):
    """Export des disponibilités en Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO
        
        tenant = await get_tenant_from_slug(tenant_slug)
        
        # Récupérer les disponibilités
        if user_id:
            disponibilites_list = await db.disponibilites.find({
                "tenant_id": tenant.id,
                "user_id": user_id
            }).to_list(length=None)
            users_list = [await db.users.find_one({"id": user_id, "tenant_id": tenant.id})]
        else:
            disponibilites_list = await db.disponibilites.find({
                "tenant_id": tenant.id
            }).to_list(length=None)
            users_list = await db.users.find({"tenant_id": tenant.id}).to_list(length=None)
        
        users_map = {u['id']: u for u in users_list}
        
        # Créer le workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Disponibilités"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-têtes
        headers = ['Date', 'Heure Début', 'Heure Fin', 'Statut', 'Type Garde']
        if not user_id:
            headers.append('Pompier')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Données
        row = 2
        for dispo in sorted(disponibilites_list, key=lambda x: x.get('date', '')):
            user = users_map.get(dispo['user_id'], {})
            pompier_nom = f"{user.get('prenom', '')} {user.get('nom', '')}" if user else "N/A"
            
            statut_fr = {
                'disponible': 'Disponible',
                'indisponible': 'Indisponible',
                'conge': 'Congé'
            }.get(dispo.get('statut', ''), dispo.get('statut', ''))
            
            data_row = [
                dispo.get('date', 'N/A'),
                dispo.get('heure_debut', 'N/A'),
                dispo.get('heure_fin', 'N/A'),
                statut_fr,
                dispo.get('type_garde_id', 'Tous') if dispo.get('type_garde_id') else 'Tous'
            ]
            
            if not user_id:
                data_row.append(pompier_nom)
            
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            row += 1
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"disponibilites_{user_id if user_id else 'tous'}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur export Excel: {str(e)}")


# POST disponibilites/resolve-conflict
@router.post("/{tenant_slug}/disponibilites/resolve-conflict")
async def resolve_disponibilite_conflict(
    tenant_slug: str,
    data: Dict[str, Any] = Body(...),
    current_user: User = Depends(get_current_user)
):
    """
    Résout un conflit lors de la création d'une disponibilité
    
    Actions possibles:
    - supprimer_conflits: Supprime les indisponibilités en conflit et crée la disponibilité
    - creer_quand_meme: Crée la disponibilité sans supprimer les conflits
    - annuler: Ne fait rien
    """
    tenant = await get_tenant_from_slug(tenant_slug)
    
    action = data.get("action")  # "supprimer_conflits", "creer_quand_meme", "annuler"
    new_item_data = data.get("new_item")
    conflict_ids = data.get("conflict_ids", [])
    
    if action == "annuler":
        return {"message": "Opération annulée", "action": "annuler"}
    
    # Récupérer les détails des conflits avant suppression pour l'historique
    conflicts_to_delete = []
    if action == "supprimer_conflits" and conflict_ids:
        for conflict_id in conflict_ids:
            conflict_doc = await db.disponibilites.find_one({"id": conflict_id, "tenant_id": tenant.id})
            if conflict_doc:
                conflicts_to_delete.append(conflict_doc)
        
        # Supprimer les conflits
        await db.disponibilites.delete_many({
            "id": {"$in": conflict_ids},
            "tenant_id": tenant.id
        })
        
        # Notifier l'utilisateur affecté si différent de l'utilisateur courant
        affected_user_id = new_item_data.get("user_id")
        if affected_user_id != current_user.id:
            notification = {
                "id": str(uuid.uuid4()),
                "tenant_id": tenant.id,
                "user_id": affected_user_id,
                "titre": "Indisponibilités modifiées",
                "message": f"{len(conflict_ids)} indisponibilité(s) supprimée(s) en raison d'un conflit avec une nouvelle disponibilité",
                "type": "disponibilite",
                "lue": False,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.notifications.insert_one(notification)
    
    # Créer la disponibilité
    dispo_dict = new_item_data.copy()
    dispo_dict["tenant_id"] = tenant.id
    dispo_dict["id"] = str(uuid.uuid4())
    dispo_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.disponibilites.insert_one(dispo_dict)
    
    # Enregistrer dans l'historique
    resolution = ConflictResolution(
        tenant_id=tenant.id,
        user_id=current_user.id,
        affected_user_id=new_item_data.get("user_id"),
        action=action,
        type_created="disponibilite",
        conflicts_deleted=conflicts_to_delete,
        created_item=dispo_dict
    )
    await db.conflict_resolutions.insert_one(resolution.dict())
    
    return {
        "message": f"Disponibilité créée avec succès. Action: {action}",
        "action": action,
        "conflicts_deleted": len(conflicts_to_delete),
        "created_item": dispo_dict
    }


# POST disponibilites/import-csv
@router.post("/{tenant_slug}/disponibilites/import-csv")
async def import_disponibilites_csv(
    tenant_slug: str,
    disponibilites_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Import en masse de disponibilités depuis un CSV/XLS"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    disponibilites = disponibilites_data.get("disponibilites", [])
    if not disponibilites:
        raise HTTPException(status_code=400, detail="Aucune disponibilité à importer")
    
    # SÉCURITÉ: Vérifier le blocage des disponibilités
    params = await db.parametres_disponibilites.find_one({"tenant_id": tenant.id})
    blocage_actif = params and params.get("blocage_dispos_active", False)
    exceptions_admin = params.get("exceptions_admin_superviseur", True) if params else True
    
    # Fonction locale pour vérifier si un mois est bloqué
    def is_month_blocked(mois_str):
        if not blocage_actif:
            return False
        
        from datetime import date as date_type
        today = datetime.now(timezone.utc).date()
        jour_blocage = params.get("jour_blocage_dispos", 15)
        
        mois_parts = mois_str.split("-")
        mois_annee = int(mois_parts[0])
        mois_mois = int(mois_parts[1])
        
        try:
            date_blocage = date_type(today.year, today.month, jour_blocage)
        except ValueError:
            import calendar
            dernier_jour = calendar.monthrange(today.year, today.month)[1]
            date_blocage = date_type(today.year, today.month, min(jour_blocage, dernier_jour))
        
        if today.month == 12:
            mois_bloque_num = 1
            annee_mois_bloque = today.year + 1
        else:
            mois_bloque_num = today.month + 1
            annee_mois_bloque = today.year
        
        est_apres_date_limite = today > date_blocage
        mois_cible_est_passe = (mois_annee < today.year) or (mois_annee == today.year and mois_mois <= today.month)
        mois_cible_est_le_mois_bloque = (mois_annee == annee_mois_bloque and mois_mois == mois_bloque_num)
        
        est_bloque = mois_cible_est_passe or (est_apres_date_limite and mois_cible_est_le_mois_bloque)
        
        # Exception pour admin/superviseur
        if est_bloque and exceptions_admin and current_user.role in ["admin", "superviseur"]:
            return False
        
        return est_bloque
    
    results = {
        "total": len(disponibilites),
        "created": 0,
        "updated": 0,
        "errors": [],
        "skipped": 0,
        "blocked": 0
    }
    
    # Précharger les utilisateurs et types de garde pour optimisation
    users_list = await db.users.find({"tenant_id": tenant.id}).to_list(1000)
    users_by_num = {u.get("numero_employe"): u for u in users_list if u.get("numero_employe")}
    users_by_name = create_user_matching_index(users_list)
    
    types_garde_list = await db.types_garde.find({"tenant_id": tenant.id}).to_list(100)
    types_garde_by_name = {tg.get("nom", "").strip().lower(): tg for tg in types_garde_list}
    
    for index, dispo_data in enumerate(disponibilites):
        try:
            # 1. Trouver l'utilisateur avec matching intelligent (fuzzy)
            # Supporter les anciennes clés (Employé) et nouvelles clés (employe) pour compatibilité
            employe_str = dispo_data.get("employe", dispo_data.get("Employé", "")).strip()
            if not employe_str:
                results["errors"].append({
                    "ligne": index + 2,
                    "erreur": "Employé manquant"
                })
                continue
            
            # Utiliser la fonction de matching intelligent (recherche fuzzy)
            # Gère les espaces doubles, accents, ordre des noms, etc.
            user_obj = find_user_intelligent(
                search_string=employe_str,
                users_by_name=users_by_name,
                users_by_num=users_by_num,
                numero_field="numero_employe"
            )
            
            if not user_obj:
                results["errors"].append({
                    "ligne": index + 2,
                    "erreur": f"Employé non trouvé avec recherche fuzzy: {employe_str}"
                })
                continue
            
            # 2. Parser les dates/heures
            debut_str = str(dispo_data.get("debut", dispo_data.get("Début", ""))).strip()
            fin_str = str(dispo_data.get("fin", dispo_data.get("Fin", ""))).strip()
            
            if not debut_str or not fin_str:
                results["errors"].append({
                    "ligne": index + 2,
                    "erreur": "Date/heure de début ou fin manquante"
                })
                continue
            
            try:
                # Parser les dates/heures - supporter plusieurs formats
                from datetime import datetime as dt
                
                # Essayer format standard: "2025-12-01 06:00"
                try:
                    debut_dt = dt.strptime(debut_str, "%Y-%m-%d %H:%M")
                    fin_dt = dt.strptime(fin_str, "%Y-%m-%d %H:%M")
                except ValueError:
                    # Essayer format avec secondes: "2025-12-01 06:00:00"
                    debut_dt = dt.strptime(debut_str, "%Y-%m-%d %H:%M:%S")
                    fin_dt = dt.strptime(fin_str, "%Y-%m-%d %H:%M:%S")
                
                date_str = debut_dt.strftime("%Y-%m-%d")
                heure_debut = debut_dt.strftime("%H:%M")
                heure_fin = fin_dt.strftime("%H:%M")
                
            except ValueError as e:
                results["errors"].append({
                    "ligne": index + 2,
                    "erreur": f"Format de date/heure invalide: {e}"
                })
                continue
            
            # SÉCURITÉ: Vérifier le blocage pour ce mois
            mois_str = debut_dt.strftime("%Y-%m")
            if is_month_blocked(mois_str):
                results["errors"].append({
                    "ligne": index + 2,
                    "erreur": f"Mois {mois_str} bloqué - date limite dépassée"
                })
                results["blocked"] += 1
                continue
            
            # 3. Mapper la sélection au statut (avec valeur par défaut "Disponible")
            selection = dispo_data.get("selection", dispo_data.get("Sélection", "Disponible")).strip().lower()
            
            # Si "Aucune", ignorer cette ligne (ne pas créer de disponibilité)
            if selection == "aucune":
                results["skipped"] += 1
                continue
            
            # Mapper les valeurs à "disponible"
            if selection in ["disponible", "disponibilité", ""]:
                statut = "disponible"
            else:
                results["errors"].append({
                    "ligne": index + 2,
                    "erreur": f"Sélection invalide: '{selection}'. Attendu: 'Disponible' ou 'Aucune'"
                })
                continue
            
            # 4. Trouver le type de garde (optionnel)
            type_garde_id = None
            quart_str = dispo_data.get("quart", dispo_data.get("Quart", "")).strip().lower()
            if quart_str:
                type_garde_obj = types_garde_by_name.get(quart_str)
                if type_garde_obj:
                    type_garde_id = type_garde_obj.get("id")
            
            # 5. Vérifier si une disponibilité existe déjà
            existing = await db.disponibilites.find_one({
                "tenant_id": tenant.id,
                "user_id": user_obj["id"],
                "date": date_str,
                "heure_debut": heure_debut,
                "heure_fin": heure_fin
            })
            
            dispo_obj = Disponibilite(
                tenant_id=tenant.id,
                user_id=user_obj["id"],
                date=date_str,
                heure_debut=heure_debut,
                heure_fin=heure_fin,
                statut=statut,
                type_garde_id=type_garde_id,
                origine="import_csv"
            )
            
            if existing:
                # Mettre à jour
                await db.disponibilites.update_one(
                    {"id": existing["id"]},
                    {"$set": dispo_obj.dict()}
                )
                results["updated"] += 1
            else:
                # Créer nouveau
                await db.disponibilites.insert_one(dispo_obj.dict())
                results["created"] += 1
                
        except Exception as e:
            results["errors"].append({
                "ligne": index + 2,
                "erreur": str(e)
            })
    
    # Créer une activité
    await creer_activite(
        tenant_id=tenant.id,
        type_activite="import_disponibilites",
        description=f"📊 {current_user.prenom} {current_user.nom} a importé {results['created'] + results['updated']} disponibilités ({results['created']} créées, {results['updated']} mises à jour, {results['skipped']} ignorées)",
        user_id=current_user.id,
        user_nom=f"{current_user.prenom} {current_user.nom}"
    )
    
    return results


# GET disponibilites/export-pdf (2)
@router.get("/{tenant_slug}/disponibilites/export-pdf")
async def export_disponibilites_pdf(
    tenant_slug: str,
    user_id: str = None,
    current_user: User = Depends(get_current_user)
):
    """Export des disponibilités en PDF"""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        from io import BytesIO
        
        tenant = await get_tenant_from_slug(tenant_slug)
        
        # Récupérer les disponibilités
        if user_id:
            disponibilites_list = await db.disponibilites.find({
                "tenant_id": tenant.id,
                "user_id": user_id
            }).to_list(length=None)
            users_list = [await db.users.find_one({"id": user_id, "tenant_id": tenant.id})]
        else:
            disponibilites_list = await db.disponibilites.find({
                "tenant_id": tenant.id
            }).to_list(length=None)
            users_list = await db.users.find({"tenant_id": tenant.id}).to_list(length=None)
        
        users_map = {u['id']: u for u in users_list}
        
        # Créer le PDF avec branding
        buffer, doc, elements = create_branded_pdf(tenant, pagesize=letter)
        styles = getSampleStyleSheet()
        modern_styles = get_modern_pdf_styles(styles)
        
        # Titre
        titre = "Disponibilités du Personnel Temps Partiel"
        if user_id and user_id in users_map:
            titre = f"Disponibilités de {users_map[user_id]['prenom']} {users_map[user_id]['nom']}"
        
        elements.append(Paragraph(titre, modern_styles['title']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Construire le tableau
        table_data = [['Date', 'Heure Début', 'Heure Fin', 'Statut', 'Type Garde', 'Pompier']]
        
        for dispo in sorted(disponibilites_list, key=lambda x: x.get('date', '')):
            user = users_map.get(dispo['user_id'], {})
            pompier_nom = f"{user.get('prenom', '')} {user.get('nom', '')}" if user else "N/A"
            
            statut_fr = {
                'disponible': 'Disponible',
                'indisponible': 'Indisponible',
                'conge': 'Congé'
            }.get(dispo.get('statut', ''), dispo.get('statut', ''))
            
            table_data.append([
                dispo.get('date', 'N/A'),
                dispo.get('heure_debut', 'N/A'),
                dispo.get('heure_fin', 'N/A'),
                statut_fr,
                dispo.get('type_garde_id', 'Tous') if dispo.get('type_garde_id') else 'Tous',
                pompier_nom if not user_id else ''
            ])
        
        # Si pas de user_id, afficher la colonne pompier, sinon la cacher
        if user_id:
            table_data = [[row[i] for i in range(5)] for row in table_data]
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), modern_styles['primary_color']),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, modern_styles['grid']),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, modern_styles['bg_light']])
        ]))
        
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"disponibilites_{user_id if user_id else 'tous'}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur export PDF: {str(e)}")


# GET disponibilites/export-excel (2)
@router.get("/{tenant_slug}/disponibilites/export-excel")
async def export_disponibilites_excel(
    tenant_slug: str,
    user_id: str = None,
    current_user: User = Depends(get_current_user)
):
    """Export des disponibilités en Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO
        
        tenant = await get_tenant_from_slug(tenant_slug)
        
        # Récupérer les disponibilités
        if user_id:
            disponibilites_list = await db.disponibilites.find({
                "tenant_id": tenant.id,
                "user_id": user_id
            }).to_list(length=None)
            users_list = [await db.users.find_one({"id": user_id, "tenant_id": tenant.id})]
        else:
            disponibilites_list = await db.disponibilites.find({
                "tenant_id": tenant.id
            }).to_list(length=None)
            users_list = await db.users.find({"tenant_id": tenant.id}).to_list(length=None)
        
        users_map = {u['id']: u for u in users_list}
        
        # Créer le workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Disponibilités"
        
        # Styles
        header_fill = PatternFill(start_color="FCA5A5", end_color="FCA5A5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Titre
        ws.merge_cells('A1:F1')
        titre = "Disponibilités du Personnel Temps Partiel"
        if user_id and user_id in users_map:
            titre = f"Disponibilités de {users_map[user_id]['prenom']} {users_map[user_id]['nom']}"
        ws['A1'] = titre
        ws['A1'].font = Font(bold=True, size=16, color="EF4444")
        ws['A1'].alignment = center_alignment
        
        # En-têtes
        row = 3
        if user_id:
            headers = ['Date', 'Heure Début', 'Heure Fin', 'Statut', 'Type Garde']
        else:
            headers = ['Date', 'Heure Début', 'Heure Fin', 'Statut', 'Type Garde', 'Pompier']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # Données
        row += 1
        for dispo in sorted(disponibilites_list, key=lambda x: x.get('date', '')):
            user = users_map.get(dispo['user_id'], {})
            pompier_nom = f"{user.get('prenom', '')} {user.get('nom', '')}" if user else "N/A"
            
            statut_fr = {
                'disponible': 'Disponible',
                'indisponible': 'Indisponible',
                'conge': 'Congé'
            }.get(dispo.get('statut', ''), dispo.get('statut', ''))
            
            ws.cell(row=row, column=1, value=dispo.get('date', 'N/A'))
            ws.cell(row=row, column=2, value=dispo.get('heure_debut', 'N/A'))
            ws.cell(row=row, column=3, value=dispo.get('heure_fin', 'N/A'))
            status_cell = ws.cell(row=row, column=4, value=statut_fr)
            ws.cell(row=row, column=5, value=dispo.get('type_garde_id', 'Tous') if dispo.get('type_garde_id') else 'Tous')
            
            if not user_id:
                ws.cell(row=row, column=6, value=pompier_nom)
            
            # Couleur statut
            if dispo.get('statut') == 'disponible':
                status_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            elif dispo.get('statut') == 'indisponible':
                status_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = center_alignment
            
            row += 1
        
        # Ajuster les largeurs de colonnes
        column_widths = {
            'A': 12,  # Date
            'B': 15,  # Type Garde
            'C': 20,  # Demandeur
            'D': 12,  # Statut
            'E': 10,  # Priorité
            'F': 20,  # Remplaçant
            'G': 30,  # Notes
            'H': 18,  # Créé le
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"disponibilites_{user_id if user_id else 'tous'}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur export Excel: {str(e)}")


# DELETE disponibilites/reinitialiser
@router.delete("/{tenant_slug}/disponibilites/reinitialiser")
async def reinitialiser_disponibilites(
    tenant_slug: str,
    reinit_data: DisponibiliteReinitialiser,
    current_user: User = Depends(get_current_user)
):
    """
    Réinitialise les disponibilités/indisponibilités pour une période donnée
    """
    # Vérifier le tenant
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Vérifier les permissions
    if current_user.role not in ["admin", "superviseur"] and current_user.id != reinit_data.user_id:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        # Calculer les dates de début et fin selon la période
        today = datetime.now(timezone.utc).date()
        
        if reinit_data.periode == "semaine":
            # Semaine courante : lundi à dimanche
            days_since_monday = today.weekday()  # 0 = lundi, 6 = dimanche
            date_debut = today - timedelta(days=days_since_monday)
            date_fin = date_debut + timedelta(days=6)
        elif reinit_data.periode == "mois":
            # Mois courant : 1er du mois à dernier jour
            date_debut = today.replace(day=1)
            # Dernier jour du mois
            if today.month == 12:
                date_fin = today.replace(day=31)
            else:
                next_month = today.replace(month=today.month + 1, day=1)
                date_fin = next_month - timedelta(days=1)
        elif reinit_data.periode == "mois_prochain":
            # Mois prochain : 1er du mois prochain à dernier jour du mois prochain
            if today.month == 12:
                # Si on est en décembre, mois prochain = janvier de l'année suivante
                date_debut = today.replace(year=today.year + 1, month=1, day=1)
                next_next_month = date_debut.replace(month=2, day=1)
                date_fin = next_next_month - timedelta(days=1)
            else:
                # Mois prochain
                date_debut = today.replace(month=today.month + 1, day=1)
                # Dernier jour du mois prochain
                if today.month == 11:
                    # Si on est en novembre, mois prochain = décembre
                    date_fin = date_debut.replace(day=31)
                else:
                    next_next_month = today.replace(month=today.month + 2, day=1)
                    date_fin = next_next_month - timedelta(days=1)
        elif reinit_data.periode == "annee":
            # Année courante : 1er janvier à 31 décembre
            date_debut = today.replace(month=1, day=1)
            date_fin = today.replace(month=12, day=31)
        elif reinit_data.periode == "personnalisee":
            # Période personnalisée : dates fournies
            if not reinit_data.date_debut or not reinit_data.date_fin:
                raise HTTPException(
                    status_code=400,
                    detail="date_debut et date_fin sont requis pour une période personnalisée"
                )
            try:
                date_debut = datetime.fromisoformat(reinit_data.date_debut).date()
                date_fin = datetime.fromisoformat(reinit_data.date_fin).date()
                
                # Validation : date_debut <= date_fin
                if date_debut > date_fin:
                    raise HTTPException(
                        status_code=400,
                        detail="date_debut doit être avant ou égale à date_fin"
                    )
                
                # Validation : plage maximale de 1 an
                if (date_fin - date_debut).days > 365:
                    raise HTTPException(
                        status_code=400,
                        detail="La plage de dates ne peut pas dépasser 1 an"
                    )
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail="Format de date invalide. Utilisez YYYY-MM-DD"
                )
        else:
            raise HTTPException(
                status_code=400,
                detail="periode doit être 'semaine', 'mois', 'mois_prochain', 'annee' ou 'personnalisee'"
            )
        
        # Construire la requête de suppression
        delete_query = {
            "user_id": reinit_data.user_id,
            "tenant_id": tenant.id,
            "date": {
                "$gte": date_debut.isoformat(),
                "$lte": date_fin.isoformat()
            }
        }
        
        # Filtre par type d'entrée (disponibilités/indisponibilités)
        if reinit_data.type_entree == "disponibilites":
            delete_query["statut"] = "disponible"
        elif reinit_data.type_entree == "indisponibilites":
            delete_query["statut"] = "indisponible"
        elif reinit_data.type_entree != "les_deux":
            raise HTTPException(
                status_code=400,
                detail="type_entree doit être 'disponibilites', 'indisponibilites' ou 'les_deux'"
            )
        
        # Si mode "generees_seulement", ne supprimer que les entrées générées automatiquement
        if reinit_data.mode == "generees_seulement":
            # Supprimer uniquement celles avec origine différente de "manuelle"
            # ET qui ont un champ origine (pour gérer les anciennes entrées)
            delete_query["$or"] = [
                {"origine": {"$exists": True, "$ne": "manuelle"}},
                {"origine": {"$exists": False}}  # Anciennes entrées sans champ origine
            ]
        elif reinit_data.mode != "tout":
            raise HTTPException(
                status_code=400,
                detail="mode doit être 'tout' ou 'generees_seulement'"
            )
        
        # Supprimer les disponibilités
        result = await db.disponibilites.delete_many(delete_query)
        
        return {
            "message": "Réinitialisation effectuée avec succès",
            "periode": reinit_data.periode,
            "mode": reinit_data.mode,
            "date_debut": date_debut.isoformat(),
            "date_fin": date_fin.isoformat(),
            "nombre_supprimees": result.deleted_count
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Erreur lors de la réinitialisation: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur lors de la réinitialisation: {str(e)}")


# DELETE disponibilites/{disponibilite_id}
@router.delete("/{tenant_slug}/disponibilites/{disponibilite_id}")
async def delete_disponibilite(tenant_slug: str, disponibilite_id: str, current_user: User = Depends(get_current_user)):
    # Vérifier le tenant
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Find the disponibilité to check ownership dans ce tenant
    disponibilite = await db.disponibilites.find_one({
        "id": disponibilite_id,
        "tenant_id": tenant.id
    })
    if not disponibilite:
        raise HTTPException(status_code=404, detail="Disponibilité non trouvée")
    
    if current_user.role not in ["admin", "superviseur"] and current_user.id != disponibilite["user_id"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    result = await db.disponibilites.delete_one({
        "id": disponibilite_id,
        "tenant_id": tenant.id
    })
    if result.deleted_count == 0:
        raise HTTPException(status_code=400, detail="Impossible de supprimer la disponibilité")
    
    # Créer une activité (seulement si l'utilisateur supprime ses propres disponibilités)
    if disponibilite["user_id"] == current_user.id:
        statut_text = "disponibilité" if disponibilite["statut"] == "disponible" else "indisponibilité"
        await creer_activite(
            tenant_id=tenant.id,
            type_activite="disponibilite_suppression",
            description=f"🗑️ {current_user.prenom} {current_user.nom} a supprimé une {statut_text} du {disponibilite['date']}",
            user_id=current_user.id,
            user_nom=f"{current_user.prenom} {current_user.nom}",
            data={"concerne_user_id": current_user.id}
        )
    
    return {"message": "Disponibilité supprimée avec succès"}


# POST auto-affecter-disponibilites
@router.post("/auto-affecter-disponibilites-temps-partiel")
async def auto_affecter_disponibilites_temps_partiel(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        # DÉTECTION AUTOMATIQUE de tous les employés temps partiel
        tous_temps_partiel = await db.users.find({
            "type_emploi": "temps_partiel",
            "statut": "Actif"
        }).to_list(1000)
        
        print(f"Trouvé {len(tous_temps_partiel)} employés temps partiel")
        
        # Supprimer les anciennes disponibilités de la semaine courante
        today = datetime.now(timezone.utc).date()
        start_week = today - timedelta(days=today.weekday())
        end_week = start_week + timedelta(days=6)
        
        await db.disponibilites.delete_many({
            "date": {
                "$gte": start_week.strftime("%Y-%m-%d"),
                "$lte": end_week.strftime("%Y-%m-%d")
            }
        })
        
        # Récupérer types de garde
        types_garde = await db.types_garde.find().to_list(100)
        
        disponibilites_created = 0
        
        # AFFECTER DISPONIBILITÉS À TOUS LES TEMPS PARTIEL DÉTECTÉS
        for index, user in enumerate(tous_temps_partiel):
            print(f"Affectation pour {user['prenom']} {user['nom']} ({user['grade']})")
            
            # Pattern de disponibilité selon l'index pour variété
            if index % 4 == 0:  # Pattern 1: Lun-Mer-Ven
                jours_disponibles = [0, 2, 4]
            elif index % 4 == 1:  # Pattern 2: Mar-Jeu-Sam  
                jours_disponibles = [1, 3, 5]
            elif index % 4 == 2:  # Pattern 3: Mer-Ven-Dim
                jours_disponibles = [2, 4, 6]
            else:  # Pattern 4: Lun-Jeu-Dim
                jours_disponibles = [0, 3, 6]
            
            for day_offset in jours_disponibles:
                date_dispo = start_week + timedelta(days=day_offset)
                date_str = date_dispo.strftime("%Y-%m-%d")
                day_name = date_dispo.strftime("%A").lower()
                
                # Créer disponibilités pour TOUS les types de garde applicables
                for type_garde in types_garde:
                    jours_app = type_garde.get("jours_application", [])
                    if jours_app and day_name not in jours_app:
                        continue
                    
                    # Créer disponibilité (stratégie intensive pour démo)
                    dispo_obj = Disponibilite(
                        user_id=user["id"],
                        date=date_str,
                        type_garde_id=type_garde["id"],
                        heure_debut=type_garde["heure_debut"],
                        heure_fin=type_garde["heure_fin"],
                        statut="disponible"
                    )
                    await db.disponibilites.insert_one(dispo_obj.dict())
                    disponibilites_created += 1
        
        return {
            "message": "Disponibilités affectées automatiquement",
            "employes_temps_partiel_detectes": len(tous_temps_partiel),
            "disponibilites_creees": disponibilites_created,
            "semaine": f"{start_week.strftime('%Y-%m-%d')} - {end_week.strftime('%Y-%m-%d')}",
            "patterns": "4 patterns différents pour variété démo",
            "employés_détectés": [f"{u['prenom']} {u['nom']} ({u['grade']})" for u in tous_temps_partiel]
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur auto-affectation: {str(e)}")

# Créer disponibilités MAXIMALES pour démo parfaite


# POST init-disponibilites-demo-complete
@router.post("/init-disponibilites-demo-complete")
async def init_disponibilites_demo_complete(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        # Supprimer toutes les disponibilités existantes
        await db.disponibilites.delete_many({})
        
        today = datetime.now(timezone.utc).date()
        start_week = today - timedelta(days=today.weekday())
        end_week = start_week + timedelta(days=6)
        
        # Récupérer TOUS les utilisateurs (temps plein ET temps partiel pour démo)
        all_users = await db.users.find({"statut": "Actif"}).to_list(100)
        types_garde = await db.types_garde.find().to_list(100)
        
        disponibilites_created = 0
        
        # STRATÉGIE DÉMO : TOUS LES EMPLOYÉS DISPONIBLES POUR TOUS LES TYPES
        for user in all_users:
            for day_offset in range(7):  # Chaque jour
                date_dispo = start_week + timedelta(days=day_offset)
                date_str = date_dispo.strftime("%Y-%m-%d")
                day_name = date_dispo.strftime("%A").lower()
                
                for type_garde in types_garde:
                    # Vérifier jours d'application
                    jours_app = type_garde.get("jours_application", [])
                    if jours_app and day_name not in jours_app:
                        continue
                    
                    # CRÉER DISPONIBILITÉ POUR TOUS (temps plein et temps partiel)
                    # Exception : respecter les heures max pour temps partiel
                    if user.get("type_emploi", "temps_plein") in ("temps_partiel", "temporaire"):
                        # Temps partiel : disponible seulement 3 jours par semaine
                        user_number = int(user["numero_employe"][-1]) if user["numero_employe"][-1].isdigit() else 0
                        
                        # Pattern par employé pour éviter épuisement
                        if user_number % 3 == 0 and day_offset in [0, 2, 4]:  # Lun-Mer-Ven
                            pass
                        elif user_number % 3 == 1 and day_offset in [1, 3, 5]:  # Mar-Jeu-Sam
                            pass  
                        elif user_number % 3 == 2 and day_offset in [2, 4, 6]:  # Mer-Ven-Dim
                            pass
                        else:
                            continue  # Skip ce jour pour cet employé
                    
                    # CRÉER DISPONIBILITÉ
                    dispo_obj = Disponibilite(
                        user_id=user["id"],
                        date=date_str,
                        type_garde_id=type_garde["id"],
                        heure_debut=type_garde["heure_debut"],
                        heure_fin=type_garde["heure_fin"],
                        statut="disponible"
                    )
                    await db.disponibilites.insert_one(dispo_obj.dict())
                    disponibilites_created += 1
        
        return {
            "message": "Disponibilités DÉMO COMPLÈTES créées",
            "semaine": f"{start_week.strftime('%Y-%m-%d')} - {end_week.strftime('%Y-%m-%d')}",
            "disponibilites_creees": disponibilites_created,
            "all_users_included": len(all_users),
            "strategy": "TOUS employés (TP+TPa) avec patterns optimisés"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")

# Créer disponibilités pour semaine courante (démo assignation auto)


# POST init-disponibilites-semaine-courante
@router.post("/init-disponibilites-semaine-courante")
async def init_disponibilites_semaine_courante(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        # Supprimer les disponibilités existantes pour la semaine courante
        today = datetime.now(timezone.utc).date()
        start_week = today - timedelta(days=today.weekday())
        end_week = start_week + timedelta(days=6)
        
        await db.disponibilites.delete_many({
            "date": {
                "$gte": start_week.strftime("%Y-%m-%d"),
                "$lte": end_week.strftime("%Y-%m-%d")
            }
        })
        
        # Récupérer tous les types de garde
        types_garde = await db.types_garde.find().to_list(100)
        # DÉTECTION AUTOMATIQUE de TOUS les employés temps partiel (peu importe le nombre)
        tous_temps_partiel = await db.users.find({
            "type_emploi": "temps_partiel",
            "statut": "Actif"
        }).to_list(1000)
        
        print(f"AUTO-DÉTECTION: {len(tous_temps_partiel)} employés temps partiel trouvés")
        
        disponibilites_created = 0
        
        # ALGORITHME OPTIMISÉ POUR TOUS VOS EMPLOYÉS TEMPS PARTIEL
        for user_index, user in enumerate(tous_temps_partiel):
            for day_offset in range(7):  # Chaque jour de la semaine courante
                date_dispo = start_week + timedelta(days=day_offset)
                date_str = date_dispo.strftime("%Y-%m-%d")
                day_name = date_dispo.strftime("%A").lower()
                
                # Pattern de disponibilité varié selon l'employé
                if user_index % 3 == 0:  # 1/3 des employés : Lun-Mer-Ven
                    jours_pattern = ['monday', 'wednesday', 'friday']
                elif user_index % 3 == 1:  # 1/3 des employés : Mar-Jeu-Sam
                    jours_pattern = ['tuesday', 'thursday', 'saturday']
                else:  # 1/3 des employés : Mer-Ven-Dim
                    jours_pattern = ['wednesday', 'friday', 'sunday']
                
                if day_name in jours_pattern:
                    # Créer disponibilités pour TOUS les types de garde applicables
                    for type_garde in types_garde:
                        jours_app = type_garde.get("jours_application", [])
                        if jours_app and day_name not in jours_app:
                            continue
                        
                        # CRÉER DISPONIBILITÉ pour vos employés (pompiers ET lieutenants)
                        dispo_obj = Disponibilite(
                            user_id=user["id"],
                            date=date_str,
                            type_garde_id=type_garde["id"],
                            heure_debut=type_garde["heure_debut"],
                            heure_fin=type_garde["heure_fin"],
                            statut="disponible"
                        )
                        await db.disponibilites.insert_one(dispo_obj.dict())
                        disponibilites_created += 1
        
        return {
            "message": "Disponibilités créées pour TOUS vos employés temps partiel",
            "employes_temps_partiel": len(tous_temps_partiel),
            "disponibilites_creees": disponibilites_created,
            "all_users_included": len(tous_temps_partiel),
            "strategy": f"AUTO-DÉTECTION: {len(tous_temps_partiel)} employés temps partiel avec patterns optimisés"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")

# Créer données de démonstration OPTIMALES pour démo client

