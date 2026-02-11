"""
Routes API pour le module Rapports
==================================

STATUT: ACTIF
Ce module gère les rapports, exports PDF/Excel, et statistiques.

Routes migrées depuis server.py:
- GET    /{tenant_slug}/rapports/statistiques-avancees
- POST   /{tenant_slug}/rapports/budgets
- POST   /{tenant_slug}/rapports/import-csv
- GET    /{tenant_slug}/rapports/budgets
- PUT    /{tenant_slug}/rapports/budgets/{budget_id}
- DELETE /{tenant_slug}/rapports/budgets/{budget_id}
- POST   /{tenant_slug}/rapports/immobilisations
- GET    /{tenant_slug}/rapports/immobilisations
- DELETE /{tenant_slug}/rapports/immobilisations/{immob_id}
- POST   /{tenant_slug}/rapports/projets-triennaux
- GET    /{tenant_slug}/rapports/projets-triennaux
- DELETE /{tenant_slug}/rapports/projets-triennaux/{projet_id}
- POST   /{tenant_slug}/rapports/interventions
- GET    /{tenant_slug}/rapports/interventions
- GET    /{tenant_slug}/rapports/dashboard-interne
- GET    /{tenant_slug}/rapports/couts-salariaux
- GET    /{tenant_slug}/rapports/tableau-bord-budgetaire
- GET    /{tenant_slug}/rapports/rapport-immobilisations
- GET    /{tenant_slug}/rapports/export-dashboard-pdf
- GET    /{tenant_slug}/rapports/export-salaires-pdf
- GET    /{tenant_slug}/rapports/export-salaires-excel
- GET    /{tenant_slug}/personnel/export-pdf
- GET    /{tenant_slug}/personnel/export-excel
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta, date
from io import BytesIO
import uuid
import logging
import csv
import io

from routes.dependencies import (
    db,
    get_current_user,
    get_tenant_from_slug,
    clean_mongo_doc,
    User
)

# Import des helpers PDF partagés
from utils.pdf_helpers import (
    create_branded_pdf,
    get_modern_pdf_styles,
    create_pdf_footer_text
)

# Imports pour PDF
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Imports pour Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(tags=["Rapports"])
logger = logging.getLogger(__name__)

# Stockage temporaire des fichiers pour téléchargement (compatible iframe sandbox)
# Les fichiers sont stockés pendant 5 minutes max
import os
import time
TEMP_EXPORT_DIR = "/tmp/exports"
os.makedirs(TEMP_EXPORT_DIR, exist_ok=True)

def cleanup_old_exports():
    """Nettoie les fichiers d'export de plus de 5 minutes"""
    try:
        now = time.time()
        for filename in os.listdir(TEMP_EXPORT_DIR):
            filepath = os.path.join(TEMP_EXPORT_DIR, filename)
            if os.path.isfile(filepath) and now - os.path.getmtime(filepath) > 300:
                os.remove(filepath)
    except Exception as e:
        logger.warning(f"Erreur nettoyage exports: {e}")


# ==================== MODÈLES PYDANTIC ====================

class Budget(BaseModel):
    """Modèle complet pour les budgets"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    annee: int
    categorie: str  # salaires, formations, equipements, carburant, entretien, autres
    budget_alloue: float
    budget_consomme: float = 0.0
    notes: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BudgetCreate(BaseModel):
    """Modèle pour création de budget"""
    annee: int
    categorie: str
    budget_alloue: float
    notes: str = ""
    # Champs legacy pour compatibilité
    montant_prevu: Optional[float] = None
    description: Optional[str] = None

class Immobilisation(BaseModel):
    """Modèle complet pour les immobilisations"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    type_immobilisation: str  # vehicule, equipement_majeur
    nom: str
    date_acquisition: str  # YYYY-MM-DD
    cout_acquisition: float
    cout_entretien_annuel: float = 0.0
    etat: str = "bon"  # bon, moyen, mauvais
    date_remplacement_prevue: Optional[str] = None
    notes: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ImmobilisationCreate(BaseModel):
    """Modèle pour création d'immobilisation"""
    type_immobilisation: str
    nom: str
    date_acquisition: str
    cout_acquisition: float
    cout_entretien_annuel: float = 0.0
    etat: str = "bon"
    date_remplacement_prevue: Optional[str] = None
    notes: str = ""
    # Champs legacy pour compatibilité
    categorie: Optional[str] = None
    valeur_acquisition: Optional[float] = None
    duree_amortissement: Optional[int] = 5

class ProjetTriennal(BaseModel):
    """Modèle complet pour les projets triennaux"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    nom: str
    description: str
    type_projet: str  # acquisition, renovation, recrutement
    annee_prevue: int
    cout_estime: float
    statut: str = "prevu"  # prevu, en_cours, termine, annule
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProjetTriennalCreate(BaseModel):
    """Modèle pour création de projet triennal"""
    nom: str
    description: str = ""
    type_projet: str
    annee_prevue: int
    cout_estime: float
    statut: str = "prevu"
    # Champs legacy pour compatibilité
    montant_total: Optional[float] = None
    annee_debut: Optional[int] = None
    annee_fin: Optional[int] = None

class Intervention(BaseModel):
    """Modèle complet pour les interventions"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tenant_id: str
    date_intervention: str
    type_intervention: str  # incendie, medical, sauvetage, autre
    duree_minutes: int
    nombre_pompiers: int
    temps_reponse_minutes: Optional[int] = None
    notes: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class InterventionCreate(BaseModel):
    """Modèle pour création d'intervention"""
    date_intervention: str
    type_intervention: str
    duree_minutes: int
    nombre_pompiers: int
    temps_reponse_minutes: Optional[int] = None
    notes: str = ""
    # Champs legacy pour compatibilité
    date: Optional[str] = None
    description: Optional[str] = None
    ressources_utilisees: Optional[List[str]] = []


# ==================== ROUTES STATISTIQUES ====================

@router.get("/{tenant_slug}/rapports/statistiques-avancees")
async def get_statistiques_avancees(tenant_slug: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    try:
        # Récupérer toutes les données nécessaires filtrées par tenant
        users = await db.users.find({"tenant_id": tenant.id}).to_list(1000)
        assignations = await db.assignations.find({"tenant_id": tenant.id}).to_list(1000)
        types_garde = await db.types_garde.find({"tenant_id": tenant.id}).to_list(1000)
        formations = await db.formations.find({"tenant_id": tenant.id}).to_list(1000)
        demandes_remplacement = await db.demandes_remplacement.find({"tenant_id": tenant.id}).to_list(1000)
        
        # Statistiques générales
        stats_generales = {
            "personnel_total": len(users),
            "personnel_actif": len([u for u in users if u.get("statut") == "Actif"]),
            "assignations_mois": len(assignations),
            "taux_couverture": 94.5,
            "formations_disponibles": len(formations),
            "remplacements_demandes": len(demandes_remplacement)
        }
        
        # Statistiques par rôle
        stats_par_role = {}
        for role in ["admin", "superviseur", "employe"]:
            users_role = [u for u in users if u.get("role") == role]
            assignations_role = [a for a in assignations if any(u["id"] == a["user_id"] and u.get("role") == role for u in users)]
            
            stats_par_role[role] = {
                "nombre_utilisateurs": len(users_role),
                "assignations_totales": len(assignations_role),
                "heures_moyennes": len(assignations_role) * 8,
                "formations_completees": sum(len(u.get("formations", [])) for u in users_role)
            }
        
        # Statistiques par employé
        stats_par_employe = []
        for user in users:
            user_assignations = [a for a in assignations if a["user_id"] == user["id"]]
            user_disponibilites = await db.disponibilites.find({"user_id": user["id"], "tenant_id": tenant.id}).to_list(100)
            
            stats_par_employe.append({
                "id": user["id"],
                "nom": f"{user.get('prenom', '')} {user.get('nom', '')}",
                "grade": user.get("grade", "N/A"),
                "role": user.get("role", "pompier"),
                "assignations": len(user_assignations),
                "disponibilites": len(user_disponibilites),
                "formations": len(user.get("formations", [])),
                "heures_totales": len(user_assignations) * 8
            })
        
        return {
            "generales": stats_generales,
            "par_role": stats_par_role,
            "par_employe": stats_par_employe
        }
    except Exception as e:
        logger.error(f"Erreur statistiques avancées: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ROUTES BUDGETS ====================

@router.post("/{tenant_slug}/rapports/budgets")
async def create_budget(
    tenant_slug: str,
    budget: BudgetCreate,
    current_user: User = Depends(get_current_user)
):
    """Créer une entrée budget"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Supporter les deux formats (nouveau et legacy)
    annee = budget.annee if hasattr(budget, 'annee') and budget.annee else datetime.now().year
    budget_alloue = budget.budget_alloue if hasattr(budget, 'budget_alloue') and budget.budget_alloue else (budget.montant_prevu or 0)
    notes = budget.notes if hasattr(budget, 'notes') and budget.notes else (budget.description or "")
    
    budget_doc = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant.id,
        "annee": annee,
        "categorie": budget.categorie,
        "budget_alloue": budget_alloue,
        "budget_consomme": 0,
        "notes": notes,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.budgets.insert_one(budget_doc)
    budget_doc.pop("_id", None)
    
    return budget_doc


@router.get("/{tenant_slug}/rapports/budgets")
async def get_budgets(
    tenant_slug: str, 
    annee: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    """Récupérer tous les budgets"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    query = {"tenant_id": tenant.id}
    if annee:
        query["annee"] = annee
    
    budgets = await db.budgets.find(query).to_list(1000)
    return [clean_mongo_doc(b) for b in budgets]


@router.put("/{tenant_slug}/rapports/budgets/{budget_id}")
async def update_budget(
    tenant_slug: str,
    budget_id: str,
    budget: BudgetCreate,
    current_user: User = Depends(get_current_user)
):
    """Modifier un budget"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Supporter les deux formats (nouveau et legacy)
    budget_alloue = budget.budget_alloue if hasattr(budget, 'budget_alloue') and budget.budget_alloue else (budget.montant_prevu or 0)
    notes = budget.notes if hasattr(budget, 'notes') and budget.notes else (budget.description or "")
    
    update_data = {
        "categorie": budget.categorie,
        "budget_alloue": budget_alloue,
        "notes": notes,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if hasattr(budget, 'annee') and budget.annee:
        update_data["annee"] = budget.annee
    
    result = await db.budgets.update_one(
        {"id": budget_id, "tenant_id": tenant.id},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Budget non trouvé")
    
    return {"message": "Budget mis à jour"}


@router.delete("/{tenant_slug}/rapports/budgets/{budget_id}")
async def delete_budget(
    tenant_slug: str,
    budget_id: str,
    current_user: User = Depends(get_current_user)
):
    """Supprimer un budget"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    result = await db.budgets.delete_one({"id": budget_id, "tenant_id": tenant.id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Budget non trouvé")
    
    return {"message": "Budget supprimé"}


# ==================== ROUTES IMMOBILISATIONS ====================

@router.post("/{tenant_slug}/rapports/immobilisations")
async def create_immobilisation(
    tenant_slug: str,
    immob: ImmobilisationCreate,
    current_user: User = Depends(get_current_user)
):
    """Créer une immobilisation"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    # Supporter les deux formats (nouveau et legacy)
    cout_acquisition = immob.cout_acquisition if hasattr(immob, 'cout_acquisition') and immob.cout_acquisition else (immob.valeur_acquisition or 0)
    type_immob = immob.type_immobilisation if hasattr(immob, 'type_immobilisation') and immob.type_immobilisation else (immob.categorie or "equipement_majeur")
    
    immob_doc = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant.id,
        "type_immobilisation": type_immob,
        "nom": immob.nom,
        "date_acquisition": immob.date_acquisition,
        "cout_acquisition": cout_acquisition,
        "cout_entretien_annuel": immob.cout_entretien_annuel if hasattr(immob, 'cout_entretien_annuel') else 0,
        "etat": immob.etat if hasattr(immob, 'etat') else "bon",
        "date_remplacement_prevue": immob.date_remplacement_prevue if hasattr(immob, 'date_remplacement_prevue') else None,
        "notes": immob.notes if hasattr(immob, 'notes') else "",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.immobilisations.insert_one(immob_doc)
    immob_doc.pop("_id", None)
    
    return immob_doc


@router.get("/{tenant_slug}/rapports/immobilisations")
async def get_immobilisations(tenant_slug: str, current_user: User = Depends(get_current_user)):
    """Récupérer toutes les immobilisations"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    immobs = await db.immobilisations.find({"tenant_id": tenant.id}).to_list(1000)
    return [clean_mongo_doc(i) for i in immobs]


@router.delete("/{tenant_slug}/rapports/immobilisations/{immob_id}")
async def delete_immobilisation(
    tenant_slug: str,
    immob_id: str,
    current_user: User = Depends(get_current_user)
):
    """Supprimer une immobilisation"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    result = await db.immobilisations.delete_one({"id": immob_id, "tenant_id": tenant.id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Immobilisation non trouvée")
    
    return {"message": "Immobilisation supprimée"}


# ==================== ROUTES PROJETS TRIENNAUX ====================

@router.post("/{tenant_slug}/rapports/projets-triennaux")
async def create_projet_triennal(
    tenant_slug: str,
    projet: ProjetTriennalCreate,
    current_user: User = Depends(get_current_user)
):
    """Créer un projet triennal"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    projet_doc = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant.id,
        "nom": projet.nom,
        "description": projet.description,
        "montant_total": projet.montant_total,
        "annee_debut": projet.annee_debut,
        "annee_fin": projet.annee_fin,
        "statut": "planifie",
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.projets_triennaux.insert_one(projet_doc)
    projet_doc.pop("_id", None)
    
    return projet_doc


@router.get("/{tenant_slug}/rapports/projets-triennaux")
async def get_projets_triennaux(tenant_slug: str, current_user: User = Depends(get_current_user)):
    """Récupérer tous les projets triennaux"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    projets = await db.projets_triennaux.find({"tenant_id": tenant.id}).to_list(1000)
    return [clean_mongo_doc(p) for p in projets]


@router.delete("/{tenant_slug}/rapports/projets-triennaux/{projet_id}")
async def delete_projet_triennal(
    tenant_slug: str,
    projet_id: str,
    current_user: User = Depends(get_current_user)
):
    """Supprimer un projet triennal"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    result = await db.projets_triennaux.delete_one({"id": projet_id, "tenant_id": tenant.id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Projet non trouvé")
    
    return {"message": "Projet supprimé"}


# ==================== ROUTES INTERVENTIONS ====================

@router.post("/{tenant_slug}/rapports/interventions")
async def create_intervention(
    tenant_slug: str,
    intervention: InterventionCreate,
    current_user: User = Depends(get_current_user)
):
    """Créer une intervention"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    intervention_doc = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant.id,
        "date": intervention.date,
        "type_intervention": intervention.type_intervention,
        "description": intervention.description,
        "duree_minutes": intervention.duree_minutes,
        "ressources_utilisees": intervention.ressources_utilisees,
        "created_at": datetime.now(timezone.utc),
        "created_by": current_user.id
    }
    
    await db.interventions.insert_one(intervention_doc)
    intervention_doc.pop("_id", None)
    
    return intervention_doc


@router.get("/{tenant_slug}/rapports/interventions")
async def get_interventions(tenant_slug: str, current_user: User = Depends(get_current_user)):
    """Récupérer toutes les interventions"""
    tenant = await get_tenant_from_slug(tenant_slug)
    
    interventions = await db.interventions.find({"tenant_id": tenant.id}).to_list(1000)
    return [clean_mongo_doc(i) for i in interventions]


# ==================== ROUTES DASHBOARD INTERNE ====================

@router.get("/{tenant_slug}/rapports/dashboard-interne")
async def get_dashboard_interne(tenant_slug: str, current_user: User = Depends(get_current_user)):
    """Données du dashboard interne"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    try:
        # Récupérer les données
        users = await db.users.find({"tenant_id": tenant.id}).to_list(1000)
        assignations = await db.assignations.find({"tenant_id": tenant.id}).to_list(10000)
        types_garde = await db.types_garde.find({"tenant_id": tenant.id}).to_list(100)
        disponibilites = await db.disponibilites.find({"tenant_id": tenant.id}).to_list(10000)
        
        # Créer map des types de garde
        types_map = {t["id"]: t for t in types_garde}
        
        # Calculer la période actuelle
        now = datetime.now()
        debut_mois = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        periode = now.strftime("%B %Y")
        
        # Calculer heures travaillées ce mois et coût salarial
        heures_mois = 0
        cout_salarial_mois = 0
        users_map = {u["id"]: u for u in users}
        
        for assignation in assignations:
            try:
                date_str = assignation.get("date", "")
                if date_str and date_str.startswith(now.strftime("%Y-%m")):
                    type_garde_id = assignation.get("type_garde_id")
                    user_id = assignation.get("user_id")
                    
                    type_garde = types_map.get(type_garde_id, {})
                    user = users_map.get(user_id, {})
                    
                    duree = type_garde.get("duree_heures", 8)
                    taux_horaire = user.get("taux_horaire", 25)
                    
                    heures_mois += duree
                    cout_salarial_mois += duree * taux_horaire
            except:
                pass
        
        # Calculer pompiers disponibles (ont au moins une disponibilité ce mois)
        users_disponibles = set()
        for dispo in disponibilites:
            try:
                date_str = dispo.get("date", "")
                if date_str and date_str.startswith(now.strftime("%Y-%m")):
                    if dispo.get("type") != "indisponibilite":
                        users_disponibles.add(dispo.get("user_id"))
            except:
                pass
        
        pompiers_disponibles = len(users_disponibles)
        total_pompiers = len(users)
        
        return {
            "periode": periode,
            "heures_travaillees_mois": heures_mois,
            "cout_salarial_mois": cout_salarial_mois,
            "pompiers_disponibles": pompiers_disponibles,
            "total_pompiers": total_pompiers,
            # Structure additionnelle pour compatibilité
            "personnel": {
                "total": total_pompiers,
                "actifs": len([u for u in users if u.get("statut") == "Actif"])
            }
        }
    except Exception as e:
        logger.error(f"Erreur dashboard interne: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ROUTES COÛTS SALARIAUX ====================

@router.get("/{tenant_slug}/rapports/couts-salariaux")
async def get_rapport_couts_salariaux(
    tenant_slug: str,
    mois: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Rapport des coûts salariaux"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    try:
        # Mois par défaut: mois courant
        if not mois:
            mois = datetime.now().strftime("%Y-%m")
        
        # Récupérer les assignations du mois
        assignations = await db.assignations.find({
            "tenant_id": tenant.id,
            "date": {"$regex": f"^{mois}"}
        }).to_list(10000)
        
        # Récupérer les utilisateurs et types de garde
        users = await db.users.find({"tenant_id": tenant.id}).to_list(1000)
        types_garde = await db.types_garde.find({"tenant_id": tenant.id}).to_list(100)
        
        users_map = {u["id"]: u for u in users}
        types_map = {t["id"]: t for t in types_garde}
        
        # Calculer les coûts par employé
        couts_par_employe = {}
        for a in assignations:
            user_id = a.get("user_id")
            type_garde_id = a.get("type_garde_id")
            
            user = users_map.get(user_id, {})
            type_garde = types_map.get(type_garde_id, {})
            
            taux_horaire = user.get("taux_horaire", 25)
            duree = type_garde.get("duree_heures", 8)
            cout = taux_horaire * duree
            
            if user_id not in couts_par_employe:
                couts_par_employe[user_id] = {
                    "nom": f"{user.get('prenom', '')} {user.get('nom', '')}",
                    "grade": user.get("grade", ""),
                    "heures_totales": 0,
                    "cout_total": 0,
                    "assignations": 0
                }
            
            couts_par_employe[user_id]["heures_totales"] += duree
            couts_par_employe[user_id]["cout_total"] += cout
            couts_par_employe[user_id]["assignations"] += 1
        
        # Totaux
        total_heures = sum(c["heures_totales"] for c in couts_par_employe.values())
        total_cout = sum(c["cout_total"] for c in couts_par_employe.values())
        
        return {
            "mois": mois,
            "par_employe": list(couts_par_employe.values()),
            "totaux": {
                "heures": total_heures,
                "cout": total_cout,
                "nb_employes": len(couts_par_employe)
            }
        }
    except Exception as e:
        logger.error(f"Erreur rapport coûts salariaux: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ROUTES TABLEAU DE BORD BUDGÉTAIRE ====================

@router.get("/{tenant_slug}/rapports/tableau-bord-budgetaire")
async def get_tableau_bord_budgetaire(tenant_slug: str, current_user: User = Depends(get_current_user)):
    """Tableau de bord budgétaire"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    try:
        budgets = await db.budgets.find({"tenant_id": tenant.id}).to_list(1000)
        
        # Grouper par catégorie
        par_categorie = {}
        for b in budgets:
            cat = b.get("categorie", "Autre")
            if cat not in par_categorie:
                par_categorie[cat] = {"prevu": 0, "realise": 0}
            par_categorie[cat]["prevu"] += b.get("montant_prevu", 0)
            par_categorie[cat]["realise"] += b.get("montant_realise", 0)
        
        # Totaux
        total_prevu = sum(c["prevu"] for c in par_categorie.values())
        total_realise = sum(c["realise"] for c in par_categorie.values())
        
        return {
            "par_categorie": par_categorie,
            "totaux": {
                "prevu": total_prevu,
                "realise": total_realise,
                "ecart": total_prevu - total_realise,
                "pourcentage_utilise": (total_realise / total_prevu * 100) if total_prevu > 0 else 0
            }
        }
    except Exception as e:
        logger.error(f"Erreur tableau bord budgétaire: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ROUTES RAPPORT IMMOBILISATIONS ====================

@router.get("/{tenant_slug}/rapports/rapport-immobilisations")
async def get_rapport_immobilisations(tenant_slug: str, current_user: User = Depends(get_current_user)):
    """Rapport détaillé des immobilisations"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    try:
        immobs = await db.immobilisations.find({"tenant_id": tenant.id}).to_list(1000)
        
        # Calculer l'amortissement et séparer véhicules/équipements
        today = datetime.now()
        vehicules = []
        equipements = []
        
        cout_acquisition_total = 0
        cout_entretien_annuel_total = 0
        ages_vehicules = []
        ages_equipements = []
        
        for immob in immobs:
            valeur = immob.get("cout_acquisition", immob.get("valeur_acquisition", 0))
            entretien = immob.get("cout_entretien_annuel", 0)
            date_acq_str = immob.get("date_acquisition", "")
            type_immob = immob.get("type_immobilisation", immob.get("categorie", "equipement_majeur"))
            
            # Calculer l'âge
            age = 0
            try:
                date_acq = datetime.strptime(date_acq_str, "%Y-%m-%d")
                age = round((today - date_acq).days / 365, 1)
            except:
                pass
            
            cout_acquisition_total += valeur
            cout_entretien_annuel_total += entretien
            
            item = {
                **clean_mongo_doc(immob),
                "cout_acquisition": valeur,
                "cout_entretien_annuel": entretien,
                "age": age
            }
            
            if type_immob == "vehicule":
                vehicules.append(item)
                if age > 0:
                    ages_vehicules.append(age)
            else:
                equipements.append(item)
                if age > 0:
                    ages_equipements.append(age)
        
        # Calculer les moyennes d'âge
        age_moyen_vehicules = round(sum(ages_vehicules) / len(ages_vehicules), 1) if ages_vehicules else 0
        age_moyen_equipements = round(sum(ages_equipements) / len(ages_equipements), 1) if ages_equipements else 0
        
        return {
            "statistiques": {
                "nombre_vehicules": len(vehicules),
                "nombre_equipements": len(equipements),
                "cout_acquisition_total": cout_acquisition_total,
                "cout_entretien_annuel_total": cout_entretien_annuel_total,
                "age_moyen_vehicules": age_moyen_vehicules,
                "age_moyen_equipements": age_moyen_equipements
            },
            "vehicules": vehicules,
            "equipements": equipements
        }
    except Exception as e:
        logger.error(f"Erreur rapport immobilisations: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ROUTES IMPORT CSV ====================

@router.post("/{tenant_slug}/rapports/import-csv")
async def import_rapports_csv(
    tenant_slug: str,
    file: UploadFile = File(...),
    type_import: str = "budgets",
    current_user: User = Depends(get_current_user)
):
    """Import CSV pour rapports (budgets, immobilisations, etc.)"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    try:
        content = await file.read()
        decoded = content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(decoded))
        
        imported = 0
        errors = []
        
        for row in reader:
            try:
                if type_import == "budgets":
                    doc = {
                        "id": str(uuid.uuid4()),
                        "tenant_id": tenant.id,
                        "categorie": row.get("categorie", ""),
                        "montant_prevu": float(row.get("montant_prevu", 0)),
                        "montant_realise": float(row.get("montant_realise", 0)),
                        "description": row.get("description", ""),
                        "created_at": datetime.now(timezone.utc)
                    }
                    await db.budgets.insert_one(doc)
                elif type_import == "immobilisations":
                    doc = {
                        "id": str(uuid.uuid4()),
                        "tenant_id": tenant.id,
                        "nom": row.get("nom", ""),
                        "categorie": row.get("categorie", ""),
                        "valeur_acquisition": float(row.get("valeur_acquisition", 0)),
                        "date_acquisition": row.get("date_acquisition", ""),
                        "duree_amortissement": int(row.get("duree_amortissement", 5)),
                        "created_at": datetime.now(timezone.utc)
                    }
                    await db.immobilisations.insert_one(doc)
                
                imported += 1
            except Exception as e:
                errors.append(str(e))
        
        return {
            "message": f"{imported} enregistrements importés",
            "imported": imported,
            "errors": errors[:10]
        }
    except Exception as e:
        logger.error(f"Erreur import CSV: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ROUTES EXPORT PDF ====================

@router.get("/{tenant_slug}/rapports/export-dashboard-pdf")
async def export_dashboard_pdf(tenant_slug: str, current_user: User = Depends(get_current_user)):
    """Export PDF du dashboard"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    try:
        # Récupérer les données
        users = await db.users.find({"tenant_id": tenant.id}).to_list(1000)
        budgets = await db.budgets.find({"tenant_id": tenant.id}).to_list(1000)
        
        # Créer le PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []
        
        # Titre
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor("#DC2626"),
            spaceAfter=30
        )
        elements.append(Paragraph("Rapport Dashboard", title_style))
        elements.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Stats personnel
        elements.append(Paragraph("Personnel", styles['Heading2']))
        data = [
            ["Métrique", "Valeur"],
            ["Total", str(len(users))],
            ["Actifs", str(len([u for u in users if u.get("statut") == "Actif"]))],
        ]
        table = Table(data, colWidths=[200, 100])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#DC2626")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        # Stats budgets
        elements.append(Paragraph("Budgets", styles['Heading2']))
        total_prevu = sum(b.get("montant_prevu", 0) for b in budgets)
        total_realise = sum(b.get("montant_realise", 0) for b in budgets)
        data = [
            ["Métrique", "Valeur"],
            ["Budget prévu", f"{total_prevu:,.2f} $"],
            ["Budget réalisé", f"{total_realise:,.2f} $"],
            ["Écart", f"{total_prevu - total_realise:,.2f} $"],
        ]
        table = Table(data, colWidths=[200, 100])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#DC2626")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=dashboard_{datetime.now().strftime('%Y%m%d')}.pdf"}
        )
    except Exception as e:
        logger.error(f"Erreur export dashboard PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{tenant_slug}/rapports/export-salaires-pdf")
async def export_salaires_pdf(
    tenant_slug: str,
    mois: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Export PDF des salaires"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not mois:
        mois = datetime.now().strftime("%Y-%m")
    
    try:
        # Récupérer les données
        assignations = await db.assignations.find({
            "tenant_id": tenant.id,
            "date": {"$regex": f"^{mois}"}
        }).to_list(10000)
        
        users = await db.users.find({"tenant_id": tenant.id}).to_list(1000)
        types_garde = await db.types_garde.find({"tenant_id": tenant.id}).to_list(100)
        
        users_map = {u["id"]: u for u in users}
        types_map = {t["id"]: t for t in types_garde}
        
        # Calculer par employé
        salaires = {}
        for a in assignations:
            user_id = a.get("user_id")
            type_garde_id = a.get("type_garde_id")
            
            user = users_map.get(user_id, {})
            type_garde = types_map.get(type_garde_id, {})
            
            taux = user.get("taux_horaire", 25)
            duree = type_garde.get("duree_heures", 8)
            
            if user_id not in salaires:
                salaires[user_id] = {
                    "nom": f"{user.get('prenom', '')} {user.get('nom', '')}",
                    "heures": 0,
                    "montant": 0
                }
            
            salaires[user_id]["heures"] += duree
            salaires[user_id]["montant"] += taux * duree
        
        # Créer le PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []
        
        # Titre
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor("#DC2626"),
            spaceAfter=30
        )
        elements.append(Paragraph(f"Rapport des Salaires - {mois}", title_style))
        elements.append(Spacer(1, 20))
        
        # Tableau
        data = [["Employé", "Heures", "Montant"]]
        for s in salaires.values():
            data.append([s["nom"], f"{s['heures']}h", f"{s['montant']:,.2f} $"])
        
        # Totaux
        total_heures = sum(s["heures"] for s in salaires.values())
        total_montant = sum(s["montant"] for s in salaires.values())
        data.append(["TOTAL", f"{total_heures}h", f"{total_montant:,.2f} $"])
        
        table = Table(data, colWidths=[250, 80, 100])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#DC2626")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#FEE2E2")),
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=salaires_{mois}.pdf"}
        )
    except Exception as e:
        logger.error(f"Erreur export salaires PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ROUTES EXPORT EXCEL ====================

@router.get("/{tenant_slug}/rapports/export-salaires-excel")
async def export_salaires_excel(
    tenant_slug: str,
    mois: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Export Excel des salaires"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    if not mois:
        mois = datetime.now().strftime("%Y-%m")
    
    try:
        # Récupérer les données
        assignations = await db.assignations.find({
            "tenant_id": tenant.id,
            "date": {"$regex": f"^{mois}"}
        }).to_list(10000)
        
        users = await db.users.find({"tenant_id": tenant.id}).to_list(1000)
        types_garde = await db.types_garde.find({"tenant_id": tenant.id}).to_list(100)
        
        users_map = {u["id"]: u for u in users}
        types_map = {t["id"]: t for t in types_garde}
        
        # Calculer par employé
        salaires = {}
        for a in assignations:
            user_id = a.get("user_id")
            type_garde_id = a.get("type_garde_id")
            
            user = users_map.get(user_id, {})
            type_garde = types_map.get(type_garde_id, {})
            
            taux = user.get("taux_horaire", 25)
            duree = type_garde.get("duree_heures", 8)
            
            if user_id not in salaires:
                salaires[user_id] = {
                    "nom": f"{user.get('prenom', '')} {user.get('nom', '')}",
                    "heures": 0,
                    "montant": 0
                }
            
            salaires[user_id]["heures"] += duree
            salaires[user_id]["montant"] += taux * duree
        
        # Créer Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Salaires"
        
        # Titre
        ws['A1'] = f"Rapport des Salaires - {mois}"
        ws['A1'].font = Font(size=14, bold=True, color="DC2626")
        ws.merge_cells('A1:C1')
        
        # En-têtes
        headers = ["Employé", "Heures", "Montant"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        
        # Données
        row = 4
        for s in salaires.values():
            ws.cell(row=row, column=1, value=s["nom"])
            ws.cell(row=row, column=2, value=s["heures"])
            ws.cell(row=row, column=3, value=s["montant"])
            row += 1
        
        # Totaux
        total_heures = sum(s["heures"] for s in salaires.values())
        total_montant = sum(s["montant"] for s in salaires.values())
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_heures).font = Font(bold=True)
        ws.cell(row=row, column=3, value=total_montant).font = Font(bold=True)
        
        # Largeur colonnes
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=salaires_{mois}.xlsx"}
        )
    except Exception as e:
        logger.error(f"Erreur export salaires Excel: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ROUTES EXPORT PERSONNEL ====================

@router.get("/{tenant_slug}/personnel/export-pdf")
async def export_personnel_pdf(
    tenant_slug: str,
    user_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Export PDF de la liste personnel ou d'un utilisateur"""
    if current_user.role == "employe":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    try:
        # Récupérer les utilisateurs
        if user_id:
            users_data = await db.users.find({"id": user_id, "tenant_id": tenant.id}).to_list(1)
        else:
            users_data = await db.users.find({"tenant_id": tenant.id}).to_list(1000)
        
        # Créer le PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []
        
        # Titre
        title = "Fiche Employé" if user_id else "Liste du Personnel"
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor("#DC2626"),
            spaceAfter=30
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 20))
        
        if user_id and users_data:
            # Fiche individuelle
            user = users_data[0]
            info = [
                ["Champ", "Valeur"],
                ["Nom", f"{user.get('prenom', '')} {user.get('nom', '')}"],
                ["Email", user.get("email", "")],
                ["Grade", user.get("grade", "")],
                ["Matricule", user.get("matricule", "")],
                ["Statut", user.get("statut", "")],
                ["Type d'emploi", user.get("type_emploi", "")],
            ]
            table = Table(info, colWidths=[150, 250])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#DC2626")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
        else:
            # Liste
            data = [["Nom", "Grade", "Statut", "Type"]]
            for u in users_data:
                data.append([
                    f"{u.get('prenom', '')} {u.get('nom', '')}",
                    u.get("grade", ""),
                    u.get("statut", ""),
                    u.get("type_emploi", "")
                ])
            
            table = Table(data, colWidths=[150, 100, 80, 100])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#DC2626")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
            ]))
            elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"fiche_employe_{user_id}.pdf" if user_id else "liste_personnel.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Erreur export personnel PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{tenant_slug}/personnel/export-excel")
async def export_personnel_excel(
    tenant_slug: str,
    user_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Export Excel de la liste personnel ou d'un utilisateur"""
    if current_user.role == "employe":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    try:
        # Récupérer les utilisateurs
        if user_id:
            users_data = await db.users.find({"id": user_id, "tenant_id": tenant.id}).to_list(1)
        else:
            users_data = await db.users.find({"tenant_id": tenant.id}).to_list(1000)
        
        # Créer Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Personnel"
        
        # Titre
        titre = "Fiche Employé" if user_id else "Liste du Personnel"
        ws['A1'] = titre
        ws['A1'].font = Font(size=14, bold=True, color="DC2626")
        ws.merge_cells('A1:F1')
        
        if not user_id:
            # Stats
            total = len(users_data)
            actifs = len([u for u in users_data if u.get("statut") == "Actif"])
            
            ws['A3'] = "Total personnel"
            ws['B3'] = total
            ws['A4'] = "Personnel actif"
            ws['B4'] = actifs
            
            # En-têtes tableau
            headers = ["Nom", "Prénom", "Grade", "Matricule", "Statut", "Type"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
            
            # Données
            for row, u in enumerate(users_data, 7):
                ws.cell(row=row, column=1, value=u.get("nom", ""))
                ws.cell(row=row, column=2, value=u.get("prenom", ""))
                ws.cell(row=row, column=3, value=u.get("grade", ""))
                ws.cell(row=row, column=4, value=u.get("matricule", ""))
                ws.cell(row=row, column=5, value=u.get("statut", ""))
                ws.cell(row=row, column=6, value=u.get("type_emploi", ""))
        else:
            # Fiche individuelle
            if users_data:
                u = users_data[0]
                ws['A3'] = "Nom"
                ws['B3'] = u.get("nom", "")
                ws['A4'] = "Prénom"
                ws['B4'] = u.get("prenom", "")
                ws['A5'] = "Email"
                ws['B5'] = u.get("email", "")
                ws['A6'] = "Grade"
                ws['B6'] = u.get("grade", "")
                ws['A7'] = "Matricule"
                ws['B7'] = u.get("matricule", "")
                ws['A8'] = "Statut"
                ws['B8'] = u.get("statut", "")
                ws['A9'] = "Type d'emploi"
                ws['B9'] = u.get("type_emploi", "")
        
        # Largeur colonnes
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"fiche_employe_{user_id}.xlsx" if user_id else "liste_personnel.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Erreur export personnel Excel: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ROUTES EXPORT DIRECT (COMPATIBLE IFRAME SANDBOX) ====================

@router.get("/{tenant_slug}/exports/download/{file_id}")
async def download_temp_export(tenant_slug: str, file_id: str):
    """Télécharge un fichier d'export temporaire (accès public avec ID unique)"""
    from fastapi.responses import FileResponse
    
    # Nettoyer les vieux fichiers
    cleanup_old_exports()
    
    # Chercher le fichier
    for filename in os.listdir(TEMP_EXPORT_DIR):
        if filename.startswith(file_id):
            filepath = os.path.join(TEMP_EXPORT_DIR, filename)
            
            # Déterminer le type MIME
            if filename.endswith('.pdf'):
                media_type = "application/pdf"
            elif filename.endswith('.xlsx'):
                media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else:
                media_type = "application/octet-stream"
            
            # Extraire le nom original du fichier
            original_name = filename.split('_', 1)[1] if '_' in filename else filename
            
            # Utiliser FileResponse pour un téléchargement direct
            return FileResponse(
                path=filepath,
                media_type=media_type,
                filename=original_name,
                headers={
                    "Access-Control-Allow-Origin": "*",
                    "Cache-Control": "no-cache"
                }
            )
    
    raise HTTPException(status_code=404, detail="Fichier non trouvé ou expiré")


@router.post("/{tenant_slug}/personnel/generate-export")
async def generate_personnel_export(
    tenant_slug: str,
    export_type: str = "pdf",
    user_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Génère un export et retourne une URL de téléchargement direct"""
    if current_user.role == "employe":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    tenant = await get_tenant_from_slug(tenant_slug)
    
    try:
        # Nettoyer les vieux fichiers
        cleanup_old_exports()
        
        # Récupérer les utilisateurs
        if user_id:
            users_data = await db.users.find({"id": user_id, "tenant_id": tenant.id}).to_list(1)
        else:
            users_data = await db.users.find({"tenant_id": tenant.id}).to_list(1000)
        
        # Générer un ID unique pour le fichier
        file_id = str(uuid.uuid4())[:8]
        
        if export_type == "pdf":
            # Créer le PDF
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            styles = getSampleStyleSheet()
            elements = []
            
            title = "Fiche Employé" if user_id else "Liste du Personnel"
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor("#DC2626"),
                spaceAfter=30
            )
            elements.append(Paragraph(title, title_style))
            elements.append(Spacer(1, 20))
            
            if user_id and users_data:
                user = users_data[0]
                info = [
                    ["Champ", "Valeur"],
                    ["Nom", f"{user.get('prenom', '')} {user.get('nom', '')}"],
                    ["Email", user.get("email", "")],
                    ["Grade", user.get("grade", "")],
                    ["Matricule", user.get("matricule", "")],
                    ["Statut", user.get("statut", "")],
                    ["Type d'emploi", user.get("type_emploi", "")],
                ]
                table = Table(info, colWidths=[150, 250])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#DC2626")),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(table)
            else:
                data = [["Nom", "Grade", "Statut", "Type"]]
                for u in users_data:
                    data.append([
                        f"{u.get('prenom', '')} {u.get('nom', '')}",
                        u.get("grade", ""),
                        u.get("statut", ""),
                        u.get("type_emploi", "")
                    ])
                
                table = Table(data, colWidths=[150, 100, 80, 100])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#DC2626")),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                ]))
                elements.append(table)
            
            doc.build(elements)
            buffer.seek(0)
            
            original_name = f"fiche_{user_id}.pdf" if user_id else "liste_personnel.pdf"
            filepath = os.path.join(TEMP_EXPORT_DIR, f"{file_id}_{original_name}")
            
            with open(filepath, 'wb') as f:
                f.write(buffer.read())
        
        else:  # Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Personnel"
            
            titre = "Fiche Employé" if user_id else "Liste du Personnel"
            ws['A1'] = titre
            ws['A1'].font = Font(size=14, bold=True, color="DC2626")
            ws.merge_cells('A1:F1')
            
            if not user_id:
                total = len(users_data)
                actifs = len([u for u in users_data if u.get("statut") == "Actif"])
                
                ws['A3'] = "Total personnel"
                ws['B3'] = total
                ws['A4'] = "Personnel actif"
                ws['B4'] = actifs
                
                headers = ["Nom", "Prénom", "Grade", "Matricule", "Statut", "Type"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=6, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
                
                for row, u in enumerate(users_data, 7):
                    ws.cell(row=row, column=1, value=u.get("nom", ""))
                    ws.cell(row=row, column=2, value=u.get("prenom", ""))
                    ws.cell(row=row, column=3, value=u.get("grade", ""))
                    ws.cell(row=row, column=4, value=u.get("matricule", ""))
                    ws.cell(row=row, column=5, value=u.get("statut", ""))
                    ws.cell(row=row, column=6, value=u.get("type_emploi", ""))
            else:
                if users_data:
                    u = users_data[0]
                    ws['A3'] = "Nom"
                    ws['B3'] = u.get("nom", "")
                    ws['A4'] = "Prénom"
                    ws['B4'] = u.get("prenom", "")
                    ws['A5'] = "Email"
                    ws['B5'] = u.get("email", "")
                    ws['A6'] = "Grade"
                    ws['B6'] = u.get("grade", "")
            
            for col in range(1, 7):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            original_name = f"fiche_{user_id}.xlsx" if user_id else "liste_personnel.xlsx"
            filepath = os.path.join(TEMP_EXPORT_DIR, f"{file_id}_{original_name}")
            wb.save(filepath)
        
        # Construire l'URL de téléchargement
        frontend_url = os.environ.get("FRONTEND_URL", "")
        download_url = f"{frontend_url}/api/{tenant_slug}/exports/download/{file_id}"
        
        return {
            "success": True,
            "download_url": download_url,
            "filename": original_name,
            "expires_in": 300  # 5 minutes
        }
        
    except Exception as e:
        logger.error(f"Erreur génération export: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
