"""
Routes Exports du module Planning
=================================
Extraits de planning.py pour maintenabilité.

Routes:
- GET /{tenant_slug}/planning/exports/pdf
- GET /{tenant_slug}/planning/exports/excel
- GET /{tenant_slug}/planning/exports/ical
- GET /{tenant_slug}/planning/rapport-heures/export-pdf
- GET /{tenant_slug}/planning/rapport-heures/export-excel
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
from io import BytesIO
import uuid
import logging

from routes.dependencies import (
    db,
    get_current_user,
    get_tenant_from_slug,
    User,
    require_permission,
    user_has_module_action
)

from utils.pdf_helpers import (
    create_branded_pdf,
    get_modern_pdf_styles,
    create_pdf_footer_text
)

router = APIRouter(tags=["Planning Exports"])
logger = logging.getLogger(__name__)


@router.get("/{tenant_slug}/planning/exports/pdf")
async def export_planning_pdf(
    tenant_slug: str, 
    periode: str,
    type: str,
    current_user: User = Depends(get_current_user)
):
    """Export du planning en PDF"""
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        from io import BytesIO
        
        tenant = await get_tenant_from_slug(tenant_slug)
        
        # Calculer la période
        if type == 'semaine':
            date_debut = datetime.strptime(periode, '%Y-%m-%d')
            date_fin = date_debut + timedelta(days=6)
        else:  # mois
            year, month = map(int, periode.split('-'))
            date_debut = datetime(year, month, 1)
            if month == 12:
                date_fin = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                date_fin = datetime(year, month + 1, 1) - timedelta(days=1)
        
        # Récupérer les données
        assignations_list = await db.assignations.find({
            "tenant_id": tenant.id,
            "date": {
                "$gte": date_debut.strftime('%Y-%m-%d'),
                "$lte": date_fin.strftime('%Y-%m-%d')
            }
        }).to_list(length=None)
        
        types_garde_list = await db.types_garde.find({"tenant_id": tenant.id}).to_list(length=None)
        users_list = await db.users.find({"tenant_id": tenant.id}).to_list(length=None)
        
        types_map = {t['id']: t for t in types_garde_list}
        users_map = {u['id']: u for u in users_list}
        
        # Créer le PDF avec branding
        buffer, doc, elements = create_branded_pdf(
            tenant,
            pagesize=landscape(letter),
            leftMargin=0.5*inch,
            rightMargin=0.5*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch
        )
        styles = getSampleStyleSheet()
        modern_styles = get_modern_pdf_styles(styles)
        
        # Titre principal
        titre = f"PLANNING DES GARDES - V4"
        elements.append(Paragraph(titre, modern_styles['title']))
        
        # Sous-titre avec période
        type_label = "Semaine" if type == "semaine" else "Mois"
        periode_str = f"{type_label} du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        elements.append(Paragraph(periode_str, modern_styles['subheading']))
        elements.append(Spacer(1, 0.1*inch))
        
        # Ligne de séparation
        from reportlab.platypus import HRFlowable
        elements.append(HRFlowable(width="100%", thickness=1, color=modern_styles['grid'], spaceAfter=0.3*inch))
        
        # Jours français
        jours_fr = ['Lun', 'Mar', 'Mer', 'Jeu', 'Ven', 'Sam', 'Dim']
        
        # Couleurs pour la grille
        HEADER_BG = colors.HexColor('#1F2937')
        PRIMARY_RED = colors.HexColor('#DC2626')
        COMPLETE_GREEN = colors.HexColor('#D1FAE5')
        PARTIAL_YELLOW = colors.HexColor('#FEF3C7')
        VACANT_RED = colors.HexColor('#FEE2E2')
        LIGHT_GRAY = colors.HexColor('#F3F4F6')
        BORDER_COLOR = colors.HexColor('#E5E7EB')
        
        # Trier les types de garde par heure
        types_garde_sorted = sorted(types_garde_list, key=lambda x: x.get('heure_debut', '00:00'))
        
        if type == 'semaine':
            # ===== FORMAT GRILLE SEMAINE =====
            from reportlab.lib.enums import TA_LEFT
            
            # Styles pour les cellules - taille augmentée et leading adapté
            header_style_white = ParagraphStyle('HeaderWhite', fontSize=9, alignment=TA_CENTER, textColor=colors.white, leading=12, wordWrap='CJK')
            garde_cell_style = ParagraphStyle('GardeCell', fontSize=8, alignment=TA_LEFT, textColor=colors.white, leading=11, wordWrap='CJK')
            day_cell_style = ParagraphStyle('DayCell', fontSize=8, alignment=TA_CENTER, leading=11, textColor=colors.HexColor('#1F2937'), wordWrap='CJK')
            
            # Debug: logger les assignations
            logging.warning(f"DEBUG PDF: {len(assignations_list)} assignations trouvées pour {date_debut.strftime('%Y-%m-%d')} à {date_fin.strftime('%Y-%m-%d')}")
            logging.warning(f"DEBUG PDF: {len(types_garde_sorted)} types de garde")
            
            # En-tête : Type de garde + 7 jours
            header_row = [Paragraph("<b>Type de garde</b>", header_style_white)]
            for i in range(7):
                d = date_debut + timedelta(days=i)
                header_row.append(Paragraph(f"<b>{jours_fr[d.weekday()]}</b><br/>{d.strftime('%d/%m')}", header_style_white))
            
            table_data = [header_row]
            cell_colors = []
            
            # Une ligne par type de garde
            for type_garde in types_garde_sorted:
                garde_nom = type_garde.get('nom', 'N/A')
                heure_debut_garde = type_garde.get('heure_debut', '??:??')
                heure_fin_garde = type_garde.get('heure_fin', '??:??')
                personnel_requis = type_garde.get('personnel_requis', 1)
                jours_app = type_garde.get('jours_application', [])
                type_garde_id = type_garde.get('id')
                
                # Première colonne - nom complet sans troncature
                row = [Paragraph(f"<b>{garde_nom}</b><br/><font size='7'>{heure_debut_garde}-{heure_fin_garde}</font>", garde_cell_style)]
                row_colors = [PRIMARY_RED]
                
                for i in range(7):
                    d = date_debut + timedelta(days=i)
                    date_str = d.strftime('%Y-%m-%d')
                    day_name = d.strftime('%A').lower()
                    
                    # Vérifier si applicable ce jour
                    if jours_app and day_name not in jours_app:
                        row.append(Paragraph("-", day_cell_style))
                        row_colors.append(LIGHT_GRAY)
                        continue
                    
                    # Trouver les assignations pour ce jour et ce type de garde
                    assignations_jour = [a for a in assignations_list 
                                        if a.get('date') == date_str and a.get('type_garde_id') == type_garde_id]
                    
                    # Construire les noms des pompiers
                    noms = []
                    for a in assignations_jour:
                        user_id = a.get('user_id')
                        if user_id and user_id in users_map:
                            u = users_map[user_id]
                            prenom = u.get('prenom', '')
                            nom = u.get('nom', '')
                            if prenom or nom:
                                noms.append(f"{prenom[:1]}. {nom}")
                    
                    if noms:
                        cell_text = "<br/>".join(noms[:4])
                        if len(noms) > 4:
                            cell_text += f"<br/><font size='6'>+{len(noms)-4}</font>"
                        row.append(Paragraph(cell_text, day_cell_style))
                        
                        if len(noms) >= personnel_requis:
                            row_colors.append(COMPLETE_GREEN)
                        else:
                            row_colors.append(PARTIAL_YELLOW)
                    else:
                        row.append(Paragraph("<font color='#B91C1C'>Vacant</font>", day_cell_style))
                        row_colors.append(VACANT_RED)
                
                table_data.append(row)
                cell_colors.append(row_colors)
            
            # Largeurs de colonnes - première colonne plus large pour les noms de garde
            page_width = landscape(letter)[0]
            available_width = page_width - 1*inch
            first_col = 1.8*inch
            day_col = (available_width - first_col) / 7
            col_widths = [first_col] + [day_col] * 7
            
            # Hauteur minimale des lignes pour permettre l'affichage des noms
            row_heights = [0.5*inch] + [0.7*inch] * (len(table_data) - 1)
            
            # Créer la table avec hauteurs de lignes
            table = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
            
            # Style de base - WORDWRAP activé pour permettre le retour à la ligne
            style_commands = [
                ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ]
            
            # Appliquer les couleurs par cellule
            for row_idx, colors_row in enumerate(cell_colors, start=1):
                for col_idx, bg_color in enumerate(colors_row):
                    style_commands.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), bg_color))
                    if col_idx == 0:
                        style_commands.append(('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), colors.white))
                        style_commands.append(('ALIGN', (col_idx, row_idx), (col_idx, row_idx), 'LEFT'))
                        style_commands.append(('LEFTPADDING', (col_idx, row_idx), (col_idx, row_idx), 8))
            
            table.setStyle(TableStyle(style_commands))
            elements.append(table)
            
            # Légende
            elements.append(Spacer(1, 0.3*inch))
            legend_style = ParagraphStyle('Legend', fontSize=9, alignment=TA_CENTER, textColor=colors.HexColor('#6B7280'))
            elements.append(Paragraph("Legende: Vert = Complet | Jaune = Partiel | Rouge = Vacant | Gris = Non applicable", legend_style))
            
        elif type == 'mois':
            # ===== FORMAT GRILLE MOIS - Une grille par semaine =====
            from reportlab.platypus import PageBreak
            
            # IMPORTANT: Trouver le lundi de la semaine contenant le 1er du mois
            # weekday() : 0 = lundi, 6 = dimanche
            jours_depuis_lundi = date_debut.weekday()
            current = date_debut - timedelta(days=jours_depuis_lundi)
            
            semaine_num = 1
            page_width = landscape(letter)[0]
            
            # Style pour titre de semaine
            semaine_style = ParagraphStyle(
                'SemaineStyle',
                parent=styles['Heading2'],
                fontSize=12,
                textColor=PRIMARY_RED,
                spaceBefore=10,
                spaceAfter=10,
                fontName='Helvetica-Bold'
            )
            
            while current <= date_fin:
                # Toujours afficher 7 jours (lundi à dimanche)
                fin_semaine = current + timedelta(days=6)
                nb_jours = 7
                
                # Titre de la semaine
                elements.append(Paragraph(
                    f"Semaine {semaine_num} - Du {current.strftime('%d/%m')} au {fin_semaine.strftime('%d/%m/%Y')}",
                    semaine_style
                ))
                
                # Styles pour le format mois - avec retour à la ligne
                from reportlab.lib.enums import TA_LEFT
                header_style_mois = ParagraphStyle('HeaderMois', fontSize=8, alignment=TA_CENTER, textColor=colors.white, leading=10, wordWrap='CJK')
                garde_cell_style_mois = ParagraphStyle('GardeCellMois', fontSize=7, alignment=TA_LEFT, textColor=colors.white, leading=9, wordWrap='CJK')
                day_cell_style_mois = ParagraphStyle('DayCellMois', fontSize=7, alignment=TA_CENTER, leading=9, textColor=colors.HexColor('#1F2937'), wordWrap='CJK')
                
                # En-tête avec des Paragraph pour le retour à la ligne
                header_row = [Paragraph("<b>Type de garde</b>", header_style_mois)]
                for i in range(nb_jours):
                    d = current + timedelta(days=i)
                    header_row.append(Paragraph(f"<b>{jours_fr[d.weekday()]}</b><br/>{d.strftime('%d')}", header_style_mois))
                
                table_data = [header_row]
                cell_colors_mois = []
                
                for type_garde in types_garde_sorted:
                    garde_nom = type_garde.get('nom', 'N/A')  # NE PLUS TRONQUER
                    heure_debut_garde = type_garde.get('heure_debut', '')
                    heure_fin_garde = type_garde.get('heure_fin', '')
                    personnel_requis = type_garde.get('personnel_requis', 1)
                    jours_app = type_garde.get('jours_application', [])
                    type_garde_id = type_garde.get('id')
                    
                    # Première colonne avec nom complet et horaires
                    row = [Paragraph(f"<b>{garde_nom}</b><br/><font size='6'>{heure_debut_garde}-{heure_fin_garde}</font>", garde_cell_style_mois)]
                    row_colors = [PRIMARY_RED]
                    
                    for i in range(nb_jours):
                        d = current + timedelta(days=i)
                        date_str = d.strftime('%Y-%m-%d')
                        day_name = d.strftime('%A').lower()
                        
                        # Vérifier si le jour est dans le mois affiché
                        if d < date_debut or d > date_fin:
                            # Jour hors du mois - afficher en gris clair
                            row.append(Paragraph("-", day_cell_style_mois))
                            row_colors.append(LIGHT_GRAY)
                            continue
                        
                        if jours_app and day_name not in jours_app:
                            row.append(Paragraph("-", day_cell_style_mois))
                            row_colors.append(LIGHT_GRAY)
                            continue
                        
                        assignations_jour = [a for a in assignations_list 
                                            if a['date'] == date_str and a['type_garde_id'] == type_garde_id]
                        
                        # Construire les noms des pompiers
                        noms = []
                        for a in assignations_jour:
                            user_id = a.get('user_id')
                            if user_id and user_id in users_map:
                                u = users_map[user_id]
                                prenom = u.get('prenom', '')
                                nom = u.get('nom', '')
                                if prenom or nom:
                                    noms.append(f"{prenom[:1]}. {nom}")
                        
                        if noms:
                            cell_text = "<br/>".join(noms[:3])  # Max 3 noms pour le mois
                            if len(noms) > 3:
                                cell_text += f"<br/><font size='5'>+{len(noms)-3}</font>"
                            row.append(Paragraph(cell_text, day_cell_style_mois))
                            
                            if len(noms) >= personnel_requis:
                                row_colors.append(COMPLETE_GREEN)
                            else:
                                row_colors.append(PARTIAL_YELLOW)
                        else:
                            row.append(Paragraph("<font color='#B91C1C'>Vacant</font>", day_cell_style_mois))
                            row_colors.append(VACANT_RED)
                    
                    table_data.append(row)
                    cell_colors_mois.append(row_colors)
                
                # Largeurs - première colonne plus large pour afficher le nom complet
                available_width = page_width - 1*inch
                first_col = 1.6*inch  # Augmenté pour les noms de garde complets
                day_col = (available_width - first_col) / nb_jours
                col_widths = [first_col] + [day_col] * nb_jours
                
                # Hauteurs de lignes - plus hautes pour afficher les noms des pompiers
                row_heights = [0.4*inch] + [0.6*inch] * (len(table_data) - 1)
                
                table = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
                
                style_commands = [
                    ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('GRID', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ('LEFTPADDING', (0, 0), (-1, -1), 3),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ]
                
                for row_idx, colors_row in enumerate(cell_colors_mois, start=1):
                    for col_idx, bg_color in enumerate(colors_row):
                        style_commands.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), bg_color))
                        if col_idx == 0:
                            style_commands.append(('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), colors.white))
                            style_commands.append(('ALIGN', (col_idx, row_idx), (col_idx, row_idx), 'LEFT'))
                            style_commands.append(('LEFTPADDING', (col_idx, row_idx), (col_idx, row_idx), 6))
                
                table.setStyle(TableStyle(style_commands))
                elements.append(table)
                elements.append(Spacer(1, 0.2*inch))
                
                # Avancer au lundi de la semaine suivante
                current = current + timedelta(days=7)
                semaine_num += 1
                
                if current <= date_fin:
                    elements.append(PageBreak())
        
        # Footer
        def add_footer(canvas, doc_obj):
            canvas.saveState()
            canvas.setStrokeColor(colors.HexColor('#e2e8f0'))
            canvas.setLineWidth(1)
            canvas.line(0.5*inch, 0.5*inch, landscape(letter)[0] - 0.5*inch, 0.5*inch)
            canvas.setFont('Helvetica', 9)
            canvas.setFillColor(colors.HexColor('#64748b'))
            footer_text = f"ProFireManager - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            canvas.drawCentredString(landscape(letter)[0] / 2, 0.35*inch, footer_text)
            canvas.setFont('Helvetica', 8)
            canvas.drawRightString(landscape(letter)[0] - 0.5*inch, 0.35*inch, f"Page {doc_obj.page}")
            canvas.restoreState()
        
        doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=planning_{type}_{periode}.pdf"}
        )
        
    except Exception as e:
        import traceback
        logging.error(f"Erreur export PDF: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Erreur export PDF: {str(e)}")


# GET export-excel
@router.get("/{tenant_slug}/planning/exports/excel")
async def export_planning_excel(
    tenant_slug: str, 
    periode: str,
    type: str,
    current_user: User = Depends(get_current_user)
):
    """Export du planning en Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO
        
        tenant = await get_tenant_from_slug(tenant_slug)
        
        # Calculer la période
        if type == 'semaine':
            date_debut = datetime.strptime(periode, '%Y-%m-%d')
            date_fin = date_debut + timedelta(days=6)
        else:
            year, month = map(int, periode.split('-'))
            date_debut = datetime(year, month, 1)
            if month == 12:
                date_fin = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                date_fin = datetime(year, month + 1, 1) - timedelta(days=1)
        
        assignations_list = await db.assignations.find({
            "tenant_id": tenant.id,
            "date": {
                "$gte": date_debut.strftime('%Y-%m-%d'),
                "$lte": date_fin.strftime('%Y-%m-%d')
            }
        }).to_list(length=None)
        
        types_garde_list = await db.types_garde.find({"tenant_id": tenant.id}).to_list(length=None)
        users_list = await db.users.find({"tenant_id": tenant.id}).to_list(length=None)
        
        types_map = {t['id']: t for t in types_garde_list}
        users_map = {u['id']: u for u in users_list}
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Planning {type}"
        
        header_fill = PatternFill(start_color="FCA5A5", end_color="FCA5A5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws.merge_cells('A1:H1')
        ws['A1'] = f"Planning des Gardes - {type.capitalize()}"
        ws['A1'].font = Font(bold=True, size=16, color="EF4444")
        ws['A1'].alignment = center_alignment
        
        ws.merge_cells('A2:H2')
        ws['A2'] = f"Du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        ws['A2'].alignment = center_alignment
        
        row = 4
        if type == 'semaine':
            headers = ['Type de Garde', 'Horaires'] + [(date_debut + timedelta(days=i)).strftime('%a %d/%m') for i in range(7)]
        else:
            headers = ['Date', 'Jour', 'Type de Garde', 'Horaires', 'Personnel', 'Requis', 'Assignés', 'Statut']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        row += 1
        
        if type == 'semaine':
            for type_garde in sorted(types_garde_list, key=lambda x: x.get('heure_debut', '')):
                ws.cell(row=row, column=1, value=type_garde['nom'])
                ws.cell(row=row, column=2, value=f"{type_garde.get('heure_debut', '')} - {type_garde.get('heure_fin', '')}")
                
                for i in range(7):
                    current_date = (date_debut + timedelta(days=i)).strftime('%Y-%m-%d')
                    assignations_jour = [a for a in assignations_list if a['date'] == current_date and a['type_garde_id'] == type_garde['id']]
                    
                    noms = [f"{users_map[a['user_id']]['prenom']} {users_map[a['user_id']]['nom']}" 
                           for a in assignations_jour if a['user_id'] in users_map]
                    
                    cell_text = '\n'.join(noms) if noms else 'Vacant'
                    cell = ws.cell(row=row, column=3+i, value=cell_text)
                    cell.alignment = center_alignment
                    cell.border = border
                    
                    if len(noms) >= type_garde.get('personnel_requis', 1):
                        cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    elif noms:
                        cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                
                row += 1
        else:
            current = date_debut
            while current <= date_fin:
                date_str = current.strftime('%Y-%m-%d')
                jour_fr = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche'][current.weekday()]
                
                for type_garde in types_garde_list:
                    assignations_jour = [a for a in assignations_list if a['date'] == date_str and a['type_garde_id'] == type_garde['id']]
                    
                    noms = [f"{users_map[a['user_id']]['prenom']} {users_map[a['user_id']]['nom']}" 
                           for a in assignations_jour if a['user_id'] in users_map]
                    
                    personnel_str = ', '.join(noms) if noms else 'Aucun'
                    requis = type_garde.get('personnel_requis', 1)
                    assignes = len(noms)
                    statut = 'Complet' if assignes >= requis else 'Partiel' if noms else 'Vacant'
                    
                    ws.cell(row=row, column=1, value=current.strftime('%d/%m/%Y'))
                    ws.cell(row=row, column=2, value=jour_fr)
                    ws.cell(row=row, column=3, value=type_garde['nom'])
                    ws.cell(row=row, column=4, value=f"{type_garde.get('heure_debut', '')} - {type_garde.get('heure_fin', '')}")
                    ws.cell(row=row, column=5, value=personnel_str)
                    ws.cell(row=row, column=6, value=requis)
                    ws.cell(row=row, column=7, value=assignes)
                    status_cell = ws.cell(row=row, column=8, value=statut)
                    
                    if statut == 'Complet':
                        status_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                    elif statut == 'Partiel':
                        status_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                    else:
                        status_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    
                    for col in range(1, 9):
                        ws.cell(row=row, column=col).border = border
                        ws.cell(row=row, column=col).alignment = center_alignment
                    
                    row += 1
                
                current += timedelta(days=1)
        
        # Définir les largeurs de colonnes fixes pour éviter les erreurs avec MergedCell
        column_widths = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        default_widths = [12, 12, 18, 15, 25, 10, 10, 12, 12]
        for i, col_letter in enumerate(column_widths):
            if i < len(default_widths):
                ws.column_dimensions[col_letter].width = default_widths[i]
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=planning_{type}_{periode}.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur export Excel: {str(e)}")



# GET export-ical - Export des gardes au format iCalendar (.ics)
@router.get("/{tenant_slug}/planning/exports/ical")
async def export_planning_ical(
    tenant_slug: str,
    date_debut: str,
    date_fin: str,
    current_user: User = Depends(get_current_user)
):
    """
    Export des gardes de l'utilisateur connecté au format iCalendar (.ics)
    Compatible avec Google Calendar, Apple Calendar, Outlook, etc.
    """
    try:
        tenant = await get_tenant_from_slug(tenant_slug)
        
        # Récupérer les assignations de l'utilisateur connecté
        assignations_list = await db.assignations.find({
            "tenant_id": tenant.id,
            "user_id": current_user.id,
            "date": {
                "$gte": date_debut,
                "$lte": date_fin
            }
        }, {"_id": 0}).to_list(length=None)
        
        # Récupérer les types de garde pour avoir les horaires
        types_garde_list = await db.types_garde.find({"tenant_id": tenant.id}, {"_id": 0}).to_list(length=None)
        types_map = {t['id']: t for t in types_garde_list}
        
        # Générer le contenu iCalendar
        ical_lines = [
            "BEGIN:VCALENDAR",
            "VERSION:2.0",
            "PRODID:-//ProFireManager//Planning Export//FR",
            "CALSCALE:GREGORIAN",
            "METHOD:PUBLISH",
            f"X-WR-CALNAME:Mes Gardes - {tenant.nom}",
            "X-WR-TIMEZONE:America/Montreal"
        ]
        
        for assignation in assignations_list:
            type_garde_id = assignation.get('type_garde_id')
            type_garde = types_map.get(type_garde_id, {})
            
            garde_nom = type_garde.get('nom', 'Garde')
            heure_debut = type_garde.get('heure_debut', '08:00')
            heure_fin = type_garde.get('heure_fin', '20:00')
            date_str = assignation.get('date')  # Format: YYYY-MM-DD
            
            # Parser la date
            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            except:
                continue
            
            # Parser les heures
            try:
                h_debut, m_debut = map(int, heure_debut.split(':'))
                h_fin, m_fin = map(int, heure_fin.split(':'))
            except:
                h_debut, m_debut = 8, 0
                h_fin, m_fin = 20, 0
            
            # Créer les datetime de début et fin
            dt_debut = date_obj.replace(hour=h_debut, minute=m_debut)
            dt_fin = date_obj.replace(hour=h_fin, minute=m_fin)
            
            # Si l'heure de fin est avant l'heure de début, c'est une garde de nuit (finit le lendemain)
            if dt_fin <= dt_debut:
                dt_fin = dt_fin + timedelta(days=1)
            
            # Générer un UID unique pour l'événement
            uid = f"{assignation.get('id', uuid.uuid4())}@profiremanager.com"
            
            # Formater les dates au format iCal (YYYYMMDDTHHMMSS)
            dtstart = dt_debut.strftime('%Y%m%dT%H%M%S')
            dtend = dt_fin.strftime('%Y%m%dT%H%M%S')
            dtstamp = datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')
            
            # Description de l'événement
            description = f"Type: {garde_nom}\\nHoraire: {heure_debut} - {heure_fin}"
            if type_garde.get('description'):
                description += f"\\n{type_garde.get('description')}"
            
            # Titre de l'événement avec le nom du tenant pour distinguer les casernes
            event_title = f"{tenant.nom} - {garde_nom}"
            
            # Ajouter l'événement
            ical_lines.extend([
                "BEGIN:VEVENT",
                f"UID:{uid}",
                f"DTSTAMP:{dtstamp}",
                f"DTSTART;TZID=America/Montreal:{dtstart}",
                f"DTEND;TZID=America/Montreal:{dtend}",
                f"SUMMARY:{event_title}",
                f"DESCRIPTION:{description}",
                "STATUS:CONFIRMED",
                "TRANSP:OPAQUE",
                "BEGIN:VALARM",
                "ACTION:DISPLAY",
                "DESCRIPTION:Rappel de garde",
                "TRIGGER:-PT1H",
                "END:VALARM",
                "END:VEVENT"
            ])
        
        ical_lines.append("END:VCALENDAR")
        
        # Joindre les lignes avec CRLF (standard iCal)
        ical_content = "\r\n".join(ical_lines)
        
        # Créer le buffer
        buffer = BytesIO(ical_content.encode('utf-8'))
        buffer.seek(0)
        
        # Nom du fichier
        filename = f"gardes_{current_user.prenom}_{current_user.nom}_{date_debut}_{date_fin}.ics"
        filename = filename.replace(' ', '_')
        
        return StreamingResponse(
            buffer,
            media_type="text/calendar",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logging.error(f"Erreur export iCal: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur export iCal: {str(e)}")


# ===== RAPPORT D'HEURES =====





@router.get("/{tenant_slug}/planning/rapport-heures/export-pdf")
async def export_rapport_heures_pdf(
    tenant_slug: str,
    date_debut: str,
    date_fin: str,
    current_user: User = Depends(get_current_user)
):
    """Génère le PDF du rapport d'heures pour impression"""
    tenant = await get_tenant_from_slug(tenant_slug)
    await require_permission(tenant.id, current_user, "planning", "exporter")
    
    # Convertir date_debut en format mois pour appeler get_rapport_heures
    mois = date_debut[:7]  # YYYY-MM
    
    # Récupérer les données du rapport en créant un faux current_user pour l'appel interne
    # On va récupérer les données directement ici
    from datetime import datetime as dt
    debut_dt = dt.strptime(date_debut, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    fin_dt = dt.strptime(date_fin, "%Y-%m-%d").replace(tzinfo=timezone.utc) + timedelta(days=1)
    
    # Récupérer toutes les assignations de la période
    assignations = await db.assignations.find({
        "tenant_id": tenant.id,
        "date": {"$gte": debut_dt.strftime("%Y-%m-%d"), "$lt": fin_dt.strftime("%Y-%m-%d")}
    }).to_list(10000)
    
    # Récupérer tous les utilisateurs (actifs ou pas - pour afficher tous ceux qui ont des assignations)
    users = await db.users.find({"tenant_id": tenant.id}).to_list(1000)
    users_dict = {u["id"]: u for u in users}
    
    # Récupérer les types de garde pour déterminer est_externe et duree
    types_garde = await db.types_garde.find({"tenant_id": tenant.id}).to_list(1000)
    types_garde_dict = {tg["id"]: tg for tg in types_garde}
    
    # Calculer les heures par employé
    heures_par_employe = {}
    for assign in assignations:
        uid = assign.get("user_id")
        if uid not in heures_par_employe:
            heures_par_employe[uid] = {"internes": 0, "externes": 0}
        
        # Récupérer la durée depuis l'assignation OU depuis le type de garde
        type_garde_id = assign.get("type_garde_id")
        type_garde = types_garde_dict.get(type_garde_id, {})
        duree = assign.get("duree_heures") or type_garde.get("duree_heures", 0)
        
        # Déterminer si externe depuis l'assignation OU depuis le type de garde
        est_externe = assign.get("est_externe") or type_garde.get("est_garde_externe", False)
        
        if est_externe:
            heures_par_employe[uid]["externes"] += duree
        else:
            heures_par_employe[uid]["internes"] += duree
    
    # Construire la liste des employés
    employes = []
    for uid, heures in heures_par_employe.items():
        user = users_dict.get(uid, {})
        employes.append({
            "nom_complet": f"{user.get('prenom', '')} {user.get('nom', '')}".strip() or "Inconnu",
            "type_emploi": user.get("type_emploi", "temps_plein"),
            "grade": user.get("grade", "N/A"),
            "heures_internes": heures["internes"],
            "heures_externes": heures["externes"],
            "total_heures": heures["internes"] + heures["externes"]
        })
    
    # Trier par nom
    employes.sort(key=lambda x: x["nom_complet"])
    
    # Statistiques
    total_internes = sum(e["heures_internes"] for e in employes)
    total_externes = sum(e["heures_externes"] for e in employes)
    nb_employes = len(employes)
    
    statistiques = {
        "nombre_employes": nb_employes,
        "total_heures_planifiees": total_internes + total_externes,
        "moyenne_heures_internes": round(total_internes / nb_employes, 1) if nb_employes > 0 else 0,
        "moyenne_heures_externes": round(total_externes / nb_employes, 1) if nb_employes > 0 else 0
    }
    
    rapport_response = {"employes": employes, "statistiques": statistiques}
    
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    
    # Utiliser la fonction helper pour créer un PDF brandé
    buffer, doc, elements = create_branded_pdf(tenant, pagesize=A4)
    styles = getSampleStyleSheet()
    modern_styles = get_modern_pdf_styles(styles)
    
    # Titre
    elements.append(Paragraph("Rapport d'Heures", modern_styles['title']))
    
    # Période
    debut_dt = datetime.strptime(date_debut, "%Y-%m-%d")
    fin_dt = datetime.strptime(date_fin, "%Y-%m-%d")
    periode_text = f"Période: {debut_dt.strftime('%d/%m/%Y')} - {fin_dt.strftime('%d/%m/%Y')}"
    elements.append(Paragraph(periode_text, modern_styles['subheading']))
    
    # Tableau des employés
    table_data = [
        ['Employé', 'Type', 'Grade', 'H. Internes', 'H. Externes', 'Total']
    ]
    
    for emp in rapport_response["employes"]:
        type_emploi = emp.get("type_emploi", "temps_plein")
        if type_emploi == "temps_plein":
            type_emploi_abbr = "TP"
        elif type_emploi == "temporaire":
            type_emploi_abbr = "Tempo"
        else:
            type_emploi_abbr = "TPart"
        table_data.append([
            emp["nom_complet"],
            type_emploi_abbr,
            emp.get("grade", "N/A"),
            f"{emp['heures_internes']}h",
            f"{emp['heures_externes']}h",
            f"{emp['total_heures']}h"
        ])
    
    table = Table(table_data, colWidths=[2.5*inch, 0.6*inch, 1.2*inch, 1*inch, 1*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), modern_styles['primary_color']),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, modern_styles['grid']),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, modern_styles['bg_light']])
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Statistiques
    stats = rapport_response["statistiques"]
    stats_text = f"""
    <b>Statistiques Globales</b><br/>
    Nombre d'employés: {stats['nombre_employes']}<br/>
    Total heures planifiées: {stats['total_heures_planifiees']}h<br/>
    Moyenne heures internes: {stats['moyenne_heures_internes']}h<br/>
    Moyenne heures externes: {stats['moyenne_heures_externes']}h
    """
    elements.append(Paragraph(stats_text, styles['Normal']))
    
    # Ajouter footer ProFireManager
    elements.append(Spacer(1, 0.5*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    footer_text = create_pdf_footer_text(tenant)
    if footer_text:
        elements.append(Paragraph(footer_text, footer_style))
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=rapport_heures_{date_debut}_{date_fin}.pdf"}
    )




@router.get("/{tenant_slug}/planning/rapport-heures/export-excel")
async def export_rapport_heures_excel(
    tenant_slug: str,
    date_debut: str,
    date_fin: str,
    current_user: User = Depends(get_current_user)
):
    """Génère l'Excel du rapport d'heures"""
    tenant = await get_tenant_from_slug(tenant_slug)
    await require_permission(tenant.id, current_user, "planning", "exporter")
    
    # Récupérer les données directement (même logique que pour le PDF)
    from datetime import datetime as dt
    debut_dt = dt.strptime(date_debut, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    fin_dt = dt.strptime(date_fin, "%Y-%m-%d").replace(tzinfo=timezone.utc) + timedelta(days=1)
    
    assignations = await db.assignations.find({
        "tenant_id": tenant.id,
        "date": {"$gte": debut_dt.strftime("%Y-%m-%d"), "$lt": fin_dt.strftime("%Y-%m-%d")}
    }).to_list(10000)
    
    users = await db.users.find({"tenant_id": tenant.id}).to_list(1000)
    users_dict = {u["id"]: u for u in users}
    
    # Récupérer les types de garde pour déterminer est_externe et duree
    types_garde = await db.types_garde.find({"tenant_id": tenant.id}).to_list(1000)
    types_garde_dict = {tg["id"]: tg for tg in types_garde}
    
    heures_par_employe = {}
    for assign in assignations:
        uid = assign.get("user_id")
        if uid not in heures_par_employe:
            heures_par_employe[uid] = {"internes": 0, "externes": 0}
        
        # Récupérer la durée depuis l'assignation OU depuis le type de garde
        type_garde_id = assign.get("type_garde_id")
        type_garde = types_garde_dict.get(type_garde_id, {})
        duree = assign.get("duree_heures") or type_garde.get("duree_heures", 0)
        
        # Déterminer si externe depuis l'assignation OU depuis le type de garde
        est_externe = assign.get("est_externe") or type_garde.get("est_garde_externe", False)
        
        if est_externe:
            heures_par_employe[uid]["externes"] += duree
        else:
            heures_par_employe[uid]["internes"] += duree
    
    employes = []
    for uid, heures in heures_par_employe.items():
        user = users_dict.get(uid, {})
        employes.append({
            "nom_complet": f"{user.get('prenom', '')} {user.get('nom', '')}".strip() or "Inconnu",
            "type_emploi": user.get("type_emploi", "temps_plein"),
            "grade": user.get("grade", "N/A"),
            "heures_internes": heures["internes"],
            "heures_externes": heures["externes"],
            "total_heures": heures["internes"] + heures["externes"]
        })
    
    employes.sort(key=lambda x: x["nom_complet"])
    
    total_internes = sum(e["heures_internes"] for e in employes)
    total_externes = sum(e["heures_externes"] for e in employes)
    nb_employes = len(employes)
    
    statistiques = {
        "nombre_employes": nb_employes,
        "total_heures_planifiees": total_internes + total_externes,
        "moyenne_heures_internes": round(total_internes / nb_employes, 1) if nb_employes > 0 else 0,
        "moyenne_heures_externes": round(total_externes / nb_employes, 1) if nb_employes > 0 else 0
    }
    
    rapport_response = {"employes": employes, "statistiques": statistiques}
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport Heures"
    
    # Styles
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Titre
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "Rapport d'Heures"
    title_cell.font = Font(size=16, bold=True, color="DC2626")
    title_cell.alignment = center_alignment
    
    # Période
    ws.merge_cells('A2:F2')
    periode_cell = ws['A2']
    debut_dt = datetime.strptime(date_debut, "%Y-%m-%d")
    fin_dt = datetime.strptime(date_fin, "%Y-%m-%d")
    periode_cell.value = f"Période: {debut_dt.strftime('%d/%m/%Y')} - {fin_dt.strftime('%d/%m/%Y')}"
    periode_cell.alignment = center_alignment
    
    # En-têtes du tableau
    headers = ['Employé', 'Type', 'Grade', 'H. Internes', 'H. Externes', 'Total']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_alignment
    
    # Données
    row = 5
    for emp in rapport_response["employes"]:
        type_emploi = emp.get("type_emploi", "temps_plein")
        if type_emploi == "temps_plein":
            type_emploi_abbr = "TP"
        elif type_emploi == "temporaire":
            type_emploi_abbr = "Tempo"
        else:
            type_emploi_abbr = "TPart"
        ws.cell(row=row, column=1, value=emp["nom_complet"])
        ws.cell(row=row, column=2, value=type_emploi_abbr)
        ws.cell(row=row, column=3, value=emp.get("grade", "N/A"))
        ws.cell(row=row, column=4, value=emp["heures_internes"])
        ws.cell(row=row, column=5, value=emp["heures_externes"])
        ws.cell(row=row, column=6, value=emp["total_heures"])
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = center_alignment
        
        row += 1
    
    # Statistiques
    stats = rapport_response["statistiques"]
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    stats_cell = ws.cell(row=row, column=1)
    stats_cell.value = "Statistiques Globales"
    stats_cell.font = Font(bold=True, size=12)
    
    row += 1
    ws.cell(row=row, column=1, value="Nombre d'employés:")
    ws.cell(row=row, column=2, value=stats['nombre_employes'])
    
    row += 1
    ws.cell(row=row, column=1, value="Total heures planifiées:")
    ws.cell(row=row, column=2, value=stats['total_heures_planifiees'])
    
    row += 1
    ws.cell(row=row, column=1, value="Moyenne heures internes:")
    ws.cell(row=row, column=2, value=stats['moyenne_heures_internes'])
    
    row += 1
    ws.cell(row=row, column=1, value="Moyenne heures externes:")
    ws.cell(row=row, column=2, value=stats['moyenne_heures_externes'])
    
    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=rapport_heures_{date_debut}_{date_fin}.xlsx"}
    )

