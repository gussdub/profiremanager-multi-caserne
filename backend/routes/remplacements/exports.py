"""
Fonctions d'export pour le module Remplacements
- Export PDF avec ReportLab
- Export Excel avec OpenPyXL
"""

from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from datetime import datetime
from typing import Any
import logging

logger = logging.getLogger(__name__)


async def export_remplacements_to_pdf(
    db,
    tenant,
    current_user,
    user_id: str = None
) -> StreamingResponse:
    """
    Export des demandes de remplacement en PDF
    
    Args:
        db: Instance de la base de données
        tenant: Objet tenant
        current_user: Utilisateur courant
        user_id: Filtrer par utilisateur (optionnel)
    
    Returns:
        StreamingResponse avec le PDF
    """
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from io import BytesIO
        
        # Import des helpers PDF depuis server
        import server
        create_branded_pdf = server.create_branded_pdf
        get_modern_pdf_styles = server.get_modern_pdf_styles
        
        # Récupérer les demandes selon les permissions
        if user_id:
            demandes_list = await db.demandes_remplacement.find({
                "tenant_id": tenant.id,
                "demandeur_id": user_id
            }).to_list(length=None)
        else:
            if current_user.role == "employe":
                demandes_list = await db.demandes_remplacement.find({
                    "tenant_id": tenant.id,
                    "demandeur_id": current_user.id
                }).to_list(length=None)
            else:
                demandes_list = await db.demandes_remplacement.find({
                    "tenant_id": tenant.id
                }).to_list(length=None)
        
        # Récupérer les données de référence
        users_list = await db.users.find({"tenant_id": tenant.id}).to_list(length=None)
        types_garde_list = await db.types_garde.find({"tenant_id": tenant.id}).to_list(length=None)
        
        users_map = {u['id']: u for u in users_list}
        types_map = {t['id']: t for t in types_garde_list}
        
        # Créer le PDF
        buffer, doc, elements = create_branded_pdf(tenant, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        modern_styles = get_modern_pdf_styles(styles)
        
        # Titre
        titre = "Demandes de Remplacement"
        if user_id and user_id in users_map:
            titre = f"Demandes de {users_map[user_id]['prenom']} {users_map[user_id]['nom']}"
        
        elements.append(Paragraph(titre, modern_styles['title']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Labels de statut
        statut_labels = {
            'en_attente': 'En attente',
            'en_cours': 'En cours',
            'accepte': 'Accepté',
            'expiree': 'Expirée',
            'annulee': 'Annulée',
            'approuve_manuellement': 'Approuvé'
        }
        
        # Données du tableau
        table_data = [['Date', 'Type Garde', 'Demandeur', 'Statut', 'Priorité', 'Remplaçant', 'Notes']]
        
        for demande in sorted(demandes_list, key=lambda x: x.get('date', ''), reverse=True):
            demandeur = users_map.get(demande['demandeur_id'], {})
            demandeur_nom = f"{demandeur.get('prenom', '')} {demandeur.get('nom', '')}" if demandeur else "N/A"
            
            type_garde = types_map.get(demande.get('type_garde_id', ''), {})
            type_nom = type_garde.get('nom', 'N/A')
            
            statut = statut_labels.get(demande.get('statut', ''), demande.get('statut', ''))
            priorite = 'Urgent' if demande.get('priorite') == 'urgent' else 'Normal'
            
            remplacant = users_map.get(demande.get('remplacant_id', ''), {})
            remplacant_nom = f"{remplacant.get('prenom', '')} {remplacant.get('nom', '')}" if remplacant else "-"
            
            raison = demande.get('raison', '')
            raison_truncated = raison[:30] + '...' if len(raison) > 30 else raison
            
            table_data.append([
                demande.get('date', ''),
                type_nom,
                demandeur_nom,
                statut,
                priorite,
                remplacant_nom,
                raison_truncated
            ])
        
        # Style du tableau
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"remplacements_{tenant.slug}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Erreur export PDF remplacements: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Erreur export PDF: {str(e)}")


async def export_remplacements_to_excel(
    db,
    tenant,
    current_user,
    user_id: str = None
) -> StreamingResponse:
    """
    Export des demandes de remplacement en Excel
    
    Args:
        db: Instance de la base de données
        tenant: Objet tenant
        current_user: Utilisateur courant
        user_id: Filtrer par utilisateur (optionnel)
    
    Returns:
        StreamingResponse avec le fichier Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        # Récupérer les demandes selon les permissions
        if user_id:
            demandes_list = await db.demandes_remplacement.find({
                "tenant_id": tenant.id,
                "demandeur_id": user_id
            }).to_list(length=None)
        else:
            if current_user.role == "employe":
                demandes_list = await db.demandes_remplacement.find({
                    "tenant_id": tenant.id,
                    "demandeur_id": current_user.id
                }).to_list(length=None)
            else:
                demandes_list = await db.demandes_remplacement.find({
                    "tenant_id": tenant.id
                }).to_list(length=None)
        
        # Récupérer les données de référence
        users_list = await db.users.find({"tenant_id": tenant.id}).to_list(length=None)
        types_garde_list = await db.types_garde.find({"tenant_id": tenant.id}).to_list(length=None)
        
        users_map = {u['id']: u for u in users_list}
        types_map = {t['id']: t for t in types_garde_list}
        
        # Créer le workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Remplacements"
        
        # Styles
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-têtes
        headers = ['Date', 'Type Garde', 'Demandeur', 'Statut', 'Priorité', 'Remplaçant', 'Raison', 'Créé le']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Labels de statut
        statut_labels = {
            'en_attente': 'En attente',
            'en_cours': 'En cours',
            'accepte': 'Accepté',
            'expiree': 'Expirée',
            'annulee': 'Annulée',
            'approuve_manuellement': 'Approuvé'
        }
        
        # Données
        for row, demande in enumerate(sorted(demandes_list, key=lambda x: x.get('date', ''), reverse=True), 2):
            demandeur = users_map.get(demande['demandeur_id'], {})
            type_garde = types_map.get(demande.get('type_garde_id', ''), {})
            remplacant = users_map.get(demande.get('remplacant_id', ''), {})
            
            data = [
                demande.get('date', ''),
                type_garde.get('nom', 'N/A'),
                f"{demandeur.get('prenom', '')} {demandeur.get('nom', '')}",
                statut_labels.get(demande.get('statut', ''), demande.get('statut', '')),
                'Urgent' if demande.get('priorite') == 'urgent' else 'Normal',
                f"{remplacant.get('prenom', '')} {remplacant.get('nom', '')}" if remplacant else "-",
                demande.get('raison', ''),
                str(demande.get('created_at', ''))[:10]
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
        
        # Ajuster la largeur des colonnes
        for column_cells in ws.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
        
        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"remplacements_{tenant.slug}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Erreur export Excel remplacements: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Erreur export Excel: {str(e)}")
